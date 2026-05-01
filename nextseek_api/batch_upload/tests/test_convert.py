"""Tests for convert.py: format detection, 4-sheet parsing, multi-file merge."""
from datetime import datetime
import json
import os
import tempfile

import openpyxl
import pytest
from pydantic import ValidationError as PydanticValidationError

from nextseek_api.batch_upload.convert import (
    convert_file,
    detect_format,
    merge_files,
    parse_traditional_file,
)
from nextseek_api.batch_upload.models import AssaySheetRow, InstructionRow


def _make_flat_xlsx(rows: list[dict]) -> str:
    """Create a flat-format xlsx (single sheet with UID, SampleType, json_metadata, assay_ids)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])
    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="test_flat_")
    os.close(fd)
    wb.save(path)
    return path


def _make_traditional_xlsx(
    instructions_rows: list[dict],
    samples_rows: list[dict],
    assay_rows: list[dict],
    ontology_rows: list[dict] = None,
) -> str:
    """Create a 4-sheet xlsx (Instructions, Samples, Assay, Ontology)."""
    wb = openpyxl.Workbook()
    wb.active.title = "Instructions"
    wi = wb["Instructions"]
    if instructions_rows:
        wi.append(list(instructions_rows[0].keys()))
        for row in instructions_rows:
            wi.append([row.get(k) for k in instructions_rows[0].keys()])

    wb.create_sheet("Samples")
    ws = wb["Samples"]
    if samples_rows:
        ws.append(list(samples_rows[0].keys()))
        for row in samples_rows:
            ws.append([row.get(k) for k in samples_rows[0].keys()])

    wb.create_sheet("Assay")
    wa = wb["Assay"]
    if assay_rows:
        wa.append(list(assay_rows[0].keys()))
        for row in assay_rows:
            wa.append([row.get(k) for k in assay_rows[0].keys()])

    wb.create_sheet("Ontology")
    wo = wb["Ontology"]
    if ontology_rows:
        wo.append(list(ontology_rows[0].keys()))
        for row in ontology_rows:
            wo.append([row.get(k) for k in ontology_rows[0].keys()])
    else:
        wo.append([])

    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="test_trad_")
    os.close(fd)
    wb.save(path)
    return path


class TestDetectFormat:
    def test_detect_format_traditional(self):
        path = _make_traditional_xlsx(
            instructions_rows=[{"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None}],
            samples_rows=[{"Name": "s1"}],
            assay_rows=[{"SampleType": "A.Sample", "AssayType": None, "Assay": 1, "Direction": 1}],
        )
        try:
            assert detect_format(path) == "traditional"
        finally:
            if os.path.isfile(path):
                os.unlink(path)

    def test_detect_format_flat(self):
        path = _make_flat_xlsx([
            {"uid": "U1", "SampleType": "A.Sample", "json_metadata": "{}", "assay_ids": "1"},
        ])
        try:
            assert detect_format(path) == "flat"
        finally:
            if os.path.isfile(path):
                os.unlink(path)


class TestBulkValidateInstructions:
    def test_bulk_validate_instructions(self):
        from nextseek_api.batch_upload.convert import _instruction_adapter

        rows = [
            {"field": "Name", "database_field": "A.Sample::Name", "field_type": "Text", "ontology": None},
            {"field": "Tissue", "database_field": "A.Sample::Tissue", "field_type": "Controlled Ontology", "ontology": "Tissue"},
        ]
        result = _instruction_adapter.validate_python(rows)
        assert len(result) == 2
        assert result[0].sample_type == "A.Sample"
        assert result[0].attribute_name == "Name"
        assert result[1].attribute_name == "Tissue"

    def test_invalid_instructions_delimiter(self):
        from nextseek_api.batch_upload.convert import _instruction_adapter

        rows = [{"field": "X", "database_field": "NoDelimiter", "field_type": "Text", "ontology": None}]
        with pytest.raises(PydanticValidationError):
            _instruction_adapter.validate_python(rows)


class TestBulkValidateAssayRows:
    def test_bulk_validate_assay_rows(self):
        from nextseek_api.batch_upload.convert import _assay_adapter

        rows = [
            {"SampleType": "A.Sample", "AssayType": None, "Assay": 1, "Direction": 1},
            {"SampleType": "A.Sample", "AssayType": "Seq", "Assay": 2, "Direction": 2},
        ]
        result = _assay_adapter.validate_python(rows)
        assert len(result) == 2
        assert result[0].assay_id == 1
        assert result[0].direction == 1
        assert result[1].assay_id == 2
        assert result[1].direction == 2


class TestConvertSamplesToInputRows:
    def test_convert_samples_to_input_rows(self):
        path = _make_traditional_xlsx(
            instructions_rows=[
                {"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None},
            ],
            samples_rows=[{"Name": "s1"}, {"Name": "s2"}],
            assay_rows=[{"SampleType": "A.Sample", "AssayType": None, "Assay": 42, "Direction": 1}],
        )
        try:
            batch = parse_traditional_file(path)
            assert len(batch.rows) == 2
            assert batch.rows[0].SampleType == "A.Sample"
            assert batch.rows[0].UID is None
            assert "Name" in batch.rows[0].json_metadata or '"Name"' in batch.rows[0].json_metadata
            assert batch.rows[0].assay_ids == [42]
            assert batch.source_files == [path]
        finally:
            if os.path.isfile(path):
                os.unlink(path)

    def test_traditional_uid_preserved_and_undeclared_columns_dropped(self):
        path = _make_traditional_xlsx(
            instructions_rows=[
                {"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None},
            ],
            samples_rows=[{"UID": "A.Sample-260331MIT-7", "Name": "s1", "Ignored": "drop-me"}],
            assay_rows=[{"SampleType": "A.Sample", "AssayType": None, "Assay": 42, "Direction": 1}],
        )
        try:
            batch = parse_traditional_file(path)
            assert len(batch.rows) == 1
            assert batch.rows[0].UID == "A.Sample-260331MIT-7"
            meta = json.loads(batch.rows[0].json_metadata)
            assert meta["UID"] == "A.Sample-260331MIT-7"
            assert meta["Name"] == "s1"
            assert "Ignored" not in meta
        finally:
            if os.path.isfile(path):
                os.unlink(path)

    def test_blank_uid_path_unchanged(self):
        path = _make_traditional_xlsx(
            instructions_rows=[
                {"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None},
            ],
            samples_rows=[{"UID": "", "Name": "s1"}],
            assay_rows=[{"SampleType": "A.Sample", "AssayType": None, "Assay": 42, "Direction": 1}],
        )
        try:
            batch = parse_traditional_file(path)
            assert len(batch.rows) == 1
            assert batch.rows[0].UID is None
            meta = json.loads(batch.rows[0].json_metadata)
            assert meta["Name"] == "s1"
            assert "UID" not in meta
        finally:
            if os.path.isfile(path):
                os.unlink(path)

    def test_rows_with_only_undeclared_fields_are_skipped(self):
        path = _make_traditional_xlsx(
            instructions_rows=[
                {"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None},
            ],
            samples_rows=[{"Ignored": "drop-me"}],
            assay_rows=[{"SampleType": "A.Sample", "AssayType": None, "Assay": 42, "Direction": 1}],
        )
        try:
            batch = parse_traditional_file(path)
            assert batch.rows == []
        finally:
            if os.path.isfile(path):
                os.unlink(path)

    def test_ontology_uses_attribute_name_when_field_differs(self):
        path = _make_traditional_xlsx(
            instructions_rows=[
                {
                    "Field": "Displayed Tissue",
                    "Database Field": "A.Sample::Tissue",
                    "Field Type": "Controlled Ontology",
                    "Ontology": "Tissue",
                },
                {"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None},
            ],
            samples_rows=[{"Name": "s1", "Displayed Tissue": "brain"}],
            assay_rows=[{"SampleType": "A.Sample", "AssayType": None, "Assay": 42, "Direction": 1}],
            ontology_rows=[{"Tissue": "Brain"}],
        )
        try:
            batch = parse_traditional_file(path)
            assert len(batch.rows) == 1
            meta = json.loads(batch.rows[0].json_metadata)
            assert meta["Displayed Tissue"] == "brain"
        finally:
            if os.path.isfile(path):
                os.unlink(path)

    def test_ontology_failure_survives_field_filtering(self):
        path = _make_traditional_xlsx(
            instructions_rows=[
                {
                    "Field": "Displayed Tissue",
                    "Database Field": "A.Sample::Tissue",
                    "Field Type": "Controlled Ontology",
                    "Ontology": "Tissue",
                },
                {"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None},
            ],
            samples_rows=[{"Name": "s1", "Displayed Tissue": "Lung"}],
            assay_rows=[{"SampleType": "A.Sample", "AssayType": None, "Assay": 42, "Direction": 1}],
            ontology_rows=[{"Tissue": "Brain"}],
        )
        try:
            batch = parse_traditional_file(path)
            assert batch.rows == []
            assert batch.ontology_result is not None
            assert batch.ontology_result.is_valid is False
            assert len(batch.ontology_result.violations) == 1
            violation = batch.ontology_result.violations[0]
            assert violation.attribute == "Tissue"
            assert violation.value == "Lung"
        finally:
            if os.path.isfile(path):
                os.unlink(path)

    def test_traditional_date_cell_is_json_serialized(self):
        path = _make_traditional_xlsx(
            instructions_rows=[
                {"Field": "Name", "Database Field": "NHP::Name", "Field Type": "Text", "Ontology": None},
                {"Field": "DateOfBirth", "Database Field": "NHP::DateOfBirth", "Field Type": "Text", "Ontology": None},
            ],
            samples_rows=[{"Name": "10225", "DateOfBirth": datetime(2020, 9, 27)}],
            assay_rows=[{"SampleType": "NHP", "AssayType": "Patient Visit", "Assay": 16, "Direction": 1}],
        )
        try:
            batch = parse_traditional_file(path)
            assert len(batch.rows) == 1
            meta = json.loads(batch.rows[0].json_metadata)
            assert meta["DateOfBirth"] == "2020-09-27"
        finally:
            if os.path.isfile(path):
                os.unlink(path)

    def test_trailing_non_instruction_row_is_ignored(self):
        path = _make_traditional_xlsx(
            instructions_rows=[
                {
                    "Field": "Name",
                    "Database Field": "NHP::Name",
                    "Field Type": "Text",
                    "Ontology": None,
                    "Required?": None,
                    "Notes / Examples": None,
                },
                {
                    "Field": None,
                    "Database Field": None,
                    "Field Type": None,
                    "Ontology": None,
                    "Required?": None,
                    "Notes / Examples": "Other data that you have is also encouraged to be added",
                },
            ],
            samples_rows=[{"Name": "10225"}],
            assay_rows=[{"SampleType": "NHP", "AssayType": "Patient Visit", "Assay": 16, "Direction": 1}],
        )
        try:
            batch = parse_traditional_file(path)
            assert len(batch.rows) == 1
            meta = json.loads(batch.rows[0].json_metadata)
            assert meta["Name"] == "10225"
        finally:
            if os.path.isfile(path):
                os.unlink(path)

    def test_multi_sample_type_rejected(self):
        """Multi-type files are now rejected with a warning."""
        path = _make_traditional_xlsx(
            instructions_rows=[
                {"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None},
                {"Field": "Name", "Database Field": "B.Sample::Name", "Field Type": "Text", "Ontology": None},
            ],
            samples_rows=[{"Name": "shared"}],
            assay_rows=[
                {"SampleType": "A.Sample", "AssayType": None, "Assay": 1, "Direction": 1},
                {"SampleType": "B.Sample", "AssayType": None, "Assay": 2, "Direction": 1},
            ],
        )
        try:
            batch = parse_traditional_file(path)
            assert len(batch.rows) == 0
            assert any("exactly one sample type" in w for w in batch.warnings)
        finally:
            if os.path.isfile(path):
                os.unlink(path)


class TestMultiFileMerge:
    def test_multi_file_merge_parallel(self):
        path1 = _make_traditional_xlsx(
            instructions_rows=[{"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None}],
            samples_rows=[{"Name": "a1"}],
            assay_rows=[{"SampleType": "A.Sample", "AssayType": None, "Assay": 1, "Direction": 1}],
        )
        path2 = _make_traditional_xlsx(
            instructions_rows=[{"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None}],
            samples_rows=[{"Name": "a2"}],
            assay_rows=[{"SampleType": "A.Sample", "AssayType": None, "Assay": 1, "Direction": 1}],
        )
        try:
            batch = merge_files([path1, path2])
            assert len(batch.rows) == 2
            assert len(batch.source_files) == 2
            names = set()
            for r in batch.rows:
                import json
                meta = json.loads(r.json_metadata)
                names.add(meta.get("Name", ""))
            assert "a1" in names and "a2" in names
        finally:
            for p in (path1, path2):
                if os.path.isfile(p):
                    os.unlink(p)

    def test_mixed_format_merge(self):
        flat_path = _make_flat_xlsx([
            {"uid": "U1", "SampleType": "A.Sample", "json_metadata": '{"Name":"flat1"}', "assay_ids": "1"},
        ])
        trad_path = _make_traditional_xlsx(
            instructions_rows=[{"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None}],
            samples_rows=[{"Name": "trad1"}],
            assay_rows=[{"SampleType": "A.Sample", "AssayType": None, "Assay": 1, "Direction": 1}],
        )
        try:
            batch = merge_files([flat_path, trad_path])
            assert len(batch.rows) >= 2
            assert len(batch.source_files) == 2
        finally:
            for p in (flat_path, trad_path):
                if os.path.isfile(p):
                    os.unlink(p)

    def test_merge_empty_paths(self):
        batch = merge_files([])
        assert len(batch.rows) == 0
        assert batch.source_files == []
        assert "No files" in (batch.warnings or [""])[0] or not batch.rows


class TestSingleTypeEnforcement:
    """Test single sample type enforcement in parse_traditional_file."""

    def test_single_type_enforcement(self):
        """More than one sample type in INSTRUCTIONS should reject."""
        path = _make_traditional_xlsx(
            instructions_rows=[
                {"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None},
                {"Field": "Name", "Database Field": "B.Sample::Name", "Field Type": "Text", "Ontology": None},
            ],
            samples_rows=[{"Name": "shared"}],
            assay_rows=[],
        )
        try:
            batch = parse_traditional_file(path)
            assert len(batch.rows) == 0
            assert any("exactly one sample type" in w for w in batch.warnings)
        finally:
            if os.path.isfile(path):
                os.unlink(path)

    def test_duplicate_field_rejection(self):
        """Duplicate Field headers in INSTRUCTIONS should reject."""
        path = _make_traditional_xlsx(
            instructions_rows=[
                {"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None},
                {"Field": "Name", "Database Field": "A.Sample::Alias", "Field Type": "Text", "Ontology": None},
            ],
            samples_rows=[{"Name": "s1"}],
            assay_rows=[],
        )
        try:
            batch = parse_traditional_file(path)
            assert len(batch.rows) == 0
            assert any("Duplicate Field header" in w for w in batch.warnings)
        finally:
            if os.path.isfile(path):
                os.unlink(path)

    def test_single_type_one_row_one_record(self):
        """Single type, multiple rows: each row becomes one InputRowModel."""
        path = _make_traditional_xlsx(
            instructions_rows=[
                {"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None},
                {"Field": "Tissue", "Database Field": "A.Sample::Tissue", "Field Type": "Text", "Ontology": None},
            ],
            samples_rows=[
                {"Name": "s1", "Tissue": "Brain"},
                {"Name": "s2", "Tissue": "Liver"},
            ],
            assay_rows=[{"SampleType": "A.Sample", "AssayType": None, "Assay": 1, "Direction": 1}],
        )
        try:
            batch = parse_traditional_file(path)
            assert len(batch.rows) == 2
            assert batch.rows[0].SampleType == "A.Sample"
            assert batch.rows[1].SampleType == "A.Sample"
            import json
            meta0 = json.loads(batch.rows[0].json_metadata)
            assert meta0["Name"] == "s1"
            assert meta0["Tissue"] == "Brain"
        finally:
            if os.path.isfile(path):
                os.unlink(path)

    def test_empty_row_skipped(self):
        """Rows with all empty values should be skipped."""
        path = _make_traditional_xlsx(
            instructions_rows=[
                {"Field": "Name", "Database Field": "A.Sample::Name", "Field Type": "Text", "Ontology": None},
            ],
            samples_rows=[
                {"Name": "s1"},
                {"Name": ""},  # empty row
                {"Name": "s3"},
            ],
            assay_rows=[{"SampleType": "A.Sample", "AssayType": None, "Assay": 1, "Direction": 1}],
        )
        try:
            batch = parse_traditional_file(path)
            assert len(batch.rows) == 2  # empty row skipped
        finally:
            if os.path.isfile(path):
                os.unlink(path)


class TestOriginalRowIndex:
    def test_traditional_parse_sets_original_row_index(self):
        path = _make_traditional_xlsx(
            instructions_rows=[
                {"Field": "Name", "Database Field": "A.Sample::Name",
                 "Field Type": "Text", "Ontology": None},
            ],
            samples_rows=[{"Name": "s1"}, {"Name": "s2"}],
            assay_rows=[{"SampleType": "A.Sample", "AssayType": None,
                         "Assay": 42, "Direction": 1}],
        )
        try:
            batch = parse_traditional_file(path)
            assert len(batch.rows) == 2
            assert batch.rows[0].original_row_index == 0
            assert batch.rows[1].original_row_index == 1
        finally:
            if os.path.isfile(path):
                os.unlink(path)

    def test_flat_parse_sets_original_row_index(self):
        path = _make_flat_xlsx([
            {"uid": "U1", "SampleType": "A.Sample",
             "json_metadata": '{"Name":"f1"}', "assay_ids": "1"},
            {"uid": "U2", "SampleType": "A.Sample",
             "json_metadata": '{"Name":"f2"}', "assay_ids": "2"},
        ])
        try:
            batch = convert_file(path)
            assert len(batch.rows) == 2
            assert batch.rows[0].original_row_index == 0
            assert batch.rows[1].original_row_index == 1
        finally:
            if os.path.isfile(path):
                os.unlink(path)
