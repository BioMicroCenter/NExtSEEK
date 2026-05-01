"""Format detection, 4-sheet conversion, and multi-file merge for batch upload."""
from __future__ import annotations

import datetime
import logging
import os
from concurrent.futures import ThreadPoolExecutor
from typing import Dict, List, Literal, Tuple

import polars as pl
from pydantic import TypeAdapter, ValidationError

from .models import (
    AssaySheetRow,
    ConvertedBatch,
    InputRowModel,
    InstructionRow,
)
from .ontology import load_ontology_from_sheet, validate_ontology_bulk

log = logging.getLogger(__name__)

try:
    import orjson

    def _json_dumps_min(obj):
        return orjson.dumps(obj).decode("utf-8")
except ImportError:
    import json

    def _json_dumps_min(obj):
        return json.dumps(obj, ensure_ascii=False, separators=(",", ":"))

# Module-level TypeAdapters (compiled once at import)
_instruction_adapter: TypeAdapter[List[InstructionRow]] = TypeAdapter(List[InstructionRow])
_assay_adapter: TypeAdapter[List[AssaySheetRow]] = TypeAdapter(List[AssaySheetRow])
_input_row_adapter: TypeAdapter[List[InputRowModel]] = TypeAdapter(List[InputRowModel])

# Traditional 4-sheet names (case-insensitive match)
_TRADITIONAL_SHEETS = {"INSTRUCTIONS", "SAMPLES", "ASSAY", "ONTOLOGY"}

# Instructions sheet: Excel column name -> model field
_INSTRUCTION_COL_MAP = {
    "field": "field",
    "database field": "database_field",
    "database_field": "database_field",
    "field type": "field_type",
    "field_type": "field_type",
    "ontology": "ontology",
}

# Assay sheet: Excel column name -> model field (for Pydantic alias)
_ASSAY_COL_MAP = {
    "sampletype": "SampleType",
    "sample type": "SampleType",
    "sample_type": "SampleType",
    "assaytype": "AssayType",
    "assay type": "AssayType",
    "assay": "Assay",
    "direction": "Direction",
}


def _get_sheet_names(xlsx_path: str) -> List[str]:
    """Read Excel sheet names without loading cell data (openpyxl read_only)."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, keep_vba=False)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def detect_format(xlsx_path: str) -> Literal["traditional", "flat"]:
    """Detect whether the workbook is traditional 4-sheet or flat format."""
    names = _get_sheet_names(xlsx_path)
    upper = {s.strip().upper() for s in names}
    if _TRADITIONAL_SHEETS <= upper:
        return "traditional"
    return "flat"


def _normalize_instruction_row(row: dict) -> dict:
    """Map Excel row to InstructionRow field names."""
    out = {}
    for k, v in row.items():
        key = (k or "").strip().lower()
        if key in _INSTRUCTION_COL_MAP:
            out[_INSTRUCTION_COL_MAP[key]] = v
    return out


def _has_instruction_mapping(row: dict) -> bool:
    """Return True when a normalized row contains real instruction mapping data."""
    return bool(str(row.get("field") or "").strip() or str(row.get("database_field") or "").strip())


def _normalize_assay_row(row: dict) -> dict:
    """Map Excel row to AssaySheetRow alias names."""
    out = {}
    for k, v in row.items():
        key = (k or "").strip().lower()
        if key in _ASSAY_COL_MAP:
            out[_ASSAY_COL_MAP[key]] = v
    return out


def _normalize_field_name(value: object) -> str:
    """Normalize traditional sheet headers with trim-only matching."""
    return str(value).strip() if value is not None else ""


def _coerce_traditional_cell_value(value: object) -> object:
    """Convert traditional Excel cell values into JSON-safe scalars."""
    if value is None:
        return None
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _build_instruction_field_maps(
    instructions: List[InstructionRow],
) -> Tuple[Dict[str, str], Dict[str, str]]:
    """Build trimmed Field lookups for metadata and ontology preparation."""
    field_name_by_normalized: Dict[str, str] = {}
    attr_name_by_normalized: Dict[str, str] = {}
    for ins in instructions:
        normalized = _normalize_field_name(ins.field)
        field_name_by_normalized[normalized] = normalized
        attr_name_by_normalized[normalized] = ins.attribute_name
    return field_name_by_normalized, attr_name_by_normalized


def _prepare_traditional_row(
    row_dict: dict,
    field_name_by_normalized: Dict[str, str],
    attr_name_by_normalized: Dict[str, str],
) -> Tuple[str | None, dict, dict]:
    """Extract row UID plus filtered metadata and ontology views."""
    row_uid: str | None = None
    metadata: dict = {}
    ontology_row: dict = {}

    for col, val in row_dict.items():
        if val is None:
            continue
        value_text = str(val).strip()
        if not value_text:
            continue

        normalized_col = _normalize_field_name(col)
        if not normalized_col:
            continue

        if normalized_col.lower() == "uid":
            row_uid = value_text
            metadata["UID"] = value_text
            continue

        canonical_field = field_name_by_normalized.get(normalized_col)
        if canonical_field is None:
            continue

        json_safe_value = _coerce_traditional_cell_value(val)
        metadata[canonical_field] = json_safe_value
        ontology_attr = attr_name_by_normalized.get(normalized_col)
        if ontology_attr:
            ontology_row[ontology_attr] = json_safe_value

    return row_uid, metadata, ontology_row


def _read_sheet(
    xlsx_path: str,
    sheet_name: str,
    raise_if_empty: bool = False,
) -> pl.DataFrame:
    """Read one sheet by name; use actual sheet name from workbook for case match."""
    return pl.read_excel(
        xlsx_path,
        sheet_name=sheet_name,
        engine="calamine",
        raise_if_empty=raise_if_empty,
    )


def parse_traditional_file(xlsx_path: str) -> ConvertedBatch:
    """Parse a traditional 4-sheet Excel file into a ConvertedBatch of InputRowModels.

    Contract:
    - Traditional files are sample-type-specific (exactly ONE sample type).
    - INSTRUCTIONS maps column headers to SampleType::Attribute.
    - If >1 sample type found in INSTRUCTIONS, the file is rejected.
    - If duplicate Field headers exist in INSTRUCTIONS, the file is rejected.
    - Non-empty SAMPLES columns whose headers match INSTRUCTIONS Field values
      become json_metadata key-value pairs.
    - A non-empty UID column is preserved on InputRowModel.UID and mirrored into
      json_metadata["UID"].
    - One SAMPLES row = one InputRowModel.
    """
    sheet_names = _get_sheet_names(xlsx_path)
    name_by_upper = {s.strip().upper(): s for s in sheet_names}

    def _sheet(s: str) -> str:
        return name_by_upper.get(s, s)

    # ── Phase 1: Read & validate INSTRUCTIONS only (cheap early rejection) ──
    df_ins = _read_sheet(xlsx_path, _sheet("INSTRUCTIONS"), raise_if_empty=True)

    ins_dicts = [
        normalized
        for d in df_ins.to_dicts()
        for normalized in [_normalize_instruction_row(d)]
        if _has_instruction_mapping(normalized)
    ]
    instructions = _instruction_adapter.validate_python(ins_dicts)

    if not instructions:
        return ConvertedBatch(
            rows=[],
            source_files=[xlsx_path],
            warnings=["No instructions found in INSTRUCTIONS sheet."],
        )

    # Concern 1: Enforce exactly ONE sample type per file (reject before reading other sheets)
    unique_types = {ins.sample_type for ins in instructions}
    if len(unique_types) != 1:
        return ConvertedBatch(
            rows=[],
            source_files=[xlsx_path],
            warnings=[
                f"Traditional files must contain exactly one sample type. "
                f"Found {len(unique_types)}: {', '.join(sorted(unique_types))}."
            ],
        )
    sample_type = unique_types.pop()

    # Concern 2: Reject duplicate Field headers (still before reading other sheets)
    seen_fields: set = set()
    for ins in instructions:
        normalized_field = _normalize_field_name(ins.field)
        if normalized_field in seen_fields:
            return ConvertedBatch(
                rows=[],
                source_files=[xlsx_path],
                warnings=[f"Duplicate Field header in INSTRUCTIONS: '{normalized_field}'."],
            )
        seen_fields.add(normalized_field)

    field_name_by_normalized, attr_name_by_normalized = _build_instruction_field_maps(instructions)

    # ── Phase 2: Read remaining sheets (only reached if INSTRUCTIONS is valid) ──
    df_samples = _read_sheet(xlsx_path, _sheet("SAMPLES"), raise_if_empty=False)
    df_assay = _read_sheet(xlsx_path, _sheet("ASSAY"), raise_if_empty=False)
    df_ont = _read_sheet(xlsx_path, _sheet("ONTOLOGY"), raise_if_empty=False)

    # Bulk-validate ASSAY
    assay_dicts = [_normalize_assay_row(d) for d in df_assay.to_dicts()]
    assay_rows = _assay_adapter.validate_python(assay_dicts)
    assay_ids: List[int] = []
    for row in assay_rows:
        if row.sample_type == sample_type:
            assay_ids.append(row.assay_id)

    # Ontology validation
    specs = load_ontology_from_sheet(instructions, df_ont)
    samples_dicts = df_samples.to_dicts()
    prepared_rows = [
        _prepare_traditional_row(row_dict, field_name_by_normalized, attr_name_by_normalized)
        for row_dict in samples_dicts
    ]
    if specs:
        row_dicts_for_ont = [ontology_row for _uid, _meta, ontology_row in prepared_rows]
        ont_result = validate_ontology_bulk(row_dicts_for_ont, specs)
        if not ont_result.is_valid:
            return ConvertedBatch(
                rows=[],
                source_files=[xlsx_path],
                ontology_result=ont_result,
                warnings=[f"Ontology validation failed: {len(ont_result.violations)} violation(s)."],
            )

    # Convert SAMPLES to InputRowModel list (one record per row)
    all_input_rows: List[dict] = []
    warnings: List[str] = []
    for idx, (row_uid, meta, _ontology_row) in enumerate(prepared_rows):
        if not meta:
            continue  # empty row

        all_input_rows.append({
            "UID": row_uid,
            "SampleType": sample_type,
            "json_metadata": _json_dumps_min(meta),
            "assay_ids": assay_ids,
            "original_row_index": idx,
        })

    try:
        validated = _input_row_adapter.validate_python(all_input_rows)
    except ValidationError as e:
        return ConvertedBatch(
            rows=[],
            source_files=[xlsx_path],
            warnings=[f"Input row validation failed: {e}"],
        )

    return ConvertedBatch(rows=validated, source_files=[xlsx_path], warnings=warnings)


def convert_file(xlsx_path: str) -> ConvertedBatch:
    """Detect format and convert one file to ConvertedBatch."""
    fmt = detect_format(xlsx_path)
    if fmt == "traditional":
        return parse_traditional_file(xlsx_path)
    # Flat: use existing stream_rows and collect into ConvertedBatch
    from .extract import stream_rows

    unknown_columns, stream = stream_rows(xlsx_path, limit=None)
    rows: List[InputRowModel] = []
    for sr in stream:
        if sr.data is not None:
            sr.data.original_row_index = sr.row_index
            rows.append(sr.data)
    return ConvertedBatch(
        rows=rows,
        source_files=[xlsx_path],
        warnings=[f"Unknown columns (ignored): {unknown_columns}"] if unknown_columns else [],
    )


def merge_files(xlsx_paths: List[str]) -> ConvertedBatch:
    """Parse multiple files (in parallel when > 1) and merge into one ConvertedBatch."""
    if not xlsx_paths:
        return ConvertedBatch(rows=[], source_files=[], warnings=["No files provided."])

    if len(xlsx_paths) == 1:
        return convert_file(xlsx_paths[0])

    max_workers = min(max(1, (os.cpu_count() or 2) - 1), len(xlsx_paths))
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        batches = list(pool.map(convert_file, xlsx_paths))

    all_rows: List[InputRowModel] = []
    all_sources: List[str] = []
    all_warnings: List[str] = []
    ontology_result = None

    for batch in batches:
        all_rows.extend(batch.rows)
        all_sources.extend(batch.source_files)
        all_warnings.extend(batch.warnings or [])
        if batch.ontology_result and not batch.ontology_result.is_valid:
            ontology_result = batch.ontology_result
            break

    return ConvertedBatch(
        rows=all_rows,
        source_files=all_sources,
        ontology_result=ontology_result,
        warnings=all_warnings,
    )
