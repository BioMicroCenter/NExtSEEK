"""The one workbook writer: README sheet first, then a sheet per sample type."""

from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from nextseek_api.services.sample_workbook import (
    COLUMN_TABLE_HEADER,
    CONTEXTDB_URL,
    DROPDOWN_SPARE_ROWS,
    EXCEL_MAX_CELL_CHARS,
    CV_SHEET,
    FLOW_MAX_WIDTH,
    FLOW_README_POINTER,
    FLOW_SHEET,
    SUMMARY_HEADER,
    build_readme_blocks,
    load_assay_titles,
    load_derivation_hops,
    load_sample_field_context,
    load_sample_type_context,
    write_samples_workbook,
)

_MOD = "nextseek_api.services.sample_workbook"


@pytest.fixture(autouse=True)
def _no_graph():
    """write_samples_workbook consults Neo4j for lineage. Tests must not reach
    for a real graph: the driver blocks rather than failing fast, which is what
    NEO4J_TIMEOUT_SECONDS bounds in production."""
    with patch(f"{_MOD}.load_derivation_hops", return_value=[]):
        yield

CONTEXT = {
    "MUS": {"name": "Mouse", "description": "A mouse sample."},
    "TIS": {"name": "Tissue", "description": "A tissue sample."},
}


@patch(f"{_MOD}.Sample_types_context")
def test_load_context_maps_code_to_name_and_description(mock_model):
    mock_model.objects.filter.return_value.values.return_value = [
        {"sample_type": "MUS", "name": "Mouse", "description": "A mouse sample."},
    ]
    assert load_sample_type_context(["MUS"]) == {
        "MUS": {"name": "Mouse", "description": "A mouse sample."}
    }


@patch(f"{_MOD}.Sample_types_context")
def test_load_context_coerces_nulls_to_empty_strings(mock_model):
    mock_model.objects.filter.return_value.values.return_value = [
        {"sample_type": "MUS", "name": None, "description": None},
    ]
    assert load_sample_type_context(["MUS"]) == {"MUS": {"name": "", "description": ""}}


@patch(f"{_MOD}.Sample_types_context")
def test_load_context_does_not_query_for_an_empty_code_list(mock_model):
    assert load_sample_type_context([]) == {}
    mock_model.objects.filter.assert_not_called()


@patch(f"{_MOD}.Sample_types_context")
def test_load_context_survives_a_missing_table(mock_model):
    """A download must not fail because the context table is absent."""
    mock_model.objects.filter.side_effect = RuntimeError("no such table")
    assert load_sample_type_context(["MUS"]) == {}


def _df():
    return pd.DataFrame([
        {"uuid": "MUS-230101ABC-1", "Name": "m1", "Sex": "F"},
        {"uuid": "TIS-230101ABC-2", "Name": "t1", "Sex": ""},
    ])


@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_readme_is_the_first_sheet(_ctx, tmp_path):
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    assert load_workbook(out).sheetnames[0] == "README"


@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_sample_type_sheets_follow_the_readme(_ctx, tmp_path):
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    assert load_workbook(out).sheetnames == ["README", "MUS", "TIS"]


@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_a1_links_to_the_contextdb(_ctx, tmp_path):
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    ws = load_workbook(out)["README"]
    assert ws["A1"].hyperlink.target == CONTEXTDB_URL


@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_helper_columns_are_dropped_from_data_sheets(_ctx, tmp_path):
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    ws = load_workbook(out)["MUS"]
    headers = [c.value for c in ws[1]]
    assert "uuid" not in headers and "sample_type" not in headers


@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_all_empty_columns_are_dropped(_ctx, tmp_path):
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    assert "Sex" not in [c.value for c in load_workbook(out)["TIS"][1]]


@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value={})
def test_workbook_still_written_when_context_table_is_empty(_ctx, _fields, tmp_path):
    """With no sample-type context the heading is the bare code and the
    description row is empty, but the section is still there."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    ws = load_workbook(out)["README"]
    assert ws["A4"].value == "MUS"          # summary table row
    assert ws["A7"].value == "MUS"          # section heading, bare code


@patch(f"{_MOD}.load_sample_type_context")
def test_supplied_context_skips_the_lookup(mock_load, tmp_path):
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out), context_by_code=CONTEXT)
    mock_load.assert_not_called()


@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_dotted_sample_type_codes_survive_extraction(_ctx, tmp_path):
    """D.SEQ and friends must not be truncated to D."""
    out = tmp_path / "w.xlsx"
    df = pd.DataFrame([{"uuid": "D.SEQ-240910LAU-3", "Name": "s1"}])
    write_samples_workbook(df, str(out))
    assert "D.SEQ" in load_workbook(out).sheetnames


def test_dbtable_sample_retrieval_data_delegates_to_the_shared_writer():
    from seek.dbtable_sample import DBtable_sample

    # __init__ opens a DB connection (dmac/dbtable.py:51) and sampleRetrievalData
    # needs none of it, so skip construction rather than pulling in django_db.
    dbs = DBtable_sample.__new__(DBtable_sample)

    df = pd.DataFrame([{"uuid": "MUS-1", "json_metadata": '{"Name": "m1"}'}])
    with patch("seek.dbtable_sample.write_samples_workbook") as mock_write:
        dbs.sampleRetrievalData(df, "/tmp/unused.xlsx")
    mock_write.assert_called_once()
    assert mock_write.call_args[0][1] == "/tmp/unused.xlsx"


def test_seek_views_sample_retrieval_data_delegates_to_the_shared_writer():
    from seek import views

    df = pd.DataFrame([{"uuid": "MUS-1", "json_metadata": '{"Name": "m1"}'}])
    with patch("seek.views.write_samples_workbook") as mock_write:
        views.sample_retrieval_data(df, "/tmp/unused.xlsx")
    mock_write.assert_called_once()
    assert mock_write.call_args[0][1] == "/tmp/unused.xlsx"


@patch(f"{_MOD}.Sample_attributes_unique")
def test_field_context_uses_the_global_row(mock_model):
    mock_model.objects.filter.return_value.values.return_value = [
        {"field_name": "Sex", "sample_type": "", "meaning": "Sex at birth."},
    ]
    assert load_sample_field_context([("MUS", "Sex")]) == {("MUS", "Sex"): "Sex at birth."}


@patch(f"{_MOD}.Sample_attributes_unique")
def test_field_context_prefers_a_sample_type_override(mock_model):
    mock_model.objects.filter.return_value.values.return_value = [
        {"field_name": "Name", "sample_type": "", "meaning": "Submitter's identifier."},
        {"field_name": "Name", "sample_type": "MUS", "meaning": "The animal's ear-tag ID."},
    ]
    assert load_sample_field_context([("MUS", "Name")]) == {
        ("MUS", "Name"): "The animal's ear-tag ID."
    }


@patch(f"{_MOD}.Sample_attributes_unique")
def test_field_context_override_does_not_leak_to_other_sample_types(mock_model):
    mock_model.objects.filter.return_value.values.return_value = [
        {"field_name": "Name", "sample_type": "", "meaning": "Submitter's identifier."},
        {"field_name": "Name", "sample_type": "MUS", "meaning": "The animal's ear-tag ID."},
    ]
    result = load_sample_field_context([("MUS", "Name"), ("TIS", "Name")])
    assert result[("TIS", "Name")] == "Submitter's identifier."


@patch(f"{_MOD}.Sample_attributes_unique")
def test_field_context_returns_blank_for_an_undefined_field(mock_model):
    mock_model.objects.filter.return_value.values.return_value = []
    assert load_sample_field_context([("MUS", "Genotype")]) == {("MUS", "Genotype"): ""}


@patch(f"{_MOD}.Sample_attributes_unique")
def test_field_context_coerces_a_null_meaning_to_a_blank(mock_model):
    mock_model.objects.filter.return_value.values.return_value = [
        {"field_name": "Sex", "sample_type": "", "meaning": None},
    ]
    assert load_sample_field_context([("MUS", "Sex")]) == {("MUS", "Sex"): ""}


@patch(f"{_MOD}.Sample_attributes_unique")
def test_field_context_does_not_query_for_an_empty_pair_list(mock_model):
    assert load_sample_field_context([]) == {}
    mock_model.objects.filter.assert_not_called()


@patch(f"{_MOD}.Sample_attributes_unique")
def test_field_context_survives_a_missing_table(mock_model):
    """A download must not fail because the definitions table is absent."""
    mock_model.objects.filter.side_effect = RuntimeError("no such table")
    assert load_sample_field_context([("MUS", "Sex")]) == {}


def test_blocks_carry_the_sample_type_name_and_description():
    blocks = build_readme_blocks([("MUS", ["UID"])], CONTEXT, {})
    assert blocks[0]["code"] == "MUS"
    assert blocks[0]["name"] == "Mouse"
    assert blocks[0]["description"] == "A mouse sample."


def test_blocks_keep_sheet_order_not_alphabetical_order():
    """The README is meant to be read beside the tab, left to right."""
    blocks = build_readme_blocks([("MUS", ["UID", "Sex", "Genotype"])], CONTEXT, {})
    assert [name for name, _ in blocks[0]["columns"]] == ["UID", "Sex", "Genotype"]


def test_blocks_follow_the_order_the_sheets_were_given():
    blocks = build_readme_blocks([("TIS", ["UID"]), ("MUS", ["UID"])], CONTEXT, {})
    assert [b["code"] for b in blocks] == ["TIS", "MUS"]


def test_blocks_attach_the_resolved_meaning():
    meanings = {("MUS", "Sex"): "Sex at birth."}
    blocks = build_readme_blocks([("MUS", ["Sex"])], CONTEXT, meanings)
    assert blocks[0]["columns"] == [("Sex", "Sex at birth.")]


def test_a_column_with_no_definition_is_listed_with_a_blank():
    """The README always indexes every column, so a gap is visible not silent."""
    blocks = build_readme_blocks([("MUS", ["Genotype"])], CONTEXT, {})
    assert blocks[0]["columns"] == [("Genotype", "")]


def test_an_undocumented_sample_type_still_gets_a_block():
    blocks = build_readme_blocks([("ZZZ", ["UID"])], CONTEXT, {})
    assert blocks[0] == {"code": "ZZZ", "name": "", "description": "", "columns": [("UID", "")]}


@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_readme_section_heading_is_bold_at_a7(_ctx, _fields, tmp_path):
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    ws = load_workbook(out)["README"]
    assert ws["A7"].value == "MUS — Mouse"
    assert ws["A7"].font.bold is True


@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_readme_description_sits_in_the_summary_table(_ctx, _fields, tmp_path):
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    ws = load_workbook(out)["README"]
    assert ws["C4"].value == "A mouse sample."


@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_readme_column_table_is_indented_into_b_and_c(_ctx, _fields, tmp_path):
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    ws = load_workbook(out)["README"]
    assert [ws["B9"].value, ws["C9"].value] == ["Column", "Meaning"]
    assert ws["B9"].font.bold is True
    assert ws["B10"].value == "Name"


@patch(f"{_MOD}.load_sample_field_context",
       return_value={("MUS", "Name"): "The animal's ear-tag ID."})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_readme_shows_the_resolved_meaning(_ctx, _fields, tmp_path):
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    ws = load_workbook(out)["README"]
    assert ws["C10"].value == "The animal's ear-tag ID."


@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_the_second_section_starts_after_a_blank_row(_ctx, _fields, tmp_path):
    """MUS has two columns (Name, Sex) at rows 7-8, so TIS heads row 10."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    ws = load_workbook(out)["README"]
    assert ws["A12"].value is None
    assert ws["A13"].value == "TIS — Tissue"


@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_readme_only_lists_columns_that_survive_the_empty_drop(_ctx, _fields, tmp_path):
    """TIS's Sex is empty in the fixture, so the TIS sheet drops it and the
    README must not claim a column the researcher cannot see.

    TIS heads row 10, description 11, blank 12, table header 13, so its single
    surviving column sits at B14."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    ws = load_workbook(out)["README"]
    assert ws["B16"].value == "Name"
    assert ws["B17"].value is None


@patch(f"{_MOD}.Sample_attributes_unique")
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_workbook_is_complete_when_the_definitions_table_is_missing(_ctx, mock_model, tmp_path):
    """Losing sample_attributes_unique costs meanings, never the download.

    This is the only Task 5 test that exercises the real loader rather than
    mocking it out, so it is what actually proves the fail-soft path end to end.
    """
    mock_model.objects.filter.side_effect = RuntimeError("no such table")
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    wb = load_workbook(out)
    assert wb.sheetnames == ["README", "MUS", "TIS"]
    ws = wb["README"]
    assert ws["B10"].value == "Name"         # still indexed
    assert ws["C10"].value in (None, "")     # meaning blank


@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_field_lookup_is_asked_only_for_columns_actually_written(_ctx, tmp_path):
    out = tmp_path / "w.xlsx"
    with patch(f"{_MOD}.load_sample_field_context", return_value={}) as lookup:
        write_samples_workbook(_df(), str(out))
    asked = set(lookup.call_args[0][0])
    assert ("TIS", "Sex") not in asked
    assert ("MUS", "Sex") in asked


@patch(f"{_MOD}.load_sample_field_context",
       return_value={("MUS", "Name"): "Line one\x0bline two."})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_a_control_character_in_a_definition_does_not_break_the_download(
    _ctx, _fields, tmp_path
):
    """openpyxl raises IllegalCharacterError on \\x0b — which is exactly what a
    line break pasted out of Word or a PDF becomes. One bad definition row must
    not cost every download of every workbook carrying that sample type."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))          # must not raise
    ws = load_workbook(out)["README"]
    assert ws["C10"].value == "Line oneline two."


@patch(f"{_MOD}.load_sample_type_context",
       return_value={"MUS": {"name": "Mouse", "description": "Desc\x07ription."},
                     "TIS": {"name": "Tissue", "description": "A tissue sample."}})
@patch(f"{_MOD}.load_sample_field_context", return_value={})
def test_a_control_character_in_a_description_is_stripped(_fields, _ctx, tmp_path):
    """Every string the README writes goes through the same sanitizer, not just
    the meanings."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    assert load_workbook(out)["README"]["C4"].value == "Description."


@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_an_over_length_definition_is_truncated_not_written_whole(_ctx, tmp_path):
    """Excel's cell limit is 32,767 characters and openpyxl does not enforce it;
    a longer value writes a file Excel reports as needing repair. `meaning` is a
    TEXT column, so a reviewer could paste well past the limit."""
    out = tmp_path / "w.xlsx"
    with patch(f"{_MOD}.load_sample_field_context",
               return_value={("MUS", "Name"): "x" * 40000}):
        write_samples_workbook(_df(), str(out))
    assert len(load_workbook(out)["README"]["C10"].value) == EXCEL_MAX_CELL_CHARS


@patch(f"{_MOD}.load_sample_field_context",
       return_value={("MUS", "Name"): "The animal's ear-tag ID."})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_the_header_cell_carries_its_definition_as_a_hover_note(_ctx, _fields, tmp_path):
    """A researcher filling the sheet in reads the header, not the README."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    header = load_workbook(out)["MUS"]["A1"]
    assert header.value == "Name"
    assert header.comment is not None
    assert "ear-tag ID" in header.comment.text


@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_an_undefined_column_gets_no_empty_hover_note(_ctx, _fields, tmp_path):
    """An empty note is worse than none: it looks like a definition that failed
    to load rather than one nobody has written yet."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    assert load_workbook(out)["MUS"]["A1"].comment is None


@patch(f"{_MOD}.Sample_attributes_unique")
def test_field_context_keeps_case_variant_names_apart(mock_model):
    """SEEK has five attribute pairs differing only by case (Figure/figure,
    Bead_Coating/Bead_coating). The table stores field_name as utf8mb4_bin so
    they stay distinct; the loader must not merge them either."""
    mock_model.objects.filter.return_value.values.return_value = [
        {"field_name": "Figure", "sample_type": "", "meaning": "Upper-case one."},
        {"field_name": "figure", "sample_type": "", "meaning": "Lower-case one."},
    ]
    result = load_sample_field_context([("D.IMG", "Figure"), ("A.IMG", "figure")])
    assert result[("D.IMG", "Figure")] == "Upper-case one."
    assert result[("A.IMG", "figure")] == "Lower-case one."


def _prov_df():
    """Two hops, one of them from a type not itself downloaded."""
    return pd.DataFrame([
        {"uuid": "D.SEQ-1", "sample_type": "D.SEQ", "Parent": "DNA-9"},
        {"uuid": "TIS-2", "sample_type": "TIS", "Parent": "MUS-3; MUS-4"},
    ])


@patch(f"{_MOD}.GraphDatabase")
def test_lineage_lookup_survives_an_unreachable_graph(mock_graph):
    mock_graph.driver.side_effect = RuntimeError("connection refused")
    assert load_derivation_hops(["D.SEQ-1"]) == []


@patch(f"{_MOD}.Samples")
def test_assay_lookup_survives_a_database_failure(mock_samples):
    """Losing the assay link must cost the labels, never the download."""
    mock_samples.objects.filter.side_effect = RuntimeError("no such table")
    assert load_assay_titles(["D.SEQ-1"]) == {}


def _cv_df():
    """A frame whose columns include two governed by GEO/SRA vocabularies."""
    return pd.DataFrame([
        {"uuid": "MUS-230101ABC-1", "Name": "m1", "DataType": "FASTQ", "LibraryDesign": "paired"},
    ])


@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_a_governed_column_offers_its_vocabulary_as_a_dropdown(_ctx, _fields, tmp_path):
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_cv_df(), str(out))
    wb = load_workbook(out)
    assert CV_SHEET in wb.sheetnames
    rules = wb["MUS"].data_validations.dataValidation
    assert {r.formula1.split("!")[0].strip("'") for r in rules} == {CV_SHEET}
    assert len(rules) == 2  # DataType and LibraryDesign, not Name or uuid


@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_dropdowns_warn_rather_than_reject(_ctx, _fields, tmp_path):
    """Downloaded data predates these vocabularies (RNA-seq for RNA-Seq), so a
    hard reject would fire on open for rows the researcher never touched."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_cv_df(), str(out))
    for rule in load_workbook(out)["MUS"].data_validations.dataValidation:
        assert rule.errorStyle == "warning"


@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_no_vocabulary_sheet_when_no_column_is_governed(_ctx, _fields, tmp_path):
    """The fixture is Name and Sex only; an empty extra sheet would be clutter."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    assert CV_SHEET not in load_workbook(out).sheetnames


def _readme_sections(ws) -> dict[str, list[str]]:
    """Parse the README back into {heading: [column names]}, without assuming
    any row arithmetic. Used to state layout invariants against the sheet
    itself rather than against hand-computed coordinates."""
    sections: dict[str, list[str]] = {}
    current = None
    in_summary = True
    for row in range(1, ws.max_row + 1):
        a, b = ws.cell(row=row, column=1), ws.cell(row=row, column=2)
        # Section headings are the bold cells of column A. The description
        # under each one is column A too, but plain; A1's link is not bold.
        if a.value and a.font.bold:
            # The summary table's bold header is not a section; skip it and the
            # sample-type rows beneath it.
            if a.value == SUMMARY_HEADER[0]:
                in_summary = True
                continue
            in_summary = False
            current = a.value.split(" — ")[0]
            sections[current] = []
        elif in_summary:
            continue
        # Column names are the plain cells of column B. The Column/Meaning
        # table header sits in B as well, but bold.
        elif b.value and not b.font.bold and current:
            assert b.value != COLUMN_TABLE_HEADER[0]
            sections[current].append(b.value)
    return sections


@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_readme_columns_match_the_sheet_header_they_describe(_ctx, _fields, tmp_path):
    """The invariant stated directly: for every tab, what the README lists is
    exactly that tab's header row after the all-empty-column drop. TIS's Sex is
    empty in the fixture, so neither the sheet nor the README carries it."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    wb = load_workbook(out)
    sections = _readme_sections(wb["README"])
    for code in ("MUS", "TIS"):
        assert sections[code] == [c.value for c in wb[code][1]]


@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_a_section_with_no_surviving_columns_has_no_table_header(_ctx, _fields, tmp_path):
    """Every TIS column is empty here, so the tab has no columns at all. A bare
    Column/Meaning header with nothing beneath it reads as a rendering bug."""
    df = pd.DataFrame([
        {"uuid": "MUS-230101ABC-1", "Name": "m1", "Sex": "F"},
        {"uuid": "TIS-230101ABC-2", "Name": "", "Sex": ""},
    ])
    out = tmp_path / "w.xlsx"
    write_samples_workbook(df, str(out))
    ws = load_workbook(out)["README"]
    assert ws["A13"].value == "TIS — Tissue"
    assert [ws.cell(row=r, column=2).value for r in (14, 15, 16)] == [None, None, None]


def _flow_df():
    """PAT -> PAV -> TIS, with TIS also downloaded."""
    return pd.DataFrame([
        {"uuid": "PAV-1", "sample_type": "PAV", "Parent": "PAT-9"},
        {"uuid": "TIS-2", "sample_type": "TIS", "Parent": "PAV-1"},
    ])


@patch(f"{_MOD}.load_assay_titles", return_value={})
@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_the_flow_sheet_sits_directly_after_the_readme(_ctx, _fields, _assays, tmp_path):
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_flow_df(), str(out))
    assert load_workbook(out).sheetnames[:2] == ["README", FLOW_SHEET]


@patch(f"{_MOD}.load_assay_titles", return_value={})
@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value={})
def test_the_tree_puts_a_root_above_its_indented_child(_ctx, _fields, _assays, tmp_path):
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_flow_df(), str(out))
    ws = load_workbook(out)[FLOW_SHEET]
    column_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "PAT" in column_a
    assert any(str(v).startswith(("├── ", "└── ")) for v in column_a if v)


@patch(f"{_MOD}.load_assay_titles", return_value={})
@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_no_flow_sheet_when_there_is_no_lineage(_ctx, _fields, _assays, tmp_path):
    """An empty sheet reads as a rendering bug -- and so does a README that
    points at a sheet the workbook does not contain. The pointer's absence was
    caught only incidentally, by an unrelated cell-position assertion; state it
    here, where the sheet's absence is the subject."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_df(), str(out))
    wb = load_workbook(out)
    assert FLOW_SHEET not in wb.sheetnames
    text = " ".join(str(c.value) for row in wb["README"].iter_rows()
                    for c in row if c.value)
    assert FLOW_README_POINTER not in text


@patch(f"{_MOD}.load_assay_titles", return_value={})
@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_the_readme_points_at_the_flow_sheet(_ctx, _fields, _assays, tmp_path):
    """A reader who never opens the tab must still learn it is there."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_flow_df(), str(out))
    text = " ".join(str(c.value) for row in load_workbook(out)["README"].iter_rows()
                    for c in row if c.value)
    assert FLOW_SHEET in text


@patch(f"{_MOD}.load_assay_titles", return_value={})
@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value={})
def test_the_flow_sheet_sizes_its_one_column_to_the_widest_line(_ctx, _fields, _assays, tmp_path):
    """The justification for a separate sheet is that it can size its own
    column. Nothing else asserts the width is applied, so it could be dropped
    silently and the tree would render clipped."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_flow_df(), str(out))
    ws = load_workbook(out)[FLOW_SHEET]
    widest = max(len(str(ws.cell(row=r, column=1).value or ""))
                 for r in range(1, ws.max_row + 1))
    assert ws.column_dimensions["A"].width == min(widest + 2, FLOW_MAX_WIDTH)


def _order_df():
    """Alphabetical order is DNA, PAT, PAV, TIS. Generation order is the
    reverse of that for PAT/PAV and puts DNA last."""
    return pd.DataFrame([
        {"uuid": "TIS-1", "sample_type": "TIS", "Parent": "PAV-1"},
        {"uuid": "DNA-1", "sample_type": "DNA", "Parent": "TIS-1"},
        {"uuid": "PAT-1", "sample_type": "PAT", "Parent": ""},
        {"uuid": "PAV-1", "sample_type": "PAV", "Parent": "PAT-1"},
    ])


@patch(f"{_MOD}.load_assay_titles", return_value={})
@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value={})
def test_sample_tabs_follow_generation_order(_ctx, _fields, _assays, tmp_path):
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_order_df(), str(out))
    tabs = [n for n in load_workbook(out).sheetnames
            if n not in ("README", FLOW_SHEET, CV_SHEET)]
    assert tabs == ["PAT", "PAV", "TIS", "DNA"]


@patch(f"{_MOD}.load_assay_titles", return_value={})
@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value={})
def test_the_readme_summary_follows_the_same_order(_ctx, _fields, _assays, tmp_path):
    """A workbook whose tabs and README disagree reads as a bug."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_order_df(), str(out))
    ws = load_workbook(out)["README"]
    codes = [ws.cell(row=r, column=1).value for r in range(4, 8)]
    assert codes == ["PAT", "PAV", "TIS", "DNA"]


@patch(f"{_MOD}.load_assay_titles", return_value={})
@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value={})
def test_a_type_with_no_lineage_sorts_last(_ctx, _fields, _assays, tmp_path):
    """ABC has no hop at all. It cannot be placed in the pipeline, so it goes
    after everything that can be."""
    df = pd.DataFrame([
        {"uuid": "TIS-1", "sample_type": "TIS", "Parent": "PAV-1"},
        {"uuid": "PAV-1", "sample_type": "PAV", "Parent": ""},
        {"uuid": "ABC-1", "sample_type": "ABC", "Parent": ""},
    ])
    out = tmp_path / "w.xlsx"
    write_samples_workbook(df, str(out))
    tabs = [n for n in load_workbook(out).sheetnames
            if n not in ("README", FLOW_SHEET, CV_SHEET)]
    assert tabs == ["PAV", "TIS", "ABC"]


@patch(f"{_MOD}.load_assay_titles", return_value={})
@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_a_malformed_uid_does_not_cost_the_whole_download(_ctx, _fields, _assays, tmp_path):
    """A UID carrying no [A-Z] run makes str.extract yield NaN, a float. It used
    to reach derivation_edges' falsy guard -- which NaN passes -- seed an edge
    keyed on a float, and blow up in sorted() as a 500 on the download. One
    unparseable row must cost that row's tab, never the workbook."""
    df = pd.DataFrame([
        {"uuid": "123-456-7", "Name": "junk", "Parent": "MUS-3"},
        {"uuid": "TIS-230101ABC-2", "Name": "t1", "Parent": "MUS-3"},
    ])
    out = tmp_path / "w.xlsx"
    write_samples_workbook(df, str(out))          # must not raise
    wb = load_workbook(out)
    assert "README" in wb.sheetnames
    assert "TIS" in wb.sheetnames


@patch(f"{_MOD}.load_assay_titles", return_value={})
@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value={})
def test_order_stays_alphabetical_without_any_lineage(_ctx, _fields, _assays, tmp_path):
    """No graph and no usable Parent column: fall back to what it did before."""
    df = pd.DataFrame([
        {"uuid": "TIS-1", "sample_type": "TIS", "Name": "t"},
        {"uuid": "DNA-1", "sample_type": "DNA", "Name": "d"},
    ])
    out = tmp_path / "w.xlsx"
    write_samples_workbook(df, str(out))
    tabs = [n for n in load_workbook(out).sheetnames
            if n not in ("README", FLOW_SHEET, CV_SHEET)]
    assert tabs == ["DNA", "TIS"]


@patch(f"{_MOD}.load_sample_field_context", return_value={})
@patch(f"{_MOD}.load_sample_type_context", return_value=CONTEXT)
def test_dropdowns_reach_past_the_filled_rows(_ctx, _fields, tmp_path):
    """A researcher adding samples must keep the dropdown, not lose it at the
    first empty row."""
    out = tmp_path / "w.xlsx"
    write_samples_workbook(_cv_df(), str(out))
    for rule in load_workbook(out)["MUS"].data_validations.dataValidation:
        last = int(str(rule.sqref).split(":")[1].lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
        assert last == 2 + DROPDOWN_SPARE_ROWS  # one data row, plus the spare


def test_a_dropdown_range_never_starts_below_its_end():
    """A frame with no data rows must not produce A2:A1, which Excel rejects.
    Called directly: write_samples_workbook always writes at least one row, so
    the guard is unreachable through the public API."""
    from openpyxl import Workbook

    from nextseek_api.services.sample_workbook import _apply_dropdowns

    ws = Workbook().active
    _apply_dropdowns(ws, ["DataType"], {"DataType": "filetype"},
                     {"filetype": "'Controlled Vocabularies'!$A$2:$A$9"}, 0)
    rule, = ws.data_validations.dataValidation
    start, end = (int(part.lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
                  for part in str(rule.sqref).split(":"))
    assert start == 2
    assert end == 2 + DROPDOWN_SPARE_ROWS


def _vocabularies():
    from nextseek_api.services.sample_workbook import _load_vocabularies
    return _load_vocabularies()


def test_every_governed_column_resolves_to_a_real_vocabulary():
    field_map, vocabularies = _vocabularies()
    assert field_map, "the file must load; a parse error costs every dropdown"
    assert all(name in vocabularies for name in field_map.values())


def test_the_house_layout_terms_are_the_ones_production_uses():
    """Production holds Paired End 3511 times and bare 'paired' 52 times. The
    dropdown used to offer only 'paired'."""
    _, vocabularies = _vocabularies()
    assert vocabularies["library_layout"] == ["Paired End", "Single End"]


def test_imaging_formats_are_offered():
    """GEO's filetype list has no term for any of these; production has
    thousands of rows of them."""
    _, vocabularies = _vocabularies()
    for term in ("TIF", "OIB", "CZI", "DICOM", "LIF", "ND2"):
        assert term in vocabularies["filetype"]


def test_illumina_instruments_are_always_prefixed():
    """GEO is inconsistent -- 'Illumina MiSeq' but bare 'NextSeq 500' -- and
    production split on exactly that seam. Offering both forms is what caused
    the split, so the bare ones are gone."""
    _, vocabularies = _vocabularies()
    models = vocabularies["instrument_model"]
    for bare in ("NextSeq 500", "NextSeq 550", "NextSeq 1000", "NextSeq 2000"):
        assert bare not in models
        assert f"Illumina {bare}" in models


def test_nextseek_only_instruments_are_present():
    _, vocabularies = _vocabularies()
    assert "Singular G4" in vocabularies["instrument_model"]
    assert "PromethION P2 Solo" in vocabularies["instrument_model"]


def test_no_vocabulary_offers_the_same_term_twice():
    """A duplicate renders as two identical dropdown entries."""
    _, vocabularies = _vocabularies()
    for name, terms in vocabularies.items():
        assert len(terms) == len(set(terms)), name


def test_no_term_would_be_mangled_on_its_way_into_a_cell():
    from nextseek_api.services.sample_workbook import _safe_cell_value
    _, vocabularies = _vocabularies()
    for terms in vocabularies.values():
        for term in terms:
            assert _safe_cell_value(term) == term


def test_the_variants_block_documents_only_real_vocabularies():
    """It is documentation, but documentation that drifts is worse than none."""
    import json
    from nextseek_api.services.sample_workbook import CV_PATH
    doc = json.loads(CV_PATH.read_text())
    documented = set(doc["variants"]) - {"_excluded"}
    assert documented <= set(doc["vocabularies"])
    assert set(doc["variants"]["_excluded"]) <= set(doc["vocabularies"])


def test_the_flow_sheet_writes_one_monospace_column():
    from openpyxl import Workbook
    from nextseek_api.services.sample_workbook import _write_flow_sheet, FLOW_SHEET

    book = Workbook()
    _write_flow_sheet(book, ["PAT", "└── PAV   [Consent]"])
    ws = book[FLOW_SHEET]
    assert ws["A1"].value == "PAT"
    assert ws["A2"].value == "└── PAV   [Consent]"
    assert ws["B1"].value is None
    assert ws["A1"].font.name == "Consolas"


def test_the_flow_sheet_is_skipped_when_there_is_no_lineage():
    from openpyxl import Workbook
    from nextseek_api.services.sample_workbook import _write_flow_sheet, FLOW_SHEET

    book = Workbook()
    _write_flow_sheet(book, [])
    assert FLOW_SHEET not in book.sheetnames
