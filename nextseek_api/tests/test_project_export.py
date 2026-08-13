"""Tests for /nextseek_api/admin/project-export/ (port of pull_all_db.py)."""

import io
from unittest.mock import patch

import openpyxl
from django.contrib.auth.models import User
from django.test import TestCase
from rest_framework.test import APIClient

from nextseek_api.services.project_export import build_sheet_tables

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
RUN_URL = "/nextseek_api/admin/project-export/run/"
DETAIL_URL = "/nextseek_api/admin/project-export/1/"

_PROJECT = {"id": 1, "title": "Published Data"}
_ROWS = [
    {"id": 11, "uuid": "TIS-250101BEH-1", "sample_type": "TIS",
     "json_metadata": '{"UID": "TIS-250101BEH-1", "Organ": "Lung", "Notes": ""}'},
    {"id": 12, "uuid": "TIS-250101BEH-2", "sample_type": "TIS",
     "json_metadata": '{"UID": "TIS-250101BEH-2", "Organ": "Liver", "Notes": ""}'},
    {"id": 13, "uuid": "AB-250101BEH-1", "sample_type": "AB",
     "json_metadata": '{"UID": "AB-250101BEH-1", "Tags": ["a", "b"]}'},
    {"id": 14, "uuid": "ABP-250101BEH-1", "sample_type": "ABP",
     "json_metadata": '{"UID": "ABP-250101BEH-1"}'},
]


def _patch_extractors(testcase, project=_PROJECT, rows=_ROWS):
    for name, value in (("resolve_project", project), ("pull_project_samples", rows)):
        patcher = patch(f"nextseek_api.services.project_export.{name}", return_value=value)
        patcher.start()
        testcase.addCleanup(patcher.stop)


class ProjectExportAuthTests(TestCase):
    databases = {"default"}

    def test_anonymous_is_rejected(self):
        self.assertEqual(
            APIClient().post(RUN_URL, {"project_id": 1}, format="json").status_code, 401
        )

    def test_staff_but_not_superuser_is_rejected(self):
        # The test this endpoint exists to get right: dmac/views.py:80,97 sets
        # is_staff=1 for every SEEK user at login, so DRF's IsAdminUser would
        # let this caller through.
        user = User.objects.create_user(
            "staffonly", password="pw", is_staff=True, is_superuser=False
        )
        client = APIClient()
        client.force_authenticate(user=user)
        self.assertEqual(
            client.post(RUN_URL, {"project_id": 1}, format="json").status_code, 403
        )

    def test_plain_user_is_rejected(self):
        client = APIClient()
        client.force_authenticate(user=User.objects.create_user("plain", password="pw"))
        self.assertEqual(
            client.post(RUN_URL, {"project_id": 1}, format="json").status_code, 403
        )


class ProjectExportSuccessTests(TestCase):
    databases = {"default"}

    def setUp(self):
        self.user = User.objects.create_user(
            "su", password="pw", is_staff=True, is_superuser=True
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)
        _patch_extractors(self)

    def test_json_response_is_grouped_by_sample_type(self):
        resp = self.client.post(RUN_URL, {"project_id": 1}, format="json")
        self.assertEqual(resp.status_code, 200)
        body = resp.json()
        self.assertEqual(body["project_id"], 1)
        self.assertEqual(body["project_title"], "Published Data")
        self.assertEqual(body["total_samples"], 4)
        self.assertEqual(body["total_sample_types"], 3)
        self.assertEqual(body["unparseable_metadata"], 0)
        self.assertEqual({g["sample_type"] for g in body["data"]}, {"TIS", "AB", "ABP"})
        self.assertIn("db_name", body["source"])

    def test_ab_does_not_absorb_abp(self):
        # The script matched sample type with str.startswith over the uuid, so
        # every ABP-* row landed on the AB sheet as well as its own.
        body = self.client.post(RUN_URL, {"project_id": 1}, format="json").json()
        groups = {g["sample_type"]: g for g in body["data"]}
        self.assertEqual(groups["AB"]["n_samples"], 1)
        self.assertEqual(groups["ABP"]["n_samples"], 1)
        self.assertEqual(groups["AB"]["samples"][0]["uuid"], "AB-250101BEH-1")

    def test_xlsx_response_headers_and_content(self):
        resp = self.client.post(
            RUN_URL, {"project_id": 1, "output_format": "xlsx"}, format="json"
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], XLSX_CONTENT_TYPE)
        self.assertIn("attachment", resp["Content-Disposition"])
        self.assertIn("project-1-Published-Data.xlsx", resp["Content-Disposition"])

        body = b"".join(resp.streaming_content)
        self.assertTrue(body.startswith(b"PK"))

        workbook = openpyxl.load_workbook(io.BytesIO(body))
        self.assertEqual(set(workbook.sheetnames), {"TIS", "AB", "ABP"})
        # List-valued metadata must survive; openpyxl raises on raw lists.
        self.assertEqual(workbook["AB"]["A1"].value, "id")
        self.assertEqual(workbook["TIS"].max_row, 3)

    def test_get_route_serves_xlsx(self):
        resp = self.client.get(DETAIL_URL, {"output_format": "xlsx"})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], XLSX_CONTENT_TYPE)

    def test_unknown_output_format_is_422(self):
        resp = self.client.post(
            RUN_URL, {"project_id": 1, "output_format": "pdf"}, format="json"
        )
        self.assertEqual(resp.status_code, 422)


class ProjectExportMissingProjectTests(TestCase):
    databases = {"default"}

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(
            user=User.objects.create_user("su2", password="pw", is_superuser=True)
        )
        _patch_extractors(self, project=None, rows=[])

    def test_unknown_project_is_404_naming_the_database(self):
        resp = self.client.post(RUN_URL, {"project_id": 9999}, format="json")
        self.assertEqual(resp.status_code, 404)
        # The script returned an empty workbook when pointed at the wrong
        # instance; the endpoint must say which database it asked.
        self.assertIn("9999", resp.json()["errors"][0]["detail"])


class BuildSheetTablesTests(TestCase):
    """Unit tests for the pure grouping/cleaning stage."""

    databases = {"default"}

    def test_all_empty_columns_are_dropped(self):
        tables, _ = build_sheet_tables(_ROWS)
        tis = next(t for t in tables if t["sample_type"] == "TIS")
        self.assertIn("Organ", tis["columns"])
        self.assertNotIn("Notes", tis["columns"])  # "" for every row

    def test_key_columns_lead_and_are_always_kept(self):
        tables, _ = build_sheet_tables(_ROWS)
        self.assertTrue(all(t["columns"][:2] == ["id", "uuid"] for t in tables))

    def test_list_valued_metadata_reaches_the_workbook(self):
        tables, _ = build_sheet_tables(_ROWS)
        ab = next(t for t in tables if t["sample_type"] == "AB")
        self.assertEqual(ab["data"][0]["Tags"], ["a", "b"])
        workbook = openpyxl.load_workbook(io.BytesIO(_generate(tables)))
        self.assertIsNotNone(workbook["AB"]["D2"].value)

    def test_unparseable_metadata_is_counted_not_dropped(self):
        rows = [{"id": 1, "uuid": "X-1", "sample_type": "X", "json_metadata": "{not json"}]
        tables, unparseable = build_sheet_tables(rows)
        self.assertEqual(unparseable, 1)
        self.assertEqual(tables[0]["data"][0]["uuid"], "X-1")

    def test_null_sample_type_falls_back_to_unknown(self):
        rows = [{"id": 1, "uuid": "X-1", "sample_type": None, "json_metadata": "{}"}]
        tables, _ = build_sheet_tables(rows)
        self.assertEqual(tables[0]["sample_type"], "UNKNOWN")

    def test_metadata_key_cannot_shadow_a_key_column(self):
        rows = [{"id": 1, "uuid": "X-1", "sample_type": "X",
                 "json_metadata": '{"uuid": "spoofed"}'}]
        tables, _ = build_sheet_tables(rows)
        record = tables[0]["data"][0]
        self.assertEqual(record["uuid"], "X-1")
        self.assertEqual(record["metadata_uuid"], "spoofed")

    def test_illegal_sheet_characters_are_sanitized(self):
        rows = [{"id": 1, "uuid": "X-1", "sample_type": "A/B:C", "json_metadata": "{}"}]
        tables, _ = build_sheet_tables(rows)
        self.assertEqual(tables[0]["label"], "A_B_C")
        self.assertEqual(tables[0]["sample_type"], "A/B:C")


def _generate(tables):
    from nextseek_api.assistant.excel_export import generate_table_xlsx
    return generate_table_xlsx(tables)
