"""The blank-template workbook: README, one headers-only sheet per type, manifest."""

from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook

from nextseek_api.services.sample_workbook import (
    CV_SHEET,
    MANIFEST_SHEET,
    README_SHEET,
    TEMPLATE_FORMAT_VERSION,
    write_template_workbook,
)
from nextseek_api.services.template_catalog import SampleTypeEntry

_MOD = "nextseek_api.services.sample_workbook"

TIS = SampleTypeEntry(code="TIS", sample_type_id=2, name="Tissue",
                      description="A tissue sample.", group="")
SEQ = SampleTypeEntry(code="D.SEQ", sample_type_id=11, name="Sequencing Data",
                      description="Reads off an instrument.", group="D.")

SPECS = {
    2: [{"title": "UID", "required": True, "pos": 0},
        {"title": "Name", "required": True, "pos": 1},
        {"title": "Weight", "required": False, "pos": 2}],
    11: [{"title": "UID", "required": True, "pos": 0},
         {"title": "File_PrimaryData", "required": False, "pos": 1}],
}


@pytest.fixture
def _no_lookups():
    """Every enrichment lookup stubbed. Individual tests override what they test."""
    with patch(f"{_MOD}.DBtable_sampleattribute") as sa, \
         patch(f"{_MOD}.load_sample_field_context", return_value={}), \
         patch(f"{_MOD}.load_relationships", return_value={}), \
         patch(f"{_MOD}._load_vocabularies", return_value=({}, {})):
        sa.return_value.getAttributeSpecsBySampleTypeIds.return_value = SPECS
        yield sa


def _write(tmp_path, entries, name="t.xlsx"):
    out = tmp_path / name
    write_template_workbook(entries, str(out))
    return load_workbook(str(out))


def test_readme_is_the_first_sheet(tmp_path, _no_lookups):
    book = _write(tmp_path, [TIS])
    assert book.sheetnames[0] == README_SHEET


def test_one_sheet_per_type_named_by_its_code_in_the_given_order(tmp_path, _no_lookups):
    book = _write(tmp_path, [SEQ, TIS])
    assert book.sheetnames[1:3] == ["D.SEQ", "TIS"]


def test_type_sheets_carry_headers_and_no_data_rows(tmp_path, _no_lookups):
    book = _write(tmp_path, [TIS])
    ws = book["TIS"]
    assert ws.max_row == 1
    assert [c.value for c in ws[1]] == ["UID*", "Name*", "Weight"]


def test_required_headers_are_starred_and_bold(tmp_path, _no_lookups):
    book = _write(tmp_path, [TIS])
    ws = book["TIS"]
    assert ws.cell(row=1, column=1).value == "UID*"
    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.cell(row=1, column=3).value == "Weight"
    assert ws.cell(row=1, column=3).font.bold is False


def test_headers_follow_pos_order(tmp_path, _no_lookups):
    """The catalog orders by pos; the writer must not re-sort."""
    _no_lookups.return_value.getAttributeSpecsBySampleTypeIds.return_value = {
        2: [{"title": "Zebra", "required": False, "pos": 0},
            {"title": "Apple", "required": False, "pos": 1}],
    }
    book = _write(tmp_path, [TIS])
    assert [c.value for c in book["TIS"][1]] == ["Zebra", "Apple"]


def test_header_cells_carry_the_definition_as_a_hover_note(tmp_path, _no_lookups):
    with patch(f"{_MOD}.load_sample_field_context",
               return_value={("TIS", "Weight"): "Mass of the tissue in mg."}):
        book = _write(tmp_path, [TIS])
    ws = book["TIS"]
    assert "Mass of the tissue in mg." in ws.cell(row=1, column=3).comment.text
    assert ws.cell(row=1, column=1).comment is None


def test_an_undefined_attribute_gets_no_note_rather_than_an_empty_one(tmp_path, _no_lookups):
    book = _write(tmp_path, [TIS])
    assert book["TIS"].cell(row=1, column=2).comment is None


def test_a_type_with_no_attributes_gets_an_explanatory_sheet(tmp_path, _no_lookups):
    from nextseek_api.services.sample_workbook import EMPTY_TYPE_NOTE

    _no_lookups.return_value.getAttributeSpecsBySampleTypeIds.return_value = {2: []}
    book = _write(tmp_path, [TIS])
    assert "TIS" in book.sheetnames
    assert book["TIS"].cell(row=1, column=1).value == EMPTY_TYPE_NOTE


def test_a_type_with_no_attributes_is_still_listed_in_the_readme(tmp_path, _no_lookups):
    _no_lookups.return_value.getAttributeSpecsBySampleTypeIds.return_value = {2: []}
    book = _write(tmp_path, [TIS])
    text = " ".join(str(c.value) for row in book[README_SHEET].iter_rows()
                    for c in row if c.value)
    assert "TIS" in text


def test_the_manifest_is_the_last_sheet_and_is_hidden(tmp_path, _no_lookups):
    book = _write(tmp_path, [TIS])
    assert book.sheetnames[-1] == MANIFEST_SHEET
    assert book[MANIFEST_SHEET].sheet_state == "hidden"


def test_the_manifest_renders_database_field_ready_for_the_instructions_sheet(
        tmp_path, _no_lookups):
    book = _write(tmp_path, [TIS])
    rows = list(book[MANIFEST_SHEET].iter_rows(values_only=True))
    body = [r for r in rows if r[0] == "TIS"]
    assert ("TIS", "TIS", "UID", "TIS::UID", 1) in body
    assert ("TIS", "TIS", "Weight", "TIS::Weight", 0) in body


def test_the_manifest_carries_the_format_version(tmp_path, _no_lookups):
    """Pinned to exact cells: the converter reads B1, not 'a 1 somewhere'."""
    book = _write(tmp_path, [TIS])
    ws = book[MANIFEST_SHEET]
    assert ws.cell(row=1, column=1).value == "format_version"
    assert ws.cell(row=1, column=2).value == TEMPLATE_FORMAT_VERSION


def test_the_manifest_header_row_is_where_the_converter_expects_it(tmp_path, _no_lookups):
    from nextseek_api.services.sample_workbook import MANIFEST_HEADER

    book = _write(tmp_path, [TIS])
    ws = book[MANIFEST_SHEET]
    assert [c.value for c in ws[2]] == MANIFEST_HEADER


def test_a_governed_column_gets_a_dropdown_and_a_vocabulary_sheet(tmp_path, _no_lookups):
    with patch(f"{_MOD}._load_vocabularies",
               return_value=({"Weight": "units"}, {"units": ["mg", "g"]})):
        book = _write(tmp_path, [TIS])
    assert CV_SHEET in book.sheetnames
    assert len(book["TIS"].data_validations.dataValidation) == 1


def test_no_governed_columns_means_no_vocabulary_sheet(tmp_path, _no_lookups):
    book = _write(tmp_path, [TIS])
    assert CV_SHEET not in book.sheetnames


def test_the_vocabulary_sheet_sits_after_the_type_sheets_not_before_them(
        tmp_path, _no_lookups):
    """Per the design doc, sheet order is README, one sheet per selected type
    in the given order, Controlled Vocabularies (only when needed), then the
    hidden manifest. _write_vocabulary_sheet must run before the type-sheet
    loop (it hands _apply_dropdowns the `ranges` it needs while the loop
    runs), so this pins the *end result* rather than the write order."""
    with patch(f"{_MOD}._load_vocabularies",
               return_value=({"Weight": "units"}, {"units": ["mg", "g"]})):
        book = _write(tmp_path, [SEQ, TIS])
    assert book.sheetnames == [
        README_SHEET, "D.SEQ", "TIS", CV_SHEET, MANIFEST_SHEET,
    ]
    # The dropdown formula references the vocabulary sheet by name, so the
    # move must not have broken it.
    rule = next(iter(book["TIS"].data_validations.dataValidation))
    assert rule.formula1 == f"'{CV_SHEET}'!$A$2:$A$3"


def test_there_is_no_flow_sheet(tmp_path, _no_lookups):
    """A blank template has no provenance to draw."""
    from nextseek_api.services.sample_workbook import FLOW_SHEET

    book = _write(tmp_path, [TIS])
    assert FLOW_SHEET not in book.sheetnames


def test_a_failing_attribute_lookup_still_produces_a_readme_workbook(tmp_path, _no_lookups):
    """A whole-batch lookup failure drops every type, but must not cost the
    workbook itself: the README is still written."""
    _no_lookups.return_value.getAttributeSpecsBySampleTypeIds.side_effect = \
        RuntimeError("db down")
    book = _write(tmp_path, [TIS, SEQ])
    assert book.sheetnames[0] == README_SHEET
    assert "TIS" not in book.sheetnames


def test_a_type_missing_from_a_successful_lookup_is_skipped_not_the_others(
        tmp_path, _no_lookups):
    """The realistic trigger for the `None` skip path: the batched lookup
    succeeds but its dict simply has no key for one requested sample_type_id
    (getAttributeSpecsBySampleTypeIds drops ids that are not positive ints),
    while the other type's id is present and covered normally."""
    _no_lookups.return_value.getAttributeSpecsBySampleTypeIds.return_value = {
        2: SPECS[2],
        # no key at all for SEQ's sample_type_id (11)
    }
    book = _write(tmp_path, [TIS, SEQ])
    assert "TIS" in book.sheetnames
    assert "D.SEQ" not in book.sheetnames


def test_relationships_reach_the_readme(tmp_path, _no_lookups):
    with patch(f"{_MOD}.load_relationships",
               return_value={"TIS": {"parents": ["MUS"], "children": ["DNA"]}}):
        book = _write(tmp_path, [TIS])
    text = " ".join(str(c.value) for row in book[README_SHEET].iter_rows()
                    for c in row if c.value)
    assert "Typically derived from: MUS" in text


def test_a_type_naming_itself_as_a_relative_is_not_shown_as_its_own_relative(
        tmp_path, _no_lookups):
    """DNA really does list itself in its own parent_sampletypes column.

    The README must never say a type is derived from, or feeds into, itself.
    """
    with patch(f"{_MOD}.load_relationships",
               return_value={"TIS": {"parents": ["TIS", "MUS"], "children": ["TIS"]}}):
        book = _write(tmp_path, [TIS])
    text = " ".join(str(c.value) for row in book[README_SHEET].iter_rows()
                    for c in row if c.value)
    assert "Typically derived from: MUS" in text
    assert "Typically derived from: TIS" not in text
    assert "Typically feeds into" not in text


def test_a_type_whose_only_relatives_are_itself_gets_no_relationship_lines(
        tmp_path, _no_lookups):
    with patch(f"{_MOD}.load_relationships",
               return_value={"TIS": {"parents": ["TIS"], "children": ["TIS"]}}):
        book = _write(tmp_path, [TIS])
    text = " ".join(str(c.value) for row in book[README_SHEET].iter_rows()
                    for c in row if c.value)
    assert "Typically derived from" not in text
    assert "Typically feeds into" not in text
