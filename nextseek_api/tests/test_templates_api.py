"""Tests for /nextseek_api/templates/ -- the Download Templates tool as an API."""

from unittest.mock import MagicMock, patch

from django.contrib.auth.models import User
from django.test import TestCase
from rest_framework.test import APIClient

from nextseek_api.services.template_catalog import SampleTypeEntry

CATALOG_URL = "/nextseek_api/templates/catalog/"
GENERATE_URL = "/nextseek_api/templates/generate/"
XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

_TIS = SampleTypeEntry(code="TIS", sample_type_id=2, name="Tissue",
                       description="A tissue sample.", group="")
_DNA = SampleTypeEntry(code="DNA", sample_type_id=42, name="DNA",
                       description="Extracted DNA.", group="")
_SEQ = SampleTypeEntry(code="D.SEQ", sample_type_id=11, name="Sequencing Data",
                       description="Reads.", group="D.")

_PAYLOAD = {
    "groups": [
        {"key": "", "label": "Experimental types", "entries": [_DNA, _TIS]},
        {"key": "D.", "label": "Data types", "entries": [_SEQ]},
    ],
    "children": {"TIS": ["DNA"]},
    "requires": {"D.SEQ": {"add": ["DNA"], "assays": ["Whole Genome Sequencing"]}},
    "companions": {"TIS": {"add": ["D.SEQ"], "assays": []}},
    "max_suggestions": 12,
}


def _patch_catalog(testcase, payload=_PAYLOAD):
    patcher = patch("nextseek_api.services.templates.build_catalog",
                    return_value=payload)
    patcher.start()
    testcase.addCleanup(patcher.stop)


class CatalogAuthTests(TestCase):
    databases = {"default"}

    def test_anonymous_is_rejected(self):
        self.assertEqual(APIClient().get(CATALOG_URL).status_code, 401)

    def test_any_authenticated_user_may_read_the_catalog(self):
        # Deliberate, and pinned so that tightening it later has to be a
        # conscious change: the catalog is schema, not sample data, and
        # seek.views.assets.templatesList already serves exactly this to any
        # logged-in user with no project-membership check.
        _patch_catalog(self)
        client = APIClient()
        client.force_authenticate(user=User.objects.create_user("plain", password="pw"))
        self.assertEqual(client.get(CATALOG_URL).status_code, 200)


class CatalogContentTests(TestCase):
    databases = {"default"}

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(
            user=User.objects.create_user("plain", password="pw")
        )
        _patch_catalog(self)

    def test_groups_keep_their_order_and_entries(self):
        body = self.client.get(CATALOG_URL).json()
        self.assertEqual([g["key"] for g in body["groups"]], ["", "D."])
        self.assertEqual([g["label"] for g in body["groups"]],
                         ["Experimental types", "Data types"])
        self.assertEqual([e["code"] for e in body["groups"][0]["entries"]],
                         ["DNA", "TIS"])

    def test_entries_are_serialised_in_full(self):
        body = self.client.get(CATALOG_URL).json()
        self.assertEqual(
            body["groups"][0]["entries"][0],
            {"code": "DNA", "sample_type_id": 42, "name": "DNA",
             "description": "Extracted DNA.", "group": ""},
        )

    def test_relations_are_passed_through(self):
        body = self.client.get(CATALOG_URL).json()
        self.assertEqual(body["children"], {"TIS": ["DNA"]})
        self.assertEqual(
            body["requires"],
            {"D.SEQ": {"add": ["DNA"], "assays": ["Whole Genome Sequencing"]}},
        )
        self.assertEqual(body["companions"],
                         {"TIS": {"add": ["D.SEQ"], "assays": []}})
        self.assertEqual(body["max_suggestions"], 12)


class CatalogEmptyTableTests(TestCase):
    databases = {"default"}

    def test_an_unpopulated_requirements_table_is_not_an_error(self):
        # The table ships empty and the picker treats that as "nothing known",
        # so a fresh install must get 200 with empty relations, never a 500.
        empty = dict(_PAYLOAD, requires={}, companions={})
        _patch_catalog(self, empty)
        client = APIClient()
        client.force_authenticate(user=User.objects.create_user("plain", password="pw"))
        resp = client.get(CATALOG_URL)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["requires"], {})
        self.assertEqual(resp.json()["companions"], {})
        self.assertTrue(resp.json()["groups"])


class CatalogMalformedRulesRowTests(TestCase):
    """End-to-end proof for the load_type_links fix: a real malformed row
    must not 500 the endpoint. Unlike the other Catalog* tests, build_catalog
    is NOT patched here -- only load_type_links' own data source is stubbed,
    so build_catalog, the real load_type_links, and the strict pydantic
    TemplateCatalogResponse all run for real and are genuinely exercised."""

    databases = {"default"}

    def test_a_malformed_rules_row_does_not_500_the_catalog_endpoint(self):
        tis = SampleTypeEntry(code="TIS", sample_type_id=2, name="Tissue",
                              description="A tissue sample.", group="")

        # add_codes is a JSON object, not a list: set() of a dict yields its
        # keys, so `set(add) <= known` would pass this through with "TIS"
        # known -- and a dict then reaches pydantic's strict List[str].
        bad_row = {"kind": "requires", "trigger_code": "TIS",
                   "add_codes": '{"TIS": 1}', "assay_titles": None}
        rules_model = MagicMock()
        rules_model.objects.filter.return_value.values.return_value = [bad_row]
        rules_model.KIND_COMPANION = "companion"
        rules_model.KIND_REQUIRES = "requires"

        with patch("nextseek_api.services.template_catalog.load_catalog",
                   return_value=[tis]), \
             patch("nextseek_api.services.template_catalog.load_relationships",
                   return_value={"TIS": {"parents": [], "children": []}}), \
             patch("nextseek_api.services.template_catalog.Sample_type_requirements",
                   rules_model):
            client = APIClient()
            client.force_authenticate(
                user=User.objects.create_user("plain", password="pw")
            )
            resp = client.get(CATALOG_URL)

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["requires"], {})


class CatalogSchemaTests(TestCase):
    databases = {"default"}

    def test_the_route_is_registered(self):
        from django.urls import reverse

        self.assertTrue(reverse("nextseek_api:templates-catalog").endswith(
            "/templates/catalog/"))

    def test_the_path_and_models_appear_in_the_openapi_schema(self):
        from drf_spectacular.generators import SchemaGenerator

        schema = SchemaGenerator().get_schema(request=None, public=True)
        self.assertIn("/nextseek_api/templates/catalog/", schema["paths"])
        operation = schema["paths"]["/nextseek_api/templates/catalog/"]["get"]
        self.assertEqual(operation["operationId"], "Templates: Catalog (GET)")
        self.assertIn("templates", operation["tags"])
        examples = operation["responses"]["200"]["content"]["application/json"]["examples"]
        self.assertGreaterEqual(len(examples), 1)
        self.assertIn("TemplateCatalogResponse", schema["components"]["schemas"])


class GenerateAuthTests(TestCase):
    databases = {"default"}

    def test_anonymous_is_rejected(self):
        self.assertEqual(
            APIClient().post(GENERATE_URL, {"codes": ["TIS"]}, format="json").status_code,
            401,
        )

    def test_staff_but_not_superuser_is_rejected(self):
        # dmac/views.py sets is_staff=1 for every SEEK user at login, so DRF's
        # IsAdminUser would let this caller straight through.
        user = User.objects.create_user("staffonly", password="pw",
                                        is_staff=True, is_superuser=False)
        client = APIClient()
        client.force_authenticate(user=user)
        self.assertEqual(
            client.post(GENERATE_URL, {"codes": ["TIS"]}, format="json").status_code,
            403,
        )

    def test_the_catalog_stays_readable_for_the_same_caller(self):
        # The whole point of the per-action gate: one class, two audiences.
        _patch_catalog(self)
        client = APIClient()
        client.force_authenticate(
            user=User.objects.create_user("plain2", password="pw")
        )
        self.assertEqual(client.get(CATALOG_URL).status_code, 200)
        self.assertEqual(
            client.post(GENERATE_URL, {"codes": ["TIS"]}, format="json").status_code,
            403,
        )


class GenerateValidationTests(TestCase):
    databases = {"default"}

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(
            user=User.objects.create_user("su", password="pw",
                                          is_staff=True, is_superuser=True)
        )
        patcher = patch("nextseek_api.services.template_catalog.load_catalog",
                        return_value=[_DNA, _TIS, _SEQ])
        patcher.start()
        self.addCleanup(patcher.stop)

    def test_empty_codes_is_rejected(self):
        resp = self.client.post(GENERATE_URL, {"codes": []}, format="json")
        self.assertEqual(resp.status_code, 422)
        self.assertIn("errors", resp.json())

    def test_a_missing_body_is_rejected(self):
        resp = self.client.post(GENERATE_URL, {}, format="json")
        self.assertEqual(resp.status_code, 422)

    def test_an_unknown_code_is_rejected_rather_than_dropped(self):
        # The page drops unknown codes so a stale bookmark still works. An API
        # caller must be told instead -- a workbook quietly missing a sheet is
        # worse than an error.
        resp = self.client.post(GENERATE_URL, {"codes": ["TIS", "D.SEQZ"]},
                                format="json")
        self.assertEqual(resp.status_code, 422)
        self.assertIn("D.SEQZ", str(resp.json()))

    def test_an_unexpected_field_is_rejected(self):
        resp = self.client.post(GENERATE_URL,
                                {"codes": ["TIS"], "output_format": "xlsx"},
                                format="json")
        self.assertEqual(resp.status_code, 422)


class GenerateWorkbookTests(TestCase):
    databases = {"default"}

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(
            user=User.objects.create_user("su", password="pw",
                                          is_staff=True, is_superuser=True)
        )
        patcher = patch("nextseek_api.services.template_catalog.load_catalog",
                        return_value=[_DNA, _TIS, _SEQ])
        patcher.start()
        self.addCleanup(patcher.stop)

    def test_the_writer_receives_the_requested_entries_in_request_order(self):
        with patch("nextseek_api.services.sample_workbook.write_template_workbook") as writer:
            resp = self.client.post(GENERATE_URL, {"codes": ["D.SEQ", "TIS"]},
                                    format="json")
        self.assertEqual(resp.status_code, 200)
        chosen = writer.call_args[0][0]
        self.assertEqual([e.code for e in chosen], ["D.SEQ", "TIS"])

    def test_repeated_codes_produce_one_entry_each(self):
        with patch("nextseek_api.services.sample_workbook.write_template_workbook") as writer:
            self.client.post(GENERATE_URL, {"codes": ["TIS", "TIS", "DNA"]},
                             format="json")
        chosen = writer.call_args[0][0]
        self.assertEqual([e.code for e in chosen], ["TIS", "DNA"])

    def test_the_response_is_a_real_xlsx_attachment(self):
        import io

        import openpyxl

        def _fake_writer(entries, destination):
            book = openpyxl.Workbook()
            del book["Sheet"]
            book.create_sheet("README")
            for entry in entries:
                book.create_sheet(entry.code)
            manifest = book.create_sheet("_NEXTSEEK")
            manifest.sheet_state = "hidden"
            book.save(destination)

        with patch("nextseek_api.services.sample_workbook.write_template_workbook",
                   side_effect=_fake_writer):
            resp = self.client.post(GENERATE_URL, {"codes": ["TIS", "D.SEQ"]},
                                    format="json")

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], XLSX_CONTENT_TYPE)
        self.assertIn("attachment", resp["Content-Disposition"])
        self.assertIn("NExtSEEK_templates_2types_", resp["Content-Disposition"])

        body = b"".join(resp.streaming_content)
        self.assertTrue(body.startswith(b"PK"))

        # Proves the BytesIO round trip end to end: openpyxl wrote into the
        # buffer the view streamed, and openpyxl can read it back.
        book = openpyxl.load_workbook(io.BytesIO(body))
        self.assertEqual(book.sheetnames, ["README", "TIS", "D.SEQ", "_NEXTSEEK"])
        self.assertEqual(book["_NEXTSEEK"].sheet_state, "hidden")


class GenerateSchemaTests(TestCase):
    databases = {"default"}

    def test_the_route_is_registered(self):
        from django.urls import reverse

        self.assertTrue(reverse("nextseek_api:templates-generate").endswith(
            "/templates/generate/"))

    def test_the_path_and_request_model_appear_in_the_openapi_schema(self):
        from drf_spectacular.generators import SchemaGenerator

        schema = SchemaGenerator().get_schema(request=None, public=True)
        self.assertIn("/nextseek_api/templates/generate/", schema["paths"])
        operation = schema["paths"]["/nextseek_api/templates/generate/"]["post"]
        self.assertEqual(operation["operationId"], "Templates: Generate (POST)")
        self.assertIn("templates", operation["tags"])
        examples = operation["requestBody"]["content"]["application/json"]["examples"]
        self.assertGreaterEqual(len(examples), 1)
        self.assertIn(XLSX_CONTENT_TYPE, operation["responses"]["200"]["content"])
        self.assertIn("TemplateGenerateRequest", schema["components"]["schemas"])
