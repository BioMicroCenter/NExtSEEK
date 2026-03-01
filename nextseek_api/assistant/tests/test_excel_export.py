"""Unit tests for Excel export utilities."""
import io

import openpyxl
from django.test import SimpleTestCase

from nextseek_api.assistant.excel_export import (
    extract_table_artifacts,
    generate_table_xlsx,
    generate_search_xlsx,
    flatten_report_to_tables,
)


class FlattenReportToTablesTests(SimpleTestCase):

    def test_geo_report_extracts_samples_table(self):
        report_output = {
            "report_type": "GEO",
            "report": {
                "study": {"title": "My Study", "organism": "Homo sapiens"},
                "samples": [
                    {"uid": "NHP-1", "title": "Sample A", "source_name": "blood"},
                    {"uid": "NHP-2", "title": "Sample B", "source_name": "tissue"},
                ],
                "protocols": {"extraction": "RNA extraction protocol"},
            },
            "narrative": "Report generated.",
            "notes": "",
        }
        tables = flatten_report_to_tables(report_output)
        samples = next(t for t in tables if t["key"] == "samples")
        self.assertEqual(samples["label"], "Samples")
        self.assertEqual(len(samples["data"]), 2)
        self.assertIn("uid", samples["columns"])

    def test_generic_report_auto_flattens(self):
        report_output = {
            "report_type": "SUBMISSION",
            "report": {
                "metadata": [
                    {"field": "PI", "value": "Dr. Smith"},
                    {"field": "Date", "value": "2026-01-01"},
                ],
            },
            "narrative": "Submission report.",
            "notes": "",
        }
        tables = flatten_report_to_tables(report_output)
        self.assertGreaterEqual(len(tables), 1)

    def test_empty_report_returns_empty(self):
        report_output = {"report_type": None, "report": {}, "narrative": "", "notes": ""}
        tables = flatten_report_to_tables(report_output)
        self.assertEqual(tables, [])


class ExtractTableArtifactsTests(SimpleTestCase):

    def test_reporter_bundle_produces_file_artifact(self):
        """GEO bundles always produce a file artifact for the workbook download."""
        bundle = {
            "mode": "reporter",
            "report_writer_output": {
                "report_type": "GEO",
                "report": {"samples": [{"uid": "X-1", "title": "S1"}], "study": {"title": "Study"}},
                "narrative": "Done.",
                "notes": "",
            },
            "report_saved_files": {"geo_seq_workbooks": ["/tmp/geo_workbook.xlsx"]},
        }
        artifacts = extract_table_artifacts(bundle)
        file_artifacts = [a for a in artifacts if a["artifact_type"] == "file"]
        self.assertEqual(len(file_artifacts), 1)
        self.assertEqual(file_artifacts[0]["key"], "geo_seq_workbooks")

    def test_reporter_bundle_no_preview_when_workbook_missing(self):
        """GEO bundles with a missing workbook file produce no preview artifact."""
        bundle = {
            "mode": "reporter",
            "report_writer_output": {
                "report_type": "GEO",
                "report": {"samples": [{"uid": "X-1", "title": "S1"}]},
                "narrative": "Done.",
                "notes": "",
            },
            "report_saved_files": {"geo_seq_workbooks": ["/tmp/nonexistent_geo_workbook.xlsx"]},
        }
        artifacts = extract_table_artifacts(bundle)
        preview_artifacts = [a for a in artifacts if a["artifact_type"] == "preview"]
        self.assertEqual(len(preview_artifacts), 0)

    def test_reporter_bundle_preview_from_real_workbook(self):
        """GEO bundles with a real workbook produce a preview artifact from the file."""
        import os
        import tempfile
        # Build a minimal workbook mimicking the GEO template structure
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Metadata"
        # Row 38 = sample headers, row 39+ = data (per _populate_geo_seq_workbook)
        ws.cell(row=38, column=1, value="sample_name")
        ws.cell(row=38, column=2, value="title")
        ws.cell(row=39, column=1, value="Sample1")
        ws.cell(row=39, column=2, value="My sample")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            wb.save(tmp_path)
            bundle = {
                "mode": "reporter",
                "report_writer_output": {
                    "report_type": "GEO",
                    "report": {},
                    "narrative": "",
                    "notes": "",
                },
                "report_saved_files": {"geo_seq_workbooks": [tmp_path]},
            }
            artifacts = extract_table_artifacts(bundle)
            preview_artifacts = [a for a in artifacts if a["artifact_type"] == "preview"]
            self.assertEqual(len(preview_artifacts), 1)
            preview = preview_artifacts[0]
            self.assertEqual(preview["key"], "geo_report_preview")
            self.assertEqual(len(preview["sheets"]), 1)
            sheet = preview["sheets"][0]
            self.assertEqual(sheet["name"], "Metadata")
            self.assertIn("sample_name", sheet["columns"])
            self.assertEqual(sheet["data"][0]["sample_name"], "Sample1")
        finally:
            os.unlink(tmp_path)

    def test_non_geo_reporter_produces_table_artifacts(self):
        """Non-GEO reporter bundles still produce inline table artifacts."""
        bundle = {
            "mode": "reporter",
            "report_writer_output": {
                "report_type": "SUBMISSION",
                "report": {"samples": [{"uid": "X-1", "title": "S1"}]},
                "narrative": "",
                "notes": "",
            },
        }
        artifacts = extract_table_artifacts(bundle)
        table_artifacts = [a for a in artifacts if a["artifact_type"] == "table"]
        self.assertGreaterEqual(len(table_artifacts), 1)

    def test_reporter_result_uses_streamlit_labels(self):
        """Impact tables use key/count columns and Streamlit-style labels."""
        bundle = {
            "mode": "sql_report",
            "reporter_result": {
                "sampletypes_table": {"NHP": 5, "PBMC": 3},
                "labs_table": {"MIT": 8},
                "rows_returned": 8,
            },
        }
        artifacts = extract_table_artifacts(bundle)
        labels = {a["label"] for a in artifacts}
        self.assertIn("SampleType Summary", labels)
        self.assertIn("Lab Summary", labels)
        for a in artifacts:
            self.assertEqual(a["columns"], ["key", "count"])
        self.assertEqual(artifacts[0].get("rows_returned"), 8)

    def test_search_bundle_returns_empty(self):
        bundle = {"mode": "new_search", "api_result_full": {"data": []}}
        artifacts = extract_table_artifacts(bundle)
        self.assertEqual(artifacts, [])

    def test_large_table_truncated(self):
        """Tables exceeding MAX_INLINE_ROWS are truncated in the SSE payload."""
        from nextseek_api.assistant.excel_export import MAX_INLINE_ROWS
        rows = [{"uid": f"X-{i}", "title": f"S{i}"} for i in range(500)]
        bundle = {
            "mode": "reporter",
            "report_writer_output": {
                "report": {"samples": rows},
            },
            "report_saved_files": {},
        }
        artifacts = extract_table_artifacts(bundle)
        table = next(a for a in artifacts if a["key"] == "samples")
        self.assertEqual(len(table["data"]), MAX_INLINE_ROWS)
        self.assertTrue(table["truncated"])
        self.assertEqual(table["total_rows"], 500)

    def test_small_table_not_truncated(self):
        """Tables within MAX_INLINE_ROWS are not truncated."""
        rows = [{"uid": f"X-{i}", "title": f"S{i}"} for i in range(50)]
        bundle = {
            "mode": "reporter",
            "report_writer_output": {
                "report": {"samples": rows},
            },
            "report_saved_files": {},
        }
        artifacts = extract_table_artifacts(bundle)
        table = next(a for a in artifacts if a["key"] == "samples")
        self.assertEqual(len(table["data"]), 50)
        self.assertNotIn("truncated", table)


class GenerateTableXlsxTests(SimpleTestCase):

    def test_single_table_produces_valid_xlsx(self):
        tables = [{"key": "samples", "label": "Samples", "columns": ["UID", "Title"], "data": [{"UID": "NHP-1", "Title": "Sample A"}, {"UID": "NHP-2", "Title": "Sample B"}]}]
        xlsx_bytes = generate_table_xlsx(tables)
        self.assertIsInstance(xlsx_bytes, bytes)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        self.assertIn("Samples", wb.sheetnames)
        ws = wb["Samples"]
        self.assertEqual(ws.cell(1, 1).value, "UID")
        self.assertEqual(ws.cell(1, 2).value, "Title")
        self.assertEqual(ws.cell(2, 1).value, "NHP-1")
        self.assertEqual(ws.cell(3, 1).value, "NHP-2")

    def test_multiple_tables_produce_multiple_sheets(self):
        tables = [
            {"key": "a", "label": "Sheet A", "columns": ["X"], "data": [{"X": 1}]},
            {"key": "b", "label": "Sheet B", "columns": ["Y"], "data": [{"Y": 2}]},
        ]
        xlsx_bytes = generate_table_xlsx(tables)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        self.assertIn("Sheet A", wb.sheetnames)
        self.assertIn("Sheet B", wb.sheetnames)

    def test_empty_tables_returns_empty_workbook(self):
        xlsx_bytes = generate_table_xlsx([])
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        self.assertGreaterEqual(len(wb.sheetnames), 1)


class GenerateSearchXlsxTests(SimpleTestCase):

    def test_jsonapi_results_to_xlsx(self):
        bundle = {
            "api_result_full": {
                "data": [
                    {"id": "1", "type": "samples", "attributes": {"title": "S1", "uid": "NHP-1"}},
                    {"id": "2", "type": "samples", "attributes": {"title": "S2", "uid": "NHP-2"}},
                ],
            },
        }
        xlsx_bytes = generate_search_xlsx(bundle)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        self.assertEqual(ws.max_row, 3)

    def test_empty_results_returns_workbook(self):
        bundle = {"api_result_full": {"data": []}}
        xlsx_bytes = generate_search_xlsx(bundle)
        self.assertIsInstance(xlsx_bytes, bytes)
