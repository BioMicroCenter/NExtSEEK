"""Tests for defensive cell-flattening in excel_export.

Report tables (PRIDE file_mapping / sample_metadata) carry nested values —
CV-param dicts {cv_label, accession, name, value} and lists of them. Before this
fix openpyxl raised "Cannot convert [] to Excel" and the React table rendered
"[object Object]". These tests pin the flattening behavior and the PRIDE file
download artifacts.
"""
from io import BytesIO

import openpyxl

from nextseek_api.assistant.excel_export import (
    _flatten_cell,
    build_artifacts,
    extract_table_artifacts,
    generate_table_xlsx,
)


def test_flatten_cell_cv_param_dict_uses_name_with_accession():
    cv = {"cv_label": "NEWT", "accession": "9606", "name": "Homo sapiens", "value": None}
    assert _flatten_cell(cv) == "Homo sapiens [9606]"


def test_flatten_cell_user_param_with_value():
    cv = {"cv_label": "PRIDE", "accession": "PRIDE:0000175", "name": "Search engine", "value": "Sequest"}
    assert _flatten_cell(cv) == "Search engine (Sequest)"


def test_flatten_cell_list_of_cv_params_joined():
    vals = [
        {"cv_label": "UNIMOD", "accession": "UNIMOD:737", "name": "TMT6plex", "value": None},
        {"cv_label": "UNIMOD", "accession": "UNIMOD:4", "name": "Carbamidomethyl", "value": None},
    ]
    out = _flatten_cell(vals)
    assert out == "TMT6plex [UNIMOD:737]; Carbamidomethyl [UNIMOD:4]"


def test_flatten_cell_empty_list_is_blank():
    assert _flatten_cell([]) == ""


def test_flatten_cell_scalars_passthrough():
    assert _flatten_cell("x") == "x"
    assert _flatten_cell(3) == 3
    assert _flatten_cell(None) is None


def test_generate_table_xlsx_with_nested_cells_does_not_raise():
    tables = [{
        "key": "sample_metadata",
        "label": "Sample Metadata",
        "columns": ["file_id", "species", "modification"],
        "data": [{
            "file_id": "1",
            "species": [{"cv_label": "NEWT", "accession": "9606", "name": "Homo sapiens", "value": None}],
            "modification": [],
        }],
    }]
    raw = generate_table_xlsx(tables)  # previously raised ValueError("Cannot convert [] to Excel")
    ws = openpyxl.load_workbook(BytesIO(raw))["Sample Metadata"]
    assert ws.cell(row=2, column=1).value == "1"
    assert ws.cell(row=2, column=2).value == "Homo sapiens [9606]"
    assert ws.cell(row=2, column=3).value in ("", None)


def test_extract_artifacts_includes_pride_file_downloads():
    bundle = {
        "mode": "reporter",
        "report_writer_output": {
            "report_type": "PRIDE",
            "report": {"file_mapping": [{"*file_id": "1", "*file_type": "SEARCH"}]},
        },
        "report_saved_files": {
            "pride_submission_px": ["/tmp/x/submission.px"],
            "pride_sdrf": ["/tmp/x/PRIDE.sdrf.tsv"],
        },
    }
    by_key = {a.get("key"): a for a in build_artifacts(bundle)}
    assert by_key.get("pride_submission_px", {}).get("artifact_type") == "file"
    assert by_key.get("pride_sdrf", {}).get("artifact_type") == "file"
