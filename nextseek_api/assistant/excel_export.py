"""
Excel export utilities for assistant artifacts.

Converts report_writer_output and search results into Excel workbooks.
Uses openpyxl (available via chat_nextseek dependency).
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Maximum number of rows to include inline in SSE event payloads.
# Full data is always available via the artifact download endpoint.
MAX_INLINE_ROWS = 200


def flatten_report_to_tables(report_writer_output: dict[str, Any]) -> list[dict[str, Any]]:
    """Convert a report_writer_output dict into a list of table artifact dicts.

    Each table has: key, label, columns, data.
    Handles known report structures (GEO samples/protocols/study) and
    falls back to generic auto-flattening for unknown types.

    The report_writer_output may be directly ``{report_type, report, ...}``
    or wrapped under a key like ``all_samples`` (per-sample report outputs).
    """
    report = report_writer_output.get("report") or {}

    # Handle nested wrapper: if there's no direct "report" key but there's a
    # sub-dict that itself has "report", unwrap it (e.g. {"all_samples": {"report": ...}}).
    if not report:
        for v in report_writer_output.values():
            if isinstance(v, dict) and "report" in v:
                report = v.get("report") or {}
                break

    if not report:
        return []

    tables: list[dict[str, Any]] = []

    for key, value in report.items():
        if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
            # List of dicts -> table rows
            columns = list(value[0].keys())
            tables.append({
                "key": key,
                "label": key.replace("_", " ").title(),
                "columns": columns,
                "data": value,
            })
        elif isinstance(value, dict) and not _is_deeply_nested(value):
            # Flat dict -> single-row table
            columns = list(value.keys())
            tables.append({
                "key": key,
                "label": key.replace("_", " ").title(),
                "columns": columns,
                "data": [value],
            })

    return tables


def _is_deeply_nested(d: dict) -> bool:
    """Check if a dict has nested dicts/lists as values (not suitable for flat table)."""
    for v in d.values():
        if isinstance(v, (dict, list)):
            return True
    return False


# Friendly labels matching Streamlit's _show_table() output.
_REPORTER_TABLE_LABELS: dict[str, str] = {
    "sampletypes_table": "SampleType Summary",
    "labs_table": "Lab Summary",
    "years_table": "Year Summary",
    "months_table": "Month Summary",
}


def _tables_from_reporter_result(rr: dict[str, Any]) -> list[dict[str, Any]]:
    """Convert reporter_result {key: count} dicts into table dicts.

    Matches Streamlit column names (``key`` / ``count``) and sort order.
    """
    tables: list[dict[str, Any]] = []
    for key, value in rr.items():
        if not key.endswith("_table"):
            continue
        if not isinstance(value, dict) or not value:
            continue
        label = _REPORTER_TABLE_LABELS.get(key, key.replace("_table", "").replace("_", " ").title())
        rows = [{"key": k, "count": v} for k, v in value.items()]
        # Months table sorted chronologically (by key); others by count descending
        if key == "months_table":
            rows.sort(key=lambda r: str(r["key"]))
        else:
            rows.sort(key=lambda r: r["count"], reverse=True)
        tables.append({
            "key": key,
            "label": label,
            "columns": ["key", "count"],
            "data": rows,
        })
    return tables


def build_tables_from_bundle(bundle: dict[str, Any]) -> list[dict[str, Any]]:
    """Build full (non-truncated) table list from a bundle.

    Checks ``report_writer_output`` first, then falls back to
    ``reporter_result`` (summary_sql / Impact mode).  Used by the
    download endpoint to generate xlsx files.
    """
    rwo = bundle.get("report_writer_output")
    if isinstance(rwo, dict):
        tables = flatten_report_to_tables(rwo)
        if tables:
            return tables

    rr = bundle.get("reporter_result")
    if isinstance(rr, dict):
        return _tables_from_reporter_result(rr)

    return []


def _is_geo_report(rwo: dict[str, Any]) -> bool:
    """Check if a report_writer_output is a GEO report."""
    if rwo.get("report_type") == "GEO":
        return True
    # Nested variant: {all_samples: {report_type: "GEO", ...}}
    for v in rwo.values():
        if isinstance(v, dict) and v.get("report_type") == "GEO":
            return True
    return False


def _read_workbook_preview(filepath: str | Path, max_rows: int = 5) -> list[dict[str, Any]]:
    """Read sheet previews from an Excel workbook file on disk.

    Returns a list of {name, columns, data, total_rows} dicts, one per sheet
    that contains data.  Uses sheet-specific row offsets to match the layout
    produced by _populate_geo_seq_workbook():

    - "Metadata" sheet: sample header at row 38, data from row 39 onward.
    - Any other sheet: generic scan — first non-empty row = headers, rows below = data.
    """
    sheets: list[dict[str, Any]] = []
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    except Exception:
        logger.warning("_read_workbook_preview: could not open %s", filepath, exc_info=True)
        return []

    try:
        for ws in wb.worksheets:
            name = ws.title

            if name == "Metadata":
                # Row 38 = sample column headers, row 39+ = sample data rows
                # (per _populate_geo_seq_workbook sample_header_row / sample_start_row)
                header_row_idx = 38
                data_start_idx = 39
                all_rows = list(ws.iter_rows(min_row=header_row_idx, values_only=True))
                if not all_rows:
                    continue
                columns = [str(c) if c is not None else "" for c in all_rows[0]]
                columns = [c for c in columns if c]  # drop trailing empty headers
                if not columns:
                    continue
                data_rows = all_rows[1:]
            else:
                # Generic: first non-empty row = headers
                all_rows = list(ws.iter_rows(values_only=True))
                # Find first row that has at least one non-None value
                header_row_offset = None
                for i, row in enumerate(all_rows):
                    if any(v is not None for v in row):
                        header_row_offset = i
                        break
                if header_row_offset is None:
                    continue
                columns = [str(c) if c is not None else "" for c in all_rows[header_row_offset]]
                columns = [c for c in columns if c]
                if not columns:
                    continue
                data_rows = all_rows[header_row_offset + 1:]

            # Count total non-empty data rows
            total_rows = sum(
                1 for row in data_rows if any(v is not None for v in row)
            )
            if total_rows == 0:
                continue

            # Build preview dicts (up to max_rows)
            preview_data: list[dict[str, Any]] = []
            col_count = len(columns)
            for row in data_rows:
                if len(preview_data) >= max_rows:
                    break
                if not any(v is not None for v in row):
                    continue
                preview_data.append({
                    columns[i]: row[i] if i < len(row) else None
                    for i in range(col_count)
                })

            sheets.append({
                "name": name,
                "columns": columns,
                "data": preview_data,
                "total_rows": total_rows,
            })
    finally:
        wb.close()

    return sheets


def extract_table_artifacts(bundle: dict[str, Any]) -> list[dict[str, Any]]:
    """Build the artifacts list from a bundle dict.

    For reporter bundles: extracts table artifacts from report_writer_output
    and file artifacts from report_saved_files.
    For search bundles: returns empty (search xlsx is generated client-side).
    """
    mode = bundle.get("mode", "")
    if mode not in ("reporter", "report_generation", "sql_report"):
        return []

    artifacts: list[dict[str, Any]] = []

    # Table artifacts from report_writer_output
    rwo = bundle.get("report_writer_output")
    if isinstance(rwo, dict):
        # Detect GEO reports — emit a preview artifact from the actual workbook file
        if _is_geo_report(rwo):
            saved = bundle.get("report_saved_files") or {}
            workbooks = saved.get("geo_seq_workbooks") or []
            workbook_path = workbooks[0] if workbooks else None

            if workbook_path and Path(workbook_path).exists():
                preview_sheets = _read_workbook_preview(workbook_path)
                if preview_sheets:
                    artifacts.append({
                        "artifact_type": "preview",
                        "key": "geo_report_preview",
                        "label": "GEO Report Preview",
                        "sheets": preview_sheets,
                    })
                else:
                    logger.warning(
                        "extract_table_artifacts: workbook at %s had no readable sheets",
                        workbook_path,
                    )
            else:
                logger.warning(
                    "extract_table_artifacts: GEO workbook not found at %r, skipping preview",
                    workbook_path,
                )
        else:
            tables = flatten_report_to_tables(rwo)
            for t in tables:
                data = t["data"]
                total_rows = len(data)
                truncated = total_rows > MAX_INLINE_ROWS
                artifact: dict[str, Any] = {
                    "artifact_type": "table",
                    "key": t["key"],
                    "label": t["label"],
                    "columns": t["columns"],
                    "data": data[:MAX_INLINE_ROWS] if truncated else data,
                }
                if truncated:
                    artifact["truncated"] = True
                    artifact["total_rows"] = total_rows
                artifacts.append(artifact)

    # Table artifacts from reporter_result (summary_sql mode)
    # reporter_result contains {key: count} dicts like sampletypes_table, labs_table, etc.
    if not artifacts:
        rr = bundle.get("reporter_result")
        if isinstance(rr, dict):
            # Add rows_returned metadata if available
            rows_returned = rr.get("rows_returned")
            for table in _tables_from_reporter_result(rr):
                data = table["data"]
                total_rows = len(data)
                truncated = total_rows > MAX_INLINE_ROWS
                artifact: dict[str, Any] = {
                    "artifact_type": "table",
                    "key": table["key"],
                    "label": table["label"],
                    "columns": table["columns"],
                    "data": data[:MAX_INLINE_ROWS] if truncated else data,
                }
                if truncated:
                    artifact["truncated"] = True
                    artifact["total_rows"] = total_rows
                if rows_returned is not None:
                    artifact["rows_returned"] = rows_returned
                artifacts.append(artifact)

    # File artifacts from saved files
    saved = bundle.get("report_saved_files") or {}
    if "geo_seq_workbooks" in saved:
        workbooks = saved["geo_seq_workbooks"]
        if isinstance(workbooks, list) and len(workbooks) > 0:
            artifacts.append({
                "artifact_type": "file",
                "key": "geo_seq_workbooks",
                "label": "GEO Submission Workbook",
                "file_format": "xlsx",
            })

    return artifacts


def generate_table_xlsx(tables: list[dict[str, Any]]) -> bytes:
    """Generate an Excel workbook with one sheet per table.

    Args:
        tables: list of dicts with keys: label, columns, data

    Returns:
        Excel file bytes.
    """
    wb = openpyxl.Workbook()

    if not tables:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Remove default sheet, we'll create named ones
    default_sheet = wb.active
    wb.remove(default_sheet)

    for table in tables:
        label = table.get("label", "Sheet")[:31]  # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=label)
        columns = table.get("columns", [])
        data = table.get("data", [])

        # Header row
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = openpyxl.styles.Font(bold=True)

        # Data rows
        for row_idx, row in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

        # Auto-width columns
        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(str(col_name))
            for row in data:
                val = str(row.get(col_name, ""))
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_search_xlsx(bundle: dict[str, Any]) -> bytes:
    """Generate Excel from a search result bundle.

    Flattens JSON:API response data into a tabular format.
    """
    api_result = bundle.get("api_result_full") or {}
    data_list = api_result.get("data") or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Search Results"

    if not data_list:
        ws.cell(row=1, column=1, value="No results")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Flatten JSON:API: merge id, type, and attributes
    rows: list[dict[str, Any]] = []
    for item in data_list:
        row: dict[str, Any] = {}
        if "id" in item:
            row["id"] = item["id"]
        if "type" in item:
            row["type"] = item["type"]
        attrs = item.get("attributes") or {}
        for k, v in attrs.items():
            # Skip deeply nested values
            if not isinstance(v, (dict, list)):
                row[k] = v
        rows.append(row)

    if not rows:
        ws.cell(row=1, column=1, value="No results")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    columns = list(rows[0].keys())

    # Header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = openpyxl.styles.Font(bold=True)

    # Data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
