"""Render a NExtSEEK 4-sheet upload workbook from CC-composed reingest rows.

One workbook == one SampleType (batch_upload/convert.py rejects INSTRUCTIONS that
declare more than one SampleType). The output parses cleanly back through
``nextseek_api.batch_upload.convert.parse_traditional_file`` — that round-trip is
the correctness oracle (see tests/test_upload_workbook.py).

Row shape (one per output sample):
    {"json_metadata": {<attr>: <value>, ...}, "assay_ids": [int, ...]}
UID is intentionally omitted (every reingested sample is [NEW]; the server mints
the UID on upload). ``json_metadata`` carries Parent / Scientist / File_PrimaryData
and any other attributes CC derived.
"""
from __future__ import annotations

import json
from typing import Any

import openpyxl


def render_upload_workbook(sample_type: str, rows: list[dict[str, Any]], out_path: str) -> None:
    """Write a 4-sheet upload workbook for one ``sample_type`` to ``out_path``.

    Emits all four sheets (Instructions / Samples / Assay / Ontology); a missing
    sheet would make convert.py silently fall back to the flat parser.
    """
    if not rows:
        raise ValueError("render_upload_workbook: no rows to render")

    # Field set = union of json_metadata keys across rows, first-seen order stable.
    fields: list[str] = []
    for row in rows:
        for key in (row.get("json_metadata") or {}):
            if key not in fields:
                fields.append(key)
    if not fields:
        raise ValueError("render_upload_workbook: rows carry no json_metadata")

    wb = openpyxl.Workbook()

    # INSTRUCTIONS: one row per Field. Database Field MUST be "SampleType::Attr"
    # (the "::" is required by InstructionRow validation). Field == attribute name.
    wi = wb.active
    wi.title = "Instructions"
    wi.append(["Field", "Database Field", "Field Type", "Ontology"])
    for field in fields:
        wi.append([field, f"{sample_type}::{field}", "Text", None])

    # SAMPLES: one column per Field (headers must equal the Field strings exactly —
    # convert.py matches trim-only, case-sensitive). No UID column: UID stays blank.
    ws = wb.create_sheet("Samples")
    ws.append(list(fields))
    for row in rows:
        meta = row.get("json_metadata") or {}
        ws.append([_cell(meta.get(field)) for field in fields])

    # ASSAY: one row per distinct assay_id, applied to this sample type (Direction 1).
    wa = wb.create_sheet("Assay")
    wa.append(["SampleType", "AssayType", "Assay", "Direction"])
    seen: set[int] = set()
    for row in rows:
        for assay_id in (row.get("assay_ids") or []):
            if assay_id not in seen:
                seen.add(assay_id)
                wa.append([sample_type, None, assay_id, 1])

    # ONTOLOGY: present but empty — reingest fields are free-text, no controlled vocab.
    wo = wb.create_sheet("Ontology")
    wo.append([])

    wb.save(out_path)


def _cell(value: Any) -> Any:
    """Serialize a metadata value into an xlsx cell (nested dict/list -> compact JSON)."""
    if isinstance(value, (dict, list)):
        return json.dumps(value, separators=(",", ":"))
    return value
