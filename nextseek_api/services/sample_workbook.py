"""The single writer for sample-download workbooks.

Every sample download in NExtSEEK ends here, so the README sheet cannot drift
between the legacy `seek` views and the `nextseek_api` endpoint. See
docs/sample-download-workflow.md for how the call paths converge on this module.
"""

from __future__ import annotations

import json
import logging
import math
from pathlib import Path
from typing import Iterable, Mapping

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from django.conf import settings
from neo4j import GraphDatabase
from openpyxl.styles import Font

from seek.dbtable_sampleattribute import DBtable_sampleattribute
from seek.models import (
    Assay_assets,
    Assays,
    Sample_attributes_unique,
    Sample_types_context,
    Samples,
)

from nextseek_api.services.sample_provenance import (
    SAMPLE_TYPE_RE,
    build_provenance_rows,
    derivation_edges,
    sample_type_depths,
)

logger = logging.getLogger(__name__)

CONTEXTDB_URL = (
    "https://github.com/BioMicroCenter/NExtSEEK/blob/main/"
    "chat_nextseek/src/chat_nextseek/context/sampletypes_db.json"
)

README_SHEET = "README"
README_LINK_TEXT = "Sample type definitions: sampletypes_db.json (GitHub)"


COLUMN_TABLE_HEADER = ["Column", "Meaning"]
# The template workbook marks which columns upload validation requires. Sample
# downloads carry real data whose required-ness is already settled, so they keep
# the two-column table and this header is unused there.
REQUIRED_TABLE_HEADER = ["Column", "Required", "Meaning"]
SUMMARY_HEADER = ["Sample Type", "Name", "Description"]
FLOW_SHEET = "How this data flowed"
FLOW_README_POINTER = f"How this data flowed: see the '{FLOW_SHEET}' sheet."
# Sample-type codes are short; assay titles are not. Alternating widths keep a
# chain readable without a 100-wide gap at every second type.
FLOW_TYPE_WIDTH = 14
FLOW_ARROW_WIDTH = 34
# A download must not wait on the graph. Provenance is worth a moment, never a
# stalled request.
NEO4J_TIMEOUT_SECONDS = 5
# SEEK suffixes assay titles by how the data is attached; the distinction is
# about SEEK bookkeeping, not about what was done to the sample.
ASSAY_TITLE_SUFFIXES = (" - Metadata", " - Data Linked", " - Data Attached")

# Hover-note geometry. openpyxl sizes comment boxes in pixels; these fit roughly
# 40 words without the reader having to drag the box open.
COMMENT_AUTHOR = "NExtSEEK"
COMMENT_WIDTH = 340
COMMENT_HEIGHT = 130

CV_SHEET = "Controlled Vocabularies"
CV_PATH = Path(__file__).with_name("controlled_vocabularies.json")

MANIFEST_SHEET = "_NEXTSEEK"
MANIFEST_HEADER = ["sheet", "code", "attribute", "database_field", "required"]
# Bumped only when the manifest's shape changes, so the part-2 converter can
# refuse a workbook it does not understand instead of misreading it.
TEMPLATE_FORMAT_VERSION = 1
# Same delimiter the INSTRUCTIONS sheet uses (seek/dbtable_sample.py:168). The
# manifest stores the finished string so the converter concatenates nothing.
MANIFEST_DBFIELD_DELIMITER = "::"
EMPTY_TYPE_NOTE = "This sample type has no attributes defined in SEEK."
REQUIRED_HEADER_MARK = "*"

# How far a dropdown reaches below the last filled row. A download is a
# starting point, not a finished sheet: a researcher adding samples must keep
# the dropdown. Full-column validation would do it too, but some tools slow
# noticeably with it applied across many columns.
DROPDOWN_SPARE_ROWS = 500


def _load_vocabularies() -> tuple[dict[str, str], dict[str, list[str]]]:
    """(column -> vocabulary name, vocabulary name -> terms).

    Fail soft, like every other lookup here: a missing or malformed file costs
    the dropdowns, never the download.
    """
    try:
        doc = json.loads(CV_PATH.read_text())
        return doc["field_map"], doc["vocabularies"]
    except Exception:
        logger.exception("controlled_vocabularies.json unreadable; no dropdowns")
        return {}, {}


def _write_vocabulary_sheet(book, needed: list[str], vocabularies) -> dict[str, str]:
    """Park each needed vocabulary in its own column and return {name: range}.

    The terms need a real sheet rather than an inline list: Excel caps an inline
    dropdown formula at 255 characters and instrument_model alone is 84 terms.
    """
    if not needed:
        return {}
    ws = book.create_sheet(CV_SHEET)
    ranges = {}
    for index, name in enumerate(sorted(needed), start=1):
        letter = get_column_letter(index)
        terms = vocabularies.get(name, [])
        ws.cell(row=1, column=index, value=name).font = Font(bold=True)
        for offset, term in enumerate(terms, start=2):
            ws.cell(row=offset, column=index, value=_safe_cell_value(term))
        ranges[name] = f"'{CV_SHEET}'!${letter}$2:${letter}${len(terms) + 1}"
        ws.column_dimensions[letter].width = 30
    return ranges

# Excel's hard per-cell limit. openpyxl does not enforce it: a longer value is
# written happily and Excel then reports the file as needing repair. `meaning`
# is a TEXT column, so 65,535 characters is reachable from the definitions table.
EXCEL_MAX_CELL_CHARS = 32767


def build_readme_blocks(
    sheets: Iterable[tuple[str, Iterable[str]]],
    context_by_code: Mapping[str, Mapping[str, str]],
    meaning_by_pair: Mapping[tuple[str, str], str],
    required_by_pair: Mapping[tuple[str, str], bool] | None = None,
    relationships_by_code: Mapping[str, Mapping[str, list]] | None = None,
) -> list[dict]:
    """One block per sheet, in the order the sheets will be written.

    `sheets` is (sample_type code, its columns in sheet order). Columns are kept
    in that order rather than sorted so the README can be read beside the tab.
    An undocumented sample type still gets a block, and a column with no
    definition is still listed, so the README always indexes the whole workbook.

    `required_by_pair` and `relationships_by_code` serve the template workbook
    and are optional. They add *separate* keys rather than widening `columns`,
    because the sample-download path and its tests depend on `columns` being
    (name, meaning) 2-tuples. Omit both and the output is what it has always
    been.
    """
    blocks = []
    for code, columns in sheets:
        entry = context_by_code.get(code) or {}
        ordered = list(columns)
        block = {
            "code": code,
            "name": entry.get("name", "") or "",
            "description": entry.get("description", "") or "",
            "columns": [
                (column, meaning_by_pair.get((code, column), "") or "")
                for column in ordered
            ],
        }
        if required_by_pair is not None:
            block["required"] = [
                bool(required_by_pair.get((code, column), False)) for column in ordered
            ]
        if relationships_by_code:
            found = relationships_by_code.get(code)
            if found:
                block["relationships"] = found
        blocks.append(block)
    return blocks


def load_sample_type_context(codes: Iterable[str]) -> dict[str, dict[str, str]]:
    """Look up code -> {name, description} in dmac.sample_types_context.

    Joins on the `sample_type` code string, not `sampletype_id`: the id column
    does not agree with `sample_types.id` across instances.
    """
    wanted = sorted({c for c in codes if c})
    if not wanted:
        return {}
    try:
        rows = Sample_types_context.objects.filter(sample_type__in=wanted).values(
            "sample_type", "name", "description"
        )
        return {
            r["sample_type"]: {
                "name": r.get("name") or "",
                "description": r.get("description") or "",
            }
            for r in rows
        }
    except Exception:
        # A missing or unreachable context table must not cost the user their
        # download; the README then lists codes with blank name/description.
        logger.exception("sample_types_context lookup failed; README will be unpopulated")
        return {}


def load_sample_field_context(
    pairs: Iterable[tuple[str, str]],
) -> dict[tuple[str, str], str]:
    """Resolve (sample_type, field_name) -> meaning against sample_attributes_unique.

    Precedence per pair: a row scoped to that sample type, else the global row
    (`sample_type == ''`), else blank. Resolving here means no caller has to
    reimplement the fallback.
    """
    wanted = [(st or "", fn) for st, fn in pairs if fn]
    if not wanted:
        return {}
    try:
        rows = list(
            Sample_attributes_unique.objects.filter(
                field_name__in=sorted({fn for _, fn in wanted})
            ).values("field_name", "sample_type", "meaning")
        )
    except Exception:
        # A missing or unreachable table must not cost the user their download;
        # every meaning then renders blank.
        logger.exception("sample_attributes_unique lookup failed; meanings will be blank")
        return {}

    global_by_field: dict[str, str] = {}
    scoped: dict[tuple[str, str], str] = {}
    for row in rows:
        meaning = row.get("meaning") or ""
        code = row.get("sample_type") or ""
        name = row.get("field_name")
        if code:
            scoped[(code, name)] = meaning
        else:
            global_by_field[name] = meaning

    return {
        (code, name): scoped.get((code, name), global_by_field.get(name, ""))
        for code, name in wanted
    }


def load_derivation_hops(uuids: Iterable[str]) -> list[tuple[str, str, str]]:
    """(parent uuid, assay title, child uuid) for the hops among these samples.

    Neo4j is the authority here, not SEEK. Its DERIVED_FROM relationship holds
    `internal_assay_title` on the *edge*, so the assay is recorded between two
    specific UIDs. SEEK's assay_assets only says which assay a sample belongs
    to, which cannot distinguish the hop that produced it from any other assay
    the same sample took part in.

    Fail-soft like every other lookup here: an unreachable graph costs the
    section, never the download -- and it must cost it *quickly*. The driver's
    default connection timeout is long enough that a graph which is merely slow
    would stall every download, so it is bounded explicitly.
    """
    wanted = sorted({u for u in uuids if u})
    if not wanted:
        return []
    query = (
        "MATCH (c:Sample)-[r:DERIVED_FROM]->(p:Sample) "
        "WHERE c.uuid IN $uuids "
        "RETURN p.uuid AS parent, r.internal_assay_title AS assay, c.uuid AS child"
    )
    try:
        # Read at call time, not import: touching settings during module import
        # pulls in the whole runtime config chain before Django is ready.
        config = settings.NEO4J_DATABASE
        with GraphDatabase.driver(
            config["URI"],
            auth=config["AUTH"],
            connection_timeout=NEO4J_TIMEOUT_SECONDS,
            max_transaction_retry_time=NEO4J_TIMEOUT_SECONDS,
        ) as driver:
            records = driver.execute_query(query, uuids=wanted).records
        return [
            (r["parent"], (r["assay"] or "").strip(), r["child"])
            for r in records
            if r["parent"] and r["child"]
        ]
    except Exception:
        logger.exception("neo4j lineage lookup failed; provenance falls back to Parent")
        return []


def load_assay_titles(uuids: Iterable[str]) -> dict[str, str]:
    """uuid -> the assay it belongs to, from SEEK's own sample/assay links.

    Only a fallback: used when Neo4j is unreachable, where a per-sample assay
    is better than no label at all. See load_derivation_hops for why the graph
    edge is the more accurate source.
    """
    wanted = sorted({u for u in uuids if u})
    if not wanted:
        return {}
    try:
        ids = dict(Samples.objects.filter(uuid__in=wanted).values_list("id", "uuid"))
        if not ids:
            return {}
        links = Assay_assets.objects.filter(
            asset_type="Sample", asset_id__in=list(ids)
        ).values_list("asset_id", "assay_id")
        titles = dict(
            Assays.objects.filter(id__in={a for _, a in links}).values_list("id", "title")
        )
    except Exception:
        logger.exception("assay lookup failed; provenance will show no assay names")
        return {}

    out: dict[str, str] = {}
    for asset_id, assay_id in links:
        title = (titles.get(assay_id) or "").strip()
        for suffix in ASSAY_TITLE_SUFFIXES:
            if title.endswith(suffix):
                title = title[: -len(suffix)]
                break
        if title and asset_id in ids:
            out.setdefault(ids[asset_id], title)
    return out


def _safe_cell_value(value: str) -> str:
    """Make a string safe to hand to openpyxl.

    Two failure modes, both fatal to the whole download and neither caught by
    the try/except around the *query*:
    - a control character (e.g. \\x0b, what a pasted Word/PDF line break becomes)
      makes `ws.cell(...)` raise IllegalCharacterError
    - a value over Excel's cell limit writes a file Excel calls corrupt

    Definitions are reviewer-authored prose, so both are reachable. Losing a
    stray character or a tail of an absurdly long definition is always better
    than losing the workbook.
    """
    return ILLEGAL_CHARACTERS_RE.sub("", value or "")[:EXCEL_MAX_CELL_CHARS]


def _write_cell(ws, row: int, column: int, value: str, bold: bool = False):
    """The single place the README writes text, so sanitizing cannot be skipped."""
    cell = ws.cell(row=row, column=column, value=_safe_cell_value(value))
    if bold:
        cell.font = Font(bold=True)
    return cell


def _write_readme(book, blocks: list[dict], *, has_flow_sheet: bool) -> None:
    ws = book.create_sheet(README_SHEET, 0)
    _write_cell(ws, 1, 1, README_LINK_TEXT)
    ws["A1"].hyperlink = CONTEXTDB_URL
    ws["A1"].style = "Hyperlink"
    # Row 2 is left blank to separate the link from the summary table.
    #
    # Sheet order: every sample type is summarised first, so a reader sees what
    # the workbook contains before meeting any column detail. The per-tab
    # column tables follow underneath.
    row = 3
    for column, label in enumerate(SUMMARY_HEADER, start=1):
        _write_cell(ws, row, column, label, bold=True)
    row += 1
    for block in blocks:
        _write_cell(ws, row, 1, block["code"])
        _write_cell(ws, row, 2, block["name"])
        _write_cell(ws, row, 3, block["description"])
        row += 1
    row += 1  # blank line between the summary table and what follows

    # The flow lives on its own sheet; the README says so, because a reader who
    # never opens the tab must still learn it is there.
    if has_flow_sheet:
        _write_cell(ws, row, 1, FLOW_README_POINTER, bold=True)
        row += 2

    for block in blocks:
        heading = f"{block['code']} — {block['name']}" if block["name"] else block["code"]
        _write_cell(ws, row, 1, heading, bold=True)
        row += 1

        # Relationships sit under the heading, before the column table: they say
        # where the sheet belongs in a pipeline, which frames everything below.
        related = block.get("relationships")
        if related:
            if related.get("parents"):
                _write_cell(ws, row, 2,
                            "Typically derived from: " + ", ".join(related["parents"]))
                row += 1
            if related.get("children"):
                _write_cell(ws, row, 2,
                            "Typically feeds into: " + ", ".join(related["children"]))
                row += 1
        row += 1  # blank line before the column table

        # A block whose columns all dropped out gets no table header: a bare
        # Column/Meaning row with nothing under it reads as a rendering bug.
        if block["columns"]:
            flags = block.get("required")
            header = REQUIRED_TABLE_HEADER if flags is not None else COLUMN_TABLE_HEADER
            for column, label in enumerate(header, start=2):
                _write_cell(ws, row, column, label, bold=True)
            row += 1
            for index, (name, meaning) in enumerate(block["columns"]):
                _write_cell(ws, row, 2, name)
                if flags is not None:
                    _write_cell(ws, row, 3, "Yes" if flags[index] else "")
                    _write_cell(ws, row, 4, meaning)
                else:
                    _write_cell(ws, row, 3, meaning)
                row += 1
        row += 1  # blank line between sections
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 100
    if any(b.get("required") is not None for b in blocks):
        # The meaning moved one column right, so widths shift with it.
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 100


def _write_flow_sheet(book, rows: list[list[str]]) -> None:
    """One chain per row, alternating type and arrow cells.

    Its own sheet rather than a README section: README's columns are sized 46 /
    34 / 100 for the summary and column tables, which puts a 100-wide gap at
    every second type of a chain.
    """
    if not rows:
        return
    ws = book.create_sheet(FLOW_SHEET, 1)
    for index, row in enumerate(rows, start=1):
        for column, value in enumerate(row, start=1):
            _write_cell(ws, index, column, value, bold=(column % 2 == 1))
    widest = max(len(row) for row in rows)
    for column in range(1, widest + 1):
        ws.column_dimensions[get_column_letter(column)].width = (
            FLOW_TYPE_WIDTH if column % 2 else FLOW_ARROW_WIDTH
        )


def _annotate_header(ws, code: str, columns: list[str], meaning_by_pair) -> None:
    """Attach each column's definition as a hover note on its header cell.

    A researcher filling the sheet in reads the header, not the README, so the
    definition is put where the question is asked. Columns with no definition
    get no note rather than an empty one.
    """
    for index, column in enumerate(columns, start=1):
        meaning = meaning_by_pair.get((code, column), "")
        if not meaning:
            continue
        note = Comment(_safe_cell_value(meaning), COMMENT_AUTHOR)
        note.width = COMMENT_WIDTH
        note.height = COMMENT_HEIGHT
        ws.cell(row=1, column=index).comment = note


def _apply_dropdowns(ws, columns: list[str], field_map, ranges, row_count: int) -> None:
    """Offer each governed column its repository vocabulary as a dropdown.

    errorStyle is 'warning', not 'stop': downloaded data already contains values
    that predate these vocabularies (RNA-seq for RNA-Seq, Paired End for
    paired), and a hard reject would fire on open for rows the researcher did
    not touch. The dropdown guides; it does not overrule what is already there.

    The range runs DROPDOWN_SPARE_ROWS past the last filled row, so the
    dropdown survives a researcher adding samples underneath.
    """
    for index, column in enumerate(columns, start=1):
        vocabulary = field_map.get(column)
        if not vocabulary or vocabulary not in ranges:
            continue
        rule = DataValidation(
            type="list", formula1=ranges[vocabulary],
            allow_blank=True, showDropDown=False, errorStyle="warning",
        )
        rule.error = f"Not a {vocabulary} term. Pick from the list, or keep this value."
        rule.errorTitle = "Outside the controlled vocabulary"
        ws.add_data_validation(rule)
        letter = get_column_letter(index)
        rule.add(f"{letter}2:{letter}{max(row_count + 1, 2) + DROPDOWN_SPARE_ROWS}")


def write_samples_workbook(parsed_df, output_path, context_by_code=None) -> None:
    """Write README as sheet 1, then one sheet per sample type.

    `parsed_df` must carry a `uuid` column; `sample_type` is derived here so the
    extraction regex lives in exactly one place. Sheets are prepared before the
    README is built, because the README must describe the columns that survive
    the all-empty drop rather than everything in the frame.
    """
    df = parsed_df.copy()
    df["sample_type"] = df["uuid"].astype(str).str.extract(SAMPLE_TYPE_RE, expand=False)

    codes = df["sample_type"].dropna().unique().tolist()
    if context_by_code is None:
        context_by_code = load_sample_type_context(codes)

    # Lineage is loaded before the sheets are prepared, because the sheet order
    # is derived from it. Same single bounded query, just earlier.
    uuids = df["uuid"].astype(str)
    hops = load_derivation_hops(uuids)
    edges = derivation_edges(df, {} if hops else load_assay_titles(uuids), hops)
    depths = sample_type_depths(edges)
    flow_rows = build_provenance_rows(edges, depths)

    prepared = []
    for sample_type, sample_type_df in df.groupby("sample_type"):
        frame = sample_type_df.drop(columns=["uuid", "sample_type"])
        frame = frame.replace("", pd.NA)
        frame = frame.dropna(axis=1, how="all")
        prepared.append((sample_type, frame))

    # Generation order, not alphabetical: a reader meets the sample types in
    # the order they were made. A type with no hop at all cannot be placed in
    # the pipeline, so it sorts after everything that can be. No lineage at all
    # leaves every depth infinite, which is a stable alphabetical sort.
    prepared.sort(key=lambda item: (depths.get(item[0], math.inf), item[0]))

    sheets = [(code, list(frame.columns)) for code, frame in prepared]
    meaning_by_pair = load_sample_field_context(
        [(code, column) for code, columns in sheets for column in columns]
    )
    blocks = build_readme_blocks(sheets, context_by_code, meaning_by_pair)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        book = writer.book
        # pandas removes openpyxl's default sheet, but guard in case that changes.
        if "Sheet" in book.sheetnames:
            del book["Sheet"]
        _write_readme(book, blocks, has_flow_sheet=bool(flow_rows))
        _write_flow_sheet(book, flow_rows)

        field_map, vocabularies = _load_vocabularies()
        needed = sorted({
            field_map[column]
            for _, columns in sheets for column in columns
            if column in field_map and field_map[column] in vocabularies
        })
        ranges = _write_vocabulary_sheet(book, needed, vocabularies)

        for code, frame in prepared:
            frame.to_excel(writer, sheet_name=code, index=False)
            sheet = writer.sheets[code]
            _annotate_header(sheet, code, list(frame.columns), meaning_by_pair)
            _apply_dropdowns(sheet, list(frame.columns), field_map, ranges, len(frame))


def load_relationships(codes, known):
    """Thin re-export so the writer has one patchable seam for relationships.

    Imported lazily: template_catalog imports load_sample_type_context from this
    module, so a top-level import here would be circular.
    """
    from nextseek_api.services.template_catalog import load_relationships as _impl

    return _impl(codes, known)


def _write_manifest(book, rows: list[list]) -> None:
    """The hidden machine-readable map of the workbook, for the part-2 converter.

    Hidden rather than absent so a researcher never has to look at it, and
    hidden rather than deleted-on-open so renaming a tab cannot orphan the
    mapping: `sheet` records where each column actually lives.
    """
    ws = book.create_sheet(MANIFEST_SHEET)
    _write_cell(ws, 1, 1, "format_version", bold=True)
    ws.cell(row=1, column=2, value=TEMPLATE_FORMAT_VERSION)
    for index, label in enumerate(MANIFEST_HEADER, start=1):
        _write_cell(ws, 2, index, label, bold=True)
    for offset, row in enumerate(rows, start=3):
        for index, value in enumerate(row, start=1):
            if isinstance(value, int):
                ws.cell(row=offset, column=index, value=value)
            else:
                _write_cell(ws, offset, index, value)
    ws.sheet_state = "hidden"


def write_template_workbook(entries, output_path) -> None:
    """Write a blank upload template: README, a headers-only sheet per type,
    then the hidden manifest.

    The same artifact `write_samples_workbook` produces, minus the data rows and
    the provenance sheet -- which is why it lives here and shares every helper.
    A blank template has no lineage, so sheets follow the order the user picked
    rather than derivation depth.
    """
    codes = [e.code for e in entries]

    try:
        specs_by_id = DBtable_sampleattribute().getAttributeSpecsBySampleTypeIds(
            [e.sample_type_id for e in entries]
        )
    except Exception:
        # Columns are this workbook's whole point, so losing them is not a soft
        # failure for the affected type -- but it must not cost the other types
        # their sheets. Every type is skipped only if every lookup failed.
        logger.exception("attribute lookup failed; affected types are skipped")
        specs_by_id = {}

    prepared = []
    for entry in entries:
        specs = specs_by_id.get(entry.sample_type_id)
        if specs is None:
            continue
        prepared.append((entry, specs))

    sheets = [(e.code, [s["title"] for s in specs]) for e, specs in prepared]
    pairs = [(code, title) for code, titles in sheets for title in titles]
    meaning_by_pair = load_sample_field_context(pairs)
    required_by_pair = {
        (e.code, s["title"]): s["required"] for e, specs in prepared for s in specs
    }

    known = set(codes)
    relationships = load_relationships(codes, known)

    context_by_code = {
        e.code: {"name": e.name, "description": e.description} for e in entries
    }
    blocks = build_readme_blocks(
        sheets, context_by_code, meaning_by_pair,
        required_by_pair=required_by_pair,
        relationships_by_code=relationships,
    )

    book = Workbook()
    if "Sheet" in book.sheetnames:
        del book["Sheet"]

    _write_readme(book, blocks, has_flow_sheet=False)

    # _write_vocabulary_sheet must run before the type-sheet loop below: it
    # returns the `ranges` dict _apply_dropdowns needs while writing each type
    # sheet. But per the design doc the vocabulary sheet belongs AFTER the type
    # sheets, immediately before the manifest. So it is created here, then
    # repositioned once the type sheets exist (only when one was actually
    # created -- with nothing to govern, _write_vocabulary_sheet returns {}
    # and creates no sheet to move).
    field_map, vocabularies = _load_vocabularies()
    needed = sorted({
        field_map[title]
        for _, titles in sheets for title in titles
        if title in field_map and field_map[title] in vocabularies
    })
    ranges = _write_vocabulary_sheet(book, needed, vocabularies)

    manifest_rows = []
    type_sheet_count = 0
    for entry, specs in prepared:
        ws = book.create_sheet(entry.code)
        type_sheet_count += 1
        if not specs:
            _write_cell(ws, 1, 1, EMPTY_TYPE_NOTE)
            ws.column_dimensions["A"].width = 60
            continue

        titles = [s["title"] for s in specs]
        for index, spec in enumerate(specs, start=1):
            label = spec["title"] + (REQUIRED_HEADER_MARK if spec["required"] else "")
            cell = ws.cell(row=1, column=index, value=_safe_cell_value(label))
            cell.font = Font(bold=bool(spec["required"]))
            ws.column_dimensions[get_column_letter(index)].width = max(
                14, min(len(label) + 4, 40)
            )
            manifest_rows.append([
                entry.code,
                entry.code,
                spec["title"],
                f"{entry.code}{MANIFEST_DBFIELD_DELIMITER}{spec['title']}",
                1 if spec["required"] else 0,
            ])

        # The header text now carries the required marker, so notes and
        # dropdowns are keyed on the bare titles the lookups know.
        _annotate_header_titles(ws, entry.code, titles, meaning_by_pair)
        _apply_dropdowns(ws, titles, field_map, ranges, 0)

    if ranges:
        # Right after creation the book is [README, Controlled Vocabularies,
        # <type sheets...>]; moving it forward by the number of type sheets
        # lands it immediately after them (the manifest is appended next).
        # Dropdown formulas reference the sheet by name, so repositioning it
        # does not disturb them.
        book.move_sheet(CV_SHEET, offset=type_sheet_count)

    _write_manifest(book, manifest_rows)
    book.save(output_path)


def _annotate_header_titles(ws, code: str, titles: list[str], meaning_by_pair) -> None:
    """_annotate_header, but keyed on bare titles while the cells show markers.

    The template's header cells read "UID*", so the shared helper's
    read-the-cell approach would miss every required column's definition.
    """
    for index, title in enumerate(titles, start=1):
        meaning = meaning_by_pair.get((code, title), "")
        if not meaning:
            continue
        note = Comment(_safe_cell_value(meaning), COMMENT_AUTHOR)
        note.width = COMMENT_WIDTH
        note.height = COMMENT_HEIGHT
        ws.cell(row=1, column=index).comment = note
