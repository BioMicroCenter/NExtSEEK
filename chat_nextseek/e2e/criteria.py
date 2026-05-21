"""Pass-criterion DSL: resolve fields from debug/session/reply + evaluate ops.

Ported from smart_test._resolve_field_v2 + _check_criterion (lines 98-438).
Adds api_artifact.* field family for filesystem-level checks.
"""
from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from e2e.catalog import PassCriterion


def _strip_html(text: Any) -> str:
    if text is None:
        return ""
    return re.sub(r"<[^>]+>", "", str(text))


def _codes(items: list[dict], key: str = "code") -> list:
    return [item.get(key) for item in items if item.get(key)]


def _api_artifact_resolve(field: str, run_root: Path | None) -> Any:
    """Handle api_artifact.<filename>[.rows_gte] resolution."""
    if run_root is None:
        return None
    sub = field[len("api_artifact."):]
    # Detect ".rows_gte" suffix
    if sub.endswith(".rows_gte"):
        fname = sub[:-len(".rows_gte")]
        target = run_root / "files" / fname
        if not target.exists():
            return 0
        if target.suffix == ".csv":
            with open(target, newline="", encoding="utf-8") as fh:
                rows = list(csv.reader(fh))
            return max(0, len(rows) - 1)  # exclude header
        if target.suffix in (".xlsx", ".xls"):
            from openpyxl import load_workbook  # noqa: PLC0415
            wb = load_workbook(target, read_only=True)
            ws = wb.active
            count = sum(1 for _ in ws.iter_rows(min_row=2)) if ws else 0
            wb.close()
            return count
        # Unknown extension: count non-blank lines
        return sum(1 for line in target.read_text().splitlines() if line.strip())
    # Existence-only check
    target = run_root / "files" / sub
    return target.exists()


def resolve_field(
    debug: dict,
    field: str,
    *,
    session: Any = None,
    last_reply: str | None = None,
    run_root: Path | None = None,
    browser_ctx: dict | None = None,
    mysql_chat_log: list[dict] | None = None,
) -> Any:
    """Resolve a pass-criterion field from debug dict + optional session/reply/run_root.

    Field families:
    - api_artifact.<filename>[.rows_gte] — filesystem checks against run_root/files/
    - wizard.*, chat_log.*, pipeline_agent.* — session-state aliases
    - ui_text.* — browser DOM checks via browser_ctx["chat_page"]
    - mysql_chat_log.* — MySQL chat_log row aliases
    - last_reply, last_target_result_id — recent-turn aliases
    - entity_*, api_ok, graph_cypher, neo4j_ok — top-level run_query debug aliases
    - dot-notation fallback — navigate nested dicts
    """
    # ── browser-tier resolvers ──────────────────────────────────────────
    if field.startswith("ui_text.") and browser_ctx is not None:
        chat_page = browser_ctx.get("chat_page")
        downloaded = browser_ctx.get("downloaded_artifacts") or set()
        sub = field[len("ui_text."):]
        if sub == "assistant_reply":
            return chat_page.latest_assistant_reply() if chat_page else ""
        if sub == "bubble_count":
            return chat_page.bubble_count() if chat_page else 0
        if sub.startswith("artifacts."):
            art_sub = sub[len("artifacts."):]
            if art_sub.endswith(".downloaded"):
                fname = art_sub[:-len(".downloaded")]
                return fname in downloaded
            return chat_page.has_artifact(art_sub) if chat_page else False
        if sub.startswith("stepper."):
            step_name = sub[len("stepper."):]
            return chat_page.stepper_status(step_name) if chat_page else None
        return None

    if field.startswith("mysql_chat_log.") and mysql_chat_log is not None:
        sub = field[len("mysql_chat_log."):]
        if sub == "length":
            return len(mysql_chat_log)
        if not mysql_chat_log:
            return None
        latest = mysql_chat_log[-1]
        if sub == "latest_reply":
            return latest.get("assistant_reply")
        if sub == "latest_mode":
            return latest.get("mode")
        if sub == "latest_query":
            return latest.get("user_query")
        return None

    if field.startswith("api_artifact."):
        return _api_artifact_resolve(field, run_root)

    # ── session-state aliases ────────────────────────────────────────────
    if field.startswith("wizard.") and session is not None:
        wz = session.get("nfcore_wizard") or {}
        sub = field[len("wizard."):]
        if sub == "active":   return bool(wz.get("active"))
        if sub == "step":     return wz.get("step")
        if sub == "pipeline": return wz.get("pipeline")
        if sub == "uid_count":
            return len(((wz.get("selection") or {}).get("uids")) or [])
        if sub == "cohort_count":
            cc = (wz.get("selection") or {}).get("cohort_criteria")
            return 1 if cc is None else (len(cc) or 1)
        if sub == "enrichment_count":
            return len(((wz.get("selection") or {}).get("enrichment_fields")) or [])
        if sub == "metadata_summary_present":
            return (wz.get("pinned_context") or {}).get("metadata_summary") is not None
        # generic dot-navigation fallback inside wizard
        wz_val: Any = wz
        for part in sub.split("."):
            if not isinstance(wz_val, dict):
                return None
            wz_val = wz_val.get(part)
        return wz_val

    if field.startswith("pipeline_agent.") and session is not None:
        pa = session.get("pipeline_agent") or {}
        sub = field[len("pipeline_agent."):]
        if sub == "active":
            return bool(session.get("pipeline_agent"))
        if sub == "phase":
            return pa.get("phase")
        # dot-nav fallback
        pa_val: Any = pa
        for part in sub.split("."):
            if not isinstance(pa_val, dict):
                return None
            pa_val = pa_val.get(part)
        return pa_val

    if field.startswith("chat_log.") and session is not None:
        log = session.get("chat_log") or []
        sub = field[len("chat_log."):]
        if sub == "length":        return len(log)
        if sub == "latest_mode":   return log[-1].get("mode") if log else None
        if sub == "latest_bundle_id": return log[-1].get("bundle_id") if log else None
        if log:
            log_val: Any = log[-1]
            for part in sub.split("."):
                if not isinstance(log_val, dict):
                    return None
                log_val = log_val.get(part)
            return log_val
        return None

    if field == "results_history.length" and session is not None:
        return len(session.get("results_history") or [])

    if field == "last_reply":
        return _strip_html(last_reply)

    if field == "last_target_result_id":
        pp = debug.get("parser_plan") or {}
        return pp.get("target_result_id")

    # ── top-level run_query debug aliases ────────────────────────────────
    aliases = {
        "entity_sampletype_codes": lambda d: _codes((d.get("entity_result") or {}).get("sampletypes", [])),
        "entity_assay_codes":      lambda d: _codes((d.get("entity_result") or {}).get("assays", [])),
        "shortlist_sampletype_codes": lambda d: d.get("shortlist_sampletype_codes") or [],
        "shortlist_assay_codes":      lambda d: d.get("shortlist_assay_codes") or [],
        "entity_projects":         lambda d: (d.get("entity_result") or {}).get("projects", []),
        "graph_cypher":            lambda d: (d.get("graph_plan") or {}).get("cypher") or "",
        "neo4j_ok":                lambda d: (d.get("graph_result") or {}).get("ok", False),
        "api_ok":                  lambda d: (d.get("api_result_meta") or {}).get("ok", False),
    }
    if field in aliases:
        return aliases[field](debug)

    # ── dot-notation fallback ────────────────────────────────────────────
    parts = field.split(".")
    dot_val: Any = debug
    aliases_per_part = {"sample_type": ("sampletype",), "sampletype": ("sample_type",)}
    for part in parts:
        if not isinstance(dot_val, dict):
            return None
        if part in dot_val:
            dot_val = dot_val[part]
            continue
        # alias fallback
        found = False
        for alt in aliases_per_part.get(part, ()):
            if alt in dot_val:
                dot_val = dot_val[alt]
                found = True
                break
        if not found:
            return None
    return dot_val


def _check_one(actual: Any, op: str, expected: Any) -> tuple[bool, str]:
    if op == "eq":
        passed = actual == expected
        return passed, (f"== {actual!r}" if passed else f"expected {expected!r}, got {actual!r}")
    if op == "contains":
        if isinstance(actual, list):
            passed = expected in actual
        elif isinstance(actual, str):
            passed = bool(expected) and expected in actual
        else:
            passed = False
        return passed, (f"contains {expected!r}" if passed else f"{actual!r} missing {expected!r}")
    if op == "nonempty":
        passed = bool(actual)
        return passed, ("non-empty" if passed else f"empty/None ({actual!r})")
    if op == "true":
        passed = actual is True
        return passed, ("is True" if passed else f"{actual!r} (expected True)")
    if op in ("gte", "lte"):
        try:
            a = int(actual); e = int(expected)
            passed = a >= e if op == "gte" else a <= e
            return passed, (f"{a} {op} {e}" if passed else f"{a} not {op} {e}")
        except (TypeError, ValueError):
            return False, f"{actual!r} not numeric for {op} {expected!r}"
    if op == "mentions":
        if isinstance(actual, str) and isinstance(expected, str):
            passed = expected.lower() in actual.lower()
        else:
            passed = False
        return passed, (f"mentions {expected!r}" if passed else f"missing {expected!r} (got {str(actual)[:120]!r})")
    if op == "matches_re":
        try:
            passed = bool(re.search(str(expected), str(actual), flags=re.IGNORECASE))
        except re.error as e:
            return False, f"invalid regex {expected!r}: {e}"
        return passed, (f"matches /{expected}/" if passed else f"{str(actual)[:120]!r} no match /{expected}/")
    return False, f"unknown op {op!r}"


def check_pass(
    debug: dict,
    criteria: list,  # PassCriterion or raw dict
    *,
    session: Any = None,
    last_reply: str | None = None,
    run_root: Path | None = None,
    browser_ctx: dict | None = None,
    console_text: str | None = None,
    mysql_chat_log: list[dict] | None = None,
) -> tuple[bool, list[dict]]:
    """Evaluate every criterion. Browser kwargs are optional and used only by
    ui_text.* / mysql_chat_log.* / trio_match criteria.

    Returns (all_passed, per_criterion_results).
    """
    from e2e.playwright.trio import trio_match as _trio  # noqa: PLC0415

    results = []
    for crit in criteria:
        if isinstance(crit, PassCriterion):
            field, op, expected = crit.field, crit.op, crit.value
        elif isinstance(crit, dict):
            field, op, expected = crit.get("field"), crit.get("op"), crit.get("value")
        else:
            continue
        if not field or not op:
            continue

        if op == "trio_match":
            chat_page = (browser_ctx or {}).get("chat_page")
            ui = chat_page.latest_assistant_reply() if chat_page else ""
            cons = console_text or ""
            mysql_reply = ""
            if mysql_chat_log:
                mysql_reply = mysql_chat_log[-1].get("assistant_reply") or ""
            passed, reason, diff = _trio(ui, cons, mysql_reply)
            results.append({
                "field": field, "op": op, "value": expected,
                "passed": passed,
                "reason": f"{field}: trio_match — {reason}",
                "diff": diff,
            })
            continue

        actual = resolve_field(
            debug, field,
            session=session, last_reply=last_reply, run_root=run_root,
            browser_ctx=browser_ctx, mysql_chat_log=mysql_chat_log,
        )
        passed, reason = _check_one(actual, op, expected)
        results.append({
            "field": field, "op": op, "value": expected,
            "passed": passed,
            "reason": f"{field}: {reason}",
        })

    all_passed = all(r["passed"] for r in results) if results else True
    return all_passed, results
