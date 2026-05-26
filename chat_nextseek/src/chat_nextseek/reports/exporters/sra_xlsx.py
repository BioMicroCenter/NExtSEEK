"""SRA submission XLSX exporter: row-based fill of SRA metadata + BioSample
template workbooks from a report JSON.

Moved out of ``helpers.py`` during Phase 2 of the src/ restructure.
"""
from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .geo_xlsx import _copy_row_format, _write_cell


def _extract_sra_section_reports(
    data: Mapping[str, Any],
    *,
    section_name: str,
) -> list[tuple[str, list[Mapping[str, Any]]]]:
    """Collect per-UID SRA report rows for a given section."""
    reports: list[tuple[str, list[Mapping[str, Any]]]] = []
    for uid, payload in data.items():
        if not isinstance(payload, Mapping):
            continue
        if (payload.get("report_type") or payload.get("report type") or "").upper() != "SRA":
            continue
        report = payload.get("report") or {}
        rows = report.get(section_name) if isinstance(report, Mapping) else None
        if isinstance(rows, list) and rows:
            reports.append((str(uid), [row for row in rows if isinstance(row, Mapping)]))
    return reports


def _worksheet_headers(ws: Worksheet, *, header_row: int) -> list[str]:
    """Read contiguous headers from a worksheet header row."""
    headers: list[str] = []
    col = 1
    while True:
        value = ws.cell(row=header_row, column=col).value
        if value in (None, ""):
            break
        headers.append(str(value))
        col += 1
    return headers


def _write_template_rows(
    wb,
    *,
    sheet_name: str,
    header_row: int,
    template_row: int,
    rows: Sequence[Mapping[str, Any]],
) -> None:
    """Populate a row-based workbook template from ordered row mappings."""
    ws = wb[sheet_name]
    headers = _worksheet_headers(ws, header_row=header_row)

    if len(rows) > 1:
        for i in range(len(rows) - 1):
            insert_at = template_row + i + 1
            ws.insert_rows(insert_at)
            _copy_row_format(ws, template_row, insert_at)

    for row_idx, row in enumerate(rows, start=template_row):
        for col_idx, header in enumerate(headers, start=1):
            _write_cell(ws, row_idx, col_idx, row.get(header))


def _export_sra_section_to_xlsx(
    report_json_path: str,
    template_xlsx_path: str,
    out_dir: str | Path,
    *,
    section_name: str,
    sheet_name: str | None,
    header_row: int,
    template_row: int,
    filename_suffix: str,
    one_workbook_per_uid: bool,
) -> list[str]:
    """Render a row-based SRA report section into workbook copies from a template."""
    json_path = Path(report_json_path)
    template_path = Path(template_xlsx_path)
    output_root = Path(out_dir)
    output_root.mkdir(parents=True, exist_ok=True)

    data = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(data, Mapping):
        return []

    template_bytes = template_path.read_bytes()
    output_paths: list[str] = []
    reports = _extract_sra_section_reports(data, section_name=section_name)

    if not reports:
        return []

    if one_workbook_per_uid:
        for uid, rows in reports:
            wb = load_workbook(BytesIO(template_bytes))
            target_sheet = sheet_name or wb.sheetnames[0]
            _write_template_rows(
                wb,
                sheet_name=target_sheet,
                header_row=header_row,
                template_row=template_row,
                rows=rows,
            )
            dest = output_root / f"{uid}_{filename_suffix}"
            wb.save(dest)
            output_paths.append(str(dest))
    else:
        merged_rows: list[Mapping[str, Any]] = []
        for _, rows in reports:
            merged_rows.extend(rows)
        if not merged_rows:
            return []
        wb = load_workbook(BytesIO(template_bytes))
        target_sheet = sheet_name or wb.sheetnames[0]
        _write_template_rows(
            wb,
            sheet_name=target_sheet,
            header_row=header_row,
            template_row=template_row,
            rows=merged_rows,
        )
        dest = output_root / f"{json_path.stem}_{filename_suffix}"
        wb.save(dest)
        output_paths.append(str(dest))

    return output_paths


def export_sra_report_to_xlsx(
    report_json_path: str,
    template_xlsx_path: str,
    out_dir: str | Path,
    *,
    one_workbook_per_uid: bool = True,
) -> list[str]:
    """
    Convert the SRA libraries report JSON into filled SRA submission workbooks.
    Returns list of output file paths.
    """
    return _export_sra_section_to_xlsx(
        report_json_path,
        template_xlsx_path,
        out_dir,
        section_name="libraries",
        sheet_name="SRA_data",
        header_row=1,
        template_row=2,
        filename_suffix="SRA_metadata_filled.xlsx",
        one_workbook_per_uid=one_workbook_per_uid,
    )


def export_sra_biosample_report_to_xlsx(
    report_json_path: str,
    template_xlsx_path: str,
    out_dir: str | Path,
    *,
    one_workbook_per_uid: bool = True,
) -> list[str]:
    """
    Convert the SRA biosamples report JSON into filled BioSample submission workbooks.
    Returns list of output file paths.
    """
    return _export_sra_section_to_xlsx(
        report_json_path,
        template_xlsx_path,
        out_dir,
        section_name="biosamples",
        sheet_name=None,
        header_row=12,
        template_row=13,
        filename_suffix="SRA_biosample_filled.xlsx",
        one_workbook_per_uid=one_workbook_per_uid,
    )
