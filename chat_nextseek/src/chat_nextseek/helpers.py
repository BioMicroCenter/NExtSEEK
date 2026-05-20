from __future__ import annotations

import ast
import calendar
import csv
import json
import os
import re
import signal
import html
from copy import copy
from io import BytesIO
from urllib.parse import quote, urlparse
from zipfile import ZipFile
from datetime import datetime
from pathlib import Path
from typing import Iterable, Mapping, Any, Sequence, Callable

import requests

from .artifacts import (
    ArtifactStore,
    build_saved_report_file_manifest,
)
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config import ChatConfig
from .schemas import ReportWriterPlan
from .session import SessionState

def slim_api_result_for_llm(api_result: dict, max_rows: int = 5, max_chars: int = 5000) -> dict:
    """
    Trim API results to keep LLM prompts small while preserving key totals and a few example rows.
    Caps row count and total serialized size, substituting a preview with truncation metadata when oversized.
    """
    data = api_result.get("data", {})
    new_data = data
    preview_items = None
    preview_key = None

    if isinstance(data, dict):
        if isinstance(data.get("rows"), list):
            preview_key = "rows"
        elif isinstance(data.get("nodes"), list):
            preview_key = "nodes"
        elif isinstance(data.get("data"), list):
            preview_key = "data"

        if preview_key:
            preview_items = data[preview_key][:max_rows]
            new_data = {**data, preview_key: preview_items}

    slimmed = {**api_result, "data": new_data}

    text = json.dumps(slimmed)
    if len(text) > max_chars:
        total = None
        total_key = "total"
        if isinstance(new_data, dict):
            for key in ("total", "total_samples", "total_nodes"):
                if new_data.get(key) is not None:
                    total = new_data.get(key)
                    total_key = key
                    break
        # Keep a small preview even when truncating for size
        preview_items = preview_items if preview_items is not None else []
        preview_label = f"{preview_key}_preview" if preview_key else "items_preview"
        slimmed = {
            **{k: v for k, v in api_result.items() if k != "data"},
            "data": {
                total_key: total,
                preview_label: preview_items,
                f"{preview_key or 'items'}_truncated": True,
                "note": f"Result truncated for LLM context (>{max_chars} chars).",
            },
        }

    return slimmed


def collect_bundle_files(bundle: dict) -> list[tuple[str, str]]:
    """Return `(label, path)` pairs for saved bundle artifacts that still exist on disk."""
    paths: list[tuple[str, str]] = []

    manifest_files = bundle.get("files") or []
    for entry in manifest_files:
        if not isinstance(entry, dict):
            continue
        path = entry.get("path")
        label = entry.get("label") or entry.get("key") or "artifact"
        if isinstance(path, str) and path and Path(path).exists():
            paths.append((str(label), path))

    if paths:
        return paths

    raw = bundle.get("raw_result_path")
    if raw and Path(raw).exists():
        paths.append(("API result (full JSON)", raw))

    graph = bundle.get("graph_debug_path")
    if graph and Path(graph).exists():
        paths.append(("Graph query debug JSON", graph))

    plan_debug = bundle.get("plan_debug_path")
    if plan_debug and Path(plan_debug).exists():
        paths.append(("Plan debug JSON", plan_debug))

    for step_id, raw_path in (bundle.get("raw_result_paths") or {}).items():
        if raw_path and Path(raw_path).exists():
            paths.append((f"API result (step {step_id})", raw_path))

    for step_id, graph_path in (bundle.get("graph_debug_paths") or {}).items():
        if graph_path and Path(graph_path).exists():
            paths.append((f"Graph debug (step {step_id})", graph_path))

    saved = bundle.get("report_saved_files") or {}
    for key, path in saved.items():
        if isinstance(path, str) and path and Path(path).exists():
            paths.append((key, path))
        elif isinstance(path, list):
            for i, p in enumerate(path):
                if isinstance(p, str) and p and Path(p).exists():
                    paths.append((f"{key}[{i}]", p))

    return paths


def normalize_api_result_for_memory(api_result: dict, min_rows_for_norm: int = 1) -> dict:
    """
    Normalize raw API rows into a compact structure the memory agent can consume reliably.
    Extracts uid/sample_type/assays/json_metadata, falling back to raw data when structure is unexpected or too sparse.
    """
    data = api_result.get("data")
    if not isinstance(data, dict):
        print("[DEBUG][MEMORY_NORM] data is not a dict; falling back to raw data.")
        return {"fallback": True, "raw_data": data}

    rows = data.get("rows")
    if not isinstance(rows, list):
        print("[DEBUG][MEMORY_NORM] No 'rows' list in data; falling back to raw data for this endpoint.")
        return {"fallback": True, "raw_data": data}

    normalized_rows = []
    for r in rows:
        if not isinstance(r, dict):
            continue

        uid = r.get("uuid")
        if not uid:
            uid = strip_html(r.get("uid"))

        sample_type = r.get("sample_type") or r.get("sampletype")
        assays = r.get("assays")
        source_id = r.get("id")

        meta = r.get("json_metadata")
        if not isinstance(meta, dict):
            meta = {}

        normalized_rows.append(
            {
                "uid": uid,
                "sample_type": sample_type,
                "assays": assays,
                "source_id": source_id,
                "metadata": meta,
            }
        )

    if len(normalized_rows) < min_rows_for_norm and len(rows) > 0:
        print(
            f"[DEBUG][MEMORY_NORM] Normalization produced only {len(normalized_rows)} rows "
            f"from {len(rows)} raw rows; falling back to raw data."
        )
        return {"fallback": True, "raw_data": data}

    print(
        f"[DEBUG][MEMORY_NORM] Normalized {len(normalized_rows)} rows "
        f"from {len(rows)} raw rows for memory agent."
    )

    return {
        "total": data.get("total"),
        "rows": normalized_rows,
        "note": "Normalized for memory agent; metadata flattened from json_metadata; HTML removed.",
    }

# ======================================================
# Graph DB helpers
# ======================================================

def tool_neo4j_query(config: ChatConfig, cypher: str, parameters: dict | None = None) -> dict:
    """
    Execute a read-only Cypher query against the configured Neo4j instance.
    Returns a structured dict: {ok, data, count, cypher, parameters, counters} on success,
    or {ok: False, error, cypher} on failure. Opens and closes a driver per call.
    """
    try:
        from neo4j import GraphDatabase  # type: ignore
    except ImportError:
        return {"ok": False, "error": "neo4j driver not installed; run 'uv add neo4j'", "data": None, "cypher": cypher}

    if not getattr(config, "NEO4J_PASSWORD", None):
        return {"ok": False, "error": "NEO4J_PASSWORD not configured", "data": None, "cypher": cypher}

    # Block any write/mutating Cypher clauses — allow read-only queries only.
    _WRITE_KEYWORDS = re.compile(
        r"\b(CREATE|MERGE|SET|DELETE|DETACH\s+DELETE|REMOVE|DROP|CALL\s+db\.|CALL\s+apoc\.schema\.|CALL\s+apoc\.periodic\.|LOAD\s+CSV)\b",
        re.IGNORECASE,
    )
    if _WRITE_KEYWORDS.search(cypher):
        print(f"[DEBUG][GRAPHDB] Blocked write query: {cypher!r}")
        return {"ok": False, "error": "Write operations are not permitted; only read (MATCH/RETURN) queries are allowed.", "data": None, "cypher": cypher}

    params = parameters or {}
    driver = None
    try:
        try:
            driver = GraphDatabase.driver(
                config.NEO4J_URI,
                auth=(config.NEO4J_USER, config.NEO4J_PASSWORD),
                notifications_min_severity="OFF",
            )
        except TypeError:
            driver = GraphDatabase.driver(
                config.NEO4J_URI,
                auth=(config.NEO4J_USER, config.NEO4J_PASSWORD),
            )
        with driver.session(database=getattr(config, "NEO4J_DATABASE", "neo4j")) as db_session:
            result = db_session.run(cypher, params)
            records = [dict(record) for record in result]
            summary = result.consume()
            counters = {}
            if summary and summary.counters:
                try:
                    counters = dict(vars(summary.counters))
                except Exception:
                    pass
            print(f"[DEBUG][GRAPHDB] Query returned {len(records)} records")
            return {
                "ok": True,
                "data": records,
                "count": len(records),
                "cypher": cypher,
                "parameters": params,
                "counters": counters,
            }
    except Exception as e:
        print(f"[DEBUG][GRAPHDB] Query failed: {e!r}")
        return {"ok": False, "error": str(e), "data": None, "cypher": cypher}
    finally:
        if driver is not None:
            try:
                driver.close()
            except Exception:
                pass


# ======================================================
# API helpers
# ======================================================

def tool_nextseek_api_request(config: ChatConfig, endpoint, method, requestBody=None, queryParameters=None):
    """
    Send an HTTP request to the NExtSEEK API with optional basic auth and schema validation.
    Logs request/response previews, parses JSON when possible, and returns a structured dict with ok/status details.
    """
    requestBody = requestBody or {}
    queryParameters = {"page_size": 1000, **(queryParameters or {})}

    base = config.NEXTSEEK_BASE_URL
    if not base:
        msg = "NEXTSEEK_BASE_URL is not set."
        print(f"[DEBUG][API] {msg}")
        return {
            "ok": False,
            "error": msg,
            "endpoint": endpoint,
            "method": method,
        }

    is_valid, error_payload = config.validate_request_body(endpoint, requestBody, method)
    if not is_valid:
        return error_payload

    url = f"{base}/{endpoint.lstrip('/')}"
    auth = (config.API_USER, config.API_PASS) if config.API_USER and config.API_PASS else None
    request_timeout = 90
    if endpoint == "/nextseek_api/samples/advanced_search/":
        request_timeout = 120

    print("[DEBUG][API] Request:")
    print(f"  METHOD: {method}")
    print(f"  URL:    {url}")
    print(f"  PARAMS: {queryParameters}")
    print(f"  BODY:   {requestBody}")
    print(f"  AUTH:   {'Basic' if auth else 'None'}")
    print(f"  TIMEOUT:{request_timeout}s")

    try:
        resp = requests.request(
            method=method,
            url=url,
            auth=auth,
            params=queryParameters,
            json=requestBody if requestBody else None,
            timeout=request_timeout,
        )

        print("[DEBUG][API] Response:")
        print(f"  STATUS: {resp.status_code}")
        preview = resp.text[:300].replace("\n", " ")
        print(f"  PREVIEW: {preview!r}")

        try:
            data = resp.json()
        except Exception:
            data = {"_raw": resp.text[:1000]}

        return {
            "ok": resp.ok,
            "url": url,
            "status_code": resp.status_code,
            "method": method,
            "query": queryParameters,
            "body": requestBody,
            "data": _sanitize_api_row_strings(data),
        }

    except Exception as e:
        print(f"[DEBUG][API] Exception: {repr(e)}")
        return {
            "ok": False,
            "error": repr(e),
            "endpoint": endpoint,
            "method": method,
        }


_API_HTML_PATTERN = re.compile(r"<[^>]+>")
_API_UID_LIKE_FIELDS = {"uid", "uuid", "title", "idlink", "idurl"}


def _sanitize_api_row_strings(payload: Any) -> Any:
    """In-place-ish: walk an API response and strip HTML wrappers from UID-bearing
    string fields. The NExtSEEK API wraps UIDs as `<a href=...>UID</a>` for the
    legacy web UI; raw text breaks downstream string filtering (memory_coder,
    samplesheet emission, chat_log previews).

    Only sanitizes whitelisted fields to avoid mangling legitimate HTML in other
    free-text fields. Returns the payload (mutated when dict/list).
    """
    if isinstance(payload, dict):
        for k, v in list(payload.items()):
            if isinstance(k, str) and k in _API_UID_LIKE_FIELDS and isinstance(v, str):
                if "<" in v and ">" in v:
                    payload[k] = _API_HTML_PATTERN.sub("", v).strip()
            elif isinstance(v, (dict, list)):
                _sanitize_api_row_strings(v)
    elif isinstance(payload, list):
        for item in payload:
            _sanitize_api_row_strings(item)
    return payload


def log_api_call(
    session,
    user_query: str,
    parser_plan: dict,
    api_plan: dict,
    api_result_full: dict,
    bundle_id: int,
):
    """
    Append a JSONL record of an API call to the session-scoped api_log_path.
    Captures query text, plans, and normalized results so console logs remain slim.
    """
    record = {
        "timestamp": datetime.now().isoformat(),
        "bundle_id": bundle_id,
        "user_query": user_query,
        "parser_plan": parser_plan,
        "api_plan": api_plan,
        "api_result_meta": {
            "ok": api_result_full.get("ok"),
            "status_code": api_result_full.get("status_code"),
            "url": api_result_full.get("url"),
        },
        "api_result_data": api_result_full.get("data"),
    }
    log_path = session.get("api_log_path") if hasattr(session, "get") else None
    if not log_path:
        return
    try:
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(json.dumps(record) + "\n")
    except Exception as e:
        print("[DEBUG][API_LOG] Failed to write API log:", repr(e))


def fix_sample_endpoint(plan: dict) -> dict:
    """
    Auto-correct admin retrieve endpoint selections when no UIDs are provided.
    Rewrites to advanced_search and annotates notes to avoid invalid admin calls while keeping other fields intact.
    """
    endpoint = plan.get("target_endpoint")
    mode = plan.get("mode")
    filters = plan.get("filters", {})
    uids = filters.get("uids") or []

    if (
        mode in ("new_search", "refine_last_search")
        and endpoint == "/nextseek_api/admin/samples/retrieve/"
        and not uids
    ):
        print("[DEBUG][PARSER_FIX] Rewriting endpoint from admin retrieve to samples/advanced_search")
        plan["target_endpoint"] = "/nextseek_api/samples/advanced_search/"
        notes = plan.get("notes", "")
        plan["notes"] = (notes + " | endpoint auto-corrected to /samples/advanced_search/").strip(" |")

    return plan


def build_recent_results_summary(session: SessionState, max_results: int = 8) -> str:
    """
    Build a short summary of recent result bundles for prompt conditioning.
    Includes bundle IDs, user queries, endpoints, and totals to guide refinement or follow-up questions.

    Now defaults to 8 bundles (up from 3) — long sessions hit a recall cliff if older
    bundles fall out of view. Parser can then pick `target_result_id` for any bundle
    in the visible window when the user uses "first", "originally", "earlier", etc.
    """
    history = session.get("results_history", [])
    if not history:
        return "No prior results in this session."

    visible = history[-max_results:]
    lines = [
        f"Recent results (most recent first; {len(visible)} of {len(history)} bundles shown):"
    ]
    for bundle in reversed(visible):
        total = None
        data = bundle.get("api_result_slim", {})
        if isinstance(data, dict):
            total = data.get("total")
            if total is None and isinstance(data.get("data"), dict):
                total = (
                    data["data"].get("total")
                    or data["data"].get("total_samples")
                    or data["data"].get("total_nodes")
                )
        lines.append(
            f"- id={bundle.get('id')}, "
            f"query={bundle.get('user_query')!r}, "
            f"endpoint={bundle.get('endpoint')}, "
            f"total={total}"
        )
    if len(history) > max_results:
        lines.append(
            f"(NOTE: {len(history) - max_results} older bundle(s) exist but are not shown. "
            "If the user references something not visible here, infer from CHAT_HISTORY narrative "
            "and set target_result_id explicitly.)"
        )
    return "\n".join(lines)


def _extract_total_and_rows(api_result_full: dict) -> tuple[int | None, int]:
    """
    Extract (total, row_count) from a NExtSEEK response, handling both wrapped and raw result shapes.
    Falls back to (None, 0) when structure is unexpected so retry heuristics remain safe.
    """
    data = api_result_full.get("data")
    if isinstance(data, dict):
        total = (
            data.get("total")
            or data.get("total_samples")
            or data.get("total_nodes")
        )
        rows = (
            data.get("rows")
            if isinstance(data.get("rows"), list)
            else data.get("nodes")
            if isinstance(data.get("nodes"), list)
            else data.get("data")
            if isinstance(data.get("data"), list)
            else None
        )
        if isinstance(rows, list):
            return total, len(rows)
        return total, 0

    # If tool returns raw dict already shaped like {"total":..., "rows":[...]}
    if isinstance(api_result_full, dict):
        total = (
            api_result_full.get("total")
            or api_result_full.get("total_samples")
            or api_result_full.get("total_nodes")
        )
        rows = (
            api_result_full.get("rows")
            if isinstance(api_result_full.get("rows"), list)
            else api_result_full.get("nodes")
            if isinstance(api_result_full.get("nodes"), list)
            else api_result_full.get("data")
            if isinstance(api_result_full.get("data"), list)
            else None
        )
        if isinstance(rows, list):
            return total, len(rows)

    return None, 0


def _should_retry_advanced_search(plan: dict, api_plan: dict, api_result_full: dict) -> bool:
    """
    Determine whether an advanced_search POST should be retried after an empty or failed result.
    Verifies endpoint/method, requires keywords, and checks for zero results or API errors.
    """
    if api_plan.get("endpoint") != "/nextseek_api/samples/advanced_search/":
        return False

    keywords = (plan.get("filters") or {}).get("keywords") or []
    if not isinstance(keywords, list) or len(keywords) < 2:
        # Still retry on errors (timeout etc.) even with fewer keywords
        if isinstance(api_result_full, dict) and api_result_full.get("ok") is False:
            return bool(keywords)
        if not keywords or not _has_expandable_keyword(keywords):
            return False

    total, row_count = _extract_total_and_rows(api_result_full)
    # Retry on empty results or API errors (timeout, connection issues)
    if isinstance(api_result_full, dict) and api_result_full.get("ok") is False:
        return True
    return (total == 0) or (row_count == 0)


def _split_retry_keyword(keyword: str) -> list[str]:
    """Split compact search phrases into useful retry terms while preserving the original elsewhere."""
    terms = [part for part in re.split(r"[\s_\-/]+", keyword.strip()) if part]
    return list(dict.fromkeys(terms))


def _has_expandable_keyword(keywords: list[str]) -> bool:
    return any(len(_split_retry_keyword(k)) > 1 for k in keywords if isinstance(k, str))


def _advanced_search_retry_attempts(keywords: list[str]) -> list[tuple[str, str]]:
    """
    Generate labeled keyword variants for retrying advanced_search.
    Produces OR-joined and single-keyword attempts so the API gets multiple matching chances.
    Also handles single-keyword retries (e.g. after a timeout on the first attempt).
    """
    kws = [k.strip() for k in keywords if isinstance(k, str) and k.strip()]
    if not kws:
        return []

    attempts: list[tuple[str, str]] = []
    if len(kws) >= 2:
        attempts.append(("OR", " OR ".join(kws)))
    elif len(kws) == 1:
        split_terms = _split_retry_keyword(kws[0])
        if len(split_terms) >= 2:
            attempts.append(("OR", " OR ".join(split_terms)))
            for term in split_terms:
                attempts.append(("SINGLE", term))
    for k in kws:
        attempts.append(("SINGLE", k))
    deduped: list[tuple[str, str]] = []
    seen: set[str] = set()
    for label, text in attempts:
        if text in seen:
            continue
        seen.add(text)
        deduped.append((label, text))
    return deduped


def _retry_advanced_search_if_empty(config: ChatConfig, plan: dict, api_plan: dict, api_result_full: dict) -> tuple[dict, dict]:
    """
    If advanced_search returns empty and multiple keywords exist, retry with OR then SINGLE.
    Returns (final_api_plan, final_api_result_full). If no retry needed, returns originals.
    """
    if not _should_retry_advanced_search(plan, api_plan, api_result_full):
        return api_plan, api_result_full

    keywords = (plan.get("filters") or {}).get("keywords") or []
    base_body = dict(api_plan.get("requestBody") or {})

    for label, search_text in _advanced_search_retry_attempts(keywords):
        retry_body = dict(base_body)
        retry_body["filter_searchText"] = search_text

        retry_api_plan = dict(api_plan)
        retry_api_plan["requestBody"] = retry_body
        retry_api_plan["notes"] = (retry_api_plan.get("notes") or "") + f" [retry={label}]"

        retry_result = tool_nextseek_api_request(
            config,
            endpoint=retry_api_plan["endpoint"],
            method=retry_api_plan["method"],
            requestBody=retry_api_plan.get("requestBody") or {},
            queryParameters=retry_api_plan.get("queryParameters") or {},
        )

        total, row_count = _extract_total_and_rows(retry_result)
        if (total and total > 0) or (row_count and row_count > 0):
            print(f"[DEBUG][API][RETRY] Success with {label}: filter_searchText={search_text!r} total={total} rows={row_count}")
            return retry_api_plan, retry_result

        print(f"[DEBUG][API][RETRY] No results with {label}: filter_searchText={search_text!r} total={total} rows={row_count}")

    # All retries failed -> keep original (so logs match original attempt)
    return api_plan, api_result_full


# ======================================================
# Entity / Parser helpers
# ======================================================

def _norm_text(text: str) -> str:
    """
    Normalize free text for matching by lowercasing, stripping non-alphanumerics, and collapsing spaces.
    Keeps comparisons stable across user input and catalog fields regardless of punctuation or capitalization.
    """
    text = (text or "").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _tokenize(text: str) -> set[str]:
    """
    Tokenize normalized text into a set with simple plural/ies reductions.
    Expands tokens to catch small inflection variants while keeping output deterministic.
    """
    tokens = set(_norm_text(text).split())
    expanded: set[str] = set()
    for t in tokens:
        expanded.add(t.rstrip("s"))
        if t.endswith("ies"):
            expanded.add(t[:-3] + "y")
    return {t for t in tokens | expanded if t}


def _doc_from_sampletype(st: dict) -> tuple[str, set[str]]:
    """
    Build a descriptive string and token set from a sampletype entry for fuzzy matching.
    Combines code, name, description, tags, and a cleaned code variant to improve overlap detection.
    """
    parts = [
        st.get("SampleType") or st.get("code") or "",
        st.get("Name") or st.get("name") or "",
        st.get("Description") or st.get("description") or "",
        st.get("Tags") or st.get("tags") or "",
    ]
    code_raw = st.get("SampleType") or st.get("code") or ""
    code_clean = re.sub(r"[^a-zA-Z0-9]+", "", str(code_raw))
    if code_clean and code_clean != code_raw:
        parts.append(code_clean)
    doc = " ".join(str(p) for p in parts if p)
    return doc, _tokenize(doc)


def _doc_from_assay(assay: dict) -> tuple[str, set[str]]:
    """
    Build a descriptive string and token set from an assay entry for matching.
    Uses name, description, and synonym-like fields so scoring can align user phrasing with catalog values.
    """
    parts = [
        assay.get("Name") or assay.get("name") or "",
        assay.get("Description") or assay.get("description") or "",
        assay.get("Tags") or assay.get("tags") or "",
        assay.get("Alternative Assay Names") or assay.get("alternative_assay_names") or "",
    ]
    doc = " ".join(str(p) for p in parts if p)
    return doc, _tokenize(doc)


def _score_pair(query_norm: str, doc_norm: str, overlap_pct: float, code_bonus: float) -> float:
    """
    Compute a blended similarity score using fuzzy matching, token overlap, and an optional code bonus.
    Falls back to difflib when rapidfuzz is unavailable to keep scoring stable across environments.
    """
    if not query_norm or not doc_norm:
        return overlap_pct + code_bonus
    try:
        from rapidfuzz import fuzz

        base = float(fuzz.token_set_ratio(query_norm, doc_norm))
    except Exception:
        # Fallback: basic ratio if rapidfuzz unavailable
        from difflib import SequenceMatcher

        base = SequenceMatcher(None, query_norm, doc_norm).ratio() * 100.0

    # Blend fuzzy similarity and token overlap; add small bonus for explicit code matches
    return base * 0.6 + overlap_pct * 0.4 + code_bonus


def shortlist_catalog(
    user_text: str,
    sampletypes: list[dict],
    assays: list[dict],
    k_st: int = 50,
    k_a: int = 50,
) -> tuple[list[dict], list[dict]]:
    """
    Return the top-k sampletypes and assays most similar to the user_text.
    Falls back gracefully if fuzzy matching lib is missing.
    """
    q_norm = _norm_text(user_text)
    q_tokens = _tokenize(user_text)
    if not q_norm:
        return (sampletypes[:k_st], assays[:k_a])

    def shortlist(items: list[dict], doc_fn, k: int, code_key: str | None = None) -> list[dict]:
        scored = []
        for item in items:
            doc, doc_tokens = doc_fn(item)
            doc_norm = _norm_text(doc)
            overlap = 0.0
            if q_tokens and doc_tokens:
                overlap = (len(q_tokens & doc_tokens) / max(len(q_tokens), 1)) * 100.0

            code_bonus = 0.0
            code_hit = False
            if code_key:
                code = (item.get(code_key) or "").lower()
                if code:
                    code_plain = code.replace(".", "")
                    q_norm_nospace = q_norm.replace(" ", "")
                    if code in q_norm or code_plain in q_norm_nospace:
                        code_bonus += 35.0
                        code_hit = True
                    if code in q_tokens or code_plain in q_tokens:
                        code_bonus += 15.0
                        code_hit = True

            score = _score_pair(q_norm, doc_norm, overlap, code_bonus)
            scored.append((code_hit, score, item))

        scored.sort(key=lambda x: (x[0], x[1]), reverse=True)
        return [item for _, _, item in scored[:k]]

    return (
        shortlist(sampletypes or [], _doc_from_sampletype, k_st, code_key="SampleType"),
        shortlist(assays or [], _doc_from_assay, k_a),
    )




# TODO (later): replace this mapping with a database-backed context/config table.

# ======================================================
# Reporter / Report Writer helpers
# ======================================================


def persist_report_file(
    label: str,
    payload: dict | list | str | None,
    base_dir: str | Path,
    subdir: str | None = None,
    *,
    kind: str = "report",
    filename: str | None = None,
    mime: str | None = None,
) -> str | None:
    """
    Persist a payload under base_dir (optional subdir) with a .json extension.
    Returns the written path or None on failure.
    """
    if payload is None:
        return None
    try:
        store = ArtifactStore(base_dir)
        out_name = filename or f"{label}.json"
        if isinstance(payload, (dict, list)):
            entry = store.write_json(
                key=label,
                label=label,
                filename=out_name,
                payload=payload,
                kind=kind,
                subdir=subdir,
            )
        else:
            entry = store.write_text(
                key=label,
                label=label,
                filename=out_name,
                payload=str(payload),
                kind=kind,
                subdir=subdir,
                mime=mime,
            )
        return entry["path"] if entry else None
    except Exception as e:
        print(f"[DEBUG][REPORTER] Failed to persist {label}:", repr(e))
        return None


def top_items(d: dict, n: int = 5) -> list[dict[str, Any]]:
    """
    Return the top n key/count pairs from a dict sorted by count descending.
    Provides a compact structure for summaries without mutating the input mapping.
    """
    if not isinstance(d, dict):
        return []
    return [
        {"key": k, "count": v}
        for k, v in sorted(d.items(), key=lambda kv: kv[1], reverse=True)[:n]
    ]


def _extract_nfcore_samplesheet_rows(merged_report: dict) -> list[dict[str, Any]]:
    """Pull samplesheet rows out of report_writer outputs (per-UID merged dict).

    Each UID maps to a ReportWriterOutput dump whose `report` body has either a
    `samplesheet` / `samplesheet_rows` / `samplesheet_template` / `rows` /
    `samples` list. We accept the first that's present.
    """
    candidates = (
        "samplesheet", "samplesheet_rows", "samplesheet_template", "rows", "samples",
    )
    rows: list[dict[str, Any]] = []
    if not isinstance(merged_report, dict):
        return rows
    for _, payload in merged_report.items():
        if not isinstance(payload, dict):
            continue
        body = payload.get("report") if isinstance(payload.get("report"), dict) else payload
        if not isinstance(body, dict):
            continue
        for key in candidates:
            section = body.get(key)
            if isinstance(section, list) and section:
                rows.extend(r for r in section if isinstance(r, dict))
                break
    return rows


def _accession_matches_criterion(
    accession: str,
    criterion: dict[str, str] | None,
    accession_metadata: dict[str, dict[str, Any]],
) -> bool:
    """True if the sample metadata for this accession satisfies every key=value
    pair in `criterion` (case-insensitive value compare). Empty criterion = always True."""
    if not criterion:
        return True
    meta = accession_metadata.get(accession) or {}
    for key, expected in criterion.items():
        if not isinstance(key, str):
            return False
        actual = meta.get(key)
        if actual is None:
            return False
        if str(actual).strip().lower() != str(expected).strip().lower():
            return False
    return True


def _handle_nfcore_artifacts(
    *,
    config: ChatConfig,
    user_query: str,
    merged_path: str,
    merged_report: dict,
    nfcore_state: dict[str, Any],
    metadata_map: dict | None = None,
) -> dict[str, Any]:
    """Emit per-cohort artifacts under a parent dir, plus a combined launch.yml.
    Optionally submits via seqerakit. Returns aggregated dict.
    """
    from .agents import seqera_agent  # local import to avoid cycle
    from .seqera import emit_nfcore_artifacts, submit_launch, write_combined_launch_yml

    cohorts: list[dict[str, Any]] = nfcore_state.get("cohorts") or []
    if not cohorts:
        cohorts = [{"label": "rnaseq", "pipeline": "rnaseq", "rationale": "fallback",
                    "enrichment_metadata_fields": [], "cohort_criterion": {}}]

    # Top-level dir name reflects multi-cohort vs single-cohort
    if len(cohorts) == 1:
        parent_dir_name = f"nfcore_{cohorts[0]['label']}"
    else:
        parent_dir_name = "nfcore_multi"
    parent_out_dir = Path(merged_path).parent / parent_dir_name
    parent_out_dir.mkdir(parents=True, exist_ok=True)

    rows = _extract_nfcore_samplesheet_rows(merged_report)
    accession_metadata = build_accession_metadata_lookup(metadata_map or {})
    tower_env = config.TOWER_ENV if config.TOWER_ENV_COMPLETE else {}

    aggregated_saved: dict[str, Any] = {}
    combined_launch: list[dict[str, Any]] = []
    cohort_summaries: list[dict[str, Any]] = []
    skipped_cohorts: list[dict[str, Any]] = []
    multi = len(cohorts) > 1

    for cohort in cohorts:
        label = cohort["label"]
        pipeline = cohort["pipeline"]
        criterion = cohort.get("cohort_criterion") or {}
        enrichment = cohort.get("enrichment_metadata_fields") or []

        # Filter rows whose accession metadata matches this cohort's criterion.
        cohort_rows: list[dict[str, Any]] = []
        for row in rows:
            acc = row.get("accession") or row.get("Accession") or row.get("ena_accession")
            if not acc:
                if not criterion:
                    cohort_rows.append(row)
                continue
            if _accession_matches_criterion(str(acc).strip(), criterion, accession_metadata):
                cohort_rows.append(row)

        # If the LLM under-emitted, fall back to synthesizing rows for every
        # accession that matches the criterion.
        if not cohort_rows:
            for acc in accession_metadata.keys():
                if _accession_matches_criterion(acc, criterion, accession_metadata):
                    meta = accession_metadata[acc]
                    sample_id = (
                        meta.get("Library_ID")
                        or meta.get("Title")
                        or meta.get("UID")
                        or acc
                    )
                    strandedness = (meta.get("Strandedness") or "auto").strip().lower() or "auto"
                    cohort_rows.append({
                        "sample": sample_id,
                        "fastq_1": "",
                        "fastq_2": "",
                        "strandedness": strandedness,
                        "accession": acc,
                    })

        # Skip empty cohorts (no rows after filtering AND no synthesizable rows).
        # This protects against the case where a user-pinned pipeline doesn't
        # match any actual data (e.g., user typed `sarek` but data is RNA-seq).
        if not cohort_rows:
            skipped_cohorts.append({
                "label": label,
                "pipeline": pipeline,
                "criterion": criterion,
                "rationale": cohort.get("rationale") or "",
                "user_pinned": bool(cohort.get("_user_pinned")),
                "reason": "0 rows matched after filtering",
            })
            print(
                f"[DEBUG][REPORTER_NFCORE][{label}] empty cohort — skipping. "
                f"criterion={criterion}, user_pinned={cohort.get('_user_pinned', False)}"
            )
            continue

        # Per-cohort launch plan (params + run name) via the seqera agent.
        launch_plan_dump: dict[str, Any] = {}
        if config.TOWER_ENV_COMPLETE:
            try:
                preview = {
                    "columns": sorted({k for r in cohort_rows[:10] for k in r.keys()}),
                    "first_row": cohort_rows[0] if cohort_rows else {},
                    "row_count": len(cohort_rows),
                    "cohort_label": label,
                    "cohort_criterion": criterion,
                }
                plan_obj = seqera_agent(
                    config=config,
                    user_query=user_query,
                    pipeline=pipeline,
                    samplesheet_preview=preview,
                    reporter_context_summary=nfcore_state.get("reporter_summary") or {},
                )
                launch_plan_dump = plan_obj.model_dump() if hasattr(plan_obj, "model_dump") else dict(plan_obj)
            except Exception as e:
                print(f"[DEBUG][REPORTER_NFCORE][{label}] seqera_agent failed:", repr(e))

        cohort_dir = parent_out_dir / label if multi else parent_out_dir
        rel_dir = label if multi else "."
        cohort_dir.mkdir(parents=True, exist_ok=True)

        emission = emit_nfcore_artifacts(
            cohort_dir,
            pipeline=pipeline,
            samplesheet_rows=cohort_rows,
            resolutions=nfcore_state.get("resolutions") or [],
            launch_plan=launch_plan_dump,
            tower_env=tower_env,
            selector_rationale=cohort.get("rationale") or "",
            enrichment_fields=enrichment,
            accession_metadata=accession_metadata,
            samplesheet_relative_dir=rel_dir,
            write_launch_yml=not multi,  # for multi-cohort we write a combined launch.yml at parent level
        )

        prefix = f"nfcore_{label}_" if multi else "nfcore_"
        for k, v in (emission.saved_files or {}).items():
            aggregated_saved[f"{prefix}{k}"] = v

        if emission.launch_entry:
            combined_launch.append(emission.launch_entry)
        if emission.fetchngs_launch_entry:
            combined_launch.append(emission.fetchngs_launch_entry)

        cohort_summaries.append({
            "label": label,
            "pipeline": pipeline,
            "rationale": cohort.get("rationale") or "",
            "criterion": criterion,
            "enrichment_fields": enrichment,
            "row_count": emission.samplesheet_row_count,
            "excluded_accessions": emission.excluded_accessions,
        })

    # Recompute multi-cohort status after skips
    effective_multi = len(cohort_summaries) > 1

    # Combined launch.yml at parent level when multi-cohort
    combined_launch_path: str | None = None
    if effective_multi and combined_launch:
        combined_launch_path = write_combined_launch_yml(parent_out_dir, combined_launch)
        if combined_launch_path:
            aggregated_saved["nfcore_launch_combined"] = combined_launch_path

    # Top-level cohort summary notes.md
    summary_notes_path = parent_out_dir / "notes.md"
    summary_notes_path.write_text(
        _build_cohort_summary_md(
            cohort_summaries,
            nfcore_state,
            multi=effective_multi,
            skipped_cohorts=skipped_cohorts,
        ),
        encoding="utf-8",
    )
    aggregated_saved["nfcore_summary_notes"] = str(summary_notes_path)

    # Optional auto-submit
    run_urls: list[str] = []
    if config.SEQERA_AUTO_LAUNCH and config.TOWER_ENV_COMPLETE:
        # Prefer combined launch when multi; else the single cohort's launch.yml
        target = combined_launch_path or aggregated_saved.get("nfcore_launch")
        if target:
            try:
                run_urls = submit_launch(target, tower_env=config.TOWER_ENV)
            except Exception as e:
                print("[DEBUG][REPORTER_NFCORE] seqerakit submit failed:", repr(e))

    return {
        "out_dir": str(parent_out_dir),
        "saved_files": aggregated_saved,
        "cohort_summaries": cohort_summaries,
        "run_urls": run_urls,
    }


def _build_cohort_summary_md(
    cohort_summaries: list[dict[str, Any]],
    nfcore_state: dict[str, Any],
    *,
    multi: bool,
    skipped_cohorts: list[dict[str, Any]] | None = None,
) -> str:
    lines = ["# nf-core run summary", ""]

    # Conflict banner: prominently surfaced at the top
    if nfcore_state.get("conflict_detected"):
        pipeline_key = nfcore_state.get("pipeline_key") or "unknown"
        recommended = sorted({
            c.get("pipeline") for c in (nfcore_state.get("cohorts") or [])
            if c.get("pipeline") != pipeline_key and not c.get("_user_pinned")
        })
        lines.append("> ⚠️ **PIPELINE DISAGREEMENT**")
        lines.append(">")
        lines.append(
            f"> You explicitly requested **nf-core/{pipeline_key}**, but the metadata "
            f"indicates **{', '.join(f'nf-core/{p}' for p in recommended) or 'a different pipeline'}** "
            "is more appropriate for this data."
        )
        lines.append(">")
        lines.append(
            "> Both samplesheets are emitted below. Pick the cohort you actually intend "
            "to run, or run both as separate Tower workflows from the top-level `launch.yml`."
        )
        lines.append("")

    if nfcore_state.get("selector_rationale"):
        lines.append(f"**Selector rationale:** {nfcore_state['selector_rationale']}")
        lines.append("")
    lines.append(f"**Cohorts emitted:** {len(cohort_summaries)}")
    if skipped_cohorts:
        lines.append(f"**Cohorts skipped (empty):** {len(skipped_cohorts)}")
    lines.append("")

    for c in cohort_summaries:
        lines.append(f"## {c['label']} → nf-core/{c['pipeline']}")
        if c.get("rationale"):
            lines.append(f"- Rationale: {c['rationale']}")
        if c.get("criterion"):
            lines.append(f"- Cohort criterion: {c['criterion']}")
        else:
            lines.append("- Cohort criterion: (all samples)")
        if c.get("enrichment_fields"):
            lines.append(f"- Enrichment columns: {c['enrichment_fields']}")
        lines.append(f"- Samplesheet rows: {c['row_count']}")
        if c.get("excluded_accessions"):
            lines.append(f"- Excluded (ENA-missing): {c['excluded_accessions']}")
        lines.append("")

    if skipped_cohorts:
        lines.append("## Skipped cohorts")
        for sc in skipped_cohorts:
            tag = " (user-pinned)" if sc.get("user_pinned") else ""
            lines.append(
                f"- `{sc['label']}` → nf-core/{sc['pipeline']}{tag}: "
                f"{sc.get('reason') or 'no rows matched'}. Criterion: {sc.get('criterion') or {}}"
            )
        lines.append("")

    if multi:
        lines.append("Run all cohorts: `seqerakit launch.yml` (top-level launch.yml).")
    else:
        lines.append("Run: `seqerakit launch.yml` (inside the cohort folder).")
    return "\n".join(lines) + "\n"


def generate_report_outputs(
    *,
    config: ChatConfig,
    user_query: str,
    parser_plan,
    reporter_plan,
    uids: list[str],
    log_dir: str | Path,
    report_writer_fn: Callable[[ChatConfig, str, ReportWriterPlan, dict | None], Any],
    per_sample_reports: bool = True,
    pre_supplied_cohorts: list[dict] | None = None,
) -> tuple[dict, dict | Any, dict[str, str], str]:
    """
    Full report-generation flow (metadata fetch, protocol fetch, report writer call, persistence).
    Returns (reporter_result, report_writer_output, saved_files, reply_text).
    """
    plan_dump = parser_plan.model_dump() if hasattr(parser_plan, "model_dump") else parser_plan or {}
    parser_report_type = getattr(parser_plan, "report_type", None) or (plan_dump.get("report_type") if isinstance(plan_dump, dict) else None)
    report_type_value = normalize_report_type(getattr(reporter_plan, "report_type", None) or parser_report_type)
    per_uid_reports: list[dict] = []
    saved_files: dict[str, Any] = {}

    # NFCORE-specific scratch state populated below if the flow is NFCORE_*
    nfcore_state: dict[str, Any] = {
        "active": False,
        "cohorts": [],            # list[dict]: {label, pipeline, rationale, enrichment_metadata_fields, cohort_criterion}
        "selector_rationale": "",
        "resolutions": [],
        "accession_rows": [],
        "reporter_summary": {},
        "metadata_summary": {},
    }

    # Fetch metadata (combined vs per-sample)
    metadata_map: dict[str | None, dict] = {}
    if per_sample_reports:
        for uid in uids or [None]:
            current_uids = [uid] if uid else []
            metadata = (
                fetch_reporter_metadata(config, current_uids)
                if current_uids
                else {"ok": False, "error": "No UID provided"}
            )
            print("[DEBUG][REPORTER] Metadata fetch result ok:", metadata.get("ok"), "uid:", uid)
            metadata = annotate_metadata_with_sampletypes(config, metadata) if metadata else metadata
            metadata_map[uid] = metadata
    else:
        metadata = fetch_reporter_metadata(config, uids) if uids else {"ok": False, "error": "No UID provided"}
        print("[DEBUG][REPORTER] Combined metadata fetch ok:", metadata.get("ok"), "uids:", uids)
        metadata = annotate_metadata_with_sampletypes(config, metadata) if metadata else metadata
        metadata_map["__all__"] = metadata

    # Collect and fetch protocols
    all_protocol_refs: dict[tuple[str, str], dict[str, str]] = {}
    for md in metadata_map.values():
        refs = extract_protocol_refs_from_metadata(md) if md else []
        if refs:
            print("[DEBUG][REPORTER_PROTOCOL] Found protocol refs:", refs)
        for ref in refs:
            key = (ref.get("source", ""), ref.get("value", ""))
            if key[1]:
                all_protocol_refs[key] = ref
    protocol_payloads = fetch_protocols(config, list(all_protocol_refs.values())) if all_protocol_refs else {}
    if protocol_payloads:
        ok_ids = [pid for pid, resp in protocol_payloads.items() if isinstance(resp, dict) and resp.get("ok")]
        print("[DEBUG][REPORTER_PROTOCOL] Protocol fetch complete. ok:", ok_ids, "total:", len(protocol_payloads))
    else:
        print("[DEBUG][REPORTER_PROTOCOL] No protocols discovered.")
    protocol_files = download_and_extract_protocol_blobs(protocol_payloads, log_dir, config=config) if protocol_payloads else {}
    if protocol_files:
        print("[DEBUG][REPORTER_PROTOCOL] Downloaded/extracted protocol files for IDs:", list(protocol_files.keys()))

    protocols_for_llm = sanitize_protocols_for_llm(protocol_payloads)

    # ── NFCORE: cohorts come from the wizard (pre_supplied_cohorts) — no LLM selector
    if report_type_value and report_type_value.startswith("NFCORE"):
        nfcore_state["active"] = True
        try:
            from .seqera import (
                extract_accessions_from_metadata,
                resolve_accessions,
            )
        except Exception as e:  # pragma: no cover
            print("[DEBUG][REPORTER_NFCORE] Failed to import nfcore deps:", repr(e))
            extract_accessions_from_metadata = None  # type: ignore
            resolve_accessions = None  # type: ignore

        # 1) Build the metadata summary once — used by the report writer + emitter.
        try:
            full_summary = build_metadata_summary(metadata_map)
            deg_summary = filter_summary_for_deg(full_summary)
            nfcore_state["metadata_summary"] = full_summary
            nfcore_state["deg_summary"] = deg_summary
            print(
                "[DEBUG][REPORTER_NFCORE] metadata_summary sample types:",
                list((full_summary.get("by_sample_type") or {}).keys()),
            )
        except Exception as e:
            print("[DEBUG][REPORTER_NFCORE] build_metadata_summary failed:", repr(e))
            full_summary = {}
            deg_summary = {}

        # 2) Cohorts: must come from the wizard. If nothing was supplied, fall back
        # to a single-cohort run keyed off the report_type's pipeline suffix.
        pipeline_key = nfcore_pipeline_from_report_type(report_type_value)
        reporter_summary = {
            "uids": uids,
            "sample_types": sorted(list((full_summary.get("by_sample_type") or {}).keys())),
        }
        nfcore_state["reporter_summary"] = reporter_summary

        cohorts: list[dict[str, Any]] = []
        rationale = ""
        if pre_supplied_cohorts:
            cohorts = [dict(c) for c in pre_supplied_cohorts]
            rationale = (
                getattr(reporter_plan, "notes", "")
                or "Cohorts collected interactively via the nf-core wizard."
            )
            print(
                f"[DEBUG][REPORTER_NFCORE] Using pre-supplied wizard cohorts ({len(cohorts)}): "
                + ", ".join(
                    f"{c.get('label')}(pipeline={c.get('pipeline')}, criterion={c.get('cohort_criterion') or {}})"
                    for c in cohorts
                )
            )

        if not cohorts:
            fallback = pipeline_key or "rnaseq"
            cohorts = [{
                "label": fallback,
                "pipeline": fallback,
                "rationale": "Fallback single cohort (no wizard cohorts supplied).",
                "enrichment_metadata_fields": [],
                "cohort_criterion": {},
                "expected_sample_count": 0,
            }]
            rationale = rationale or cohorts[0]["rationale"]
            print(f"[DEBUG][REPORTER_NFCORE] No pre-supplied cohorts — falling back to single cohort '{fallback}'.")

        nfcore_state["cohorts"] = cohorts
        nfcore_state["selector_rationale"] = rationale
        nfcore_state["conflict_detected"] = False
        nfcore_state["pipeline_key"] = pipeline_key

        # 3) For schema/template loading, pick the first cohort's pipeline as the
        # canonical report_type_value (so report_writer loads SOME nf-core
        # template). Cohort-specific filtering happens in _handle_nfcore_artifacts.
        primary_pipeline = cohorts[0]["pipeline"]
        report_type_value = f"NFCORE_{primary_pipeline.upper()}"

        # 3) ENA accession resolution
        try:
            accessions: list[str] = []
            for md in metadata_map.values():
                accessions.extend(extract_accessions_from_metadata(md) if extract_accessions_from_metadata else [])
            seen: set[str] = set()
            ordered: list[str] = []
            for acc in accessions:
                if acc and acc not in seen:
                    seen.add(acc)
                    ordered.append(acc)
            resolutions = resolve_accessions(ordered) if (resolve_accessions and ordered) else []
            nfcore_state["resolutions"] = resolutions
            rows: list[dict[str, Any]] = []
            for r in resolutions:
                if r.missing:
                    continue
                for run in r.runs:
                    rows.append({
                        "accession": r.accession,
                        "run_accession": run.run_accession,
                        "fastq_1": run.fastq_1,
                        "fastq_2": run.fastq_2,
                        "library_layout": run.layout,
                    })
            nfcore_state["accession_rows"] = rows
            print(
                f"[DEBUG][REPORTER_NFCORE] Resolved {len(rows)} runs from "
                f"{sum(1 for r in resolutions if not r.missing)}/{len(resolutions)} accessions"
            )
        except Exception as e:
            print("[DEBUG][REPORTER_NFCORE] ENA resolver failed:", repr(e))

    # ── NFCORE bypass: skip the report_writer LLM entirely ─────────────────
    # Rationale: the seqera emitter (_handle_nfcore_artifacts) already synthesizes
    # samplesheet rows directly from accession_metadata via its fallback path. The
    # report_writer was producing a JSON shape that the emitter then re-parsed and
    # validated against templates — pure indirection that cost us a 5.1M-token
    # prompt on a 195-UID NDMA-mice flow. The wizard's explicit cohort_criteria +
    # enrichment_fields give us everything we need without an LLM call here.
    if nfcore_state["active"]:
        print("[DEBUG][REPORTER_NFCORE] Bypassing report_writer; emitter will synthesize rows from accession_metadata.")
        per_uid_reports = [{
            "uid": None,
            "metadata": metadata_map.get("__all__") if not per_sample_reports else None,
            "protocols": protocol_payloads,
            "report_writer_output": {
                "report_type": report_type_value,
                "report": {"samplesheet": []},
                "narrative": "Samplesheet synthesized from accession metadata (no LLM call).",
                "notes": "NFCORE bypass — emitter handles row construction.",
            },
        }]
        merged_report = {"all_samples": per_uid_reports[0]["report_writer_output"]}
    elif per_sample_reports:
        loop_uids = uids or [None]
        for idx, uid in enumerate(loop_uids):
            meta = metadata_map.get(uid)
            reporter_context = getattr(reporter_plan, "reporter_context", {}) or {}
            reporter_context = {
                **(reporter_context or {}),
                "uids": [uid] if uid else [],
                "metadata": meta,
                "protocols": protocols_for_llm,
                "protocol_files": protocol_files,
                "parser_plan": plan_dump,
                "reporter_plan": reporter_plan.model_dump() if hasattr(reporter_plan, "model_dump") else {},
            }
            if nfcore_state["active"]:
                reporter_context["nfcore_cohorts"] = nfcore_state["cohorts"]
                reporter_context["accession_rows"] = nfcore_state["accession_rows"]
                reporter_context["nfcore_selector_rationale"] = nfcore_state["selector_rationale"]
                reporter_context["nfcore_metadata_summary"] = nfcore_state.get("deg_summary") or nfcore_state.get("metadata_summary") or {}
            report_writer_plan = ReportWriterPlan(
                report_type=report_type_value,
                reporter_context=reporter_context,
                notes=getattr(reporter_plan, "notes", ""),
            )
            template = load_report_template(config, report_writer_plan.report_type)
            template_for_llm = {k: v for k, v in (template or {}).items() if k != "schema"}
            print("[DEBUG][REPORT_WRITER] Using template keys (schema stripped):", list(template_for_llm.keys()), "uid:", uid)
            report_writer_output = report_writer_fn(config, user_query, report_writer_plan, template_for_llm)
            per_uid_reports.append(
                {
                    "uid": uid,
                    "metadata": meta,
                    "protocols": protocol_payloads,
                    "report_writer_output": report_writer_output.model_dump()
                    if hasattr(report_writer_output, "model_dump")
                    else report_writer_output,
                }
            )
        merged_report = {
            entry.get("uid") or f"item_{i}": entry.get("report_writer_output", {}) for i, entry in enumerate(per_uid_reports)
        }
    else:
        meta = metadata_map.get("__all__")
        reporter_context = getattr(reporter_plan, "reporter_context", {}) or {}
        reporter_context = {
            **(reporter_context or {}),
            "uids": uids,
            "metadata": meta,
            "protocols": protocols_for_llm,
            "protocol_files": protocol_files,
            "parser_plan": plan_dump,
            "reporter_plan": reporter_plan.model_dump() if hasattr(reporter_plan, "model_dump") else {},
        }
        report_writer_plan = ReportWriterPlan(
            report_type=report_type_value,
            reporter_context=reporter_context,
            notes=getattr(reporter_plan, "notes", ""),
        )
        template = load_report_template(config, report_writer_plan.report_type)
        template_for_llm = {k: v for k, v in (template or {}).items() if k != "schema"}
        print("[DEBUG][REPORT_WRITER] Using template keys (schema stripped):", list(template_for_llm.keys()), "uid: ALL")
        combined_output = report_writer_fn(config, user_query, report_writer_plan, template_for_llm)
        per_uid_reports.append(
            {
                "uid": None,
                "metadata": meta,
                "protocols": protocol_payloads,
                "report_writer_output": combined_output.model_dump()
                if hasattr(combined_output, "model_dump")
                else combined_output,
            }
        )
        merged_report = {"all_samples": combined_output.model_dump() if hasattr(combined_output, "model_dump") else combined_output}

    reporter_result = {
        "reports": per_uid_reports,
        "merged_report": merged_report,
    }
    report_writer_output = merged_report

    # Persist report payloads and metadata
    extracted_protocols = {}
    for pid, files_list in (protocol_files or {}).items():
        texts = []
        for f in files_list or []:
            if f.get("text"):
                texts.append({"filename": f.get("filename"), "text": f.get("text")})
        if texts:
            extracted_protocols[pid] = texts

    meta_map = {}
    for idx, entry in enumerate(per_uid_reports):
        uid_key = entry.get("uid") or f"item_{idx}"
        combined_meta = entry.get("metadata") or {}
        combined_meta = dict(combined_meta)
        combined_meta["protocols"] = extracted_protocols
        meta_map[uid_key] = combined_meta
        saved = persist_report_file(
            f"report_writer_output_{uid_key}",
            entry.get("report_writer_output"),
            log_dir,
            kind="report",
        )
        if saved:
            saved_files[f"report_writer_output_{uid_key}"] = saved

    report_type_label = normalize_report_type(report_type_value) or "REPORT"
    merged_filename = f"merged_report_{report_type_label}"
    merged_path = persist_report_file(merged_filename, merged_report, log_dir, kind="report")
    if merged_path:
        saved_files["merged_report"] = merged_path

    if report_type_label == "GEO" and merged_path:
        try:
            geo_workbooks = export_geo_report_to_seq_xlsx(
                merged_path,
                str(config.SEQ_TEMPLATE_PATH),
                Path(merged_path).parent,
                one_workbook_per_uid=False,
            )
            if geo_workbooks:
                saved_files["geo_seq_workbooks"] = geo_workbooks
                print("[DEBUG][REPORTER_GEO] Exported GEO submission workbooks:", geo_workbooks)
        except Exception as e:
            print("[DEBUG][REPORTER_GEO] Failed to export GEO XLSX:", repr(e))
    elif report_type_label.startswith("NFCORE") and merged_path and nfcore_state["active"]:
        try:
            nfcore_artifacts = _handle_nfcore_artifacts(
                config=config,
                user_query=user_query,
                merged_path=merged_path,
                merged_report=merged_report,
                nfcore_state=nfcore_state,
                metadata_map=metadata_map,
            )
            for k, v in (nfcore_artifacts.get("saved_files") or {}).items():
                saved_files[k if k.startswith("nfcore_") else f"nfcore_{k}"] = v
            run_urls = nfcore_artifacts.get("run_urls") or []
            if run_urls:
                saved_files["nfcore_tower_run_urls"] = run_urls
                print("[DEBUG][REPORTER_NFCORE] Tower run URLs:", run_urls)
            print(
                "[DEBUG][REPORTER_NFCORE] Emitted nf-core artifacts at",
                nfcore_artifacts.get("out_dir"),
            )
        except Exception as e:
            print("[DEBUG][REPORTER_NFCORE] Failed to emit nf-core artifacts:", repr(e))
    elif report_type_label == "SRA" and merged_path:
        try:
            sra_workbooks = export_sra_report_to_xlsx(
                merged_path,
                str(Path(config.BASE_DIR) / "reports" / "SRA_metadata.xlsx"),
                Path(merged_path).parent,
                one_workbook_per_uid=False,
            )
            if sra_workbooks:
                saved_files["sra_submission_workbooks"] = sra_workbooks
                print("[DEBUG][REPORTER_SRA] Exported SRA submission workbooks:", sra_workbooks)
        except Exception as e:
            print("[DEBUG][REPORTER_SRA] Failed to export SRA workbook:", repr(e))
        try:
            biosample_workbooks = export_sra_biosample_report_to_xlsx(
                merged_path,
                str(Path(config.BASE_DIR) / "reports" / "SRA_biosample.xlsx"),
                Path(merged_path).parent,
                one_workbook_per_uid=False,
            )
            if biosample_workbooks:
                saved_files["sra_biosample_workbooks"] = biosample_workbooks
                print("[DEBUG][REPORTER_SRA] Exported SRA BioSample workbooks:", biosample_workbooks)
        except Exception as e:
            print("[DEBUG][REPORTER_SRA] Failed to export SRA BioSample workbook:", repr(e))

    meta_path = persist_report_file("report_metadata", meta_map, log_dir, kind="report")
    if meta_path:
        saved_files["metadata"] = meta_path
    if protocol_payloads:
        proto_path = persist_report_file("protocols", protocol_payloads, log_dir, subdir="protocols", kind="protocol")
        if proto_path:
            saved_files["protocols"] = proto_path
    if protocol_files:
        proto_files_path = persist_report_file("protocol_files", protocol_files, log_dir, subdir="protocols", kind="protocol")
        if proto_files_path:
            saved_files["protocol_files"] = proto_files_path

    reply_lines = ["Generated report payload is available in the Reporter result panel."]
    for entry in per_uid_reports:
        narrative = ((entry.get("report_writer_output") or {}).get("narrative")) if entry else None
        if narrative:
            reply_lines.insert(0, narrative.strip())
            break
    reply = "\n\n".join(reply_lines)

    return reporter_result, report_writer_output, saved_files, reply


def _normalize_project_id(config: ChatConfig, project: int | str | None) -> int | None:
    """
    Normalize a project identifier from int or string to canonical integer ID.
    Accepts numeric strings or known project names via PROJECT_NAME_TO_ID, raising on unknown names.
    """
    if project is None:
        return None
    if isinstance(project, int):
        return project
    key = project.strip().upper()
    if not key:
        return None
    if key.isdigit():
        return int(key)
    if key not in config.PROJECT_NAME_TO_ID:
        # Fuzzy fallback: accept if any canonical key is contained in the input or vice versa.
        fuzzy_match = next(
            (
                (canonical, pid)
                for canonical, pid in config.PROJECT_NAME_TO_ID.items()
                if canonical in key or key in canonical
            ),
            None,
        )
        if fuzzy_match:
            print(f"[WARN][PROJECT] '{project}' fuzzy-matched to '{fuzzy_match[0]}' (id={fuzzy_match[1]})")
            return fuzzy_match[1]
        raise ValueError(
            f"Unknown project '{project}'. Expected one of: {sorted(config.PROJECT_NAME_TO_ID.keys())} "
            f"or a numeric project_id."
        )
    return config.PROJECT_NAME_TO_ID[key]

def _normalize_years(years: Iterable[int | str]) -> list[str]:
    """
    Convert an iterable of years into deduped two-digit strings (e.g., '24', '25').
    Accepts 2- or 4-digit inputs as ints or strings and raises on malformed values to keep SQL filters safe.
    """
    out: list[str] = []
    for y in years:
        s = str(y).strip()
        if re.fullmatch(r"\d{4}", s):
            out.append(s[2:])
        elif re.fullmatch(r"\d{2}", s):
            out.append(s)
        else:
            raise ValueError(f"Invalid year value: {y!r}. Use 2024 or 24 (or strings).")
    seen = set()
    deduped = []
    for yy in out:
        if yy not in seen:
            seen.add(yy)
            deduped.append(yy)
    return deduped

def _parse_month(s: str) -> tuple[int, int]:
    """
    Parse a month string in formats like 'YYYY-MM', 'YYYY/MM', or 'YYYYMM' (also 'YY-MM' -> assumes 20YY).
    Returns (year, month) and raises ValueError on invalid inputs so callers can surface clear errors.
    """
    t = s.strip()
    m = re.fullmatch(r"(\d{4})[-/](\d{2})", t) or re.fullmatch(r"(\d{4})(\d{2})", t)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
    else:
        m2 = re.fullmatch(r"(\d{2})[-/](\d{2})", t) or re.fullmatch(r"(\d{2})(\d{2})", t)
        if not m2:
            raise ValueError(f"Invalid month format: {s!r}. Use '2024-01' (preferred).")
        year = 2000 + int(m2.group(1))
        month = int(m2.group(2))

    if not (1 <= month <= 12):
        raise ValueError(f"Invalid month in {s!r}.")
    return year, month

def _month_range_to_yymmdd_bounds(month_range: tuple[str, str]) -> tuple[str, str]:
    """
    Convert a (start_month, end_month) tuple into inclusive YYMMDD bounds (e.g., '2024-01' -> '240101').
    Swaps bounds when reversed to keep downstream SQL BETWEEN clauses robust.
    """
    y1, m1 = _parse_month(month_range[0])
    y2, m2 = _parse_month(month_range[1])

    start = f"{y1 % 100:02d}{m1:02d}01"
    last_day = calendar.monthrange(y2, m2)[1]
    end = f"{y2 % 100:02d}{m2:02d}{last_day:02d}"

    # If user accidentally swaps, we still handle it
    if start > end:
        start, end = end, start
    return start, end

def _parse_day(s: str) -> tuple[int, int, int]:
    """
    Accepts 'YYYY-MM-DD' or 'YYYY/MM/DD' or 'YYYYMMDD' (also 'YYMMDD' -> assumes 20YY).
    Returns (year, month, day).
    """
    t = s.strip()
    m = re.fullmatch(r"(\d{4})[-/](\d{2})[-/](\d{2})", t) or re.fullmatch(r"(\d{4})(\d{2})(\d{2})", t)
    if m:
        year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
    else:
        m2 = re.fullmatch(r"(\d{2})(\d{2})(\d{2})", t)
        if not m2:
            raise ValueError(f"Invalid day format: {s!r}. Use '2024-04-22' (preferred).")
        year, month, day = 2000 + int(m2.group(1)), int(m2.group(2)), int(m2.group(3))

    # Basic validity check
    datetime(year, month, day)  # will raise if invalid
    return year, month, day

def _day_range_to_yymmdd_bounds(day_range: tuple[str, str]) -> tuple[str, str]:
    """
    Convert a (start_day, end_day) tuple into inclusive YYMMDD bounds for SQL filtering.
    Handles reversed inputs by swapping so BETWEEN logic stays correct.
    """
    (y1, m1, d1) = _parse_day(day_range[0])
    (y2, m2, d2) = _parse_day(day_range[1])
    start = f"{y1 % 100:02d}{m1:02d}{d1:02d}"
    end = f"{y2 % 100:02d}{m2:02d}{d2:02d}"
    if start > end:
        start, end = end, start
    return start, end

def run_project_sample_report(
    config: ChatConfig,
    project: int | str | None,
    years: list[int | str] | None = None,
    month_range: tuple[str, str] | None = None,
    day_range: tuple[str, str] | None = None,
    outputs_root: str | Path = "outputs",
) -> dict:
    """
    Project-scoped sample UUID reporting with optional date filters extracted from UUID.
    If project is None, the report runs across all projects.

    UUID format: SampleType-YYMMDDLAB-Incrementer (e.g., TIS-240422DFC-6)
    Assay-derived/sample-like UIDs may include dots in the sample type (e.g., D.FCS-240306SAS-10).

    Date extraction used in SQL:
      date6 = LEFT(SUBSTRING_INDEX(SUBSTRING_INDEX(s.uuid,'-',2),'-',-1), 6)  -> 'YYMMDD'

    Filters:
      - years: [2024, 2025] -> filters date6 year part ('24','25')
      - month_range: ('2024-01','2025-12') -> date6 BETWEEN '240101' AND '251231'
      - day_range: ('2024-04-22','2024-06-30') -> date6 BETWEEN '240422' AND '240630'

    Output JSON includes:
      - sampletypes_table: counts by sample type
      - labs_table: counts by lab code
      - years_table: counts by YY (e.g., {"24": 123, "25": 456})
      - months_table: counts by YYMM (e.g., {"2401": 50, "2402": 61, ...})
    """
    project_id = _normalize_project_id(config, project)

    conn = config._db_conn or config._connect_db(env="prod")
    if conn is None:
        return {"ok": False, "error": "DB connection failed"}

    # SQL expression for YYMMDD extracted from UUID
    date6_expr = "LEFT(SUBSTRING_INDEX(SUBSTRING_INDEX(s.uuid, '-', 2), '-', -1), 6)"

    conditions: list[str] = []
    params: list = []

    if project_id is not None:
        conditions.append("ps.project_id = %s")
        params.append(project_id)

    if years:
        yy_list = _normalize_years(years)
        placeholders = ", ".join(["%s"] * len(yy_list))
        conditions.append(f"LEFT({date6_expr}, 2) IN ({placeholders})")
        params.extend(yy_list)

    if month_range:
        start6, end6 = _month_range_to_yymmdd_bounds(month_range)
        conditions.append(f"{date6_expr} BETWEEN %s AND %s")
        params.extend([start6, end6])

    if day_range:
        start6, end6 = _day_range_to_yymmdd_bounds(day_range)
        conditions.append(f"{date6_expr} BETWEEN %s AND %s")
        params.extend([start6, end6])

    where_clause = " AND ".join(conditions) if conditions else "1=1"
    query = f"""
    SELECT
      ps.project_id,
      ps.sample_id,
      s.uuid
    FROM
      seek_production.projects_samples ps
    JOIN
      seek_production.samples s
        ON ps.sample_id = s.id
    WHERE
      {where_clause};
    """.strip()

    outputs_root = Path(outputs_root)
    outputs_root.mkdir(exist_ok=True)

    cursor = None
    try:
        cursor = conn.cursor(dictionary=True)
        print("[REPORTER][SQL] Running project sample report query", {"query": query, "params": params})
        cursor.execute(query, params)
        rows = cursor.fetchall() or []
        uuids = [r["uuid"] for r in rows if r.get("uuid")]

        # --- Build summary tables from UID parsing ---
        # Format assumed: <SAMPLETYPE>-<YYMMDD><LAB>-<INCREMENT>
        uid_re = re.compile(
            r"^(?P<sampletype>[^-]+)-(?P<yymmdd>\d{6})(?P<lab>[A-Za-z]+)-(?P<inc>\d+)$"
        )

        sampletype_counts: dict[str, int] = {}
        lab_counts: dict[str, int] = {}
        year_counts: dict[str, int] = {}
        month_counts: dict[str, int] = {}
        unparsable_count = 0

        for uid in uuids:
            m = uid_re.match(str(uid))
            if not m:
                unparsable_count += 1
                continue

            stype = m.group("sampletype")
            lab = m.group("lab")
            yymmdd = m.group("yymmdd")  # "YYMMDD"
            yy = yymmdd[:2]
            yymm = yymmdd[:4]

            sampletype_counts[stype] = sampletype_counts.get(stype, 0) + 1
            lab_counts[lab] = lab_counts.get(lab, 0) + 1
            year_counts[yy] = year_counts.get(yy, 0) + 1
            month_counts[yymm] = month_counts.get(yymm, 0) + 1

        # Sort tables for readability
        sampletypes_table = dict(sorted(sampletype_counts.items(), key=lambda kv: (-kv[1], kv[0])))
        labs_table = dict(sorted(lab_counts.items(), key=lambda kv: (-kv[1], kv[0])))
        years_table = dict(sorted(year_counts.items(), key=lambda kv: (-kv[1], kv[0])))
        # Months are often nicer sorted chronologically; YYMM sorts lexicographically as desired
        months_table = dict(sorted(month_counts.items(), key=lambda kv: kv[0]))

        # --- Write report artifact (JSON) ---
        # Write directly to outputs_root (no reports/ subdirectory)
        ts = datetime.now().strftime("%Y%m%dT%H%M%S")
        project_label = project_id if project_id is not None else "all"
        report_filename = f"project_{project_label}_{ts}.uuids.json"

        payload = {
            "project_id": project_id,
            "generated_at": datetime.now().isoformat(),
            "filters": {
                "project": project,
                "years": years,
                "month_range": month_range,
                "day_range": day_range,
            },
            "rows_returned": len(rows),
            "uuids": uuids,
            "sampletypes_table": sampletypes_table,
            "labs_table": labs_table,
            "years_table": years_table,
            "months_table": months_table,
            "parsing_notes": {
                "uid_format": "<SAMPLETYPE>-<YYMMDD><LAB>-<INCREMENT>",
                "regex": uid_re.pattern,
                "unparsable_uids": unparsable_count,
            },
        }

        report_entry = ArtifactStore(outputs_root).write_json(
            key="uuid_report_file",
            label="Samples report JSON",
            filename=report_filename,
            payload=payload,
            kind="report",
        )
        report_path = report_entry["path"] if report_entry else None

        # When a specific project was requested and zero rows came back,
        # probe whether the project exists in THIS DB and whether it has
        # ANY rows. This lets the chatter distinguish "no samples in the
        # requested period" from "this DB has no data for that project at
        # all" (common when MYSQL_HOST_PROD is aliased to a local dev DB).
        db_diagnostic: dict[str, Any] = {}
        if project_id is not None and len(rows) == 0:
            try:
                probe_cursor = conn.cursor(dictionary=True)
                probe_cursor.execute(
                    "SELECT (SELECT COUNT(*) FROM seek_production.projects "
                    "WHERE id=%s) AS project_exists, "
                    "(SELECT COUNT(*) FROM seek_production.projects_samples "
                    "WHERE project_id=%s) AS total_for_project",
                    [project_id, project_id],
                )
                probe = probe_cursor.fetchone() or {}
                probe_cursor.close()
                project_exists = bool(probe.get("project_exists"))
                total_for_project = int(probe.get("total_for_project") or 0)
                db_diagnostic = {
                    "project_exists_in_db": project_exists,
                    "total_rows_for_project": total_for_project,
                    "likely_missing_data": (not project_exists) or total_for_project == 0,
                }
            except Exception as probe_err:
                db_diagnostic = {"probe_error": repr(probe_err)}

        return {
            "ok": True,
            "project_id": project_id,
            "rows_returned": len(rows),
            "uuids_saved": len(uuids),
            "uuid_report_file": report_path,
            "uuid_preview": uuids[:10],
            "sampletypes_table": sampletypes_table,
            "labs_table": labs_table,
            "years_table": years_table,
            "months_table": months_table,
            "unparsable_uids": unparsable_count,
            "db_diagnostic": db_diagnostic,
        }

    except Exception as e:
        return {"ok": False, "error": repr(e)}


def run_project_protocols_report(
    config,
    project: int | str | None = None,
    years: list[int | str] | None = None,
    month_range: tuple[str, str] | None = None,
    day_range: tuple[str, str] | None = None,
    outputs_root: str | Path = "outputs",
) -> dict:
    """
    Project-scoped SOP/protocol reporting with optional date filters extracted from title.
    If project is None, runs across all projects.

    Title format: P.<LAB>-<YYMMDD>-<rest>  (e.g. P.SAS-240827-V1_RSTR_BMDM_protocol.docx)

    Date extraction used in SQL:
      date6 = LEFT(SUBSTRING_INDEX(SUBSTRING_INDEX(sop.title, '-', 2), '-', -1), 6)  -> 'YYMMDD'

    Filters:
      - years: [2024, 2025] -> filters date6 year part ('24','25')
      - month_range: ('2024-01','2025-12') -> date6 BETWEEN '240101' AND '251231'
      - day_range: ('2024-04-22','2024-06-30') -> date6 BETWEEN '240422' AND '240630'

    Output includes:
      - labs_table: counts by lab code (e.g. {"SAS": 12, "DFC": 5})
      - years_table: counts by YY
      - months_table: counts by YYMM
    """
    project_id = _normalize_project_id(config, project)

    conn = config._db_conn or config._connect_db(env="prod")
    if conn is None:
        return {"ok": False, "error": "DB connection failed"}

    # Title format: P.<LAB>-<YYMMDD>-<rest>
    # Second '-'-delimited segment is always YYMMDD
    date6_expr = "LEFT(SUBSTRING_INDEX(SUBSTRING_INDEX(sop.title, '-', 2), '-', -1), 6)"

    conditions: list[str] = []
    params: list = []

    if project_id is not None:
        conditions.append("ps.project_id = %s")
        params.append(project_id)

    if years:
        yy_list = _normalize_years(years)
        placeholders = ", ".join(["%s"] * len(yy_list))
        conditions.append(f"LEFT({date6_expr}, 2) IN ({placeholders})")
        params.extend(yy_list)

    if month_range:
        start6, end6 = _month_range_to_yymmdd_bounds(month_range)
        conditions.append(f"{date6_expr} BETWEEN %s AND %s")
        params.extend([start6, end6])

    if day_range:
        start6, end6 = _day_range_to_yymmdd_bounds(day_range)
        conditions.append(f"{date6_expr} BETWEEN %s AND %s")
        params.extend([start6, end6])

    where_clause = " AND ".join(conditions) if conditions else "1=1"
    query = f"""
    SELECT
      ps.project_id,
      ps.sop_id,
      sop.title
    FROM
      seek_production.projects_sops ps
    JOIN
      seek_production.sops sop
        ON ps.sop_id = sop.id
    WHERE
      {where_clause};
    """.strip()

    outputs_root = Path(outputs_root)
    outputs_root.mkdir(exist_ok=True)

    cursor = None
    try:
        cursor = conn.cursor(dictionary=True)
        print("[REPORTER][SQL] Running project protocols report query", {"query": query, "params": params})
        cursor.execute(query, params)
        rows = cursor.fetchall() or []
        titles = [r["title"] for r in rows if r.get("title")]

        # Parse: P.<LAB>-<YYMMDD>-<rest>
        title_re = re.compile(r"^P\.(?P<lab>[^-]+)-(?P<yymmdd>\d{6})-(?P<rest>.*)$")

        lab_counts: dict[str, int] = {}
        year_counts: dict[str, int] = {}
        month_counts: dict[str, int] = {}
        unparsable_count = 0

        for title in titles:
            m = title_re.match(str(title))
            if not m:
                unparsable_count += 1
                continue

            lab = m.group("lab")
            yymmdd = m.group("yymmdd")
            yy = yymmdd[:2]
            yymm = yymmdd[:4]

            lab_counts[lab] = lab_counts.get(lab, 0) + 1
            year_counts[yy] = year_counts.get(yy, 0) + 1
            month_counts[yymm] = month_counts.get(yymm, 0) + 1

        labs_table = dict(sorted(lab_counts.items(), key=lambda kv: (-kv[1], kv[0])))
        years_table = dict(sorted(year_counts.items(), key=lambda kv: (-kv[1], kv[0])))
        months_table = dict(sorted(month_counts.items(), key=lambda kv: kv[0]))

        ts = datetime.now().strftime("%Y%m%dT%H%M%S")
        project_label = project_id if project_id is not None else "all"
        report_filename = f"project_{project_label}_{ts}.protocols.json"

        payload = {
            "project_id": project_id,
            "generated_at": datetime.now().isoformat(),
            "filters": {
                "project": project,
                "years": years,
                "month_range": month_range,
                "day_range": day_range,
            },
            "rows_returned": len(rows),
            "titles": titles,
            "labs_table": labs_table,
            "years_table": years_table,
            "months_table": months_table,
            "parsing_notes": {
                "title_format": "P.<LAB>-<YYMMDD>-<rest>",
                "regex": title_re.pattern,
                "unparsable_titles": unparsable_count,
            },
        }

        report_entry = ArtifactStore(outputs_root).write_json(
            key="protocols_report",
            label="Protocols report JSON",
            filename=report_filename,
            payload=payload,
            kind="report",
        )
        report_path = report_entry["path"] if report_entry else None

        return {
            "ok": True,
            "summary_mode": "protocols",
            "project_id": project_id,
            "rows_returned": len(rows),
            "titles_saved": len(titles),
            "report_file": report_path,
            "titles_preview": titles[:10],
            "labs_table": labs_table,
            "years_table": years_table,
            "months_table": months_table,
            "unparsable_titles": unparsable_count,
        }

    except Exception as e:
        return {"ok": False, "error": repr(e)}


def run_project_published_report(  # noqa: C901
    config,
    project: int | str | None = None,
    years: list[int | str] | None = None,
    month_range: tuple[str, str] | None = None,
    day_range: tuple[str, str] | None = None,
    outputs_root: str | Path = "outputs",
) -> dict:
    """
    Project-scoped published samples and protocols report.

    Published samples and studies are determined via Neo4j: the production graph contains
    only data that has been published/submitted to public repositories.
    Traversal: (inv:Investigation) <-[:IN_INVESTIGATION]- (study:Study) <-[:IN_STUDY]- (s:Sample)
    Filtering by project uses case-insensitive CONTAINS on inv.title (normalized, spaces stripped).
    Date filtering uses the same YYMMDD substring from the sample UUID.

    Published protocols are determined by intersection:
      - Protocols for this project+date in seek_production
      - Protocol titles that also exist in seek_development (dev = published/submitted)

    Returns counts of published samples (by type/lab/year/month), study count, and protocol count.
    """
    project_id = _normalize_project_id(config, project)
    outputs_root = Path(outputs_root)
    outputs_root.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%dT%H%M%S")
    project_label = project_id if project_id is not None else "all"

    # ── 1. Published samples + studies via Neo4j ──────────────────────────────
    samples_result: dict = {}

    # Build Cypher date filters based on UUID substring YYMMDD
    # UUID format: <TYPE>-<YYMMDD><LAB>-<INC>
    cypher_conditions: list[str] = []
    cypher_params: dict = {}

    if project is not None:
        # Normalize project hint: lowercase, strip spaces for CONTAINS match
        hint = re.sub(r"\s+", "", str(project).lower())
        cypher_conditions.append(
            "replace(toLower(inv.title), ' ', '') CONTAINS $proj_hint"
        )
        cypher_params["proj_hint"] = hint

    if years:
        yy_list = _normalize_years(years)
        cypher_conditions.append("substring(split(s.uuid, '-')[1], 0, 2) IN $yy_list")
        cypher_params["yy_list"] = yy_list

    if month_range:
        start6, end6 = _month_range_to_yymmdd_bounds(month_range)
        cypher_conditions.append(
            "substring(split(s.uuid, '-')[1], 0, 6) >= $date6_start "
            "AND substring(split(s.uuid, '-')[1], 0, 6) <= $date6_end"
        )
        cypher_params["date6_start"] = start6
        cypher_params["date6_end"] = end6

    if day_range:
        start6, end6 = _day_range_to_yymmdd_bounds(day_range)
        cypher_conditions.append(
            "substring(split(s.uuid, '-')[1], 0, 6) >= $date6_start "
            "AND substring(split(s.uuid, '-')[1], 0, 6) <= $date6_end"
        )
        cypher_params["date6_start"] = start6
        cypher_params["date6_end"] = end6

    where_clause = (" WHERE " + " AND ".join(cypher_conditions)) if cypher_conditions else ""
    cypher = (
        f"MATCH (inv:Investigation)<-[:IN_INVESTIGATION]-(study:Study)<-[:IN_STUDY]-(s:Sample)"
        f"{where_clause} "
        f"RETURN s.uuid AS uuid, study.title AS study_title, s.type AS sampletype"
    )

    print("[REPORTER][NEO4J] Running published samples query", {"cypher": cypher, "params": cypher_params})
    neo4j_result = tool_neo4j_query(config, cypher, cypher_params)

    if not neo4j_result.get("ok"):
        samples_result = {"ok": False, "error": neo4j_result.get("error", "Neo4j query failed")}
    else:
        rows = neo4j_result.get("data") or []
        uids = [r["uuid"] for r in rows if r.get("uuid")]
        study_set = {r["study_title"] for r in rows if r.get("study_title")}

        uid_re = re.compile(
            r"^(?P<sampletype>[^-]+)-(?P<yymmdd>\d{6})(?P<lab>[A-Za-z]+)-(?P<inc>\d+)(-\w+)*$"
        )
        sampletype_counts: dict[str, int] = {}
        lab_counts: dict[str, int] = {}
        year_counts: dict[str, int] = {}
        month_counts: dict[str, int] = {}
        unparsable_count = 0

        for uid in uids:
            m = uid_re.match(str(uid))
            if not m:
                unparsable_count += 1
                continue
            stype = m.group("sampletype")
            lab = m.group("lab")
            yymmdd = m.group("yymmdd")
            yy = yymmdd[:2]
            yymm = yymmdd[:4]
            sampletype_counts[stype] = sampletype_counts.get(stype, 0) + 1
            lab_counts[lab] = lab_counts.get(lab, 0) + 1
            year_counts[yy] = year_counts.get(yy, 0) + 1
            month_counts[yymm] = month_counts.get(yymm, 0) + 1

        samples_result = {
            "ok": True,
            "rows_returned": len(uids),
            "study_count": len(study_set),
            "studies": sorted(study_set),
            "sampletypes_table": dict(sorted(sampletype_counts.items(), key=lambda kv: (-kv[1], kv[0]))),
            "labs_table": dict(sorted(lab_counts.items(), key=lambda kv: (-kv[1], kv[0]))),
            "years_table": dict(sorted(year_counts.items(), key=lambda kv: (-kv[1], kv[0]))),
            "months_table": dict(sorted(month_counts.items(), key=lambda kv: kv[0])),
            "unparsable_uids": unparsable_count,
        }

    # ── 2. Published protocols via prod ∩ dev MySQL ───────────────────────────
    protocols_result: dict = {}
    try:
        # Step A: prod titles for this project + date range
        prod_conn = config._db_conn or config._connect_db(env="prod")
        if prod_conn is None:
            protocols_result = {"ok": False, "error": "Prod DB connection failed"}
        else:
            date6_expr = "LEFT(SUBSTRING_INDEX(SUBSTRING_INDEX(sop.title, '-', 2), '-', -1), 6)"
            prod_conditions: list[str] = []
            prod_params: list = []

            if project_id is not None:
                prod_conditions.append("ps.project_id = %s")
                prod_params.append(project_id)

            if years:
                yy_list = _normalize_years(years)
                placeholders = ", ".join(["%s"] * len(yy_list))
                prod_conditions.append(f"LEFT({date6_expr}, 2) IN ({placeholders})")
                prod_params.extend(yy_list)

            if month_range:
                start6, end6 = _month_range_to_yymmdd_bounds(month_range)
                prod_conditions.append(f"{date6_expr} BETWEEN %s AND %s")
                prod_params.extend([start6, end6])

            if day_range:
                start6, end6 = _day_range_to_yymmdd_bounds(day_range)
                prod_conditions.append(f"{date6_expr} BETWEEN %s AND %s")
                prod_params.extend([start6, end6])

            prod_where = " AND ".join(prod_conditions) if prod_conditions else "1=1"
            prod_query = f"""
SELECT sop.title
FROM seek_production.projects_sops ps
JOIN seek_production.sops sop ON ps.sop_id = sop.id
WHERE {prod_where};
""".strip()

            print("[REPORTER][SQL] Published protocols — prod query", {"query": prod_query, "params": prod_params})
            cursor = prod_conn.cursor(dictionary=True)
            cursor.execute(prod_query, prod_params)
            prod_titles: set[str] = {r["title"] for r in (cursor.fetchall() or []) if r.get("title")}

            # Step B: all dev titles (no project/date filter — dev = published)
            dev_conn = config._connect_db(env="dev")
            if dev_conn is None:
                protocols_result = {"ok": False, "error": "Dev DB connection failed"}
            else:
                dev_query = "SELECT title FROM seek_production.sops WHERE title IS NOT NULL;"
                print("[REPORTER][SQL] Published protocols — dev query")
                dev_cursor = dev_conn.cursor(dictionary=True)
                dev_cursor.execute(dev_query)
                dev_titles: set[str] = {r["title"] for r in (dev_cursor.fetchall() or []) if r.get("title")}

                published_titles = sorted(prod_titles & dev_titles)

                title_re = re.compile(r"^P\.(?P<lab>[^-]+)-(?P<yymmdd>\d{6})-(?P<rest>.*)$")
                lab_counts_p: dict[str, int] = {}
                year_counts_p: dict[str, int] = {}
                month_counts_p: dict[str, int] = {}
                unparsable_p = 0

                for title in published_titles:
                    m = title_re.match(title)
                    if not m:
                        unparsable_p += 1
                        continue
                    lab = m.group("lab")
                    yymmdd = m.group("yymmdd")
                    yy = yymmdd[:2]
                    yymm = yymmdd[:4]
                    lab_counts_p[lab] = lab_counts_p.get(lab, 0) + 1
                    year_counts_p[yy] = year_counts_p.get(yy, 0) + 1
                    month_counts_p[yymm] = month_counts_p.get(yymm, 0) + 1

                protocols_result = {
                    "ok": True,
                    "rows_returned": len(published_titles),
                    "titles": published_titles,
                    "labs_table": dict(sorted(lab_counts_p.items(), key=lambda kv: (-kv[1], kv[0]))),
                    "years_table": dict(sorted(year_counts_p.items(), key=lambda kv: (-kv[1], kv[0]))),
                    "months_table": dict(sorted(month_counts_p.items(), key=lambda kv: kv[0])),
                    "unparsable_titles": unparsable_p,
                }
    except Exception as e:
        protocols_result = {"ok": False, "error": repr(e)}

    # ── 3. Write artifact + return ─────────────────────────────────────────────
    report_filename = f"project_{project_label}_{ts}.published.json"
    payload = {
        "project_id": project_id,
        "generated_at": datetime.now().isoformat(),
        "filters": {"project": project, "years": years, "month_range": month_range, "day_range": day_range},
        "samples": samples_result,
        "protocols": protocols_result,
    }
    report_entry = ArtifactStore(outputs_root).write_json(
        key="published_report",
        label="Published report JSON",
        filename=report_filename,
        payload=payload,
        kind="report",
    )
    report_path = report_entry["path"] if report_entry else None

    ok = samples_result.get("ok", False) or protocols_result.get("ok", False)
    return {
        "ok": ok,
        "summary_mode": "published",
        "project_id": project_id,
        "report_file": report_path,
        "samples": samples_result,
        "protocols": protocols_result,
    }


def annotate_metadata_with_sampletypes(config, metadata: dict) -> dict:
    """
    Enrich a metadata payload with human-friendly sample type names/descriptions from the cached lookup.
    Traverses entries safely and returns the original structure when shape mismatches occur.
    Keeps the function tolerant of missing fields so reporter flows do not fail on partial data.
    """
    _SAMPLETYPE_LOOKUP: dict[str, dict[str, str | None]] = {
        row.get("SampleType"): {"name": row.get("Name"), "description": row.get("Description")}
        for row in (config.MIN_SAMPLETYPES or [])
        if isinstance(row, dict) and row.get("SampleType")
    }
    try:
        data_block = metadata.get("data", {})
        entries = data_block.get("data", [])
        if not isinstance(entries, list):
            return metadata
        new_entries = []
        for entry in entries:
            if not isinstance(entry, dict):
                new_entries.append(entry)
                continue
            code = entry.get("sample_type")
            info = _SAMPLETYPE_LOOKUP.get(code, {})
            enriched = dict(entry)
            if info:
                enriched["sample_type_name"] = info.get("name")
                enriched["sample_type_description"] = info.get("description")
            new_entries.append(enriched)
        new_data_block = dict(data_block)
        new_data_block["data"] = new_entries
        new_metadata = dict(metadata)
        new_metadata["data"] = new_data_block
        return new_metadata
    except Exception as e:
        print("[DEBUG][REPORTER_META] Failed to annotate sampletypes:", repr(e))
        return metadata


def fetch_reporter_metadata(config, uids: list[str]) -> dict:
    """
    Fetch full sample metadata (and lineage) for provided UIDs using the admin retrieve endpoint.
    Returns the raw request result so callers can inspect ok/status and enrich downstream reports.
    Logs debug hints but tolerates failures by returning an error payload instead of raising.
    """
    if not uids:
        return {"ok": False, "error": "No UIDs provided for metadata retrieval"}

    print("[DEBUG][REPORTER_META] Fetching metadata for UIDs:", uids)
    try:
        result = tool_nextseek_api_request(
            config,
            endpoint="/nextseek_api/admin/samples/retrieve/",
            method="POST",
            requestBody={"identifiers": uids},
            queryParameters={},
        )
        print("[DEBUG][REPORTER_META] Metadata fetch ok:", result.get("ok"), "status:", result.get("status_code"))
        return result
    except Exception as e:
        print("[DEBUG][REPORTER_META] Metadata fetch exception:", repr(e))
        return {"ok": False, "error": repr(e)}


def load_report_template(config, report_type: str | None) -> dict:
    """
    Load a report template JSON from reports/ based on report_type name.
    Returns an empty dict on missing files or parse errors to let report writer proceed with defaults.
    """
    template_basename = get_report_template_basename(report_type)
    if not template_basename:
        return {}
    try:
        path = Path(config.BASE_DIR) / "reports" / f"{template_basename}.json"
        if path.exists():
            return json.loads(path.read_text())
    except Exception as e:
        print("[DEBUG][REPORT_WRITER] Failed to load report template:", repr(e))
    return {}


def normalize_report_type(report_type: str | None) -> str | None:
    """
    Normalize user/model-facing report type aliases into canonical internal labels.
    Returns None for empty values so callers can preserve "not specified" semantics.
    """
    if not isinstance(report_type, str):
        return None
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", report_type).strip("_").upper()
    if not cleaned:
        return None

    alias_map = {
        "GEO": "GEO",
        "PRIDE": "PRIDE",
        "SRA": "SRA",
        "NFCORE": "NFCORE",
        "NF_CORE": "NFCORE",
        "NFCORE_RNASEQ": "NFCORE_RNASEQ",
        "NFCORE_RNASEQ_SAMPLESHEET": "NFCORE_RNASEQ",
        "NF_CORE_RNASEQ": "NFCORE_RNASEQ",
        "NFCORE_SCRNASEQ": "NFCORE_SCRNASEQ",
        "NFCORE_SCRNASEQ_SAMPLESHEET": "NFCORE_SCRNASEQ",
        "NF_CORE_SCRNASEQ": "NFCORE_SCRNASEQ",
        "NFCORE_ATACSEQ": "NFCORE_ATACSEQ",
        "NF_CORE_ATACSEQ": "NFCORE_ATACSEQ",
        "NFCORE_CHIPSEQ": "NFCORE_CHIPSEQ",
        "NF_CORE_CHIPSEQ": "NFCORE_CHIPSEQ",
        "NFCORE_SAREK": "NFCORE_SAREK",
        "NF_CORE_SAREK": "NFCORE_SAREK",
        "NFCORE_METHYLSEQ": "NFCORE_METHYLSEQ",
        "NF_CORE_METHYLSEQ": "NFCORE_METHYLSEQ",
        "NFCORE_AMPLISEQ": "NFCORE_AMPLISEQ",
        "NF_CORE_AMPLISEQ": "NFCORE_AMPLISEQ",
        "NFCORE_FETCHNGS": "NFCORE_FETCHNGS",
        "NF_CORE_FETCHNGS": "NFCORE_FETCHNGS",
    }
    return alias_map.get(cleaned, cleaned)


def nfcore_pipeline_from_report_type(report_type: str | None) -> str | None:
    """If report_type is NFCORE_<PIPELINE>, return the lowercase pipeline key.
    Generic NFCORE returns None (caller should run pipeline selector).
    """
    canonical = normalize_report_type(report_type)
    if not canonical or not canonical.startswith("NFCORE"):
        return None
    if canonical == "NFCORE":
        return None
    return canonical[len("NFCORE_"):].lower()


def get_report_template_basename(report_type: str | None) -> str | None:
    """
    Map canonical report types to the JSON template basename stored in reports/.
    """
    canonical = normalize_report_type(report_type)
    if not canonical:
        return None
    static = {
        "GEO": "GEO-updated",
        "PRIDE": "pride",
        "SRA": "SRA",
    }
    if canonical in static:
        return static[canonical]
    if canonical.startswith("NFCORE_"):
        return f"nfcore/{canonical[len('NFCORE_'):].lower()}"
    if canonical == "NFCORE":
        return "nfcore/rnaseq"  # generic default; pipeline selector overrides upstream
    return canonical.lower()


def _scalar_for_summary(value: Any) -> str | None:
    """Coerce a metadata value to a short string for summary display.
    Returns None for empties so they don't count as populated."""
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    if isinstance(value, (list, tuple)):
        if not value:
            return None
        return json.dumps(value, default=str)[:120]
    if isinstance(value, dict):
        if not value:
            return None
        return json.dumps(value, default=str)[:120]
    return str(value)[:120]


def _entries_from_metadata(md: Any) -> list:
    """Find the list of sample-type blocks regardless of one or two levels of wrapping.

    tool_nextseek_api_request returns {"ok":..., "data": <body>} where body is
    {"total_samples":..., "data":[{sample_type, samples:[...]}]}.
    """
    if not isinstance(md, dict):
        return []
    inner = md.get("data")
    if isinstance(inner, dict):
        entries = inner.get("data")
        if isinstance(entries, list):
            return entries
    if isinstance(inner, list):
        return inner
    return []


def build_metadata_summary(metadata_map: dict | None) -> dict[str, Any]:
    """Walk per-UID metadata bundles and produce a sample-type-keyed summary
    of distinct field names + value-variation hints.

    Lineage edges are computed by following each sample's `Parent` metadata
    field up the chain (NExtSEEK's lineage representation). A `uid_index`
    is also returned to support post-hoc lineage filtering.

    Output shape:
        {
          "by_sample_type": {
            "<sample_type>": {
              "n_samples": int,
              "fields": {<name>: {n_populated, n_distinct, examples}}
            }
          },
          "lineage_edges": ["<parent_st> → <child_st>", ...],
          "_uid_index": {uid: {sample_type, parent_uid, metadata}}   # used by filter helpers
        }
    """
    by_st: dict[str, dict] = {}
    uid_index: dict[str, dict[str, Any]] = {}

    def _record(sample_type: str, sample_meta: dict | None) -> None:
        st = by_st.setdefault(sample_type, {"n_samples": 0, "fields": {}})
        st["n_samples"] += 1
        if not isinstance(sample_meta, dict):
            return
        for k, v in sample_meta.items():
            if not isinstance(k, str) or k.startswith("_"):
                continue
            scalar = _scalar_for_summary(v)
            field = st["fields"].setdefault(k, {"n_populated": 0, "_seen": set(), "examples": []})
            if scalar is not None:
                field["n_populated"] += 1
                if scalar not in field["_seen"]:
                    field["_seen"].add(scalar)
                    if len(field["examples"]) < 3:
                        field["examples"].append(scalar)

    # Pass 1: collect every sample's (sample_type, parent_uid, metadata)
    for md in (metadata_map or {}).values():
        for st_block in _entries_from_metadata(md):
            if not isinstance(st_block, dict):
                continue
            top_st = st_block.get("sample_type") or "unknown"
            for sample in st_block.get("samples") or []:
                if not isinstance(sample, dict):
                    continue
                meta = sample.get("metadata") or {}
                uid = (
                    (isinstance(meta, dict) and meta.get("UID"))
                    or sample.get("uuid")
                    or sample.get("uid")
                )
                parent_uid = meta.get("Parent") if isinstance(meta, dict) else None
                if uid:
                    uid_index[uid] = {
                        "sample_type": top_st,
                        "parent_uid": parent_uid,
                        "metadata": meta,
                    }
                _record(top_st, meta)
                # Also descend into any `children` array (legacy / other
                # responses that nest lineage there)
                for related in sample.get("children") or []:
                    if not isinstance(related, dict):
                        continue
                    cmeta = related.get("metadata") or {}
                    cst = (
                        related.get("sample_type")
                        or related.get("sampletype")
                        or related.get("type")
                        or "unknown"
                    )
                    cuid = (
                        (isinstance(cmeta, dict) and cmeta.get("UID"))
                        or related.get("uuid")
                    )
                    cparent = cmeta.get("Parent") if isinstance(cmeta, dict) else None
                    if cuid:
                        uid_index.setdefault(cuid, {
                            "sample_type": cst,
                            "parent_uid": cparent,
                            "metadata": cmeta,
                        })
                    _record(cst, cmeta)

    # Pass 2: derive lineage edges via Parent links — child.parent_uid → parent_st → child_st
    lineage_edges: set[str] = set()
    for uid, entry in uid_index.items():
        parent_uid = entry.get("parent_uid")
        if not parent_uid:
            continue
        parent = uid_index.get(parent_uid)
        if not parent:
            continue
        parent_st = parent.get("sample_type")
        child_st = entry.get("sample_type")
        if parent_st and child_st and parent_st != child_st:
            lineage_edges.add(f"{parent_st} → {child_st}")

    # Finalize: drop the internal _seen sets and compute n_distinct
    for st_data in by_st.values():
        for fname, fdata in st_data["fields"].items():
            fdata["n_distinct"] = len(fdata.pop("_seen"))
    return {
        "by_sample_type": by_st,
        "lineage_edges": sorted(lineage_edges),
        "_uid_index": uid_index,
    }


_SEQUENCING_SAMPLE_TYPE_HINTS = ("D.SEQ", "D.SEQUENCING", "SEQ")


def _is_sequencing_type(sample_type: str) -> bool:
    """Heuristic: which NExtSEEK sample types represent SEQUENCING data
    (where LibraryStrategy / accessions live)? Currently D.SEQ. Extend the
    hint list if other sequencing data types appear."""
    if not isinstance(sample_type, str):
        return False
    st = sample_type.upper()
    return any(st == h or st.startswith(f"{h}.") for h in _SEQUENCING_SAMPLE_TYPE_HINTS)


def filter_summary_to_sequencing_lineage(summary: dict[str, Any]) -> dict[str, Any]:
    """Return a slimmer summary with only the sequencing sample types and their
    LINEAGE ANCESTORS (samples that appear upstream via the `Parent` chain).

    For pipeline selection, only D.SEQ + ancestors are useful — biology fields
    live on the upstream samples (TIS, NHP, RNA Sample, etc.). Downstream
    artifacts (A.SCXP, A.FLOW, D.IMG, BAC, etc.) and parallel sample types
    contribute noise.

    If no sequencing sample types are present, returns the original summary
    unchanged (defensive fallback).
    """
    by_st = (summary or {}).get("by_sample_type") or {}
    uid_index = (summary or {}).get("_uid_index") or {}

    seq_types = {st for st in by_st.keys() if _is_sequencing_type(st)}
    if not seq_types:
        return summary

    # Walk Parent chains up from every sequencing sample to collect ancestor types
    keep_types: set[str] = set(seq_types)
    for uid, entry in uid_index.items():
        if entry.get("sample_type") not in seq_types:
            continue
        cursor = entry.get("parent_uid")
        seen: set[str] = set()
        while cursor and cursor not in seen:
            seen.add(cursor)
            parent = uid_index.get(cursor)
            if not parent:
                break
            keep_types.add(parent.get("sample_type"))
            cursor = parent.get("parent_uid")

    filtered_by_st = {
        st: data for st, data in by_st.items() if st in keep_types
    }
    # Keep only edges entirely within the kept set
    filtered_edges = [
        e for e in (summary.get("lineage_edges") or [])
        if all(side.strip() in keep_types for side in e.split("→"))
    ]
    return {
        "by_sample_type": filtered_by_st,
        "lineage_edges": filtered_edges,
        "_uid_index": uid_index,
    }


def build_accession_metadata_lookup(metadata_map: dict | None) -> dict[str, dict[str, Any]]:
    """For each sample, walk its lineage UPWARD via the `Parent` metadata
    field, flattening biology fields from upstream samples (TIS, NHP, PAV…)
    onto the leaf sample's flat metadata. Index by every accession found.

    Returns: {accession: flat_metadata_dict}.

    Flatten precedence: leaf sample's own metadata wins; upstream parents
    fill in missing keys but never overwrite. This way `UID` / `Strandedness`
    on D.SEQ stay D.SEQ's, while biology fields like `Treatment1` (which
    live on TIS or NHP upstream) get pulled in.

    Also walks legacy `children` arrays for older API responses that nest
    lineage there.
    """
    out: dict[str, dict[str, Any]] = {}
    if not isinstance(metadata_map, dict):
        return out

    # Build a uid → (sample_dict, parent_uid) index across the entire bundle
    uid_to_sample: dict[str, dict[str, Any]] = {}

    def _walk_collect(sample: Any) -> None:
        if not isinstance(sample, dict):
            return
        meta = sample.get("metadata")
        uid = (
            (isinstance(meta, dict) and meta.get("UID"))
            or sample.get("uuid")
            or sample.get("uid")
        )
        if uid:
            uid_to_sample[uid] = sample
        for child in sample.get("children") or []:
            _walk_collect(child)

    for md in metadata_map.values():
        for st_block in _entries_from_metadata(md):
            if not isinstance(st_block, dict):
                continue
            for sample in st_block.get("samples") or []:
                _walk_collect(sample)

    def _flatten_with_parents(sample: dict) -> dict[str, Any]:
        flat: dict[str, Any] = {}
        seen_uids: set[str] = set()
        cursor: dict[str, Any] | None = sample
        while isinstance(cursor, dict):
            meta = cursor.get("metadata") or {}
            cur_uid = meta.get("UID") if isinstance(meta, dict) else None
            if cur_uid and cur_uid in seen_uids:
                break  # cycle guard
            if cur_uid:
                seen_uids.add(cur_uid)
            if isinstance(meta, dict):
                for k, v in meta.items():
                    if isinstance(k, str) and k not in flat and v not in (None, ""):
                        flat[k] = v
            # Walk legacy `children` (some bundles nest lineage there)
            for child in cursor.get("children") or []:
                if isinstance(child, dict):
                    cmeta = child.get("metadata") or {}
                    if isinstance(cmeta, dict):
                        for k, v in cmeta.items():
                            if isinstance(k, str) and k not in flat and v not in (None, ""):
                                flat[k] = v
            # Walk upward via Parent UID
            parent_uid = (
                meta.get("Parent") if isinstance(meta, dict) else None
            )
            if not parent_uid or parent_uid in seen_uids:
                break
            cursor = uid_to_sample.get(parent_uid)
        return flat

    for sample in uid_to_sample.values():
        flat = _flatten_with_parents(sample)
        accessions: set[str] = set()
        for v in flat.values():
            if isinstance(v, str):
                for m in re.findall(
                    r"\b(?:SRR|SRX|SRP|SRS|ERR|ERX|ERP|ERS|DRR|DRX|DRP|DRS|GSE|GSM)\d+\b",
                    v,
                ):
                    accessions.add(m)
        for acc in accessions:
            out.setdefault(acc, flat)
    return out


def filter_summary_for_deg(summary: dict[str, Any]) -> dict[str, Any]:
    """Return a slimmer summary keeping only fields that are good DEG candidates.

    Drops:
    - Fields with no populated values (n_distinct == 0).
    - Uniform fields (n_distinct == 1) — always useless for contrasts, regardless
      of sample count.
    - Per-sample-unique fields (n_distinct == n_samples) — only when n_samples >= 3,
      because with very small n every field tends to look unique.
    """
    filtered: dict[str, dict] = {}
    for st, st_data in (summary or {}).get("by_sample_type", {}).items():
        n = st_data.get("n_samples", 0)
        keep: dict[str, dict] = {}
        for fname, fdata in (st_data.get("fields") or {}).items():
            nd = fdata.get("n_distinct", 0)
            if nd <= 1:
                # No variation → skip
                continue
            if n >= 3 and nd == n:
                # Per-sample unique → likely an ID, not a contrast variable
                continue
            keep[fname] = fdata
        if keep:
            filtered[st] = {"n_samples": n, "fields": keep}
    return {
        "by_sample_type": filtered,
        "lineage_edges": (summary or {}).get("lineage_edges") or [],
    }


def extract_protocol_refs_from_metadata(metadata: dict) -> list[dict[str, str]]:
    """
    Walk metadata dict and collect supported protocol references from any key named 'Protocol'.
    Supported references:
    - fairdata-dev / fairdata URLs pointing at /sops/{id-or-name}
    - fairdomhub URLs pointing at /sops/{id-or-name}
    - direct protocol names beginning with 'P.'
    Unsupported external/vendor URLs are ignored.
    """
    refs: dict[tuple[str, str], dict[str, str]] = {}

    def _add(source: str, value: str, raw_value: str) -> None:
        clean = value.strip()
        if not clean:
            return
        refs[(source, clean)] = {"source": source, "value": clean, "raw": raw_value}

    def _classify(value: Any) -> None:
        if isinstance(value, (int, float)):
            _add("fairdata-dev", str(value), str(value))
            return
        if not isinstance(value, str):
            return

        raw = value.strip()
        if not raw:
            return

        parsed = urlparse(raw)
        host = (parsed.netloc or "").lower()
        path = parsed.path or ""
        sop_match = re.search(r"/sops/([^/?#]+)", path, flags=re.IGNORECASE)

        if host in {"fairdata-dev.mit.edu", "fairdata.mit.edu"} and sop_match:
            _add(host, sop_match.group(1), raw)
            return
        if host == "fairdomhub.org" and sop_match:
            _add(host, sop_match.group(1), raw)
            return
        if re.match(r"^P\.[A-Za-z0-9._-]+$", raw):
            _add("protocol_name", raw, raw)
            return

    def _walk(node):
        if isinstance(node, dict):
            for k, v in node.items():
                if isinstance(k, str) and k.lower() == "protocol":
                    _classify(v)
                _walk(v)
        elif isinstance(node, list):
            for item in node:
                _walk(item)

    _walk(metadata)
    return [refs[key] for key in sorted(refs)]


def _request_protocol_record(config: ChatConfig, base_url: str, protocol_ref: str) -> dict:
    """
    Fetch a protocol record from a specific host using the SOP detail endpoint.
    Accepts numeric ids or protocol names.
    """
    host = (urlparse(base_url).netloc or "").lower()
    if host == "fairdomhub.org":
        url = f"{base_url.rstrip('/')}/sops/{quote(protocol_ref, safe='')}/"
    else:
        url = f"{base_url.rstrip('/')}/nextseek_api/sops/{quote(protocol_ref, safe='')}/"
    auth = None
    headers: dict[str, str] = {}
    auth_mode = "None"

    if host == "fairdomhub.org":
        fdh_api = os.getenv("FDH_API")
        if fdh_api:
            headers["Authorization"] = f"Bearer {fdh_api}"
            headers["Accept"] = "application/json"
            auth_mode = "Bearer(FDH_API)"
    elif config.API_USER and config.API_PASS:
        auth = (config.API_USER, config.API_PASS)
        auth_mode = "Basic"

    print("[DEBUG][API] Request:")
    print("  METHOD: GET")
    print(f"  URL:    {url}")
    print("  PARAMS: {'page_size': 1000}")
    print("  BODY:   {}")
    print(f"  AUTH:   {auth_mode}")
    print("  TIMEOUT:90s")

    try:
        resp = requests.get(url, auth=auth, headers=headers or None, params={"page_size": 1000}, timeout=90)
        print("[DEBUG][API] Response:")
        print(f"  STATUS: {resp.status_code}")
        preview = resp.text[:300].replace("\n", " ")
        print(f"  PREVIEW: {preview!r}")
        try:
            data = resp.json()
        except Exception:
            data = {"_raw": resp.text[:1000]}
        return {
            "ok": resp.ok,
            "url": url,
            "status_code": resp.status_code,
            "method": "GET",
            "query": {"page_size": 1000},
            "body": {},
            "data": data,
            "source_base_url": base_url.rstrip("/"),
            "protocol_ref": protocol_ref,
        }
    except Exception as e:
        print(f"[DEBUG][API] Exception: {repr(e)}")
        return {
            "ok": False,
            "error": repr(e),
            "url": url,
            "method": "GET",
            "source_base_url": base_url.rstrip("/"),
            "protocol_ref": protocol_ref,
        }


def fetch_protocols(config, protocol_refs: list[dict[str, str]]) -> dict:
    """
    Fetch protocol details for classified metadata references.
    fairdata-dev/fairdata hosts are queried directly; fairdomhub uses its own host;
    P.* names are resolved against the configured NExtSEEK API host.
    Returns a mapping keyed by the protocol reference value.
    """
    results: dict[str, dict] = {}
    host_map = {
        "fairdata-dev.mit.edu": getattr(config, "NEXTSEEK_BASE_URL", "") or "https://nextseek-dev.mit.edu",
        "fairdata.mit.edu": getattr(config, "NEXTSEEK_BASE_URL", "") or "https://nextseek-dev.mit.edu",
        "fairdomhub.org": "https://fairdomhub.org",
        "fairdata-dev": getattr(config, "NEXTSEEK_BASE_URL", "") or "https://fairdata-dev.mit.edu",
        "protocol_name": getattr(config, "NEXTSEEK_BASE_URL", "") or "https://fairdata-dev.mit.edu",
    }

    for ref in protocol_refs or []:
        source = ref.get("source", "")
        value = ref.get("value", "")
        if not value:
            continue
        base_url = host_map.get(source)
        if not base_url:
            print("[DEBUG][REPORTER_PROTOCOL] Skipping unsupported protocol reference:", ref)
            continue
        try:
            resp = _request_protocol_record(config, base_url, value)
            resp["protocol_source"] = source
            resp["protocol_raw"] = ref.get("raw")
            results[value] = resp
            print("[DEBUG][REPORTER_PROTOCOL] Fetched protocol", value, "source:", source, "ok:", resp.get("ok"))
        except Exception as e:
            results[value] = {"ok": False, "error": repr(e), "protocol_source": source, "protocol_raw": ref.get("raw")}
            print("[DEBUG][REPORTER_PROTOCOL] Failed to fetch protocol", value, "source:", source, "err:", repr(e))
    return results


def _extract_docx_text(content: bytes) -> str | None:
    """
    Extract plain text from a DOCX binary by reading word/document.xml.
    Strips tags, unescapes entities, and returns None on failure so callers can continue gracefully.
    """
    try:
        with ZipFile(BytesIO(content)) as zf:
            with zf.open("word/document.xml") as f:
                xml = f.read().decode("utf-8", errors="ignore")
        # Strip tags and unescape XML entities
        text = re.sub(r"<[^>]+>", " ", xml)
        text = html.unescape(text)
        text = re.sub(r"\s+", " ", text).strip()
        return text
    except Exception as e:
        print("[DEBUG][REPORTER_PROTOCOL] docx extract failed:", repr(e))
        return None


def _extract_pdf_text(content: bytes) -> str | None:
    """
    Extract text from a PDF binary using PyPDF2 when available.
    Returns None when the library is missing or extraction fails, logging debug hints for diagnostics.
    """
    try:
        import PyPDF2
    except Exception:
        print("[DEBUG][REPORTER_PROTOCOL] PyPDF2 not available; skipping PDF text extraction.")
        return None
    try:
        reader = PyPDF2.PdfReader(BytesIO(content))
        parts = []
        for page in reader.pages:
            try:
                parts.append(page.extract_text() or "")
            except Exception:
                continue
        text = "\n".join(parts)
        text = re.sub(r"\s+", " ", text).strip()
        return text or None
    except Exception as e:
        print("[DEBUG][REPORTER_PROTOCOL] PDF extract failed:", repr(e))
        return None


def sanitize_protocols_for_llm(protocol_payloads: dict) -> dict:
    """
    Sanitize protocol payloads for LLM consumption:
    - Replace localhost URLs with fairdata-dev.mit.edu
    - Remove internal fields not needed by the LLM
    Returns a cleaned copy of the protocol payloads.
    Keeps URLs model-safe while preserving useful content for prompt context.
    """
    if not protocol_payloads:
        return {}

    def fix_url(url: str | None) -> str | None:
        if not url:
            return url
        if "localhost" in url or "127.0.0.1" in url:
            path_match = re.search(r"https?://[^/]+(/.*)", url)
            if path_match:
                return f"https://fairdata-dev.mit.edu{path_match.group(1)}"
        return url

    def sanitize_dict(d: dict) -> dict:
        result = {}
        for k, v in d.items():
            if isinstance(v, dict):
                result[k] = sanitize_dict(v)
            elif isinstance(v, list):
                result[k] = [sanitize_dict(i) if isinstance(i, dict) else fix_url(i) if isinstance(i, str) and ("localhost" in i or "127.0.0.1" in i) else i for i in v]
            elif isinstance(v, str) and ("localhost" in v or "127.0.0.1" in v):
                result[k] = fix_url(v)
            else:
                result[k] = v
        return result

    return {pid: sanitize_dict(resp) if isinstance(resp, dict) else resp for pid, resp in protocol_payloads.items()}


def download_and_extract_protocol_blobs(protocol_payloads: dict, base_dir: str | Path, config=None) -> dict:
    """
    For each protocol response, download attached files (content_blobs), save them under base_dir/protocols/files,
    and attempt to extract text (docx/pdf). Returns a mapping id -> list of file metadata with text.
    """
    store = ArtifactStore(base_dir)
    results: dict[str, list[dict]] = {}
    session = requests.Session()

    for pid, resp in (protocol_payloads or {}).items():
        files_out: list[dict] = []
        source_base_url = resp.get("source_base_url") if isinstance(resp, dict) else None
        source_host = (urlparse(source_base_url).netloc or "").lower() if source_base_url else ""
        if source_host == "fairdomhub.org":
            session.auth = None
            session.headers.pop("Authorization", None)
            fdh_api = os.getenv("FDH_API")
            if fdh_api:
                session.headers["Authorization"] = f"Bearer {fdh_api}"
        else:
            session.headers.pop("Authorization", None)
            session.auth = (config.API_USER, config.API_PASS) if config and config.API_USER and config.API_PASS else None
        # Response structure: {"ok": ..., "data": {"data": {"id": ..., "attributes": {"content_blobs": [...]}}}}
        attrs = resp.get("data", {}).get("data", {}).get("attributes", {}) if isinstance(resp, dict) else {}
        blobs = attrs.get("content_blobs") or []
        for idx, blob in enumerate(blobs):
            link = blob.get("link") or blob.get("url")
            # Fix localhost URLs - content blobs are served from fairdata-dev.mit.edu
            if link and ("localhost" in link or "127.0.0.1" in link):
                path_match = re.search(r"https?://[^/]+(/.*)", link)
                if path_match and source_base_url:
                    link = f"{source_base_url}{path_match.group(1)}"
            fname = blob.get("original_filename") or f"{pid}_{idx}"
            entry: dict = {"filename": fname, "content_type": blob.get("content_type"), "link": link}
            if not link:
                files_out.append({"filename": fname, "ok": False, "error": "No link in content_blob"})
                continue
            try:
                content_resp = None
                attempted = []
                for candidate in (f"{link}/download", f"{link}?download=1", link):
                    attempted.append(candidate)
                    r = session.get(candidate, timeout=30)
                    ctype_hdr = (r.headers.get("Content-Type") or "").lower()
                    looks_json = ctype_hdr.startswith("application/vnd.api+json") or r.content[:1] in (b"{", b"[")
                    is_ok = r.status_code == 200 and not looks_json
                    if is_ok:
                        content_resp = r
                        entry["status_code"] = r.status_code
                        entry["response_content_type"] = r.headers.get("Content-Type")
                        entry["download_url_used"] = candidate
                        break
                    entry["last_status"] = r.status_code
                    entry["last_response_content_type"] = r.headers.get("Content-Type")

                if content_resp is None:
                    entry.update({"ok": False, "error": f"Unable to download blob; tried {attempted}"})
                    files_out.append(entry)
                    continue

                artifact_entry = store.write_bytes(
                    key=f"protocol_blob_{pid}_{idx}",
                    label=fname,
                    filename=fname,
                    payload=content_resp.content,
                    kind="protocol",
                    subdir="files",
                    mime=blob.get("content_type"),
                )
                dest_path = artifact_entry["path"] if artifact_entry else None
                text = None
                text_error = None
                ctype = (blob.get("content_type") or "").lower()
                try:
                    if "pdf" in ctype or fname.lower().endswith(".pdf") or content_resp.content[:4] == b"%PDF":
                        text = _extract_pdf_text(content_resp.content)
                    elif "word" in ctype or fname.lower().endswith(".docx"):
                        text = _extract_docx_text(content_resp.content)
                    else:
                        # Fallback: try docx, then pdf
                        text = _extract_docx_text(content_resp.content) or _extract_pdf_text(content_resp.content)
                except Exception as e:
                    text_error = repr(e)

                # Truncate text to ~3000 tokens max
                text_truncated = False
                if text:
                    PROTOCOL_TOKEN_LIMIT = 3000
                    token_count = estimate_tokens_from_text(text)
                    if token_count > PROTOCOL_TOKEN_LIMIT:
                        # Truncate: ~4 chars per token
                        max_chars = PROTOCOL_TOKEN_LIMIT * 4
                        text = text[:max_chars] + "\n\n[... truncated, exceeded 3000 token limit ...]"
                        text_truncated = True

                entry.update(
                    {
                        "path": dest_path,
                        "md5": blob.get("md5sum"),
                        "sha1": blob.get("sha1sum"),
                        "size": blob.get("size"),
                        "ok": True,
                        "text": text,
                        "text_truncated": text_truncated,
                        "text_error": text_error,
                    }
                )
                files_out.append(entry)
            except Exception as e:
                entry.update({"ok": False, "error": repr(e)})
                files_out.append(entry)
        if files_out:
            results[str(pid)] = files_out
    return results


# ======================================================
# GEO report -> SEQ template export
# ======================================================

def _copy_row_format(ws: Worksheet, source_row: int, target_row: int) -> None:
    """
    Copy styling and height from source_row to target_row on the same sheet.
    Preserves number formats, comments, and hyperlinks so cloned rows match the template.
    Useful when expanding list sections without rebuilding formatting manually.
    """
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for cell in ws[source_row]:
        tgt = ws.cell(row=target_row, column=cell.col_idx)
        tgt._style = copy(cell._style)
        tgt.number_format = cell.number_format
        tgt._comment = cell._comment
        tgt.hyperlink = cell.hyperlink


def _write_cell(ws: Worksheet, row: int, col: int, value: Any) -> None:
    """
    Write a value into a worksheet cell unless the value is None.
    Keeps template defaults intact when optional fields are absent.
    """
    if value is None:
        return
    ws.cell(row=row, column=col).value = value


def _write_list_down(
    ws: Worksheet,
    start_row: int,
    col: int,
    values: Sequence[Any] | None,
    *,
    max_rows: int | None = None,
) -> None:
    """
    Write a sequence of values down a column, optionally packing overflow into the final row.
    Preserves existing cell content when condensing and skips empty values so templates stay clean.
    """
    if not values:
        return
    for idx, val in enumerate(values):
        if val in (None, ""):
            continue
        row = start_row + idx
        if max_rows and idx >= max_rows:
            row = start_row + max_rows - 1
            existing = ws.cell(row=row, column=col).value or ""
            separator = "\n" if existing else ""
            ws.cell(row=row, column=col).value = f"{existing}{separator}{val}"
            continue
        ws.cell(row=row, column=col).value = val


def _build_header_map(ws: Worksheet, header_row: int) -> dict[str, list[int]]:
    """
    Build a mapping of normalized header labels to column indices for a given header row.
    Supports duplicate headers by returning lists so repeated fields can be filled in order.
    """
    mapping: dict[str, list[int]] = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        label = _normalize_sheet_label(cell.value)
        mapping.setdefault(label, []).append(cell.col_idx)
    return mapping


def _normalize_geo_key(value: Any) -> str:
    """
    Normalize GEO template keys so starred/unstarred variants map to the same logical field.
    Collapses whitespace and lowercases labels to tolerate minor template or model variations.
    """
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"^[*#\s]+", "", text)
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_sheet_label(value: Any) -> str:
    """
    Normalize worksheet labels while preserving leading marker characters like '*'.
    This keeps starred and unstarred template columns distinct.
    """
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _geo_get(mapping: Mapping[str, Any] | None, *keys: str) -> Any:
    """
    Read a GEO field from a mapping using exact and normalized-key fallback.
    This lets the exporter accept both literal template keys ('*title') and logical keys ('title').
    """
    if not isinstance(mapping, Mapping):
        return None

    normalized = {_normalize_geo_key(key): value for key, value in mapping.items() if isinstance(key, str)}
    for key in keys:
        if key in mapping:
            return mapping[key]
        norm_key = _normalize_geo_key(key)
        if norm_key in normalized:
            return normalized[norm_key]
    return None


def _find_first_row_with_label(ws: Worksheet, label: str, *, col: int = 1) -> int | None:
    """
    Find the first row whose target column matches a label after GEO-key normalization.
    """
    target = _normalize_geo_key(label)
    for row_idx in range(1, ws.max_row + 1):
        if _normalize_geo_key(ws.cell(row=row_idx, column=col).value) == target:
            return row_idx
    return None


def _find_sample_header_row(ws: Worksheet) -> int:
    """
    Locate the sample header row from the required GEO sample columns.
    Falls back to the current template row if discovery fails.
    """
    required = {
        _normalize_geo_key("*library name"),
        _normalize_geo_key("*title"),
        _normalize_geo_key("*library strategy"),
        _normalize_geo_key("*organism"),
    }
    for row_idx in range(1, ws.max_row + 1):
        labels = {
            _normalize_geo_key(ws.cell(row=row_idx, column=col).value)
            for col in range(1, ws.max_column + 1)
            if ws.cell(row=row_idx, column=col).value not in (None, "")
        }
        if required.issubset(labels):
            return row_idx
    return 38


def _find_paired_end_header_row(ws: Worksheet) -> int:
    """
    Locate the paired-end table header row from the file-name columns.
    """
    required = {
        _normalize_geo_key("file name 1"),
        _normalize_geo_key("file name 2"),
        _normalize_geo_key("file name 3"),
        _normalize_geo_key("file name 4"),
    }
    for row_idx in range(1, ws.max_row + 1):
        labels = {
            _normalize_geo_key(ws.cell(row=row_idx, column=col).value)
            for col in range(1, 5)
            if ws.cell(row=row_idx, column=col).value not in (None, "")
        }
        if required.issubset(labels):
            return row_idx
    return 76


def _select_study_summary(study: Mapping[str, Any] | None) -> str | None:
    """
    Return the best available study summary, preferring 'summary (abstract)' then 'summary'.
    Keeps GEO population tolerant of missing fields while still returning None when nothing is present.
    """
    if not isinstance(study, Mapping):
        return None
    return _geo_get(study, "*summary (abstract)", "summary (abstract)", "summary") or None


def _populate_geo_seq_workbook(wb, report: Mapping[str, Any]) -> None:
    """
    Fill the GEO SEQ template workbook in-place from a single report entry.
    Writes study metadata, sample records, protocols, paired-end entries, and checksum info while preserving formats.
    Mutates the workbook directly so callers can immediately save it to disk.
    """
    meta_sheet = wb["Metadata"]
    study = report.get("study") or {}
    samples = report.get("samples") or []
    protocols = report.get("protocols") or {}
    paired_end_experiments = report.get("paired_end_experiments") or []
    checksums = (report.get("checksums") or {})

    # ---- Study section ----
    title_row = _find_first_row_with_label(meta_sheet, "*title") or 12
    summary_row = _find_first_row_with_label(meta_sheet, "*summary (abstract)") or 13
    design_row = _find_first_row_with_label(meta_sheet, "*experimental design") or 14
    contributor_row = _find_first_row_with_label(meta_sheet, "contributor") or 15
    supplementary_row = _find_first_row_with_label(meta_sheet, "supplementary file") or 22

    _write_cell(meta_sheet, title_row, 2, _geo_get(study, "*title", "title"))
    _write_cell(meta_sheet, summary_row, 2, _select_study_summary(study))
    _write_cell(meta_sheet, design_row, 2, _geo_get(study, "*experimental design", "experimental design"))

    contributors = _geo_get(study, "contributor") or []
    _write_list_down(meta_sheet, start_row=contributor_row, col=2, values=contributors, max_rows=7)

    supplementary = _geo_get(study, "supplementary file") or []
    _write_list_down(meta_sheet, start_row=supplementary_row, col=2, values=supplementary, max_rows=16)

    # ---- Samples section ----
    sample_header_row = _find_sample_header_row(meta_sheet)
    header_map = _build_header_map(meta_sheet, sample_header_row)
    sample_start_row = sample_header_row + 1
    sample_rows_available = max(0, 52 - sample_start_row + 1)

    extra_sample_rows = max(0, len(samples) - sample_rows_available)
    if extra_sample_rows:
        template_row = 52
        insert_at = template_row + 1
        for offset in range(extra_sample_rows):
            meta_sheet.insert_rows(insert_at + offset)
            _copy_row_format(meta_sheet, template_row, insert_at + offset)

    def set_sample_field(row_idx: int, header_key: str, value: Any, occurrence: int = 0) -> None:
        key = _normalize_sheet_label(header_key)
        cols = header_map.get(key)
        if not cols or occurrence >= len(cols) or value in (None, ""):
            return
        meta_sheet.cell(row=row_idx, column=cols[occurrence]).value = value

    for idx, sample in enumerate(samples):
        if not isinstance(sample, Mapping):
            continue
        row_idx = sample_start_row + idx
        set_sample_field(row_idx, "*library name", _geo_get(sample, "*library name", "library name"))
        set_sample_field(row_idx, "*title", _geo_get(sample, "*title", "title"))
        set_sample_field(row_idx, "*library strategy", _geo_get(sample, "*library strategy", "library strategy"))
        set_sample_field(row_idx, "*organism", _geo_get(sample, "*organism", "organism"))
        set_sample_field(row_idx, "**tissue", _geo_get(sample, "**tissue", "tissue"))
        set_sample_field(row_idx, "**cell line", _geo_get(sample, "**cell line", "cell line"))
        set_sample_field(row_idx, "**cell type", _geo_get(sample, "**cell type", "cell type"))
        set_sample_field(row_idx, "genotype", _geo_get(sample, "genotype"))
        set_sample_field(row_idx, "treatment", _geo_get(sample, "treatment"))
        set_sample_field(row_idx, "batch", _geo_get(sample, "batch"))
        set_sample_field(row_idx, "*molecule", _geo_get(sample, "*molecule", "molecule"))
        set_sample_field(row_idx, "*single or paired-end", _geo_get(sample, "*single or paired-end", "single or paired-end"))
        set_sample_field(row_idx, "*instrument model", _geo_get(sample, "*instrument model", "instrument model"))
        set_sample_field(row_idx, "description", _geo_get(sample, "description"))
        set_sample_field(row_idx, "processed data file", _geo_get(sample, "processed data file"), occurrence=0)
        set_sample_field(row_idx, "processed data file", _geo_get(sample, "processed data file (2)"), occurrence=1)
        set_sample_field(row_idx, "*raw file", _geo_get(sample, "*raw file", "raw file"), occurrence=0)
        set_sample_field(row_idx, "raw file", _geo_get(sample, "raw file"), occurrence=0)
        set_sample_field(row_idx, "raw file", _geo_get(sample, "raw file (2)"), occurrence=1)
        set_sample_field(row_idx, "raw file", _geo_get(sample, "raw file (3)"), occurrence=2)
        set_sample_field(row_idx, "raw file", _geo_get(sample, "raw file (4)"), occurrence=3)

    # ---- Protocols section ----
    growth_row = _find_first_row_with_label(meta_sheet, "growth protocol") or 57
    treatment_row = _find_first_row_with_label(meta_sheet, "treatment protocol") or 58
    extract_row = _find_first_row_with_label(meta_sheet, "*extract protocol") or 59
    library_row = _find_first_row_with_label(meta_sheet, "*library construction protocol") or 60
    base_data_processing_row = _find_first_row_with_label(meta_sheet, "*data processing step") or 62

    _write_cell(meta_sheet, growth_row, 2, _geo_get(protocols, "growth protocol"))
    _write_cell(meta_sheet, treatment_row, 2, _geo_get(protocols, "treatment protocol"))
    _write_cell(meta_sheet, extract_row, 2, _geo_get(protocols, "*extract protocol", "extract protocol"))
    _write_cell(meta_sheet, library_row, 2, _geo_get(protocols, "*library construction protocol", "library construction protocol"))

    data_processing_steps: list[Any] = []
    primary_data_processing = _geo_get(protocols, "*data processing step", "data processing step")
    if primary_data_processing not in (None, ""):
        if isinstance(primary_data_processing, list):
            data_processing_steps.extend(primary_data_processing)
        else:
            data_processing_steps.append(primary_data_processing)
    extra_processing = _geo_get(protocols, "data processing step")
    if isinstance(extra_processing, list):
        data_processing_steps.extend([step for step in extra_processing if step not in (None, "")])
    elif extra_processing not in (None, "") and extra_processing != primary_data_processing:
        data_processing_steps.append(extra_processing)

    data_processing_rows_available = 1
    probe_row = base_data_processing_row + 1
    while _normalize_geo_key(meta_sheet.cell(row=probe_row, column=1).value) == _normalize_geo_key("data processing step"):
        data_processing_rows_available += 1
        probe_row += 1

    extra_dp_rows = max(0, len(data_processing_steps) - data_processing_rows_available)
    for i in range(extra_dp_rows):
        insert_at = base_data_processing_row + data_processing_rows_available + i
        meta_sheet.insert_rows(insert_at)
        _copy_row_format(meta_sheet, base_data_processing_row + data_processing_rows_available - 1, insert_at)

    first_dp_label = meta_sheet.cell(row=base_data_processing_row, column=1).value or "*data processing step"
    for idx, step in enumerate(data_processing_steps):
        row_idx = base_data_processing_row + idx
        label = first_dp_label if idx == 0 else "data processing step"
        _write_cell(meta_sheet, row_idx, 1, label)
        _write_cell(meta_sheet, row_idx, 2, step)

    genome_build_row = (_find_first_row_with_label(meta_sheet, "*genome build/assembly") or 67) + extra_dp_rows
    processed_format_row = (_find_first_row_with_label(meta_sheet, "*processed data files format and content") or 68) + extra_dp_rows
    _write_cell(meta_sheet, genome_build_row, 2, _geo_get(protocols, "*genome build/assembly", "genome build/assembly"))
    processed_val = _geo_get(
        protocols,
        "*processed data files format and content",
        "processed data files format and content",
    )
    if isinstance(processed_val, list):
        processed_val = "\n".join([str(v) for v in processed_val if v not in (None, "")])
    _write_cell(meta_sheet, processed_format_row, 2, processed_val)

    # ---- Paired-end experiments ----
    paired_header_row = _find_paired_end_header_row(meta_sheet)
    paired_data_start_row = paired_header_row + 1
    paired_label_row_template = paired_data_start_row

    if len(paired_end_experiments) > 1:
        for i in range(len(paired_end_experiments) - 1):
            insert_at = paired_data_start_row + i + 1
            meta_sheet.insert_rows(insert_at)
            _copy_row_format(meta_sheet, paired_label_row_template, insert_at)

    for idx, entry in enumerate(paired_end_experiments):
        if not isinstance(entry, Mapping):
            continue
        row_idx = paired_data_start_row + idx
        meta_sheet.cell(row=row_idx, column=1).value = entry.get("file name 1")
        meta_sheet.cell(row=row_idx, column=2).value = entry.get("file name 2")
        meta_sheet.cell(row=row_idx, column=3).value = entry.get("file name 3")
        meta_sheet.cell(row=row_idx, column=4).value = entry.get("file name 4")

    # ---- Checksums sheet ----
    checksums_sheet = wb["MD5 Checksums"]
    raw_files = (checksums.get("raw_data_files") or []) if isinstance(checksums, Mapping) else []
    checksums_header_row = _find_first_row_with_label(checksums_sheet, "file name") or 8
    raw_start_row = checksums_header_row + 1
    for idx, item in enumerate(raw_files):
        if not isinstance(item, Mapping):
            continue
        row_idx = raw_start_row + idx
        _write_cell(checksums_sheet, row_idx, 1, item.get("file name"))
        _write_cell(checksums_sheet, row_idx, 2, item.get("file checksum"))

    processed_files = (checksums.get("processed_data_files") or []) if isinstance(checksums, Mapping) else []
    if processed_files:
        processed_section_row = raw_start_row + max(len(raw_files), 1) + 2
        _write_cell(checksums_sheet, processed_section_row, 1, "PROCESSED FILES")
        _write_cell(checksums_sheet, processed_section_row + 1, 1, "file name")
        _write_cell(checksums_sheet, processed_section_row + 1, 2, "file checksum")
        for idx, item in enumerate(processed_files):
            if not isinstance(item, Mapping):
                continue
            row_idx = processed_section_row + 2 + idx
            _write_cell(checksums_sheet, row_idx, 1, item.get("file name"))
            _write_cell(checksums_sheet, row_idx, 2, item.get("file checksum"))


def export_geo_report_to_seq_xlsx(
    report_json_path: str,
    template_xlsx_path: str,
    out_dir: str,
    *,
    one_workbook_per_uid: bool = True,
) -> list[str]:
    """
    Convert a GEO report JSON into filled GEO submission Excel workbooks.
    Returns list of output file paths.
    """
    json_path = Path(report_json_path)
    template_path = Path(template_xlsx_path)
    output_root = Path(out_dir)
    output_root.mkdir(parents=True, exist_ok=True)

    data = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(data, Mapping):
        return []

    template_bytes = template_path.read_bytes()
    output_paths: list[str] = []

    reports: list[tuple[str, Mapping[str, Any]]] = []
    for uid, payload in data.items():
        if not isinstance(payload, Mapping):
            continue
        if (payload.get("report_type") or payload.get("report type") or "").upper() != "GEO":
            continue
        report = payload.get("report") or {}
        if isinstance(report, Mapping):
            reports.append((str(uid), report))

    if not reports:
        return []

    def merge_reports(entries: list[Mapping[str, Any]]) -> Mapping[str, Any]:
        merged: dict[str, Any] = {}
        studies = [r.get("study") for r in entries if isinstance(r, Mapping)]
        if studies:
            merged_study = dict(studies[0] or {})
            for study in studies[1:]:
                if not isinstance(study, Mapping):
                    continue
                for key, val in study.items():
                    if key in {"contributor", "supplementary file"} and isinstance(val, list):
                        existing = merged_study.get(key) or []
                        if not isinstance(existing, list):
                            existing = [existing]
                        merged_study[key] = list(existing) + list(val)
                    elif key not in merged_study or merged_study.get(key) in (None, ""):
                        merged_study[key] = val
            merged["study"] = merged_study

        merged_samples: list[Mapping[str, Any]] = []
        for r in entries:
            if isinstance(r, Mapping) and isinstance(r.get("samples"), list):
                merged_samples.extend(r["samples"])
        if merged_samples:
            merged["samples"] = merged_samples

        merged_protocols: dict[str, Any] = {}
        for r in entries:
            proto = r.get("protocols")
            if not isinstance(proto, Mapping):
                continue
            if not merged_protocols:
                merged_protocols = dict(proto)
                continue
            for key, val in proto.items():
                if key == "data processing step" and isinstance(val, list):
                    existing = merged_protocols.get(key) or []
                    if not isinstance(existing, list):
                        existing = [existing] if existing else []
                    merged_protocols[key] = list(existing) + list(val)
                elif merged_protocols.get(key) in (None, ""):
                    merged_protocols[key] = val
        if merged_protocols:
            merged["protocols"] = merged_protocols

        merged_pairs: list[Mapping[str, Any]] = []
        for r in entries:
            if isinstance(r, Mapping) and isinstance(r.get("paired_end_experiments"), list):
                merged_pairs.extend(r["paired_end_experiments"])
        if merged_pairs:
            merged["paired_end_experiments"] = merged_pairs

        merged_checksums: dict[str, Any] = {}
        for r in entries:
            csum = r.get("checksums")
            if not isinstance(csum, Mapping):
                continue
            if "raw_data_files" in csum and isinstance(csum["raw_data_files"], list):
                existing = merged_checksums.get("raw_data_files") or []
                if not isinstance(existing, list):
                    existing = []
                merged_checksums["raw_data_files"] = existing + list(csum["raw_data_files"])
            for key, val in csum.items():
                if key == "raw_data_files":
                    continue
                if key not in merged_checksums:
                    merged_checksums[key] = val
        if merged_checksums:
            merged["checksums"] = merged_checksums

        return merged

    if one_workbook_per_uid:
        for uid, report in reports:
            wb = load_workbook(BytesIO(template_bytes))
            _populate_geo_seq_workbook(wb, report)
            filename = f"{uid}_GEO_template_filled.xlsx"
            dest = output_root / filename
            wb.save(dest)
            output_paths.append(str(dest))
    else:
        merged_report = merge_reports([r for _, r in reports])
        wb = load_workbook(BytesIO(template_bytes))
        _populate_geo_seq_workbook(wb, merged_report)
        dest = output_root / f"{json_path.stem}_GEO_template_filled.xlsx"
        wb.save(dest)
        output_paths.append(str(dest))

    return output_paths


def _extract_sra_section_reports(
    data: Mapping[str, Any],
    *,
    section_name: str,
) -> list[tuple[str, list[Mapping[str, Any]]]]:
    """Collect per-UID SRA report rows for a given section."""
    reports: list[tuple[str, list[Mapping[str, Any]]]] = []
    for uid, payload in data.items():
        if not isinstance(payload, Mapping):
            continue
        if (payload.get("report_type") or payload.get("report type") or "").upper() != "SRA":
            continue
        report = payload.get("report") or {}
        rows = report.get(section_name) if isinstance(report, Mapping) else None
        if isinstance(rows, list) and rows:
            reports.append((str(uid), [row for row in rows if isinstance(row, Mapping)]))
    return reports


def _worksheet_headers(ws: Worksheet, *, header_row: int) -> list[str]:
    """Read contiguous headers from a worksheet header row."""
    headers: list[str] = []
    col = 1
    while True:
        value = ws.cell(row=header_row, column=col).value
        if value in (None, ""):
            break
        headers.append(str(value))
        col += 1
    return headers


def _write_template_rows(
    wb,
    *,
    sheet_name: str,
    header_row: int,
    template_row: int,
    rows: Sequence[Mapping[str, Any]],
) -> None:
    """Populate a row-based workbook template from ordered row mappings."""
    ws = wb[sheet_name]
    headers = _worksheet_headers(ws, header_row=header_row)

    if len(rows) > 1:
        for i in range(len(rows) - 1):
            insert_at = template_row + i + 1
            ws.insert_rows(insert_at)
            _copy_row_format(ws, template_row, insert_at)

    for row_idx, row in enumerate(rows, start=template_row):
        for col_idx, header in enumerate(headers, start=1):
            _write_cell(ws, row_idx, col_idx, row.get(header))


def _export_sra_section_to_xlsx(
    report_json_path: str,
    template_xlsx_path: str,
    out_dir: str | Path,
    *,
    section_name: str,
    sheet_name: str | None,
    header_row: int,
    template_row: int,
    filename_suffix: str,
    one_workbook_per_uid: bool,
) -> list[str]:
    """Render a row-based SRA report section into workbook copies from a template."""
    json_path = Path(report_json_path)
    template_path = Path(template_xlsx_path)
    output_root = Path(out_dir)
    output_root.mkdir(parents=True, exist_ok=True)

    data = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(data, Mapping):
        return []

    template_bytes = template_path.read_bytes()
    output_paths: list[str] = []
    reports = _extract_sra_section_reports(data, section_name=section_name)

    if not reports:
        return []

    if one_workbook_per_uid:
        for uid, rows in reports:
            wb = load_workbook(BytesIO(template_bytes))
            target_sheet = sheet_name or wb.sheetnames[0]
            _write_template_rows(
                wb,
                sheet_name=target_sheet,
                header_row=header_row,
                template_row=template_row,
                rows=rows,
            )
            dest = output_root / f"{uid}_{filename_suffix}"
            wb.save(dest)
            output_paths.append(str(dest))
    else:
        merged_rows: list[Mapping[str, Any]] = []
        for _, rows in reports:
            merged_rows.extend(rows)
        if not merged_rows:
            return []
        wb = load_workbook(BytesIO(template_bytes))
        target_sheet = sheet_name or wb.sheetnames[0]
        _write_template_rows(
            wb,
            sheet_name=target_sheet,
            header_row=header_row,
            template_row=template_row,
            rows=merged_rows,
        )
        dest = output_root / f"{json_path.stem}_{filename_suffix}"
        wb.save(dest)
        output_paths.append(str(dest))

    return output_paths


def export_sra_report_to_xlsx(
    report_json_path: str,
    template_xlsx_path: str,
    out_dir: str | Path,
    *,
    one_workbook_per_uid: bool = True,
) -> list[str]:
    """
    Convert the SRA libraries report JSON into filled SRA submission workbooks.
    Returns list of output file paths.
    """
    return _export_sra_section_to_xlsx(
        report_json_path,
        template_xlsx_path,
        out_dir,
        section_name="libraries",
        sheet_name="SRA_data",
        header_row=1,
        template_row=2,
        filename_suffix="SRA_metadata_filled.xlsx",
        one_workbook_per_uid=one_workbook_per_uid,
    )


def export_sra_biosample_report_to_xlsx(
    report_json_path: str,
    template_xlsx_path: str,
    out_dir: str | Path,
    *,
    one_workbook_per_uid: bool = True,
) -> list[str]:
    """
    Convert the SRA biosamples report JSON into filled BioSample submission workbooks.
    Returns list of output file paths.
    """
    return _export_sra_section_to_xlsx(
        report_json_path,
        template_xlsx_path,
        out_dir,
        section_name="biosamples",
        sheet_name=None,
        header_row=12,
        template_row=13,
        filename_suffix="SRA_biosample_filled.xlsx",
        one_workbook_per_uid=one_workbook_per_uid,
    )


def _coerce_scalar_csv_value(value: Any) -> str:
    """Convert JSON-like values into CSV-safe scalar strings."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        if all(not isinstance(item, (dict, list)) for item in value):
            return ";".join("" if item is None else str(item) for item in value)
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def _normalize_rows_for_csv(rows: Any) -> list[dict[str, Any]]:
    """Normalize a report section into a list of row dicts suitable for CSV export."""
    if rows is None:
        return []
    if isinstance(rows, Mapping):
        return [dict(rows)]
    if isinstance(rows, list):
        normalized: list[dict[str, Any]] = []
        for item in rows:
            if isinstance(item, Mapping):
                normalized.append(dict(item))
            elif item is not None:
                normalized.append({"value": item})
        return normalized
    return [{"value": rows}]


def _extract_report_section_rows(report: Mapping[str, Any], candidates: Sequence[str]) -> list[dict[str, Any]]:
    """Return the first matching report section that looks like tabular row data."""
    for key in candidates:
        value = report.get(key)
        rows = _normalize_rows_for_csv(value)
        if rows:
            return rows
    return []


def _ordered_csv_columns(rows: Sequence[Mapping[str, Any]], preferred: Sequence[str]) -> list[str]:
    """Build CSV column order with required columns first, then observed extras in row order."""
    columns: list[str] = []
    seen: set[str] = set()
    for col in preferred:
        if col not in seen:
            columns.append(col)
            seen.add(col)
    for row in rows:
        for key in row.keys():
            if key not in seen:
                columns.append(key)
                seen.add(key)
    return columns


def _write_csv_rows(path: Path, rows: Sequence[Mapping[str, Any]], columns: Sequence[str]) -> str:
    """Write ordered rows to CSV."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({col: _coerce_scalar_csv_value(row.get(col)) for col in columns})
    return str(path)


def run_reporter_summary(
    config,
    reporter_plan,
    log_dir: "str | Path | None",
) -> "tuple[dict, dict[str, str], dict]":
    """
    Execute the summary reporter pipeline (samples / protocols / published / RPPR).

    Returns (reporter_result, saved_files, reporter_summary) where:
      - reporter_result  — raw dict from run_project_*_report helpers
      - saved_files      — {key: file_path} of output files written to log_dir
      - reporter_summary — condensed dict suitable for passing to a chatter/LLM

    This is shared between run_query (orchestrator) and _plan_tool_reporter (agents)
    to avoid duplicating the execution logic.
    """
    project = reporter_plan.project
    if isinstance(project, str) and not project.strip():
        project = None
    years = reporter_plan.years or []
    month_range = reporter_plan.month_range
    day_range = reporter_plan.day_range
    summary_mode = reporter_plan.summary_mode or "samples"

    print(f"[DEBUG][REPORTER] Summary mode: {summary_mode}, project: {project}, years: {years}")

    try:
        if summary_mode == "RPPR":
            samples_result = run_project_sample_report(
                config, project, years=years, month_range=month_range,
                day_range=day_range, outputs_root=log_dir or "outputs",
            )
            protocols_result = run_project_protocols_report(
                config, project, years=years, month_range=month_range,
                day_range=day_range, outputs_root=log_dir or "outputs",
            )
            published_result = run_project_published_report(
                config, project, years=years, month_range=month_range,
                day_range=day_range, outputs_root=log_dir or "outputs",
            )
            reporter_result: dict = {
                "ok": samples_result.get("ok"),
                "summary_mode": "RPPR",
                "samples": samples_result,
                "protocols": protocols_result,
                "published": published_result,
                "rows_returned": samples_result.get("rows_returned"),
                "filters": samples_result.get("filters") or {},
            }
        elif summary_mode == "protocols":
            reporter_result = run_project_protocols_report(
                config, project, years=years, month_range=month_range,
                day_range=day_range, outputs_root=log_dir or "outputs",
            )
        elif summary_mode == "published":
            reporter_result = run_project_published_report(
                config, project, years=years, month_range=month_range,
                day_range=day_range, outputs_root=log_dir or "outputs",
            )
        else:  # "samples" (default)
            reporter_result = run_project_sample_report(
                config, project, years=years, month_range=month_range,
                day_range=day_range, outputs_root=log_dir or "outputs",
            )
    except Exception as e:
        reporter_result = {"ok": False, "error": repr(e)}

    # ── Register output files ─────────────────────────────────────────
    saved_files: dict[str, str] = {}

    def _register_file(key: str, path: "str | None") -> None:
        from pathlib import Path as _Path
        if path and _Path(path).exists():
            saved_files[key] = path

    if summary_mode == "RPPR":
        _register_file("samples_report", reporter_result.get("samples", {}).get("uuid_report_file"))
        _register_file("protocols_report", reporter_result.get("protocols", {}).get("report_file"))
        _register_file("published_report", reporter_result.get("published", {}).get("report_file"))
    elif summary_mode == "protocols":
        _register_file("protocols_report", reporter_result.get("report_file"))
    elif summary_mode == "published":
        _register_file("published_report", reporter_result.get("report_file"))
    else:
        _register_file("samples_report", reporter_result.get("uuid_report_file"))

    summary_path = persist_report_file("reporter_result", reporter_result, log_dir or "outputs", kind="report")
    if summary_path:
        saved_files["reporter_result"] = summary_path

    # ── Build condensed reporter_summary for LLM chatter ─────────────
    def _sub_summary(r: dict) -> dict:
        return {
            "rows_returned": r.get("rows_returned"),
            "top_sampletypes": top_items(r.get("sampletypes_table"), 5),
            "top_labs": top_items(r.get("labs_table"), 5),
            "years": top_items(r.get("years_table"), 10),
            "top_months": top_items(r.get("months_table"), 12),
            "db_diagnostic": r.get("db_diagnostic") or {},
        }

    reporter_summary: dict = {
        "summary_mode": summary_mode,
        "project": project,
        "project_id": reporter_result.get("project_id"),
        "filters": reporter_result.get("filters") or {},
    }
    if summary_mode == "RPPR":
        reporter_summary["samples"] = _sub_summary(reporter_result.get("samples") or {})
        reporter_summary["protocols"] = {
            "rows_returned": (reporter_result.get("protocols") or {}).get("rows_returned"),
            "top_labs": top_items((reporter_result.get("protocols") or {}).get("labs_table"), 5),
            "years": top_items((reporter_result.get("protocols") or {}).get("years_table"), 10),
        }
        reporter_summary["published"] = {
            "samples": _sub_summary((reporter_result.get("published") or {}).get("samples") or {}),
            "protocols_count": ((reporter_result.get("published") or {}).get("protocols") or {}).get("rows_returned"),
            "study_count": ((reporter_result.get("published") or {}).get("samples") or {}).get("study_count"),
        }
    elif summary_mode == "protocols":
        reporter_summary.update({
            "rows_returned": reporter_result.get("rows_returned"),
            "top_labs": top_items(reporter_result.get("labs_table"), 5),
            "years": top_items(reporter_result.get("years_table"), 10),
            "top_months": top_items(reporter_result.get("months_table"), 12),
        })
    elif summary_mode == "published":
        pub_samples = reporter_result.get("samples") or {}
        pub_protocols = reporter_result.get("protocols") or {}
        reporter_summary.update({
            "samples": _sub_summary(pub_samples),
            "study_count": pub_samples.get("study_count"),
            "studies": pub_samples.get("studies"),
            "protocols_count": pub_protocols.get("rows_returned"),
            "top_protocol_labs": top_items(pub_protocols.get("labs_table"), 5),
        })
    else:
        reporter_summary.update(_sub_summary(reporter_result))

    print(f"[DEBUG][REPORTER] result ok={reporter_result.get('ok')}, rows={reporter_result.get('rows_returned')}, files={list(saved_files.keys())}")
    return reporter_result, saved_files, reporter_summary


def enumerate_lineage_leaves(
    metadata_bundle: dict,
    *,
    accepted_types: list[str],
) -> list[dict[str, str]]:
    """Walk a NExtSEEK metadata bundle and return every sample whose
    ``sample_type`` is in ``accepted_types``.

    Returns a list of ``{uid, sample_type, assay, source_uid}`` dicts. ``source_uid``
    is the top-level (root) sample UID this leaf was reached from, used by the
    sanity step to report "X source UIDs have zero matching leaves".

    Empty ``accepted_types`` short-circuits to ``[]`` — used by pipelines with
    ``samplesheet_input_kind="accession"`` (e.g. fetchngs) where lineage isn't
    walked at all.

    The bundle shape mirrors the reporter API response:
        {"data": [
            {"sample_type": "NHP", "samples": [
                {"uuid": "NHP-1", "metadata": {...}, "children": [
                    {"uuid": "TIS-1", "metadata": {...}, "children": [
                        {"uuid": "D.SEQ-1", "metadata": {"UID": "...", "assay_name": "RNA-seq"}},
                    ]},
                ]},
            ]},
        ]}

    Assay value is read from ``metadata["assay_name"]`` with fallbacks to
    ``"assay"`` and ``"AssayName"``; missing assay defaults to "".
    """
    accepted = set(accepted_types or [])
    if not accepted:
        return []

    out: list[dict[str, str]] = []

    def _walk(sample: dict, source_uid: str, leaf_sample_type: str | None) -> None:
        if not isinstance(sample, dict):
            return
        uid = (sample.get("metadata") or {}).get("UID") or sample.get("uuid") or ""
        st = leaf_sample_type or ""
        # Only emit when sample_type is in accepted set.
        if st in accepted and uid:
            md = sample.get("metadata") or {}
            assay = (
                md.get("assay_name")
                or md.get("assay")
                or md.get("AssayName")
                or ""
            )
            out.append({
                "uid": str(uid),
                "sample_type": st,
                "assay": str(assay),
                "source_uid": source_uid,
            })
        for child in sample.get("children") or []:
            # Children's sample_type comes from the inner-block annotation OR
            # from each child's own metadata. Reporter response stamps it on
            # the surrounding block; for nested children we re-read from
            # metadata.sample_type when available.
            child_st = (child.get("metadata") or {}).get("sample_type") or st
            _walk(child, source_uid, child_st)

    for block in (metadata_bundle or {}).get("data") or []:
        if not isinstance(block, dict):
            continue
        block_type = block.get("sample_type")
        for sample in block.get("samples") or []:
            root_uid = (sample.get("metadata") or {}).get("UID") or sample.get("uuid") or ""
            _walk(sample, source_uid=str(root_uid), leaf_sample_type=block_type)

    return out


# Moved to helpers_new in Phase 2 — re-exported for backward compat
from .helpers_new.prompts import load_prompt, log_usage, log_prompt  # noqa: E402,F401
from .helpers_new.json_io import _extract_required_paths, estimate_tokens_from_text, safe_parse_json  # noqa: E402,F401
from .helpers_new.text import strip_html, strip_html_recursive, load_file_for_memory, load_json_for_memory  # noqa: E402,F401
from .helpers_new.tools.memory_code import (  # noqa: E402,F401
    _json_type_name,
    _compact_json_value,
    _merge_skeletons,
    _build_skeleton_node,
    _find_record_arrays,
    build_memory_data_profile,
    _build_field_index,
    MemoryCodeSafetyError,
    MemoryCodeTimeoutError,
    _validate_memory_code,
    execute_memory_code,
)

