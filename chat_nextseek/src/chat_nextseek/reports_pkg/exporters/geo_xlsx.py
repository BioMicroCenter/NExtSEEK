"""GEO submission XLSX exporter: fills a GEO SEQ template workbook from a
report JSON and writes per-UID or merged workbooks.

Moved out of ``helpers.py`` during Phase 2 of the src/ restructure.
"""
from __future__ import annotations

import json
import re
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _copy_row_format(ws: Worksheet, source_row: int, target_row: int) -> None:
    """
    Copy styling and height from source_row to target_row on the same sheet.
    Preserves number formats, comments, and hyperlinks so cloned rows match the template.
    Useful when expanding list sections without rebuilding formatting manually.
    """
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for cell in ws[source_row]:
        tgt = ws.cell(row=target_row, column=cell.col_idx)
        tgt._style = copy(cell._style)
        tgt.number_format = cell.number_format
        tgt._comment = cell._comment
        tgt.hyperlink = cell.hyperlink


def _write_cell(ws: Worksheet, row: int, col: int, value: Any) -> None:
    """
    Write a value into a worksheet cell unless the value is None.
    Keeps template defaults intact when optional fields are absent.
    """
    if value is None:
        return
    ws.cell(row=row, column=col).value = value


def _write_list_down(
    ws: Worksheet,
    start_row: int,
    col: int,
    values: Sequence[Any] | None,
    *,
    max_rows: int | None = None,
) -> None:
    """
    Write a sequence of values down a column, optionally packing overflow into the final row.
    Preserves existing cell content when condensing and skips empty values so templates stay clean.
    """
    if not values:
        return
    for idx, val in enumerate(values):
        if val in (None, ""):
            continue
        row = start_row + idx
        if max_rows and idx >= max_rows:
            row = start_row + max_rows - 1
            existing = ws.cell(row=row, column=col).value or ""
            separator = "\n" if existing else ""
            ws.cell(row=row, column=col).value = f"{existing}{separator}{val}"
            continue
        ws.cell(row=row, column=col).value = val


def _build_header_map(ws: Worksheet, header_row: int) -> dict[str, list[int]]:
    """
    Build a mapping of normalized header labels to column indices for a given header row.
    Supports duplicate headers by returning lists so repeated fields can be filled in order.
    """
    mapping: dict[str, list[int]] = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        label = _normalize_sheet_label(cell.value)
        mapping.setdefault(label, []).append(cell.col_idx)
    return mapping


def _normalize_geo_key(value: Any) -> str:
    """
    Normalize GEO template keys so starred/unstarred variants map to the same logical field.
    Collapses whitespace and lowercases labels to tolerate minor template or model variations.
    """
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"^[*#\s]+", "", text)
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_sheet_label(value: Any) -> str:
    """
    Normalize worksheet labels while preserving leading marker characters like '*'.
    This keeps starred and unstarred template columns distinct.
    """
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _geo_get(mapping: Mapping[str, Any] | None, *keys: str) -> Any:
    """
    Read a GEO field from a mapping using exact and normalized-key fallback.
    This lets the exporter accept both literal template keys ('*title') and logical keys ('title').
    """
    if not isinstance(mapping, Mapping):
        return None

    normalized = {_normalize_geo_key(key): value for key, value in mapping.items() if isinstance(key, str)}
    for key in keys:
        if key in mapping:
            return mapping[key]
        norm_key = _normalize_geo_key(key)
        if norm_key in normalized:
            return normalized[norm_key]
    return None


def _find_first_row_with_label(ws: Worksheet, label: str, *, col: int = 1) -> int | None:
    """
    Find the first row whose target column matches a label after GEO-key normalization.
    """
    target = _normalize_geo_key(label)
    for row_idx in range(1, ws.max_row + 1):
        if _normalize_geo_key(ws.cell(row=row_idx, column=col).value) == target:
            return row_idx
    return None


def _find_sample_header_row(ws: Worksheet) -> int:
    """
    Locate the sample header row from the required GEO sample columns.
    Falls back to the current template row if discovery fails.
    """
    required = {
        _normalize_geo_key("*library name"),
        _normalize_geo_key("*title"),
        _normalize_geo_key("*library strategy"),
        _normalize_geo_key("*organism"),
    }
    for row_idx in range(1, ws.max_row + 1):
        labels = {
            _normalize_geo_key(ws.cell(row=row_idx, column=col).value)
            for col in range(1, ws.max_column + 1)
            if ws.cell(row=row_idx, column=col).value not in (None, "")
        }
        if required.issubset(labels):
            return row_idx
    return 38


def _find_paired_end_header_row(ws: Worksheet) -> int:
    """
    Locate the paired-end table header row from the file-name columns.
    """
    required = {
        _normalize_geo_key("file name 1"),
        _normalize_geo_key("file name 2"),
        _normalize_geo_key("file name 3"),
        _normalize_geo_key("file name 4"),
    }
    for row_idx in range(1, ws.max_row + 1):
        labels = {
            _normalize_geo_key(ws.cell(row=row_idx, column=col).value)
            for col in range(1, 5)
            if ws.cell(row=row_idx, column=col).value not in (None, "")
        }
        if required.issubset(labels):
            return row_idx
    return 76


def _select_study_summary(study: Mapping[str, Any] | None) -> str | None:
    """
    Return the best available study summary, preferring 'summary (abstract)' then 'summary'.
    Keeps GEO population tolerant of missing fields while still returning None when nothing is present.
    """
    if not isinstance(study, Mapping):
        return None
    return _geo_get(study, "*summary (abstract)", "summary (abstract)", "summary") or None


def _populate_geo_seq_workbook(wb, report: Mapping[str, Any]) -> None:
    """
    Fill the GEO SEQ template workbook in-place from a single report entry.
    Writes study metadata, sample records, protocols, paired-end entries, and checksum info while preserving formats.
    Mutates the workbook directly so callers can immediately save it to disk.
    """
    meta_sheet = wb["Metadata"]
    study = report.get("study") or {}
    samples = report.get("samples") or []
    protocols = report.get("protocols") or {}
    paired_end_experiments = report.get("paired_end_experiments") or []
    checksums = (report.get("checksums") or {})

    # ---- Study section ----
    title_row = _find_first_row_with_label(meta_sheet, "*title") or 12
    summary_row = _find_first_row_with_label(meta_sheet, "*summary (abstract)") or 13
    design_row = _find_first_row_with_label(meta_sheet, "*experimental design") or 14
    contributor_row = _find_first_row_with_label(meta_sheet, "contributor") or 15
    supplementary_row = _find_first_row_with_label(meta_sheet, "supplementary file") or 22

    _write_cell(meta_sheet, title_row, 2, _geo_get(study, "*title", "title"))
    _write_cell(meta_sheet, summary_row, 2, _select_study_summary(study))
    _write_cell(meta_sheet, design_row, 2, _geo_get(study, "*experimental design", "experimental design"))

    contributors = _geo_get(study, "contributor") or []
    _write_list_down(meta_sheet, start_row=contributor_row, col=2, values=contributors, max_rows=7)

    supplementary = _geo_get(study, "supplementary file") or []
    _write_list_down(meta_sheet, start_row=supplementary_row, col=2, values=supplementary, max_rows=16)

    # ---- Samples section ----
    sample_header_row = _find_sample_header_row(meta_sheet)
    header_map = _build_header_map(meta_sheet, sample_header_row)
    sample_start_row = sample_header_row + 1
    sample_rows_available = max(0, 52 - sample_start_row + 1)

    extra_sample_rows = max(0, len(samples) - sample_rows_available)
    if extra_sample_rows:
        template_row = 52
        insert_at = template_row + 1
        for offset in range(extra_sample_rows):
            meta_sheet.insert_rows(insert_at + offset)
            _copy_row_format(meta_sheet, template_row, insert_at + offset)

    def set_sample_field(row_idx: int, header_key: str, value: Any, occurrence: int = 0) -> None:
        key = _normalize_sheet_label(header_key)
        cols = header_map.get(key)
        if not cols or occurrence >= len(cols) or value in (None, ""):
            return
        meta_sheet.cell(row=row_idx, column=cols[occurrence]).value = value

    for idx, sample in enumerate(samples):
        if not isinstance(sample, Mapping):
            continue
        row_idx = sample_start_row + idx
        set_sample_field(row_idx, "*library name", _geo_get(sample, "*library name", "library name"))
        set_sample_field(row_idx, "*title", _geo_get(sample, "*title", "title"))
        set_sample_field(row_idx, "*library strategy", _geo_get(sample, "*library strategy", "library strategy"))
        set_sample_field(row_idx, "*organism", _geo_get(sample, "*organism", "organism"))
        set_sample_field(row_idx, "**tissue", _geo_get(sample, "**tissue", "tissue"))
        set_sample_field(row_idx, "**cell line", _geo_get(sample, "**cell line", "cell line"))
        set_sample_field(row_idx, "**cell type", _geo_get(sample, "**cell type", "cell type"))
        set_sample_field(row_idx, "genotype", _geo_get(sample, "genotype"))
        set_sample_field(row_idx, "treatment", _geo_get(sample, "treatment"))
        set_sample_field(row_idx, "batch", _geo_get(sample, "batch"))
        set_sample_field(row_idx, "*molecule", _geo_get(sample, "*molecule", "molecule"))
        set_sample_field(row_idx, "*single or paired-end", _geo_get(sample, "*single or paired-end", "single or paired-end"))
        set_sample_field(row_idx, "*instrument model", _geo_get(sample, "*instrument model", "instrument model"))
        set_sample_field(row_idx, "description", _geo_get(sample, "description"))
        set_sample_field(row_idx, "processed data file", _geo_get(sample, "processed data file"), occurrence=0)
        set_sample_field(row_idx, "processed data file", _geo_get(sample, "processed data file (2)"), occurrence=1)
        set_sample_field(row_idx, "*raw file", _geo_get(sample, "*raw file", "raw file"), occurrence=0)
        set_sample_field(row_idx, "raw file", _geo_get(sample, "raw file"), occurrence=0)
        set_sample_field(row_idx, "raw file", _geo_get(sample, "raw file (2)"), occurrence=1)
        set_sample_field(row_idx, "raw file", _geo_get(sample, "raw file (3)"), occurrence=2)
        set_sample_field(row_idx, "raw file", _geo_get(sample, "raw file (4)"), occurrence=3)

    # ---- Protocols section ----
    growth_row = _find_first_row_with_label(meta_sheet, "growth protocol") or 57
    treatment_row = _find_first_row_with_label(meta_sheet, "treatment protocol") or 58
    extract_row = _find_first_row_with_label(meta_sheet, "*extract protocol") or 59
    library_row = _find_first_row_with_label(meta_sheet, "*library construction protocol") or 60
    base_data_processing_row = _find_first_row_with_label(meta_sheet, "*data processing step") or 62

    _write_cell(meta_sheet, growth_row, 2, _geo_get(protocols, "growth protocol"))
    _write_cell(meta_sheet, treatment_row, 2, _geo_get(protocols, "treatment protocol"))
    _write_cell(meta_sheet, extract_row, 2, _geo_get(protocols, "*extract protocol", "extract protocol"))
    _write_cell(meta_sheet, library_row, 2, _geo_get(protocols, "*library construction protocol", "library construction protocol"))

    data_processing_steps: list[Any] = []
    primary_data_processing = _geo_get(protocols, "*data processing step", "data processing step")
    if primary_data_processing not in (None, ""):
        if isinstance(primary_data_processing, list):
            data_processing_steps.extend(primary_data_processing)
        else:
            data_processing_steps.append(primary_data_processing)
    extra_processing = _geo_get(protocols, "data processing step")
    if isinstance(extra_processing, list):
        data_processing_steps.extend([step for step in extra_processing if step not in (None, "")])
    elif extra_processing not in (None, "") and extra_processing != primary_data_processing:
        data_processing_steps.append(extra_processing)

    data_processing_rows_available = 1
    probe_row = base_data_processing_row + 1
    while _normalize_geo_key(meta_sheet.cell(row=probe_row, column=1).value) == _normalize_geo_key("data processing step"):
        data_processing_rows_available += 1
        probe_row += 1

    extra_dp_rows = max(0, len(data_processing_steps) - data_processing_rows_available)
    for i in range(extra_dp_rows):
        insert_at = base_data_processing_row + data_processing_rows_available + i
        meta_sheet.insert_rows(insert_at)
        _copy_row_format(meta_sheet, base_data_processing_row + data_processing_rows_available - 1, insert_at)

    first_dp_label = meta_sheet.cell(row=base_data_processing_row, column=1).value or "*data processing step"
    for idx, step in enumerate(data_processing_steps):
        row_idx = base_data_processing_row + idx
        label = first_dp_label if idx == 0 else "data processing step"
        _write_cell(meta_sheet, row_idx, 1, label)
        _write_cell(meta_sheet, row_idx, 2, step)

    genome_build_row = (_find_first_row_with_label(meta_sheet, "*genome build/assembly") or 67) + extra_dp_rows
    processed_format_row = (_find_first_row_with_label(meta_sheet, "*processed data files format and content") or 68) + extra_dp_rows
    _write_cell(meta_sheet, genome_build_row, 2, _geo_get(protocols, "*genome build/assembly", "genome build/assembly"))
    processed_val = _geo_get(
        protocols,
        "*processed data files format and content",
        "processed data files format and content",
    )
    if isinstance(processed_val, list):
        processed_val = "\n".join([str(v) for v in processed_val if v not in (None, "")])
    _write_cell(meta_sheet, processed_format_row, 2, processed_val)

    # ---- Paired-end experiments ----
    paired_header_row = _find_paired_end_header_row(meta_sheet)
    paired_data_start_row = paired_header_row + 1
    paired_label_row_template = paired_data_start_row

    if len(paired_end_experiments) > 1:
        for i in range(len(paired_end_experiments) - 1):
            insert_at = paired_data_start_row + i + 1
            meta_sheet.insert_rows(insert_at)
            _copy_row_format(meta_sheet, paired_label_row_template, insert_at)

    for idx, entry in enumerate(paired_end_experiments):
        if not isinstance(entry, Mapping):
            continue
        row_idx = paired_data_start_row + idx
        meta_sheet.cell(row=row_idx, column=1).value = entry.get("file name 1")
        meta_sheet.cell(row=row_idx, column=2).value = entry.get("file name 2")
        meta_sheet.cell(row=row_idx, column=3).value = entry.get("file name 3")
        meta_sheet.cell(row=row_idx, column=4).value = entry.get("file name 4")

    # ---- Checksums sheet ----
    checksums_sheet = wb["MD5 Checksums"]
    raw_files = (checksums.get("raw_data_files") or []) if isinstance(checksums, Mapping) else []
    checksums_header_row = _find_first_row_with_label(checksums_sheet, "file name") or 8
    raw_start_row = checksums_header_row + 1
    for idx, item in enumerate(raw_files):
        if not isinstance(item, Mapping):
            continue
        row_idx = raw_start_row + idx
        _write_cell(checksums_sheet, row_idx, 1, item.get("file name"))
        _write_cell(checksums_sheet, row_idx, 2, item.get("file checksum"))

    processed_files = (checksums.get("processed_data_files") or []) if isinstance(checksums, Mapping) else []
    if processed_files:
        processed_section_row = raw_start_row + max(len(raw_files), 1) + 2
        _write_cell(checksums_sheet, processed_section_row, 1, "PROCESSED FILES")
        _write_cell(checksums_sheet, processed_section_row + 1, 1, "file name")
        _write_cell(checksums_sheet, processed_section_row + 1, 2, "file checksum")
        for idx, item in enumerate(processed_files):
            if not isinstance(item, Mapping):
                continue
            row_idx = processed_section_row + 2 + idx
            _write_cell(checksums_sheet, row_idx, 1, item.get("file name"))
            _write_cell(checksums_sheet, row_idx, 2, item.get("file checksum"))


def export_geo_report_to_seq_xlsx(
    report_json_path: str,
    template_xlsx_path: str,
    out_dir: str,
    *,
    one_workbook_per_uid: bool = True,
) -> list[str]:
    """
    Convert a GEO report JSON into filled GEO submission Excel workbooks.
    Returns list of output file paths.
    """
    json_path = Path(report_json_path)
    template_path = Path(template_xlsx_path)
    output_root = Path(out_dir)
    output_root.mkdir(parents=True, exist_ok=True)

    data = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(data, Mapping):
        return []

    template_bytes = template_path.read_bytes()
    output_paths: list[str] = []

    reports: list[tuple[str, Mapping[str, Any]]] = []
    for uid, payload in data.items():
        if not isinstance(payload, Mapping):
            continue
        if (payload.get("report_type") or payload.get("report type") or "").upper() != "GEO":
            continue
        report = payload.get("report") or {}
        if isinstance(report, Mapping):
            reports.append((str(uid), report))

    if not reports:
        return []

    def merge_reports(entries: list[Mapping[str, Any]]) -> Mapping[str, Any]:
        merged: dict[str, Any] = {}
        studies = [r.get("study") for r in entries if isinstance(r, Mapping)]
        if studies:
            merged_study = dict(studies[0] or {})
            for study in studies[1:]:
                if not isinstance(study, Mapping):
                    continue
                for key, val in study.items():
                    if key in {"contributor", "supplementary file"} and isinstance(val, list):
                        existing = merged_study.get(key) or []
                        if not isinstance(existing, list):
                            existing = [existing]
                        merged_study[key] = list(existing) + list(val)
                    elif key not in merged_study or merged_study.get(key) in (None, ""):
                        merged_study[key] = val
            merged["study"] = merged_study

        merged_samples: list[Mapping[str, Any]] = []
        for r in entries:
            if isinstance(r, Mapping) and isinstance(r.get("samples"), list):
                merged_samples.extend(r["samples"])
        if merged_samples:
            merged["samples"] = merged_samples

        merged_protocols: dict[str, Any] = {}
        for r in entries:
            proto = r.get("protocols")
            if not isinstance(proto, Mapping):
                continue
            if not merged_protocols:
                merged_protocols = dict(proto)
                continue
            for key, val in proto.items():
                if key == "data processing step" and isinstance(val, list):
                    existing = merged_protocols.get(key) or []
                    if not isinstance(existing, list):
                        existing = [existing] if existing else []
                    merged_protocols[key] = list(existing) + list(val)
                elif merged_protocols.get(key) in (None, ""):
                    merged_protocols[key] = val
        if merged_protocols:
            merged["protocols"] = merged_protocols

        merged_pairs: list[Mapping[str, Any]] = []
        for r in entries:
            if isinstance(r, Mapping) and isinstance(r.get("paired_end_experiments"), list):
                merged_pairs.extend(r["paired_end_experiments"])
        if merged_pairs:
            merged["paired_end_experiments"] = merged_pairs

        merged_checksums: dict[str, Any] = {}
        for r in entries:
            csum = r.get("checksums")
            if not isinstance(csum, Mapping):
                continue
            if "raw_data_files" in csum and isinstance(csum["raw_data_files"], list):
                existing = merged_checksums.get("raw_data_files") or []
                if not isinstance(existing, list):
                    existing = []
                merged_checksums["raw_data_files"] = existing + list(csum["raw_data_files"])
            for key, val in csum.items():
                if key == "raw_data_files":
                    continue
                if key not in merged_checksums:
                    merged_checksums[key] = val
        if merged_checksums:
            merged["checksums"] = merged_checksums

        return merged

    if one_workbook_per_uid:
        for uid, report in reports:
            wb = load_workbook(BytesIO(template_bytes))
            _populate_geo_seq_workbook(wb, report)
            filename = f"{uid}_GEO_template_filled.xlsx"
            dest = output_root / filename
            wb.save(dest)
            output_paths.append(str(dest))
    else:
        merged_report = merge_reports([r for _, r in reports])
        wb = load_workbook(BytesIO(template_bytes))
        _populate_geo_seq_workbook(wb, merged_report)
        dest = output_root / f"{json_path.stem}_GEO_template_filled.xlsx"
        wb.save(dest)
        output_paths.append(str(dest))

    return output_paths
