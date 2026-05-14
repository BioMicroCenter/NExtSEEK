# Add scripted test cases here for the Streamlit runner (see sidebar in app.py).
# Top-level keys are test case suites (test_case_1, test_case_2, ...).
# Each suite contains case IDs mapping to prompt/reasoning/output.
from __future__ import annotations

import argparse
import ast
import json
import os
import re
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill

TEST_CASES: dict[str, dict[str, dict[str, str]]] = {
    "test_case_1": {
        "mice_ndma": {
            "prompt": "Find me mice treated with NDMA.",
            "reasoning": "general advanced search",
            "output": "",
        },
        "monkeys_flow_and_seq": {
            "prompt": "Find me monkeys that have flow and sequencing data.",
            "reasoning": "find parents by child type test",
            "output": "",
        },
        "cd8_antibodies": {
            "prompt": "Find me all CD8 antibodies in the database.",
            "reasoning": "advanced search",
            "output": "",
        },
        "sample_tree": {
            "prompt": "Show me all samples derived from CEL-250319WHI-1-PUB?",
            "reasoning": "sample tree endpoint test",
            "output": "",
        },
        "retrieval": {
            "prompt": "Retrieve all samples associated with: NHP-220630FLY-5-PUB",
            "reasoning": "sample RETRIEVAL endpoint test",
            "output": "",
        },
        "tumor_patient_4": {
            "prompt": "Show me tumor data for patient 4.",
            "reasoning": "advanced search",
            "output": "",
        },
        "tumor_dfci4": {
            "prompt": "Try that search again but instead with DFCI4.",
            "reasoning": "revise previous search",
            "output": "",
        },
        "what_monkeys_exist": {
            "prompt": "What monkeys exist in the database?",
            "reasoning": "advanced search",
            "output": "",
        },
        "monkeys_cd8_depleted": {
            "prompt": "Which of those monkeys are depleted of CD8?",
            "reasoning": "ask about last results",
            "output": "",
        },
        "list_assays_access": {
            "prompt": "Show me all assays I have access to.",
            "reasoning": "",
            "output": "",
        },
        "omero_fibrin_images": {
            "prompt": "What fibrin images exist?",
            "reasoning": "advanced search test, tests edge cases",
            "output": "",
        },
        "gpt_assay": {
            "prompt": "What GPT data exists in the database?",
            "reasoning": "advanced search test, tests edge cases",
            "output": "",
        },
        "reporter_impact": {
            "prompt": "How many samples were uploaded for impact from 2023 to 2025?",
            "reasoning": "Test SQL reporting for a project",
            "output": "",
        },
        "reporter_language": {
            "prompt": "How many samples were uploaded last month?",
            "reasoning": "Test human lanugage sql reporting ",
            "output": "",
        },
        "GEO": {
            "prompt": "Build me a GEO Submission for D.SEQ-221031SHA-67-PUB and D.SEQ-221031SHA-65-PUB",
            "reasoning": "GEO Submission functionality",
            "output": "",
        },
    },
    "test_case_2": {
        "nonexistent_sampletype": {
            "prompt": "Find me all zebrafish samples in the database.",
            "reasoning": "negative control: zebrafish is not in the sampletype catalog. Entity agent should return empty sampletypes or attempt a weak match.",
            "output": "",
        },
        "pat_pbmc": {
            "prompt": "Find me all PBMCs associated with BTC Patients",
            "reasoning": "multi-step planning required — cannot be executed in a single pass. Correct approach: (1) graph_query to find TIS samples in the GBM_BTC investigation, (2) filter those UIDs by PBMC-related downstream samples via a second graph or API call. PBMC is not a standalone sampletype code; it maps to a downstream TIS derivation. Skipped until multi-step planning is implemented.",
            "output": "",
        },
        "domain_knowledge": {
            "prompt": "Explain what NDMA is and why it's carcinogenic",
            "reasoning": "unsupported: general domain knowledge question, not a data query. Parser should route to unsupported.",
            "output": "",
        },
        "sys_capabilities": {
            "prompt": "What can you do? List your capabilities",
            "reasoning": "system question: parser should route to system_question mode.",
            "output": "",
        },
        "ambiguous_facs": {
            "prompt": "Show me all FACS data for the monkeys",
            "reasoning": "ambiguous: FACS could map to D.FLOW (raw) or A.FLOW (analysis). Entity agent should prefer D.FLOW per raw-first rule. Also tests NHP + D.FLOW combo -> parents_by_child_types.",
            "output": "",
        },
        "ambiguous_proteomics": {
            "prompt": "What proteomics data exists in the database?",
            "reasoning": "ambiguous: could map to D.MSP, A.MSP, or A.IMP. Entity should prefer D.MSP (raw). Tests fuzzy matching.",
            "output": "",
        },
        "informal_scrna": {
            "prompt": "Find me scRNA-seq clustering results",
            "reasoning": "informal name: should map to A.SCCL (Single Cell Clustering Analysis). 'results' implies analyzed, so A.* is correct here.",
            "output": "",
        },
        "colloquial_western": {
            "prompt": "Do we have any western blot data?",
            "reasoning": "colloquial name: should map to D.WBLT. Tests catalog matching on informal assay name.",
            "output": "",
        },
        "derived_vs_raw_seq": {
            "prompt": "Show me analyzed sequencing data",
            "reasoning": "user explicitly says 'analyzed' so should prefer A.ALN or A.GEX over D.SEQ. Tests the D.* vs A.* preference override.",
            "output": "",
        },
        "count_vs_search": {
            "prompt": "How many mice are in the database?",
            "reasoning": "graph_query: global count — 'how many X are in the database' with no time or project scope should use graph COUNT. MATCH (s:Sample {type:'MUS'}) RETURN count(s). Do NOT route to reporter (no time span) or new_search (no attribute filtering needed — just a global aggregate).",
            "output": "",
        },
        "multi_sampletype": {
            "prompt": "Find all tissue and cell samples from IMPACT",
            "reasoning": "graph_query: IMPACT is a named Investigation — any query scoped to a named investigation should route to graph_query regardless of how many sampletypes are involved. Graph should match (s:Sample)-[:IN_STUDY]->(st:Study)-[:IN_INVESTIGATION]->(i:Investigation) WHERE s.type IN ['TIS','CEL'] AND toLower(i.title) CONTAINS 'impact'.",
            "output": "",
        },
        "graph_investigation_studies": {
            "prompt": "What studies are in the Griffith project?",
            "reasoning": "graph_query: Study->Investigation traversal. No REST endpoint for listing studies in a project.",
            "output": "",
        },
        "graph_project_aggregation": {
            "prompt": "What projects have mouse samples?",
            "reasoning": "graph_query: project-level aggregation across graph (Sample->Study->Investigation). Tests graph routing for 'what projects have X'.",
            "output": "",
        },
        "graph_count_in_study": {
            "prompt": "Show me samples in the GBM study?",
            "reasoning": "graph_query: count scoped to a named Study. 'how many' with study scope -> graph, not reporter (no time span, no upload intent).",
            "output": "",
        },
        "ask_graph_unique_types": {
            "prompt": "What unique sample types are in those results?",
            "reasoning": "ask_about_last_results: asks about previously returned graph query results. Tests memory agent handling of graph-sourced data.",
            "output": "",
        },
        "graph_nhp_in_investigation": {
            "prompt": "Show me all NHPs in the SRP project",
            "reasoning": "graph_query: multi-hop traversal — NHP samples in a named Investigation. SRP is a known investigation title.",
            "output": "",
        },
        "graph_protocol_vocab": {
            "prompt": "Find samples that underwent single cell sequencing",
            "reasoning": "graph_query: protocol/assay vocabulary traversal via DERIVED_FROM.internal_assay_title. No sampletype code resolved; action verb 'underwent' signals graph path.",
            "output": "",
        },
        "graph_count_global": {
            "prompt": "How many TIS samples exist across all studies?",
            "reasoning": "graph_query: global count aggregation across all studies via graph traversal. No study/project name given but 'across all studies' signals graph path.",
            "output": "",
        },
        "graph_count_studies_in_project": {
            "prompt": "How many studies are there in the IMPACT project?",
            "reasoning": "graph_query: count of Study nodes scoped to the Impact Investigation via IN_INVESTIGATION traversal. Counting graph-related nodes → graph_query, not reporter (no time span, no upload intent).",
            "output": "",
        },
        # === Graph - DERIVED_FROM single-hop (assay filter on relationship) ===
        "graph_dna_short_read": {
            "prompt": "Find all DNA samples associated with Short Read Sequencing",
            "reasoning": "graph_query: DERIVED_FROM single-hop — filter on r.internal_assay_title = 'Short Read Sequencing'. No REST endpoint can filter by assay on the DERIVED_FROM relationship.",
            "output": "",
        },
        "graph_tis_tissue_collection": {
            "prompt": "Show me TIS samples that have been processed via Tissue Collection",
            "reasoning": "graph_query: DERIVED_FROM single-hop — filter on both sample type (TIS) and r.internal_assay_title = 'Tissue Collection'. Tests combined type + assay filter on relationship.",
            "output": "",
        },
        # === Graph - DERIVED_FROM multi-hop ancestor ===
        "graph_mice_with_seq": {
            "prompt": "Find me all mice that have sequencing data",
            "reasoning": "graph_query: multi-hop DERIVED_FROM — find MUS ancestors of D.SEQ descendants using variable-length path [:DERIVED_FROM*1..]. No REST endpoint for ancestor traversal.",
            "output": "",
        },
        # === Graph - project/investigation aggregation + assay ===
        "graph_projects_short_read": {
            "prompt": "What projects have samples that undergo Short Read Sequencing?",
            "reasoning": "graph_query: project-level aggregation combining DERIVED_FROM assay filter with multi-hop Sample->Study->Investigation traversal. More complex than graph_project_aggregation.",
            "output": "",
        },
        "graph_investigations_nhp_seq": {
            "prompt": "Which investigations have NHP samples with sequencing data?",
            "reasoning": "graph_query: multi-hop — NHP samples with D.SEQ descendants grouped by Investigation. Combines ancestor traversal with project aggregation.",
            "output": "",
        },
        # === Graph - vocabulary matching (assay name fuzzy match) ===
        "graph_flow_cytometry_vocab": {
            "prompt": "Show me samples associated with flow cytometry protocols",
            "reasoning": "graph_query: DERIVED_FROM vocab match — 'flow cytometry' fuzzy-matches 'Flow Cytometry' in internal_assay_titles. No sampletype restriction.",
            "output": "",
        },
        "graph_mass_spec_vocab": {
            "prompt": "Show me samples processed via single cell",
            "reasoning": "graph_query: DERIVED_FROM vocab match — 'single cell' fuzzy-matches internal_assay_title values (e.g. 'Single Cell Sequencing'). Tests protocol vocabulary lookup on relationship metadata.",
            "output": "",
        },
        # === System ===
        "sys_howto_graph": {
            "prompt": "Can I search for samples by which study they belong to?",
            "reasoning": "system_question: asking about a specific capability (study-scoped search). Should explain the graph_query path exists for this.",
            "output": "",
        },
    },
    # =========================================================================
    # test_case_3 — Wizard + memory (LLM-driven nf-core wizard + chat_log /
    # results_history recall). Runs as a single linear conversation; each turn
    # inherits session state from the prior turn (chat_log persists,
    # results_history accumulates, wizard state lives across its scenario).
    #
    # Total: ~55 turns. With ~15s avg/turn + the runner's 30s cooldown, runtime
    # is ~40–50 min. UIDs below are taken from testing.json and from real
    # results we've observed in the dev DB.
    #
    # Scenarios:
    #   A. Wizard happy path (seeded pipeline + "use last search")
    #   B. Wizard rabbit-hole (pipeline questions + cohort exploration)
    #   C. Wizard cancel mid-flight
    #   D. Memory recall against search bundles ("how many?", "which males?")
    #   E. Memory recall across mixed bundle types (graph + api)
    #   F. Wizard restart from mid-step
    #   G. Wizard multi-cohort split via LLM proposal flow
    # =========================================================================
    "test_case_3": {
        # ── Scenario A: wizard happy path, seeded pipeline + use-last-search ──
        "A1_seed_search": {
            "prompt": "Find me mice treated with NDMA.",
            "reasoning": "Scenario A setup. Seeds results_history with an api search bundle the wizard's 'use last search' branch will pull from.",
            "output": "",
        },
        "A2_start_wizard": {
            "prompt": "Build me an nf-core samplesheet for those.",
            "reasoning": "Scenario A turn 1. Parser routes mode=reporter report_type=NFCORE → wizard intercepts, sends intro + asks for pipeline. NO end-to-end execution.",
            "output": "",
        },
        "A3_pick_rnaseq": {
            "prompt": "rnaseq",
            "reasoning": "Scenario A turn 2. wizard_agent advances pipeline step with extracted={'pipeline':'rnaseq'}; state.step→samples.",
            "output": "",
        },
        "A4_use_last_search": {
            "prompt": "use the UIDs we just filtered above",
            "reasoning": "Scenario A turn 3. wizard_agent extracts use_last_search=true; _resolve_last_search_uids walks results_history; HTML-wrapped UIDs are stripped to raw form.",
            "output": "",
        },
        "A5_single_cohort": {
            "prompt": "single cohort, no split",
            "reasoning": "Scenario A turn 4. cohorts step advance with cohort_criteria=[]. Triggers metadata prefetch on entry. State.step→enrichment.",
            "output": "",
        },
        "A6_no_enrichment": {
            "prompt": "skip enrichment columns",
            "reasoning": "Scenario A turn 5. enrichment step advance with []. State.step→confirm.",
            "output": "",
        },
        "A7_confirm_emit": {
            "prompt": "yes go ahead and build it",
            "reasoning": "Scenario A turn 6. confirm step advance; orchestrator runs _execute_nfcore_wizard → generate_report_outputs with pre_supplied_cohorts. Pipeline_selector_agent must NOT be invoked.",
            "output": "",
        },
        # ── Scenario B: wizard rabbit-hole — questions before answers ──────
        "B1_seed_dseq_search": {
            "prompt": "Find me D.SEQ samples in project IMPACT",
            "reasoning": "Scenario B setup. Graph bundle (mode=graph_query). Note: api_result_full will be None for this bundle — memory branch must handle that without crashing.",
            "output": "",
        },
        "B2_start_wizard_generic": {
            "prompt": "Build me an nf-core samplesheet",
            "reasoning": "Scenario B turn 1. No pipeline hinted; wizard intro should not auto-seed a pipeline.",
            "output": "",
        },
        "B3_ask_difference": {
            "prompt": "what's the difference between rnaseq and scrnaseq?",
            "reasoning": "Scenario B turn 2 — RABBIT HOLE. wizard_agent action=stay; reply must explain the distinction AND re-pose the pipeline question. State.step stays = 'pipeline'.",
            "output": "",
        },
        "B4_pipeline_choice": {
            "prompt": "what about scrnaseq vs sarek?",
            "reasoning": "Scenario B turn 3 — KEEP RABBIT-HOLING. wizard_agent action=stay again; reply differentiates the pipelines. Still pipeline step.",
            "output": "",
        },
        "B5_commit_pipeline": {
            "prompt": "okay let's go with rnaseq",
            "reasoning": "Scenario B turn 4. Advance pipeline; State.step→samples.",
            "output": "",
        },
        "B6_use_last_search": {
            "prompt": "use the IMPACT samples we just found",
            "reasoning": "Scenario B turn 5. Should match the use_last_search aliasing → pulls graph bundle UIDs from results_history.",
            "output": "",
        },
        "B7_what_cohort_info": {
            "prompt": "what cohort information exists for these samples?",
            "reasoning": "Scenario B turn 6 — THE KEY TEST. Triggers metadata prefetch on cohorts step entry. wizard_agent action=stay; reply must enumerate ACTUAL fields + value distributions from the fetched metadata (not abstract suggestions).",
            "output": "",
        },
        "B8_explore_options": {
            "prompt": "of those fields, which would actually make biologically meaningful cohorts?",
            "reasoning": "Scenario B turn 7 — KEEP RABBIT-HOLING ON COHORTS. wizard_agent action=stay; reply should reason about field variability from METADATA_SUMMARY and suggest 2-3 splits.",
            "output": "",
        },
        "B9_pick_single": {
            "prompt": "actually let's just keep it as a single cohort, this is for an initial run",
            "reasoning": "Scenario B turn 8. Advance cohorts step with cohort_criteria=[]. State.step→enrichment.",
            "output": "",
        },
        "B10_skip_enrichment": {
            "prompt": "none",
            "reasoning": "Scenario B turn 9. Advance enrichment step. State.step→confirm.",
            "output": "",
        },
        "B11_confirm": {
            "prompt": "yes",
            "reasoning": "Scenario B turn 10. Confirm → executes generate_report_outputs.",
            "output": "",
        },
        # ── Scenario C: wizard cancel mid-flight ───────────────────────────
        "C1_search_nhp": {
            "prompt": "Find me NHP samples from study GBM",
            "reasoning": "Scenario C setup. Graph bundle.",
            "output": "",
        },
        "C2_start_wizard": {
            "prompt": "build me a samplesheet for nf-core",
            "reasoning": "Scenario C turn 1. Wizard intercepts.",
            "output": "",
        },
        "C3_cancel_explicit": {
            "prompt": "cancel",
            "reasoning": "Scenario C turn 2. Deterministic cancel short-circuit; wizard cleared without LLM call. Subsequent turns should go through normal parser.",
            "output": "",
        },
        "C4_normal_query_after_cancel": {
            "prompt": "What investigations exist in the database?",
            "reasoning": "Scenario C turn 3. Verify wizard state truly cleared — this should route to system_question or similar, NOT through any wizard logic.",
            "output": "",
        },
        # ── Scenario D: memory recall against search bundles ───────────────
        "D1_search_for_memory": {
            "prompt": "Find me all D.SEQ samples for the SHA lab",
            "reasoning": "Scenario D setup. New api_result_full bundle with rich row data.",
            "output": "",
        },
        "D2_how_many": {
            "prompt": "How many were returned in that search?",
            "reasoning": "Scenario D turn 1. Parser should route mode=ask_about_last_results with target_result_id pointing at the most recent bundle. memory_coder writes Python over the full bundle (not just the chat_log first-5 summary).",
            "output": "",
        },
        "D3_filter_male": {
            "prompt": "Of those, which ones are from the SHA lab specifically? Give me UIDs.",
            "reasoning": "Scenario D turn 2. Chained memory question; tests pronoun resolution via chat_log's recent_turns.",
            "output": "",
        },
        "D4_first_three": {
            "prompt": "Just show me the first 3 UIDs from that search.",
            "reasoning": "Scenario D turn 3. Trivial memory question — could be served from chat_log's first_uids alone, but parser will likely route to ask_about_last_results.",
            "output": "",
        },
        # ── Scenario E: memory across mixed bundle types ───────────────────
        "E1_graph_dseq_impact": {
            "prompt": "Find me D.SEQ samples in project IMPACT",
            "reasoning": "Scenario E bundle 1 (graph_query).",
            "output": "",
        },
        "E2_graph_nhp_gbm": {
            "prompt": "Find me NHP samples from study GBM",
            "reasoning": "Scenario E bundle 2 (graph_query).",
            "output": "",
        },
        "E3_api_ndma_mice": {
            "prompt": "Find me mice treated with NDMA",
            "reasoning": "Scenario E bundle 3 (new_search via api).",
            "output": "",
        },
        "E4_recall_first_bundle": {
            "prompt": "How many D.SEQ IMPACT samples did the first search return?",
            "reasoning": "Scenario E turn 1 — by-bundle-id recall. parser sees chat_log and should pick target_result_id pointing at the FIRST bundle, not the most recent.",
            "output": "",
        },
        "E5_recall_middle_bundle": {
            "prompt": "What about those NHP samples — how many were there?",
            "reasoning": "Scenario E turn 2. Pronoun 'those' should resolve to the NHP/GBM bundle (middle one), not the most recent NDMA mice bundle.",
            "output": "",
        },
        "E6_recall_latest_bundle": {
            "prompt": "Of the NDMA mice, which are male?",
            "reasoning": "Scenario E turn 3. Explicit reference back to the most recent bundle; memory_coder filters api rows by Sex field.",
            "output": "",
        },
        # ── Scenario F: wizard restart ─────────────────────────────────────
        "F1_seed": {
            "prompt": "Find me NHP samples from study GBM",
            "reasoning": "Scenario F setup.",
            "output": "",
        },
        "F2_start_wizard": {
            "prompt": "Build an nfcore samplesheet for these",
            "reasoning": "Scenario F turn 1. Wizard intercept.",
            "output": "",
        },
        "F3_pick_rnaseq": {
            "prompt": "rnaseq",
            "reasoning": "Scenario F turn 2.",
            "output": "",
        },
        "F4_use_last": {
            "prompt": "use the samples we just found",
            "reasoning": "Scenario F turn 3.",
            "output": "",
        },
        "F5_restart": {
            "prompt": "actually wait, let me restart this",
            "reasoning": "Scenario F turn 4 — RESTART. wizard_agent action=restart; state resets pipeline/uids/cohort_criteria/enrichment_fields/metadata_summary; State.step→pipeline.",
            "output": "",
        },
        "F6_pick_sarek_after_restart": {
            "prompt": "this time let's do sarek",
            "reasoning": "Scenario F turn 5. After restart, must NOT remember the prior rnaseq choice; advances with pipeline=sarek.",
            "output": "",
        },
        "F7_cancel_after_restart": {
            "prompt": "okay actually cancel I changed my mind",
            "reasoning": "Scenario F turn 6 — combo restart-then-cancel. wizard_agent action=cancel (or deterministic short-circuit on 'cancel' substring).",
            "output": "",
        },
        # ── Scenario G: wizard multi-cohort split via LLM proposal flow ────
        "G1_seed_dseq": {
            "prompt": "Find me D.SEQ samples in project IMPACT",
            "reasoning": "Scenario G setup.",
            "output": "",
        },
        "G2_start_wizard": {
            "prompt": "build me an nfcore samplesheet for those samples",
            "reasoning": "Scenario G turn 1.",
            "output": "",
        },
        "G3_pick_rnaseq": {
            "prompt": "use rnaseq",
            "reasoning": "Scenario G turn 2.",
            "output": "",
        },
        "G4_use_last_search": {
            "prompt": "use the previous search results",
            "reasoning": "Scenario G turn 3.",
            "output": "",
        },
        "G5_ask_about_cohorts": {
            "prompt": "what fields could I split cohorts on?",
            "reasoning": "Scenario G turn 4. Metadata prefetch runs on cohorts entry; wizard_agent stays, lists fields + values.",
            "output": "",
        },
        "G6_propose_split": {
            "prompt": "let's split by LibraryStrategy if there's variation",
            "reasoning": "Scenario G turn 5 — PROPOSAL FLOW. wizard_agent should NOT auto-advance; should propose the concrete cohorts (e.g. RNA-Seq / amplicon if both exist) and ask the user to confirm.",
            "output": "",
        },
        "G7_refine_proposal": {
            "prompt": "actually just keep the RNA-Seq ones",
            "reasoning": "Scenario G turn 6 — REFINEMENT. wizard_agent should adjust the proposal to a single filtered cohort, then ask if that's the final answer.",
            "output": "",
        },
        "G8_confirm_cohorts": {
            "prompt": "yes that's right",
            "reasoning": "Scenario G turn 7. Advance with the refined cohort_criteria. State.step→enrichment.",
            "output": "",
        },
        "G9_pick_enrichment": {
            "prompt": "add Treatment1 and Timepoint as enrichment columns",
            "reasoning": "Scenario G turn 8. Advance with enrichment_fields=['Treatment1','Timepoint']. The LLM should validate against the metadata_summary keys before advancing.",
            "output": "",
        },
        "G10_confirm_emit": {
            "prompt": "yes build it",
            "reasoning": "Scenario G turn 9. Confirm → executes generate_report_outputs with multi-cohort-like spec.",
            "output": "",
        },
        # ── Final regression checks ────────────────────────────────────────
        "Z1_memory_after_long_thread": {
            "prompt": "Going way back — how many NDMA mice did we originally find?",
            "reasoning": "Final test. After ~50 turns of mixed wizard + searches, the chat_log tail-5 should NOT contain the very first NDMA search. Parser must rely on results_history + bundle-id reasoning rather than chat_log alone.",
            "output": "",
        },
        "Z2_clean_after_all_scenarios": {
            "prompt": "Find me all assays I have access to.",
            "reasoning": "Final test. A fresh new_search after all the wizard activity must work cleanly — no leftover wizard state, no parser confusion from the long history.",
            "output": "",
        },
    },
    # =========================================================================
    # wizard_e2e — Regression suite for the 10 fixes shipped in commit 95eab50
    # ("Real chat memory + LLM-driven nf-core wizard + memory hardening").
    #
    # Each scenario targets one or more specific bugs from the first test run's
    # post-mortem. To verify a fix, scrub the run's console.txt + chat_log for
    # the "what to check" hints under each scenario header. Runs as one shared
    # session — chat_log, results_history, and wizard state persist across all
    # turns the way a real conversation would.
    #
    # Coverage map (bug# → scenario):
    #   #1  NFCORE 5.1M-token blowup       → W1
    #   #2  Wizard escape hatch             → W2
    #   #3  Cohort metadata sample 5→15     → W5
    #   #4  Fatal → chat_log persist        → (covered by no-crash in W1)
    #   #6  Memory artifact unique names    → W4 (multi-recall on same bundle)
    #   #7  Memory_coder no hallucination   → W3
    #   #8  wizard_agent log_prompt         → confirm prompts.json has entries
    #   #9  HTML strip at bundle write      → confirm any bundle's row.uid
    #   #10 Deeper bundle visibility (3→8)  → W4
    #
    # Total: 36 turns. ~30s cooldown each + ~15s avg/turn + 1 long wizard execute
    # (W1 emit) ≈ 35–45 min wall time on mixed profile.
    # =========================================================================
    "wizard_e2e": {
        # ── W1: NFCORE bypass survives large UID set (Bug #1, #4 implicitly) ──
        # PRE-FIX BEHAVIOR: report_writer received per-sample lineage metadata
        # for all 195 NDMA mice → 5,119,285-token prompt → Opus ValidationException.
        # POST-FIX: emitter synthesizes rows from accession_metadata; no LLM call
        # in the NFCORE path of generate_report_outputs. Should emit a samplesheet.
        # CHECK: console.txt grep "Bypassing report_writer" + no FATAL line.
        "W1_seed_search": {
            "prompt": "Find me mice treated with NDMA.",
            "reasoning": "W1 setup. Seeds 195-UID api bundle.",
            "output": "",
        },
        "W1_start_wizard": {
            "prompt": "Build me an nf-core samplesheet for those mice.",
            "reasoning": "W1 turn 1. Wizard activates; intro emitted.",
            "output": "",
        },
        "W1_pick_rnaseq": {
            "prompt": "rnaseq",
            "reasoning": "W1 turn 2. Pipeline selected.",
            "output": "",
        },
        "W1_all_195_uids": {
            "prompt": "use the UIDs we just found",
            "reasoning": "W1 turn 3. Pulls ALL 195 UIDs via _resolve_last_search_uids. Pre-fix this is what caused the metadata fetch + report_writer to blow up.",
            "output": "",
        },
        "W1_single_cohort": {
            "prompt": "single cohort",
            "reasoning": "W1 turn 4.",
            "output": "",
        },
        "W1_no_enrichment": {
            "prompt": "no enrichment columns",
            "reasoning": "W1 turn 5.",
            "output": "",
        },
        "W1_confirm_emit": {
            "prompt": "yes go ahead and build it",
            "reasoning": "W1 turn 6. CRITICAL: must succeed end-to-end without 5M token crash. Look for 'Bypassing report_writer' debug line and an emitted /report/nfcore_rnaseq/samplesheet.csv. If this crashes the run dies (no further scenarios run).",
            "output": "",
        },
        # ── W2: Wizard escape hatch (Bug #2) ──────────────────────────────────
        # PRE-FIX: After the wizard restarted (or any state where the LLM
        # response was "want to restart?"), off-topic follow-up queries got
        # action=stay and the user was trapped. 4+ trapped turns in the prior run.
        # POST-FIX: Prompt rule + heuristic counter. Off-topic message →
        # action=cancel, wizard cleared, parser handles the actual query.
        # CHECK: W2_off_topic turn should NOT show wizard_active=True afterward.
        "W2_seed_search": {
            "prompt": "Find me D.SEQ samples in project IMPACT",
            "reasoning": "W2 setup. New bundle.",
            "output": "",
        },
        "W2_start_wizard": {
            "prompt": "Build me an nf-core samplesheet",
            "reasoning": "W2 turn 1. Wizard activates.",
            "output": "",
        },
        "W2_off_topic_search": {
            "prompt": "Actually, find me all assays I have access to.",
            "reasoning": "W2 turn 2 — OFF-TOPIC. wizard_agent should action=cancel (the prompt update + heuristic detect 'find me' as a task verb). Wizard must release control.",
            "output": "",
        },
        "W2_normal_after_escape": {
            "prompt": "What investigations exist in the database?",
            "reasoning": "W2 turn 3 — verify wizard truly cleared. Should route to normal parser (new_search), NOT through any wizard logic.",
            "output": "",
        },
        # ── W3: Memory_coder no hallucination (Bug #7, #9) ────────────────────
        # PRE-FIX: D3 returned 3 GRI-prefix UIDs claiming "from the SHA lab" but
        # those samples had no Lab metadata field. Substring scan against
        # HTML-wrapped URLs matched the string "SHA" accidentally.
        # POST-FIX: Prompt requires exact-match on categorical fields. Data
        # profile includes field_index. API rows are HTML-stripped at bundle
        # write. Expected answer: either correct exact matches OR a structured
        # "no Lab field present" reply.
        # CHECK: memory artifact JSON `memory_coder_bundle_*_<ts>.json` —
        # `computed_result.field_used` should reference json_metadata.Lab (or
        # similar), and either return real Lab=SHA samples or report
        # "field not present".
        "W3_seed_dseq_sha": {
            "prompt": "Find me all D.SEQ samples for the SHA lab",
            "reasoning": "W3 setup. Same loose-keyword API search that triggered the hallucination last run.",
            "output": "",
        },
        "W3_memory_filter_sha": {
            "prompt": "Of those, which are actually from the SHA lab? Give me UIDs with their Lab field value.",
            "reasoning": "W3 turn 1 — KEY TEST. memory_coder must NOT substring-match. Must either (a) report no Lab field, or (b) return samples whose json_metadata.Lab == 'SHA' exactly.",
            "output": "",
        },
        "W3_memory_count_unique_labs": {
            "prompt": "How many distinct labs are represented across those 264 samples?",
            "reasoning": "W3 turn 2. Verifies memory_coder uses field_index (Lab as categorical) to enumerate distinct values.",
            "output": "",
        },
        # ── W4: Deep memory recall + artifact uniqueness (Bug #6, #10) ────────
        # PRE-FIX: build_recent_results_summary capped at 3 bundles; bundles
        # older than that were invisible to parser. Memory artifacts overwrote
        # each other for follow-ups against the same bundle.
        # POST-FIX: cap bumped to 8; older bundles still mentioned via
        # CHAT_HISTORY narrative. Each memory_coder artifact gets a UTC-ts suffix.
        # CHECK: ls outputs/.../memory/ — should see multiple unique JSON files
        # for the same bundle_id. parser should resolve "original" to the early
        # NDMA bundle even after 8 intervening searches.
        "W4_search1_seq_in_impact": {
            "prompt": "Find me D.SEQ samples in project IMPACT",
            "reasoning": "W4 bundle 1 of 8 — earliest.",
            "output": "",
        },
        "W4_search2_nhp_metnet": {
            "prompt": "Find me NHP samples in the MetNet project",
            "reasoning": "W4 bundle 2.",
            "output": "",
        },
        "W4_search3_dna_samples": {
            "prompt": "Find me DNA samples",
            "reasoning": "W4 bundle 3.",
            "output": "",
        },
        "W4_search4_rna_samples": {
            "prompt": "Find me RNA samples",
            "reasoning": "W4 bundle 4.",
            "output": "",
        },
        "W4_search5_cell_lines": {
            "prompt": "Find me cell line samples",
            "reasoning": "W4 bundle 5.",
            "output": "",
        },
        "W4_search6_tissue": {
            "prompt": "Find me tissue samples in IMPACT",
            "reasoning": "W4 bundle 6.",
            "output": "",
        },
        "W4_search7_protocols": {
            "prompt": "What protocols exist for the IMPACT project?",
            "reasoning": "W4 bundle 7 (reporter or graph_query).",
            "output": "",
        },
        "W4_search8_studies": {
            "prompt": "Find me studies in MetNet",
            "reasoning": "W4 bundle 8.",
            "output": "",
        },
        "W4_recall_earliest": {
            "prompt": "How many D.SEQ samples did the very first search in this session return?",
            "reasoning": "W4 deep recall — bundle from 8 turns back. Parser should set target_result_id to W4_search1 bundle id even though it's at the FAR end of recent_summary (now 8-bundle window).",
            "output": "",
        },
        "W4_recall_middle_again": {
            "prompt": "And how many DNA samples did we find?",
            "reasoning": "W4 follow-up on bundle 3 — verify memory artifact gets unique filename (doesn't overwrite the recall_earliest artifact).",
            "output": "",
        },
        # ── W5: Wizard rabbit-hole with real metadata sample (Bug #3) ─────────
        # PRE-FIX: cohort step prefetched metadata for only 5 UIDs. For D.SEQ
        # IMPACT (250 samples), 5 randomly happened to all be LibraryStrategy=WGS,
        # making the wizard report "no variation" incorrectly.
        # POST-FIX: prefetch bumped to 15. Better representativeness.
        # CHECK: wizard_agent reply at W5_what_fields should list multiple
        # actual fields with example values; if data really has variation in
        # Scientist or LibraryStrategy, it should now surface that.
        "W5_seed_search": {
            "prompt": "Find me D.SEQ samples in project IMPACT",
            "reasoning": "W5 setup.",
            "output": "",
        },
        "W5_start_wizard": {
            "prompt": "Build me an nf-core samplesheet",
            "reasoning": "W5 turn 1.",
            "output": "",
        },
        "W5_pick_rnaseq": {
            "prompt": "rnaseq",
            "reasoning": "W5 turn 2.",
            "output": "",
        },
        "W5_use_last": {
            "prompt": "use the previous search",
            "reasoning": "W5 turn 3.",
            "output": "",
        },
        "W5_what_fields": {
            "prompt": "What metadata fields are available for these samples? I want to see real variation.",
            "reasoning": "W5 turn 4 — KEY TEST. Triggers 15-UID prefetch on cohorts entry. wizard should list actual fields with example values from real metadata. Compare against pre-fix run which only saw 5 UIDs.",
            "output": "",
        },
        "W5_pick_split": {
            "prompt": "If Scientist has variation, split by Scientist; otherwise just use single cohort.",
            "reasoning": "W5 turn 5. Lets the LLM decide based on actual prefetched data. Tests the proposal-then-confirm flow.",
            "output": "",
        },
        "W5_confirm_split_proposal": {
            "prompt": "yes that's right, use those cohorts",
            "reasoning": "W5 turn 6. Confirms whatever proposal the LLM made.",
            "output": "",
        },
        "W5_no_enrichment": {
            "prompt": "skip enrichment",
            "reasoning": "W5 turn 7.",
            "output": "",
        },
        "W5_cancel_before_emit": {
            "prompt": "cancel before emitting",
            "reasoning": "W5 turn 8. Cancel from confirm step; verifies deterministic cancel path on natural-language. Avoids a 2nd long emit run (W1 already covered emission).",
            "output": "",
        },
        # ── W6: Bundle skip transparency (edge case) ──────────────────────────
        # PRE-FIX: "use last search" silently picked an old bundle if the most
        # recent was zero-result; user had no way to know they got the wrong UIDs.
        # POST-FIX: _last_search_context exposes skipped_bundles. wizard_agent
        # is instructed (via step_goal) to mention them in its reply.
        # CHECK: W6_use_last_after_zero reply should explicitly say something
        # like "the most recent search returned zero, so I'm using <earlier query>".
        "W6_seed_zero": {
            "prompt": "Find me NHP samples in a nonexistent investigation called XYZQQQ",
            "reasoning": "W6 setup — should return zero results.",
            "output": "",
        },
        "W6_start_wizard": {
            "prompt": "Build me an nf-core samplesheet",
            "reasoning": "W6 turn 1. Wizard activates.",
            "output": "",
        },
        "W6_pick_rnaseq": {
            "prompt": "rnaseq",
            "reasoning": "W6 turn 2.",
            "output": "",
        },
        "W6_use_last_after_zero": {
            "prompt": "use the last search results",
            "reasoning": "W6 turn 3 — KEY TEST. Most recent bundle was zero-result. Wizard should walk past it and tell the user explicitly which earlier bundle was used. Reply must mention 'most recent returned zero' or similar wording.",
            "output": "",
        },
        "W6_cancel_clean_exit": {
            "prompt": "cancel",
            "reasoning": "W6 final — deterministic cancel.",
            "output": "",
        },
    },
}

ROOT = Path(__file__).resolve().parent
REFERENCE_DIR = ROOT / "outputs" / "reference_set"
CODE_SCAN_TARGETS: tuple[Path, ...] = (
    ROOT / "app.py",
    ROOT / "agents.py",
    ROOT / "helpers.py",
    ROOT / "config.py",
    ROOT / "cli.py",
    ROOT / "llm_clients.py",
    ROOT / "schemas" / "schema_helper.py",
    ROOT / "schemas" / "tools.py",
    ROOT / "schemas" / "router.py",
    ROOT / "schemas" / "entity.py",
    ROOT / "schemas" / "chat.py",
    ROOT / "test.py",
)

GROUND_TRUTH_PATH = ROOT / "src" / "chat_nextseek" / "context" / "testing_ground_truth.json"

# Cell background colors for validation results
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# Inline fonts for rich text coloring
FONT_GREEN = InlineFont(color="006400")  # Dark green
FONT_RED = InlineFont(color="8B0000")    # Dark red
FONT_BOLD = InlineFont(b=True)

EXCLUDED_VALIDATION_KEYS: set[str] = set()


def _parse_block(lines: list[str], start_idx: int, prefix: str) -> tuple[Any, int]:
    """
    Parse a block (JSON or text) starting after the prefix on start_idx line.
    Returns (parsed_value, next_index_after_block).
    Attempts JSON parsing first, falls back to text.
    """
    block: list[str] = []
    line = lines[start_idx]

    # Get content after the prefix
    if prefix in line:
        tail = line[len(prefix):].strip()
        # If there's a colon after the prefix text, split on that
        # e.g., "Parsed plan: {..." -> we want the part after ": "
        if tail and tail[0] != '{' and tail[0] != '[' and ": " in tail:
            tail = tail.split(": ", 1)[1].strip()
        if tail:
            block.append(tail)

    i = start_idx + 1
    while i < len(lines):
        # Stop at any bracketed debug/log line
        if lines[i].startswith("["):
            break
        block.append(lines[i])
        i += 1

    text = "\n".join(block).strip()
    if not text:
        return None, i

    # Try JSON first
    try:
        return json.loads(text), i
    except Exception:
        return text, i


# Patterns to capture: (line_prefix, agent_key, is_list)
# is_list=True means multiple occurrences should be appended to a list
AGENT_PATTERNS: list[tuple[str, str, bool]] = [
    ("[DEBUG][ENTITY] Parsed entity result:", "entity", False),
    ("[DEBUG][PARSER] Parsed plan:", "parser_plan", False),
    ("[DEBUG][API_AGENT] Parsed API plan:", "api_plan", True),
    ("[DEBUG][REPORTER] Parsed reporter plan:", "reporter_plan", False),
    ("[REPORTER][SQL] Running project sample report query", "reporter_sql", True),
    ("[DEBUG][REPORT_WRITER] Parsed report writer output:", "report_writer", False),
    ("[DEBUG][CHATTER] Final answer (post-processed):", "chatter", False),
    ("[DEBUG][CHATTER] Raw answer:", "chatter_raw", False),
    ("[DEBUG][MEMORY][MEMORY] Answer:", "memory", False),
    ("[DEBUG][API] Request:", "api_request", True),
    ("[DEBUG][API] Response:", "api_response", True),
]

def iter_test_cases() -> Iterable[dict[str, str]]:
    """
    Yield flattened test case entries with "id", "prompt", "reasoning", "output".
    Includes a "suite" key for the top-level grouping.
    """
    for suite_name, cases in TEST_CASES.items():
        for case_id, payload in cases.items():
            entry = {
                "id": case_id,
                "suite": suite_name,
                "prompt": payload.get("prompt", ""),
                "reasoning": payload.get("reasoning", ""),
                "output": payload.get("output", ""),
            }
            yield entry


def parse_console(path: Path) -> dict[str, dict[str, Any]]:
    """
    Parse console.txt into {query: {agent_key: output, ...}}.
    Dynamically captures all agent outputs based on AGENT_PATTERNS.
    """
    content = path.read_text(encoding="utf-8")
    lines = content.splitlines()
    results: dict[str, dict[str, Any]] = {}
    current: str | None = None

    for idx, line in enumerate(lines):
        # Detect new query boundary
        if line.startswith("[DEBUG][ENTITY] User query: "):
            marker = "User query:"
            if marker in line:
                current = line.split(marker, 1)[1].strip()
                results.setdefault(current, {})
            continue

        if current is None:
            continue

        # Capture timing lines: [TIMING][AGENT] X.XXs
        if line.startswith("[TIMING]["):
            try:
                agent_name = line.split("][")[1].split("]")[0]
                seconds_str = line.split("] ", 1)[1].replace("s", "").strip()
                seconds = float(seconds_str)
                results[current].setdefault("timing", {})[agent_name] = seconds
            except Exception:
                pass
            continue

        # Check all agent patterns
        for prefix, key, is_list in AGENT_PATTERNS:
            if line.startswith(prefix):
                parsed, _ = _parse_block(lines, idx, prefix)
                if parsed is not None:
                    if is_list:
                        results[current].setdefault(key, []).append(parsed)
                    else:
                        results[current][key] = parsed
                break

    return results

def _get_test_id_for_prompt(prompt: str) -> str:
    """
    Map a prompt string to its configured test ID for consistent sheet naming.
    Falls back to a sanitized slice of the prompt when the test case is absent so reports still generate.
    Keeps prompt-to-id resolution centralized for both parsing and workbook creation.
    """
    for case in iter_test_cases():
        if case.get("prompt") == prompt:
            return case.get("id", prompt[:30])
    # Fallback: use first 30 chars of prompt as ID
    return prompt[:30].replace(" ", "_").replace("?", "").replace(".", "")


def load_ground_truth() -> dict[str, dict[str, Any]]:
    """Load ground truth from src/chat_nextseek/context/testing_ground_truth.json."""
    if not GROUND_TRUTH_PATH.exists():
        return {}
    return json.loads(GROUND_TRUTH_PATH.read_text(encoding="utf-8"))


def _extract_codes(items: list) -> list[str]:
    """Extract sorted list of 'code' values from list of dicts."""
    return sorted([item.get("code", "") for item in items if isinstance(item, dict)])


def _validate_entity(actual: dict | None, expected: dict) -> dict[str, bool]:
    """Validate entity fields against ground truth."""
    results = {}
    if "sampletypes" in expected:
        actual_codes = _extract_codes((actual or {}).get("sampletypes", []))
        expected_codes = _extract_codes(expected["sampletypes"])
        results["entity.sampletypes"] = actual_codes == expected_codes
    if "keywords" in expected:
        actual_keywords = sorted((actual or {}).get("keywords", []) or [])
        expected_keywords = sorted(expected.get("keywords", []) or [])
        results["entity.keywords"] = actual_keywords == expected_keywords
    return results


def _validate_parser_plan(actual: dict | None, expected: dict) -> dict[str, bool]:
    """Validate parser_plan fields: mode, target_endpoint, report_mode, report_type."""
    results = {}
    actual = actual or {}
    for field in ["mode", "target_endpoint", "report_mode", "report_type"]:
        if field in expected:
            results[f"parser.{field}"] = actual.get(field) == expected[field]
    return results


def _validate_api_plan(actual: list | dict | None, expected: dict) -> dict[str, bool]:
    """
    Validate api_plan fields: endpoint, method, and requestBody subkeys.
    api_plan is typically a list with one or more plan objects.
    """
    results = {}

    # Extract the first plan from the list
    if isinstance(actual, list) and actual:
        actual = actual[0]
    actual = actual or {}

    # Validate endpoint
    if "endpoint" in expected:
        results["api_plan.endpoint"] = actual.get("endpoint") == expected["endpoint"]

    # Validate method
    if "method" in expected:
        results["api_plan.method"] = actual.get("method") == expected["method"]

    # Validate full plan (core fields only) when expected is provided
    results["api_plan.full"] = _normalize_body(_normalize_api_plan(actual)) == _normalize_body(_normalize_api_plan(expected))

    # Validate requestBody fields
    if "requestBody" in expected:
        actual_body = actual.get("requestBody", {})
        expected_body = expected["requestBody"]

        for field, expected_val in expected_body.items():
            actual_val = actual_body.get(field)
            if isinstance(expected_val, list) and isinstance(actual_val, list):
                # Compare lists order-independently
                results[f"api_plan.requestBody.{field}"] = sorted(str(x) for x in expected_val) == sorted(str(x) for x in actual_val)
            else:
                results[f"api_plan.requestBody.{field}"] = actual_val == expected_val

        # Also validate full request body when present in ground truth
        results["api_plan.requestBody_full"] = _normalize_body(actual_body) == _normalize_body(expected_body)

    return results


def _normalize_body(value: Any) -> Any:
    """Normalize request bodies for stable comparison."""
    if isinstance(value, dict):
        return {k: _normalize_body(v) for k, v in sorted(value.items())}
    if isinstance(value, list):
        normalized = [_normalize_body(v) for v in value]
        try:
            return sorted(normalized, key=lambda x: json.dumps(x, sort_keys=True))
        except Exception:
            return normalized
    return value


def _normalize_api_plan(plan: dict | None) -> dict:
    """Keep only stable API plan fields for validation."""
    plan = plan or {}
    return {
        "endpoint": plan.get("endpoint"),
        "method": plan.get("method"),
        "requestBody": plan.get("requestBody", {}),
    }


def _filter_validation_results(
    validation_results: dict[str, dict[str, dict[str, bool]]],
    exclude_keys: set[str],
) -> dict[str, dict[str, dict[str, bool]]]:
    """Remove selected validation keys before rendering the questions sheet."""
    if not exclude_keys:
        return validation_results
    filtered: dict[str, dict[str, dict[str, bool]]] = {}
    for test_id, test_data in validation_results.items():
        filtered[test_id] = {}
        for run_name, run_validation in test_data.items():
            filtered[test_id][run_name] = {
                k: v for k, v in run_validation.items() if k not in exclude_keys
            }
    return filtered


def _apply_sans_serif_font(wb: Workbook, font_name: str = "Calibri") -> None:
    """Apply a sans-serif font to all cells, preserving style flags."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                # Rich text cells keep their own inline fonts.
                if isinstance(cell.value, CellRichText):
                    continue
                f = cell.font or Font()
                cell.font = Font(
                    name=font_name,
                    size=f.sz or 11,
                    bold=f.b,
                    italic=f.i,
                    underline=f.u,
                    color=f.color,
                )


def validate_against_ground_truth(
    data: dict[str, dict[str, dict[str, Any]]],
    ground_truth: dict[str, dict[str, Any]]
) -> dict[str, dict[str, dict[str, bool]]]:
    """
    Validate test results against ground truth.

    Returns: {test_id: {run_name: {field_path: passed}}}
    """
    results: dict[str, dict[str, dict[str, bool]]] = {}

    for test_id, test_data in data.items():
        if test_id not in ground_truth:
            continue

        expected = ground_truth[test_id]
        results[test_id] = {}

        for run_name, outputs in test_data.items():
            run_results: dict[str, bool] = {}

            if "entity" in expected:
                run_results.update(_validate_entity(outputs.get("entity"), expected["entity"]))
            if "parser_plan" in expected:
                run_results.update(_validate_parser_plan(outputs.get("parser_plan"), expected["parser_plan"]))
            if "api_plan" in expected:
                run_results.update(_validate_api_plan(outputs.get("api_plan"), expected["api_plan"]))

            results[test_id][run_name] = run_results

    return results


def generate_comparison_report(date_folder: str) -> dict[str, dict[str, dict[str, Any]]]:
    """
    Parse all console.txt files in outputs/reference_set/{date_folder}/.

    Returns: {test_id: {run_name: {agent: output, ...}, ...}, ...}
    """
    folder_path = REFERENCE_DIR / date_folder
    if not folder_path.exists():
        raise FileNotFoundError(f"Reference folder not found: {folder_path}")

    # Collect all run folders
    run_folders = sorted([
        p for p in folder_path.iterdir()
        if p.is_dir() and (p / "console.txt").exists()
    ])

    if not run_folders:
        raise FileNotFoundError(f"No run folders with console.txt found in {folder_path}")

    # Structure: {test_id: {run_name: {agent: output}}}
    comparison_data: dict[str, dict[str, dict[str, Any]]] = {}

    for run_folder in run_folders:
        # Strip leading YYMMDD_HHMMSS_ date prefix from folder name if present
        run_name = re.sub(r"^\d{6}_\d{6}_", "", run_folder.name)
        console_path = run_folder / "console.txt"
        parsed = parse_console(console_path)

        # Map prompt -> test_id and organize data
        for prompt, agents in parsed.items():
            test_id = _get_test_id_for_prompt(prompt)
            if test_id not in comparison_data:
                comparison_data[test_id] = {}
            comparison_data[test_id][run_name] = agents

    return comparison_data


def create_xlsx_report(
    data: dict[str, dict[str, dict[str, Any]]],
    output_path: Path,
    validation_results: dict[str, dict[str, dict[str, bool]]] | None = None
) -> None:
    """
    Create an xlsx report with one sheet per test question.
    Each sheet only includes agents that were actually used for that question.
    If validation_results provided, adds validation columns to questions sheet.

    Args:
        data: {test_id: {run_name: {agent: output, ...}, ...}, ...}
        output_path: Path to write the xlsx file
        validation_results: {test_id: {run_name: {field_path: passed}}}
    """
    if validation_results:
        validation_results = _filter_validation_results(validation_results, EXCLUDED_VALIDATION_KEYS)

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Create "questions" sheet first (will appear first in workbook)
    # We'll add hyperlinks after creating all sheets
    questions_data: list[tuple[str, str]] = []
    data_ids = set(data.keys())
    for case in iter_test_cases():
        qid = case.get("id", "")
        prompt = case.get("prompt", "")
        if qid in data_ids:
            questions_data.append((qid, prompt))

    # Preferred ordering for agent keys
    preferred_order = [
        "entity", "parser_plan", "api_plan", "api_request", "api_response",
        "reporter_plan", "reporter_sql", "report_writer", "memory", "chatter_raw", "chatter"
    ]

    def sort_agent_keys(keys: set[str]) -> list[str]:
        """
        Sort agent keys into a stable, human-friendly order for workbook display.
        Preferred keys appear first, with any new or unexpected keys appended alphabetically.
        """
        result = []
        remaining = set(keys)
        for key in preferred_order:
            if key in remaining:
                result.append(key)
                remaining.discard(key)
        result.extend(sorted(remaining))  # Any new keys alphabetically
        return result

    # Get all run names across all test cases
    all_runs: set[str] = set()
    for test_data in data.values():
        all_runs.update(test_data.keys())
    run_names = sorted(all_runs)

    all_keys_seen: set[str] = set()

    # Load ground truth for displaying expected values on test sheets
    ground_truth = load_ground_truth()

    for test_id, test_data in sorted(data.items()):
        # Discover agent keys for THIS test question only
        sheet_agent_keys: set[str] = set()
        for run_data in test_data.values():
            sheet_agent_keys.update(k for k in run_data.keys() if k != "timing")

        agent_keys = sort_agent_keys(sheet_agent_keys)
        all_keys_seen.update(agent_keys)

        # Truncate sheet name to 31 chars (Excel limit)
        sheet_name = test_id[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Header row: Agent | ground_truth | run1 | run2 | ...
        ws.cell(row=1, column=1, value="Agent").font = Font(bold=True, name="Calibri")
        ws.cell(row=1, column=2, value="ground_truth").font = Font(bold=True, name="Calibri")
        for col_idx, run_name in enumerate(run_names, start=3):
            cell = ws.cell(row=1, column=col_idx, value=run_name)
            cell.font = Font(bold=True, name="Calibri")

        # Get ground truth for this test
        test_gt = ground_truth.get(test_id, {})

        # Data rows: one per agent (only agents used in this test)
        for row_idx, agent_key in enumerate(agent_keys, start=2):
            ws.cell(row=row_idx, column=1, value=agent_key).font = Font(bold=True, name="Calibri")

            # Ground truth column - show expected value for this agent
            gt_value = ""
            if agent_key == "entity" and "entity" in test_gt:
                gt_value = json.dumps(test_gt["entity"], indent=2, ensure_ascii=False)
            elif agent_key == "parser_plan" and "parser_plan" in test_gt:
                gt_value = json.dumps(test_gt["parser_plan"], indent=2, ensure_ascii=False)
            elif agent_key == "api_plan" and "api_plan" in test_gt:
                gt_value = json.dumps(test_gt["api_plan"], indent=2, ensure_ascii=False)

            gt_cell = ws.cell(row=row_idx, column=2, value=gt_value)
            gt_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # LLM run columns
            for col_idx, run_name in enumerate(run_names, start=3):
                agent_output = test_data.get(run_name, {}).get(agent_key)
                if agent_output is None:
                    cell_value = ""
                else:
                    try:
                        cell_value = json.dumps(agent_output, indent=2, ensure_ascii=False)
                    except Exception:
                        cell_value = str(agent_output)

                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Timing row at the bottom
        timing_row = len(agent_keys) + 2
        ws.cell(row=timing_row, column=1, value="timing").font = Font(bold=True, name="Calibri")
        ws.cell(row=timing_row, column=2, value="")
        for col_idx, run_name in enumerate(run_names, start=3):
            timing_data = test_data.get(run_name, {}).get("timing")
            if timing_data:
                cell_value = json.dumps(timing_data, indent=2)
            else:
                cell_value = ""
            cell = ws.cell(row=timing_row, column=col_idx, value=cell_value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Adjust column widths
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 40  # ground_truth column
        for col_idx in range(3, len(run_names) + 3):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 60

    # Now create the "performance" sheet with hyperlinks to each test sheet
    qs = wb.create_sheet(title="performance", index=0)

    # Headers: question_id, question, then validation columns per run
    qs.cell(row=1, column=1, value="question_id").font = Font(bold=True, name="Calibri")
    qs.cell(row=1, column=2, value="question").font = Font(bold=True, name="Calibri")

    # Add validation column headers if we have validation results
    if validation_results:
        for col_offset, run_name in enumerate(run_names):
            cell = qs.cell(row=1, column=3 + col_offset, value=run_name)
            cell.font = Font(bold=True, name="Calibri")

    for row_idx, (qid, prompt) in enumerate(questions_data, start=2):
        # question_id with hyperlink
        cell = qs.cell(row=row_idx, column=1, value=qid)
        sheet_name = qid[:31]
        if sheet_name in wb.sheetnames:
            cell.hyperlink = f"#'{sheet_name}'!A1"
            cell.font = Font(color="0000FF", underline="single", name="Calibri")

        # question text
        qs.cell(row=row_idx, column=2, value=prompt)

        # Validation columns for each run - with colored text per field
        if validation_results and qid in validation_results:
            test_validation = validation_results[qid]

            for col_offset, run_name in enumerate(run_names):
                col_idx = 3 + col_offset

                if run_name in test_validation:
                    run_validation = test_validation[run_name]

                    # Build rich text with colored lines for each field
                    passed = sum(1 for v in run_validation.values() if v)
                    total = len(run_validation)

                    # Create rich text blocks
                    blocks = [TextBlock(FONT_BOLD, f"{passed}/{total} correct\n")]
                    for field, field_passed in run_validation.items():
                        font = FONT_GREEN if field_passed else FONT_RED
                        status = "PASS" if field_passed else "FAIL"
                        blocks.append(TextBlock(font, f"{field}: {status}\n"))

                    cell = qs.cell(row=row_idx, column=col_idx)
                    cell.value = CellRichText(blocks)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    qs.cell(row=row_idx, column=col_idx, value="N/A")
        elif validation_results:
            # No ground truth for this test - leave blank cells
            pass

    # Add summary row at the bottom
    if validation_results:
        summary_row = len(questions_data) + 3  # +2 for header, +1 for spacing

        # Calculate totals per run
        run_totals: dict[str, dict[str, Any]] = {run: {"passed": 0, "total": 0, "failures": {}} for run in run_names}

        for test_id, test_data in validation_results.items():
            for run_name, run_validation in test_data.items():
                if run_name in run_totals:
                    for field, field_passed in run_validation.items():
                        run_totals[run_name]["total"] += 1
                        if field_passed:
                            run_totals[run_name]["passed"] += 1
                        else:
                            # Track failures by field
                            if field not in run_totals[run_name]["failures"]:
                                run_totals[run_name]["failures"][field] = []
                            run_totals[run_name]["failures"][field].append(test_id)

        # Write summary row
        summary_cell = qs.cell(row=summary_row, column=1, value="SUMMARY")
        summary_cell.font = Font(bold=True, name="Calibri")
        qs.cell(row=summary_row, column=2, value="")

        for col_offset, run_name in enumerate(run_names):
            col_idx = 3 + col_offset
            totals = run_totals[run_name]
            passed = totals["passed"]
            total = totals["total"]
            pct = (passed / total * 100) if total > 0 else 0

            # Build summary with failures grouped by field
            failures_by_field = totals["failures"]
            failure_lines = []
            for field, test_ids in sorted(failures_by_field.items()):
                failure_lines.append(f"{field}: {test_ids}")

            # Create rich text with summary and failures
            blocks = [TextBlock(FONT_BOLD, f"{passed}/{total} ({pct:.0f}%)\n\n")]
            if failure_lines:
                blocks.append(TextBlock(FONT_RED, "Failures:\n"))
                for line in failure_lines:
                    blocks.append(TextBlock(FONT_RED, f"  {line}\n"))

            cell = qs.cell(row=summary_row, column=col_idx)
            cell.value = CellRichText(blocks)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    qs.column_dimensions["A"].width = 25
    qs.column_dimensions["B"].width = 80
    if validation_results:
        for col_idx in range(3, 3 + len(run_names)):
            col_letter = qs.cell(row=1, column=col_idx).column_letter
            qs.column_dimensions[col_letter].width = 30

    # ======================================================
    # "timing" sheet — per-question full agent timing + avg column
    # ======================================================
    TIMING_AGENT_ORDER = ["ENTITY", "PARSER", "API_AGENT", "REPORTER", "MEMORY", "CHATTER", "TOTAL"]

    # Collect full timing dicts: {qid: {run_name: {agent: float}}}
    timing_grid: dict[str, dict[str, dict[str, float]]] = {}
    for qid, _ in questions_data:
        timing_grid[qid] = {}
        for run_name in run_names:
            t = data.get(qid, {}).get(run_name, {}).get("timing") or {}
            if t:
                timing_grid[qid][run_name] = t

    # Per-question per-agent averages across runs: {qid: {agent: float}}
    def _q_agent_avgs(qid: str) -> dict[str, float]:
        avgs: dict[str, float] = {}
        all_agents: set[str] = set()
        for t in timing_grid[qid].values():
            all_agents.update(t.keys())
        for agent in all_agents:
            vals = [timing_grid[qid][r][agent] for r in timing_grid[qid] if agent in timing_grid[qid][r]]
            if vals:
                avgs[agent] = sum(vals) / len(vals)
        return avgs

    # Pre-compute avg TOTAL per run to sort columns
    run_avg_times: dict[str, float | None] = {}
    for run_name in run_names:
        totals = [timing_grid[qid].get(run_name, {}).get("TOTAL") for qid, _ in questions_data]
        valid = [v for v in totals if v is not None]
        run_avg_times[run_name] = sum(valid) / len(valid) if valid else None

    # Sort runs fastest (lowest avg TOTAL) → slowest
    timing_run_order = sorted(run_names, key=lambda r: (run_avg_times[r] is None, run_avg_times[r] or 0))

    # Columns: A=ID, B=Question, C=Average, D+=runs
    AVG_COL = 3
    RUN_START_COL = 4

    ts = wb.create_sheet(title="timing")
    ts.cell(row=1, column=1, value="ID").font = Font(bold=True, name="Calibri")
    ts.cell(row=1, column=2, value="Question").font = Font(bold=True, name="Calibri")
    avg_hdr = ts.cell(row=1, column=AVG_COL, value="Average")
    avg_hdr.font = Font(bold=True, name="Calibri")
    avg_hdr.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for col_idx, run_name in enumerate(timing_run_order, start=RUN_START_COL):
        cell = ts.cell(row=1, column=col_idx, value=run_name)
        cell.font = Font(bold=True, name="Calibri")
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ts.row_dimensions[1].height = 40

    def _fmt_timing_cell(t: dict[str, float], avgs: dict[str, float]) -> CellRichText:
        """Build rich-text cell: each agent line colored green/red vs avg."""
        blocks: list[TextBlock] = []
        agents_present = [a for a in TIMING_AGENT_ORDER if a in t]
        for agent in agents_present:
            val = t[agent]
            avg = avgs.get(agent)
            diff = val - avg if avg is not None else None
            label = f"{agent}: {val:.2f}s"
            if diff is not None and abs(diff) > 0.01:
                sign = "+" if diff > 0 else "-"
                label += f" ({sign}{abs(diff):.2f})"
                font = InlineFont(color="8B0000" if diff > 0 else "006400")  # red=slower, green=faster
            else:
                font = InlineFont()
            if agent == "TOTAL":
                font = InlineFont(b=True, color=getattr(font, "color", None))
            blocks.append(TextBlock(font, label + "\n"))
        return CellRichText(blocks) if blocks else CellRichText([TextBlock(InlineFont(), "")])

    def _fmt_avg_cell(avgs: dict[str, float]) -> str:
        lines = [f"{a}: {avgs[a]:.2f}s" for a in TIMING_AGENT_ORDER if a in avgs]
        return "\n".join(lines)

    for row_idx, (qid, prompt) in enumerate(questions_data, start=2):
        id_cell = ts.cell(row=row_idx, column=1, value=qid)
        sheet_name = qid[:31]
        if sheet_name in wb.sheetnames:
            id_cell.hyperlink = f"#'{sheet_name}'!A1"
            id_cell.font = Font(color="0000FF", underline="single", bold=True, name="Calibri")
        else:
            id_cell.font = Font(bold=True, name="Calibri")
        ts.cell(row=row_idx, column=2, value=prompt).alignment = Alignment(wrap_text=True, vertical="top")

        q_avgs = _q_agent_avgs(qid)

        # Average column
        avg_cell = ts.cell(row=row_idx, column=AVG_COL, value=_fmt_avg_cell(q_avgs))
        avg_cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Per-run columns — color background by TOTAL, text by agent diffs
        row_totals = [timing_grid[qid].get(r, {}).get("TOTAL") for r in timing_run_order]
        valid_totals = [v for v in row_totals if v is not None]
        row_min = min(valid_totals) if valid_totals else None
        row_max = max(valid_totals) if valid_totals else None

        for col_idx, run_name in enumerate(timing_run_order, start=RUN_START_COL):
            t = timing_grid[qid].get(run_name)
            cell = ts.cell(row=row_idx, column=col_idx)
            if t:
                cell.value = _fmt_timing_cell(t, q_avgs)
                total = t.get("TOTAL")
                if total is not None and row_min is not None and row_max is not None and row_min != row_max:
                    if total == row_min:
                        cell.fill = FILL_GREEN
                    elif total == row_max:
                        cell.fill = FILL_RED
            else:
                cell.value = ""
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Average row (per-run avg across all questions)
    avg_row = len(questions_data) + 2
    ts.cell(row=avg_row, column=1, value="Average").font = Font(bold=True, name="Calibri")
    ts.cell(row=avg_row, column=2, value="")

    # Overall average column (avg of per-question avgs)
    overall_agent_avgs: dict[str, list[float]] = {}
    for qid, _ in questions_data:
        for agent, val in _q_agent_avgs(qid).items():
            overall_agent_avgs.setdefault(agent, []).append(val)
    overall_avgs = {a: sum(v) / len(v) for a, v in overall_agent_avgs.items()}
    ts.cell(row=avg_row, column=AVG_COL, value=_fmt_avg_cell(overall_avgs)).alignment = Alignment(wrap_text=True, vertical="top")

    avg_total_vals = [run_avg_times[r] for r in timing_run_order if run_avg_times[r] is not None]
    for col_idx, run_name in enumerate(timing_run_order, start=RUN_START_COL):
        # Build per-agent averages for this run across all questions
        run_agent_vals: dict[str, list[float]] = {}
        for qid, _ in questions_data:
            for agent, val in (timing_grid[qid].get(run_name) or {}).items():
                run_agent_vals.setdefault(agent, []).append(val)
        run_agent_avgs = {a: sum(v) / len(v) for a, v in run_agent_vals.items()}

        cell = ts.cell(row=avg_row, column=col_idx)
        cell.value = _fmt_timing_cell(run_agent_avgs, overall_avgs) if run_agent_avgs else ""
        cell.font = Font(bold=True, name="Calibri")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        avg = run_avg_times[run_name]
        if avg is not None and avg_total_vals and min(avg_total_vals) != max(avg_total_vals):
            if avg == min(avg_total_vals):
                cell.fill = FILL_GREEN
            elif avg == max(avg_total_vals):
                cell.fill = FILL_RED

    ts.column_dimensions["A"].width = 22
    ts.column_dimensions["B"].width = 45
    ts.column_dimensions[ts.cell(row=1, column=AVG_COL).column_letter].width = 22
    for col_idx in range(RUN_START_COL, len(timing_run_order) + RUN_START_COL):
        ts.column_dimensions[ts.cell(row=1, column=col_idx).column_letter].width = 25

    # ======================================================
    # "summary" sheet — leaderboard: accuracy + timing + score
    # ======================================================
    ss = wb.create_sheet(title="summary")
    ss.cell(row=1, column=1, value="Run").font = Font(bold=True, name="Calibri")
    ss.cell(row=1, column=2, value="Accuracy %").font = Font(bold=True, name="Calibri")
    ss.cell(row=1, column=3, value="Avg Total Time (s)").font = Font(bold=True, name="Calibri")
    ss.cell(row=1, column=4, value="Score (acc%^1.3 / time)").font = Font(bold=True, name="Calibri")
    ss.cell(row=1, column=5, value="Passed / Total").font = Font(bold=True, name="Calibri")

    # Build per-run accuracy from validation_results
    run_accuracy: dict[str, dict[str, Any]] = {r: {"passed": 0, "total": 0} for r in run_names}
    if validation_results:
        for test_id, test_data in validation_results.items():
            for run_name, run_val in test_data.items():
                if run_name in run_accuracy:
                    for field_passed in run_val.values():
                        run_accuracy[run_name]["total"] += 1
                        if field_passed:
                            run_accuracy[run_name]["passed"] += 1

    # Build rows: (run, acc_pct, avg_time, score)
    summary_rows: list[tuple[str, float | None, float | None, float | None, str]] = []
    for run_name in run_names:
        acc = run_accuracy[run_name]
        acc_pct = (acc["passed"] / acc["total"] * 100) if acc["total"] > 0 else None
        avg_time = run_avg_times.get(run_name)
        score = ((acc_pct ** 1.3) / avg_time) if (acc_pct is not None and avg_time) else None
        passed_str = f"{acc['passed']}/{acc['total']}" if acc["total"] > 0 else "N/A"
        summary_rows.append((run_name, acc_pct, avg_time, score, passed_str))

    # Sort by score descending (None last)
    summary_rows.sort(key=lambda x: (x[3] is None, -(x[3] or 0)))

    scores = [r[3] for r in summary_rows if r[3] is not None]
    times = [r[2] for r in summary_rows if r[2] is not None]
    accs = [r[1] for r in summary_rows if r[1] is not None]

    for row_idx, (run_name, acc_pct, avg_time, score, passed_str) in enumerate(summary_rows, start=2):
        ss.cell(row=row_idx, column=1, value=run_name).font = Font(bold=True, name="Calibri")

        acc_cell = ss.cell(row=row_idx, column=2, value=round(acc_pct, 1) if acc_pct is not None else "N/A")
        acc_cell.alignment = Alignment(horizontal="center")
        if accs and acc_pct is not None and min(accs) != max(accs):
            acc_cell.fill = FILL_GREEN if acc_pct == max(accs) else (FILL_RED if acc_pct == min(accs) else PatternFill())

        time_cell = ss.cell(row=row_idx, column=3, value=round(avg_time, 2) if avg_time is not None else "N/A")
        time_cell.alignment = Alignment(horizontal="center")
        if times and avg_time is not None and min(times) != max(times):
            time_cell.fill = FILL_GREEN if avg_time == min(times) else (FILL_RED if avg_time == max(times) else PatternFill())

        score_cell = ss.cell(row=row_idx, column=4, value=round(score, 3) if score is not None else "N/A")
        score_cell.alignment = Alignment(horizontal="center")
        if scores and score is not None and min(scores) != max(scores):
            score_cell.fill = FILL_GREEN if score == max(scores) else (FILL_RED if score == min(scores) else PatternFill())

        ss.cell(row=row_idx, column=5, value=passed_str).alignment = Alignment(horizontal="center")

    ss.column_dimensions["A"].width = 30
    ss.column_dimensions["B"].width = 15
    ss.column_dimensions["C"].width = 20
    ss.column_dimensions["D"].width = 22
    ss.column_dimensions["E"].width = 15

    # Reorder sheets: summary first, timing second, performance third, question sheets after
    sheet_names = wb.sheetnames
    summary_idx = sheet_names.index("summary") if "summary" in sheet_names else -1
    timing_idx = sheet_names.index("timing") if "timing" in sheet_names else -1
    perf_idx = sheet_names.index("performance") if "performance" in sheet_names else -1
    if summary_idx > 0:
        wb.move_sheet("summary", offset=-summary_idx)
    if timing_idx >= 0:
        wb.move_sheet("timing", offset=1 - wb.sheetnames.index("timing"))
    if perf_idx >= 0:
        wb.move_sheet("performance", offset=2 - wb.sheetnames.index("performance"))

    _apply_sans_serif_font(wb, font_name="Calibri")

    wb.save(output_path)
    print(f"Wrote comparison report to {output_path}")
    print(f"Agent keys found across all tests: {sort_agent_keys(all_keys_seen)}")


def _collect_top_level_functions(path: Path) -> list[tuple[str, int]]:
    """
    Return (name, lineno) for top-level functions in the given Python file.
    """
    text = path.read_text(encoding="utf-8")
    tree = ast.parse(text, filename=str(path))
    defs: list[tuple[str, int]] = []

    def visit(node: ast.AST, depth: int) -> None:
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)) and depth == 0:
            defs.append((node.name, node.lineno))
        for child in ast.iter_child_nodes(node):
            child_depth = depth + 1 if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)) else depth
            visit(child, child_depth)

    visit(tree, 0)
    return defs


def _collect_function_usages(path: Path, target_names: set[str]) -> set[str]:
    """
    Collect function names referenced (Name or Attribute attr) in a file.
    """
    text = path.read_text(encoding="utf-8")
    tree = ast.parse(text, filename=str(path))
    used: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Name) and isinstance(node.ctx, ast.Load):
            if node.id in target_names:
                used.add(node.id)
        elif isinstance(node, ast.Attribute):
            if node.attr in target_names:
                used.add(node.attr)
        elif isinstance(node, ast.ImportFrom):
            for alias in node.names:
                if alias.name in target_names or (alias.asname and alias.asname in target_names):
                    used.add(alias.name)
                if alias.asname and alias.asname in target_names:
                    used.add(alias.asname)
    return used


def run_test_suite(
    suite_name: str,
    models: list[str] | None = None,
    planner: bool = False,
    only: str | None = None,
) -> int:
    """
    Run a test suite (e.g., test_case_1) across specified models.
    Creates output folders like Streamlit: outputs/YYMMDD_HHMMSS_mode_model/

    Args:
        suite_name: Name of the test suite (test_case_1, test_case_2, etc.)
        models: List of models to run (default: ["gcp", "anth", "oai"])
        planner: If True, use run_query_plan (planner pipeline) instead of run_query
        only: Comma-separated case-key prefixes to filter (e.g. "W3,W4" → all W3_* + W4_*).
              When None/empty, runs every case in the suite.

    Returns:
        0 on success, 1 on error
    """
    import sys
    import time
    from datetime import datetime

    if suite_name not in TEST_CASES:
        print(f"[test] Unknown test suite: {suite_name}")
        print(f"[test] Available suites: {list(TEST_CASES.keys())}")
        return 1

    if models is None:
        models = ["mixed"]

    suite = TEST_CASES[suite_name]
    all_prompts = [(case_id, case["prompt"]) for case_id, case in suite.items()]

    if only:
        prefixes = tuple(p.strip() for p in only.split(",") if p.strip())
        prompts = [(cid, q) for (cid, q) in all_prompts if cid.startswith(prefixes) or cid in prefixes]
        if not prompts:
            print(f"[test] --only={only!r} matched no cases in suite '{suite_name}'")
            print(f"[test] Available case keys: {[c for c, _ in all_prompts]}")
            return 1
        print(f"[test] --only={only!r} → running {len(prompts)}/{len(all_prompts)} cases: {[c for c, _ in prompts]}")
    else:
        prompts = all_prompts

    pipeline = "planner" if planner else "standard"
    print(f"[test] Running suite '{suite_name}' with {len(prompts)} test cases")
    print(f"[test] Models: {models}")
    print(f"[test] Pipeline: {pipeline}")

    outputs_dir = ROOT / "outputs"
    outputs_dir.mkdir(exist_ok=True)

    created_dirs: list[Path] = []

    import importlib
    import sys

    for mode in models:
        print(f"\n{'='*60}")
        print(f"[test] Starting model: {mode}")
        print(f"{'='*60}\n")

        # Set mode via env before reloading modules
        os.environ["NEXTSEEK_MODE"] = mode

        # Remove cached modules so reload picks up new env
        modules_to_reload = [
            "chat_nextseek.config",
            "chat_nextseek.llm_clients",
            "chat_nextseek.schemas.schema_helper",
            "chat_nextseek.helpers",
            "chat_nextseek.agents",
            "chat_nextseek.orchestrator",
        ]
        for mod_name in modules_to_reload:
            if mod_name in sys.modules:
                del sys.modules[mod_name]

        # Now import fresh
        from chat_nextseek.config import ChatConfig
        from chat_nextseek.orchestrator import handle_query, handle_query_plan
        from chat_nextseek.session import SQLiteSessionState
        from chat_nextseek.tee import Tee
        query_fn = handle_query_plan if planner else handle_query

        config = ChatConfig({})

        LLM_MODEL = config.LLM_MODEL
        
        # Create output folder like Streamlit does
        ts = datetime.now().strftime("%y%m%d_%H%M%S")
        log_dir = outputs_dir / f"{ts}_{mode}_{LLM_MODEL}"
        log_dir.mkdir(exist_ok=True)
        created_dirs.append(log_dir)

        console_log_path = log_dir / "console.txt"

        # Set up Tee logging
        original_stdout = sys.stdout
        original_stderr = sys.stderr
        tee_out = Tee(sys.stdout, str(console_log_path))
        tee_err = Tee(sys.stderr, str(console_log_path))
        sys.stdout = tee_out
        sys.stderr = tee_err

        try:
            # Log config
            print(f"[test] Output directory: {log_dir}")
            try:
                snapshot = config.get_config_snapshot()
                print("[CONFIG] Snapshot:\n" + json.dumps(snapshot, indent=2))
            except Exception as e:
                print(f"[CONFIG] Failed to log config: {e}")

            # Create fresh session for this model run
            db_path = log_dir / "session.sqlite"
            session = SQLiteSessionState(str(db_path), "test-runner")
            session["log_dir"] = str(log_dir)
            session["results_history"] = []

            # Run each test case with retry logic for timeouts
            max_retries = 3
            for case_id, prompt in prompts:
                # One-liner continuity snapshot — confirms session memory is carrying across turns.
                _rh = session.get("results_history") or []
                _cl = session.get("chat_log") or []
                _wz = session.get("nfcore_wizard") or {}
                print(
                    f"\n[test][state] before {case_id}: "
                    f"results_history={len(_rh)} "
                    f"chat_log={len(_cl)} "
                    f"wizard_active={bool(_wz.get('active'))} "
                    f"wizard_step={_wz.get('step') if _wz.get('active') else '-'}"
                )
                print(f"{'-'*40}")
                print(f"[test] Running: {case_id}")
                print(f"[test] Prompt: {prompt}")
                print(f"{'-'*40}\n")

                from chat_nextseek.llm_clients import LLMTimeoutError

                for attempt in range(1, max_retries + 1):
                    try:
                        query_fn(session, config, prompt)
                        break  # Success, move to next test case
                    except LLMTimeoutError as e:
                        print(f"[test] Timeout in {case_id} (attempt {attempt}/{max_retries}): {e}")
                        if attempt < max_retries:
                            print(f"[test] Retrying {case_id}...")
                        else:
                            print(f"[test] Max retries reached for {case_id}, skipping")
                    except Exception as e:
                        print(f"[test] Error in {case_id}: {e}")
                        import traceback
                        traceback.print_exc()
                        break  # Non-timeout errors don't retry

                # Cooldown between test cases to avoid provider bursts.
                try:
                    print("[test] waiting 30 seconds")
                    time.sleep(30)
                except Exception:
                    pass

            print(f"\n[test] Completed {mode}: {log_dir}")

        finally:
            # Restore original stdout/stderr
            sys.stdout = original_stdout
            sys.stderr = original_stderr
            # Flush tee files
            tee_out.flush()
            tee_err.flush()

    print(f"\n{'='*60}")
    print(f"[test] All runs complete. Output directories:")
    for d in created_dirs:
        print(f"  {d}")
    print(f"{'='*60}")

    return 0


def run_unused_function_scan() -> int:
    """
    Scan top-level functions across CODE_SCAN_TARGETS and report any that are never referenced.
    Returns 0 always; prints findings to stdout.
    """
    defs: dict[str, list[tuple[Path, int]]] = {}
    for path in CODE_SCAN_TARGETS:
        if not path.exists():
            continue
        for name, lineno in _collect_top_level_functions(path):
            defs.setdefault(name, []).append((path, lineno))

    all_names = set(defs.keys())
    used: set[str] = set()
    for path in CODE_SCAN_TARGETS:
        if not path.exists():
            continue
        used.update(_collect_function_usages(path, all_names))

    unused = {name: locations for name, locations in defs.items() if name not in used}

    if not unused:
        print("All top-level functions are referenced at least once in scanned files.")
        return 0

    print("Unused top-level functions (by name, with locations):")
    for name, locations in sorted(unused.items()):
        loc_str = ", ".join(f"{p.relative_to(ROOT)}:{lineno}" for p, lineno in locations)
        print(f"- {name}: {loc_str}")
    print("\nNote: usage scan is static and may over/under-count in dynamic or reflective code.")
    return 0


def main():
    """
    CLI entrypoint for comparison reports (mode=report) or static code scan (mode=code).
    Mode defaults to report; code scan prints unused top-level functions.
    """
    parser = argparse.ArgumentParser(description="Test comparison tool for NExtSEEK")
    parser.add_argument(
        "-m", "--mode",
        choices=["report", "code"],
        default="report",
        help="Mode: report (comparison) or code (unused function scan).",
    )
    parser.add_argument(
        "mode_pos",
        nargs="?",
        choices=["report", "code"],
        help="Optional positional mode to allow `python test.py code`.",
    )
    parser.add_argument(
        "-i", "--input",
        type=str,
        help="Date folder name in outputs/reference_set/ (e.g., 260115) when mode=report.",
    )
    args = parser.parse_args()

    mode = args.mode_pos or args.mode or "report"

    if mode == "code":
        return run_unused_function_scan()

    if not args.input:
        print("Mode 'report' requires -i/--input (e.g., 260115).")
        return

    data = generate_comparison_report(args.input)

    # Load ground truth and validate
    ground_truth = load_ground_truth()
    validation_results = None
    if ground_truth:
        validation_results = validate_against_ground_truth(data, ground_truth)
        print(f"Validated {len(validation_results)} test cases against ground truth")
    else:
        print("No ground truth file found, skipping validation")

    output_path = REFERENCE_DIR / args.input / "comparison_report.xlsx"
    create_xlsx_report(data, output_path, validation_results)


if __name__ == "__main__":
    main()
