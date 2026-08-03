from __future__ import annotations

import csv
from pathlib import Path, PurePosixPath
from typing import Any

from nessie_tests.pathsetup import ensure_e2e_importable

ensure_e2e_importable()
from e2e.criteria import check_pass, resolve_field  # noqa: E402
from nessie_tests import route_observer as ro

# Criteria in the base catalog address produced files as `api_artifact.<name>`.
# e2e resolves those against a `run_root` on disk, but nessie drives the async
# HTTP endpoint, which reuses ONE run root per gunicorn process rather than
# creating one per turn — so there is no per-turn root to hand it, and every
# such criterion resolved to None and could never pass. We resolve them from the
# turn's own emitted artifact paths instead, which is both correct per-turn and
# independent of where the server happens to write.
ARTIFACT_PREFIX = "api_artifact."
ARTIFACT_INDEX_KEY = "nessie_artifact_index"

# Row counts that land EXACTLY on a cypher LIMIT are almost certainly truncated.
# A capped result is indistinguishable from a complete one to every other
# criterion — `graph.what_mice_are_in_the_impact_st` passed every assertion in
# the 2026-07-24 run while returning exactly 250 rows. Both the historical cap
# and the current one are listed so an old deployment is still caught.
# Re-exported for back-compat; the definition moved to nessie_tests.limits so
# consistency.py can use it without importing e2e.criteria and openpyxl.
from nessie_tests.limits import GRAPH_LIMIT_SENTINELS  # noqa: E402,F401

# Provider-outage detection. Re-exported rather than reimplemented: consistency.py
# needs the same verdict and cannot import this module (openpyxl, e2e.criteria), so
# the detector itself lives in the dependency-free nessie_tests.outage.
from nessie_tests.outage import (  # noqa: E402,F401
    OUTAGE_REASON, PROVIDER_OUTAGE_MARKER, is_provider_outage,
)


def classify_turn_status(passed: bool, last_reply: str | None) -> str:
    """Map one turn's criterion outcome to a manifest status.

    An outaged turn is ``error`` no matter what its criteria did.

    * Not ``failed``, because nothing was tested: the provider chain gave up
      before the parser ran, so the reply is an infrastructure message and
      scoring it as a product regression is how ten of the eighteen reds in the
      2026-08-03 seed-6 run came to be triaged by hand.
    * Not ``passed`` either, even when a criterion is satisfied. ``route`` is
      resolved from ``route_decided``, which fires BEFORE the LLM call, and
      ``last_reply nonempty`` is satisfied by the error text itself — so an
      outaged turn can look green while having exercised nothing. Recording that
      as a pass is the same lie in the other direction.

    ``error`` is the status the runner already uses for infrastructure
    (runner.py:171), so no new status is introduced; what IS new is the
    ``outage`` flag the runner sets alongside it, which is what exempts the entry
    from the gate. A non-outage error stays gate-failing.
    """
    if is_provider_outage(last_reply):
        return "error"
    return "passed" if passed else "failed"


def _last(payload, name):
    data = None
    for ev in payload.get("progress") or []:
        if ev.get("event") == name:
            data = ev.get("data") or {}
    return data


def build_observed_debug(payload: dict) -> dict:
    # The live turn carries api_result_meta/graph_result on query_complete.debug;
    # that primary path is authoritative. (A former search_complete api_ok/neo4j_ok
    # backfill was dead — the live search_complete event emits {source, ok, count},
    # not api_ok/neo4j_ok — so it is intentionally omitted.)
    return dict((_last(payload, "query_complete") or {}).get("debug") or {})


def _collect_paths(value: Any, out: list[str]) -> None:
    if isinstance(value, str):
        out.append(value)
    elif isinstance(value, dict):
        for item in value.values():
            _collect_paths(item, out)
    elif isinstance(value, (list, tuple)):
        for item in value:
            _collect_paths(item, out)


def build_artifact_index(debug: dict, payload: dict) -> dict[str, str]:
    """Map produced-artifact basename -> absolute path for this turn.

    Sources, all on the query_complete event except the first:

    * ``debug.report_saved_files`` — reporter/pipeline writes (NS).
    * ``files`` — anything the NS turn attached.
    * ``artifacts`` — emitted by BOTH the Container-CC publish path
      (``cc_engine.py:789``) and the NS reporter path (``orchestrator.py:813-839``
      via ``extract_table_artifacts``). Only ``artifact_type == "file"`` entries
      are indexed: the reporter also emits ``"table"`` and ``"preview"`` entries
      whose ``label`` is a human string ("GEO Report Preview", "Sample Types")
      with no file behind it, and ``api_artifact.<name>`` means "a file with this
      basename was produced", so indexing those would make it true for an inline
      table. Neither producer sets ``path``, so ``label`` is the fallback.
    * ``cc_raw_files`` — Container-CC scratch/raw writes, plain path strings.

    The last two exist because a CC turn emits NEITHER of the first two: without
    them every ``api_artifact.<name>`` criterion resolved False on every CC turn,
    so the three file-producing families could never be scored at all.

    KNOWN LIMIT — a multi-deliverable CC turn only ever exposes ``artifacts.zip``.
    ``_publish_artifacts`` (``cc_engine.py:954-961``) zips whenever it finds more
    than one deliverable and emits a SINGLE artifact labelled ``artifacts.zip``;
    the member filenames never reach ``query_complete``. So a reingest turn that
    writes a workbook plus anything else resolves ``api_artifact.upload.xlsx``
    False and only ``api_artifact.artifacts.zip`` True. There is no harness-side
    fix — the names are not in the payload. CC criteria must assert
    ``api_artifact.artifacts.zip`` for multi-file turns; only a single-deliverable
    turn can assert a real basename.
    """
    paths: list[str] = []
    _collect_paths(debug.get("report_saved_files") or {}, paths)
    qc = _last(payload, "query_complete") or {}

    def _add(entries, *keys: str, files_only: bool = False) -> None:
        for entry in entries or []:
            if isinstance(entry, dict):
                if files_only and entry.get("artifact_type") not in (None, "file"):
                    continue
                candidate = next((entry.get(k) for k in keys if entry.get(k)), None)
                if candidate:
                    paths.append(str(candidate))
            elif isinstance(entry, str):
                paths.append(entry)

    _add(qc.get("files"), "path", "name")
    _add(qc.get("artifacts"), "path", "label", files_only=True)
    _add(qc.get("cc_raw_files"), "path", "name")
    return {PurePosixPath(p).name: p for p in paths if p}


def _count_rows(path: Path) -> int:
    if not path.exists():
        return 0
    if path.suffix == ".csv":
        with open(path, newline="", encoding="utf-8") as fh:
            return max(0, len(list(csv.reader(fh))) - 1)  # exclude header
    if path.suffix in (".xlsx", ".xls"):
        from openpyxl import load_workbook  # noqa: PLC0415
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        count = sum(1 for _ in ws.iter_rows(min_row=2)) if ws else 0
        wb.close()
        return count
    return sum(1 for line in path.read_text().splitlines() if line.strip())


def resolve_artifact(index: dict[str, str], field: str) -> Any:
    """Resolve one `api_artifact.<name>[.rows_gte]` field from the turn's index."""
    sub = field[len(ARTIFACT_PREFIX):]
    if sub.endswith(".rows_gte"):
        name = sub[: -len(".rows_gte")]
        path = index.get(name)
        # A CC/reporter artifact carries no path and is indexed under its bare
        # label, so Path(label) would resolve against the harness cwd (/app in
        # the container lane) and count rows out of an unrelated same-named file
        # — `samplesheet.csv` is asserted 3x across 2 variants. A value with no
        # separator is a label, not a path: return the 0 this branch returned
        # before artifacts were indexed at all, without touching the filesystem.
        if path is None or "/" not in path:
            return 0
        return _count_rows(Path(path))
    return sub in index


def augment_debug(debug: dict, obs: ro.RouteObservation, bundle_summary: dict | None = None,
                  artifact_index: dict[str, str] | None = None) -> dict:
    debug = dict(debug)
    debug["route"] = obs.route
    debug["engine"] = obs.engine
    debug["route_source"] = obs.source
    if bundle_summary is not None:
        debug["bundle"] = bundle_summary
    if artifact_index is not None:
        debug[ARTIFACT_INDEX_KEY] = artifact_index
    debug["graph_not_truncated"] = _graph_not_truncated(debug)
    debug["graph_truncation_disclosed"] = _graph_truncation_disclosed(debug)
    debug["report_produced_output"] = _report_produced_output(debug)
    debug["graph_outcome_observed"] = _graph_outcome_observed(debug)
    debug["api_outcome_observed"] = _api_outcome_observed(debug)
    debug["outcome_observed"] = _outcome_observed(debug)
    return debug


def _outcome_observed(debug: dict) -> bool:
    """True when the turn produced an outcome by ANY engine.

    The family floor is attached at corpus-BUILD time keyed on `v.family`
    (`corpus.py:45`), and the family names are engine-shaped: `search_advanced`
    means the REST advanced_search endpoint, `graph_query` means Cypher. The parser
    is free to answer the same question with either, and in the 2026-08-03 seed-6
    run three `search_advanced` cases routed NS correctly, answered correctly via
    the graph, and went red on `api_ok` / `api_outcome_observed` — both False on a
    graph turn by construction. The operator's note on all three was "this was
    correct".

    `apply_family_floor` cannot know which engine will run, so the fix has to live
    here, where the outcome is already in hand. The floor now asserts THAT an
    outcome was produced; asserting WHICH engine produced it is a per-case decision
    and the three engine-specific booleans stay available for exactly that.

    Strictly the disjunction of those three — nothing else. It must stay that way:
    a floor that is true unconditionally is worse than no floor, because it reports
    green. `test_a_turn_that_produced_no_outcome_at_all_still_fails_the_floor` and
    `test_outcome_observed_is_exactly_the_disjunction_of_the_three` pin it.
    """
    return (_graph_outcome_observed(debug)
            or _api_outcome_observed(debug)
            or _report_produced_output(debug))


def _graph_outcome_observed(debug: dict) -> bool:
    """True when the graph turn reported a row count, INCLUDING zero.

    The floor's job is to catch a criterion that cannot see the answer, not to
    assert what the answer is. `gte 1` conflates the two and calls an honest zero a
    failure, which is wrong for the 14 GBM graph variants whose ground truth is
    zero. `nonempty` is worse: it is `bool(actual)`, so it rejects 0 as well.
    """
    return (debug.get("graph_result") or {}).get("count") is not None


def _api_outcome_observed(debug: dict) -> bool:
    """True when the REST turn reported a row count, or answered from a recall.

    A follow-up that resolves against a previously recalled bundle carries
    `source_mode` and no `row_count` of its own, and that is correct rather than a
    gap. The floor lands on the LAST turn, so without this branch such a follow-up
    would be a guaranteed false failure.

    CORRECTION: this docstring used to name `retrieve.then_inspect` as "the only
    multi-turn variant in any floored family". That variant has since been
    RETIRED; `tree.then_ask_about` is now the only one
    (`test_it_is_still_the_only_multi_turn_variant_in_a_floored_family`). The
    branch itself is unchanged and still needed — only the example was stale.
    """
    meta = debug.get("api_result_meta") or {}
    return meta.get("row_count") is not None or meta.get("source_mode") is not None


def _report_produced_output(debug: dict) -> bool:
    """True when a reporting turn produced either of its two possible result shapes.

    `orchestrator.py` splits the reporting family in two: `reporter_mode ==
    "report_generation"` goes to `generate_report_outputs`, whose evidence is
    `report_saved_files`; everything else goes to `run_reporter_summary`, whose
    evidence is `reporter_result.ok`. The floor asserted only the second, so all 24
    SRA/GEO/PRIDE variants carried a criterion their own code path never sets.

    Asserting the disjunction is strictly better than tagging those 24 variants: it
    needs no per-case bookkeeping, and it stays meaningful either way. If generation
    does write files the cases pass; if it silently writes nothing they still fail,
    but now for a true reason rather than an unsatisfiable one.
    """
    result = debug.get("reporter_result")
    if isinstance(result, dict) and result.get("ok") is True:
        return True
    return bool(debug.get("report_saved_files"))


def _graph_truncation_disclosed(debug: dict) -> bool:
    """True when the graph result is either complete or HONEST about being capped.

    `graph_not_truncated` asserts a cap never happened. That is the wrong invariant
    for a floor: `graph.tissue_cell_impact` has 10,688 legitimate rows against a
    5,000 limit, so it can never satisfy it no matter how correct the answer is.
    Meanwhile the failure that actually burned us — the 2026-07-27 run reporting
    exactly 5,000 as if it were the total — is a DISCLOSURE failure, not a
    truncation one.

    So: a capped result passes only if it also reports a real total exceeding the
    rows it returned. That is precisely what the total probe added this wave makes
    checkable, and it turns the floor into something a correct answer can satisfy.

    True for non-graph turns, so the criterion stays inert on REST families.
    """
    graph = debug.get("graph_result") or {}
    if not graph:
        return True
    count, total = graph.get("count"), graph.get("total")
    if "truncated" in graph:
        if not graph.get("truncated"):
            return True
        return isinstance(total, int) and isinstance(count, int) and total > count
    # Results produced before the total probe existed carry no `truncated` flag, so
    # a count landing exactly on a known LIMIT is the only remaining signal, and an
    # undisclosed one by definition.
    return not (isinstance(count, int) and count in GRAPH_LIMIT_SENTINELS)


def _graph_not_truncated(debug: dict) -> bool:
    """True when the graph result is not a capped sample.

    Prefers the real `truncated` flag, which the Neo4j tool now sets by comparing the
    row count against the query's own trailing LIMIT. The sentinel comparison below is
    only a fallback for results produced before that flag existed — it can never be
    more than a guess, and it went stale the moment the limit moved from 250 to 5000.

    True for non-graph turns too, so the criterion is only meaningful where a graph
    query actually ran.
    """
    graph = debug.get("graph_result") or {}
    if "truncated" in graph:
        return not bool(graph.get("truncated"))
    count = graph.get("count")
    return not (isinstance(count, int) and count in GRAPH_LIMIT_SENTINELS)


def _criterion_parts(crit) -> tuple[str | None, str | None, Any]:
    if isinstance(crit, dict):
        return crit.get("field"), crit.get("op"), crit.get("value")
    return getattr(crit, "field", None), getattr(crit, "op", None), getattr(crit, "value", None)


def observe_values(debug: dict, criteria: list, *, last_reply: str | None = None) -> dict[str, Any]:
    """Resolve every criterion field to its OBSERVED value.

    check_pass reports only pass/fail plus a prose reason, so a manifest built
    from it cannot answer "what did it actually return?". Resolving alongside it
    keeps the run self-triaging without touching the vendored e2e DSL.
    """
    index = debug.get(ARTIFACT_INDEX_KEY) or {}
    observed: dict[str, Any] = {}
    for crit in criteria:
        field, _op, _value = _criterion_parts(crit)
        if not field:
            continue
        try:
            if field.startswith(ARTIFACT_PREFIX):
                observed[field] = resolve_artifact(index, field)
            else:
                observed[field] = resolve_field(debug, field, last_reply=last_reply)
        except Exception as exc:  # a resolver blowing up must not fail the run
            observed[field] = f"<unresolved: {type(exc).__name__}>"
    return observed


def _split_local_criteria(criteria: list) -> tuple[list, list]:
    """Partition into (locally evaluated criteria, delegated criteria)."""
    local, rest = [], []
    for crit in criteria:
        field, _op, _value = _criterion_parts(crit)
        (local if (field or "").startswith(ARTIFACT_PREFIX) else rest).append(crit)
    return local, rest


# Field families this harness cannot observe over HTTP. `evaluate_turn` calls
# `check_pass` with no `session=`, `browser_ctx=` or `mysql_chat_log=`, so every one
# of these resolves to None and fails unconditionally — `pipeline.reject_non_directive`
# asserts `eq False` and fails only because `None != False`. 27 such criteria span 12
# variants, nearly all in the expensive pipeline_nfcore family, so most seeds draw one
# or two guaranteed red herrings.
#
# Skipping is strictly better than failing: a criterion that cannot be evaluated is not
# evidence either way, and letting it fail buries real failures in noise.
#
# These are recoverable, not impossible. bundle.py already imports ChatSession, so a
# `session_reader(session_id)` threaded into check_pass(session=...) would make all 27
# meaningful. Until then they are recorded as skipped rather than silently failing.
_UNOBSERVABLE_FIELD_PREFIXES = ("pipeline_agent.", "chat_log.", "ui_text.")
_UNOBSERVABLE_OPS = ("trio_match",)
UNOBSERVABLE_REASON = "field family not observable over HTTP"

# The four DERIVED outcome fields `augment_debug` computes, all of them off keys
# that live on `query_complete.debug`. A Container-CC `query_complete` carries no
# `debug` key AT ALL — `cc_engine` emits {reply, mode, artifacts, cc_raw_files} —
# so `graph_result`, `api_result_meta`, `reporter_result` and `report_saved_files`
# are absent and every one of these four resolves False by construction,
# regardless of what the CC turn actually did.
#
# Part 1 of this fix made the family floor engine-AGNOSTIC by flooring on
# `outcome_observed` instead of `api_ok`/`neo4j_ok`. Necessary but not sufficient:
# `outcome_observed` is exactly the disjunction of the other three, so it is
# constant-false on a CC turn too, and every CC-routed case in a floored family
# stayed an automatic red that proved nothing. `green.refine_recall` failed for
# precisely this reason in the 2026-08-03 seed-6 run.
#
# The scope is deliberately NARROW: only these four derived fields, and only when
# the OBSERVED route is container_cc. `api_ok`, `neo4j_ok`, `parser_plan.*`,
# `api_plan.*` and `graph_result.*` are inline, case-level assertions someone
# wrote by hand — a case carrying them is claiming a particular engine answered
# it. A CC-routed case carrying an inline `api_ok` therefore STILL FAILS, and
# that is intended: widening the skip to cover it is a corpus decision with a
# much bigger blast radius, not a harness one.
CC_ROUTE = "container_cc"
CC_UNOBSERVABLE_FIELDS = frozenset({
    "api_outcome_observed", "graph_outcome_observed",
    "report_produced_output", "outcome_observed",
})
# Names the route, so a manifest reader can tell this skip from the HTTP one
# above. Two skips with the same reason string would be untriageable.
CC_UNOBSERVABLE_REASON = f"NS outcome field not observable on a {CC_ROUTE} turn"

# What the runner records when a case evaluated ZERO criteria. Lives here, next
# to the two reasons that cause it, so the three read as one story.
NO_ASSERTIONS_REASON = (
    "every criterion this case evaluated was skipped as unobservable, so the "
    "case asserted nothing about the product"
)


def unobservable_reason(field: str | None, op: str | None,
                        route: str | None = None) -> str | None:
    """Why this criterion cannot be evaluated on this turn, or None if it can.

    ``route`` is optional and defaults to None so the HTTP-family check keeps
    working for callers that have no route in hand; it is only consulted for the
    container_cc case.
    """
    if (field or "").startswith(_UNOBSERVABLE_FIELD_PREFIXES) or op in _UNOBSERVABLE_OPS:
        return UNOBSERVABLE_REASON
    if route == CC_ROUTE and field in CC_UNOBSERVABLE_FIELDS:
        return CC_UNOBSERVABLE_REASON
    return None


def is_unobservable(field: str | None, op: str | None, route: str | None = None) -> bool:
    return unobservable_reason(field, op, route) is not None


def any_criterion_evaluated(results: list[dict]) -> bool:
    """Did this turn actually TEST anything?

    False both when every result was skipped and when there were no criteria at
    all — the two are the same claim about the turn, and the runner turns a case
    made entirely of such turns into ``no_assertions`` rather than ``passed``.
    """
    return any(not r.get("skipped") for r in results)


def evaluate_turn(payload, criteria, obs, *, last_reply=None, bundle_summary=None):
    raw_debug = build_observed_debug(payload)
    debug = augment_debug(raw_debug, obs, bundle_summary,
                          artifact_index=build_artifact_index(raw_debug, payload))

    # The route is read off `debug` rather than plumbed through from the runner:
    # `augment_debug` has just set it from the SAME observation the criteria are
    # about to be evaluated against, so there is no second source to drift from.
    route = debug.get("route")
    paired = [(c, unobservable_reason(*_criterion_parts(c)[:2], route=route)) for c in criteria]
    skipped = [(c, why) for c, why in paired if why]
    criteria = [c for c, why in paired if not why]

    local_criteria, delegated = _split_local_criteria(criteria)
    passed, results = check_pass(debug, delegated, last_reply=last_reply)

    for crit, why in skipped:
        field, op, expected = _criterion_parts(crit)
        results.append({"field": field, "op": op, "value": expected,
                        "passed": True, "skipped": True,
                        "reason": f"{field}: SKIPPED — {why}"})

    index = debug.get(ARTIFACT_INDEX_KEY) or {}
    for crit in local_criteria:
        field, op, expected = _criterion_parts(crit)
        is_artifact = field.startswith(ARTIFACT_PREFIX)
        actual = (resolve_artifact(index, field) if is_artifact
                  else resolve_field(debug, field, last_reply=last_reply))
        if op == "neq":
            ok = actual != expected
            reason = (f"is {actual!r} (must not be {expected!r})" if not ok
                      else f"{actual!r} != {expected!r}")
        elif op == "true":
            ok = actual is True
            reason = "produced" if ok else f"not produced (have: {sorted(index) or 'nothing'})"
        elif op == "gte":
            try:
                ok = int(actual) >= int(expected)
            except (TypeError, ValueError):
                ok = False
            reason = f"{actual} (needed >= {expected})"
        else:
            ok, reason = False, f"unsupported op {op!r} for a locally-evaluated criterion"
        results.append({"field": field, "op": op, "value": expected,
                        "passed": ok, "reason": f"{field}: {reason}"})
        passed = passed and ok

    return passed, results, observe_values(debug, criteria, last_reply=last_reply)
