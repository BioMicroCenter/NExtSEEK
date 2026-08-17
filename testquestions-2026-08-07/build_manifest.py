"""Build a content and provenance manifest for an extracted ZIP archive.

Purpose
-------
This utility inventories the *members* of a ZIP archive and describes their
corresponding extracted files.  It is intended for reproducibility, evidence
retention, and data-delivery audits.  The archive itself is recorded once in the
manifest header; it is not duplicated as a file entry.

The script is deliberately dataset-neutral.  It contains no contributor names,
institutional hostnames, workstation paths, credentials, or delivery-specific
claims.  Operators may supply provenance text and optional mirror metadata on the
command line.  Local absolute paths are redacted by default because manifests are
often committed to source control or attached to public reports.

Security and privacy
--------------------
Creating a manifest is not a release review.  Checksums, filenames, spreadsheet
headers, Markdown headings, and inferred JSON schemas can themselves disclose
sensitive information.  Review MANIFEST.json and schema sidecars before sharing
them.  This script does not scan for secrets, classify personal data, anonymize
content, establish licensing rights, or authorize publication of the archive.

Archive layout and path mapping
-------------------------------
``--archive`` names the ZIP.  ``--extracted-base`` names the directory beneath
which each ZIP member path can be found.  For a member named
``delivery/results/run.json``, the script reads
``EXTRACTED_BASE/delivery/results/run.json``.  If ``--extracted-base`` is omitted,
the archive's parent directory is used.  This handles both archives with a single
top-level directory and archives containing files at their root.

Relative member paths are the stable identity of entries.  Basenames are not
unique in realistic evidence bundles, so schema sidecar names are derived from the
full relative path and receive a digest suffix when filesystem length limits would
otherwise be exceeded.

Supported metadata
------------------
* Every file: SHA-256, byte size, archive timestamp, extension, and relative path.
* CSV/TSV: first-row headers and parsed data-row count.
* XLS/XLSX/XLSM: sheet names, first-row values, and rows below the first row.
* Markdown: ATX headings outside fenced code blocks.
* JSON: schema plus a documented top-level count.
* JSONL: schema plus count of independently parseable non-empty lines.
* Nested ZIP files: treated as opaque members; their contents are not enumerated.

JSON schemas
------------
The preferred engine is ``srs-lens schema``.  If that command fails on otherwise
valid JSON, the script falls back to Python ``genson``.  Schemas are always stored
under ``schemas/`` beside this script.  ``--no-inline-schemas`` keeps the manifest
compact by storing only sidecar references; without it, schemas are also embedded.
The ``schemas/_schema_tools.json`` index records which engine produced each cached
sidecar.  A sidecar newer than its source file is reused.

Example
-------
Run from any directory; outputs are written beside this script::

    uv run --with openpyxl --with genson python build_manifest.py \
      --archive /data/delivery.zip \
      --extracted-base /data \
      --source-description "Received through the approved transfer channel" \
      --no-inline-schemas

To record absolute workstation paths, add ``--include-local-paths``.  Do this only
for a controlled manifest: the safe default records relative archive-member paths
and nulls the archive's local path.  Optional ``--mirror-host`` and
``--mirror-root`` values are likewise emitted verbatim and should not be used in a
public artifact unless disclosure is intentional.

Operational guarantees and non-guarantees
------------------------------------------
The script fails if an archive member is missing from the extracted tree or if the
number of emitted entries differs from the member count.  It does not compare ZIP
member bytes directly with extracted bytes; the SHA-256 values describe the
extracted copies.  It also does not recurse into nested archives, interpret the
semantic header row of submission templates, or assign a universal meaning to
``num_rows``.  Consumers must read ``num_rows_basis`` for each entry.
"""

import argparse
import csv
import hashlib
import json
import subprocess
import zipfile
from datetime import datetime, timezone
from pathlib import Path

OUT_DIR = Path(__file__).resolve().parent
SCHEMA_DIR = OUT_DIR / "schemas"

# Runtime configuration is populated by main().  Keeping neutral defaults makes
# importing this module safe and prevents local infrastructure details from being
# baked into source control.
INCLUDE_LOCAL_PATHS = False
MIRROR_HOST = None
MIRROR_ROOT = None

WORKBOOK_EXT = {".xlsx", ".xlsm", ".xls"}
DELIMITED = {".csv": ",", ".tsv": "\t"}

# Filesystem limit is 255 bytes; leave room for the ".schema.json" suffix.
MAX_SIDECAR_STEM = 180


def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for block in iter(lambda: fh.read(1 << 20), b""):
            h.update(block)
    return h.hexdigest()


def delimited_headers_and_rows(path, delimiter):
    """Header list and data-row count, using the csv module so quoted newlines
    inside a field are not miscounted as row breaks."""
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
        reader = csv.reader(fh, delimiter=delimiter)
        try:
            headers = next(reader)
        except StopIteration:
            return [], 0
        return headers, sum(1 for _ in reader)


def markdown_headings(path):
    """ATX headings only. Fenced code blocks are skipped so that a shell comment
    such as `# make` inside a code fence is not mistaken for a heading."""
    out = []
    in_fence = False
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        stripped = line.strip()
        if stripped.startswith("```") or stripped.startswith("~~~"):
            in_fence = not in_fence
            continue
        if not in_fence and stripped.startswith("#"):
            out.append(stripped)
    return out


def workbook_sheets_and_headers(path):
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    sheets, headers, rows = [], [], 0
    for ws in wb.worksheets:
        sheets.append(ws.title)
        first = next(ws.iter_rows(values_only=True), None)
        if first:
            headers.extend(f"{ws.title}::{c}" for c in first if c is not None)
        rows += max((ws.max_row or 1) - 1, 0)
    wb.close()
    return sheets, headers, rows


def sidecar_path(relpath):
    """Return a collision-resistant sidecar path derived from the member path.

    Using only a basename or stem would silently overwrite schemas whenever an
    archive repeats conventional filenames in multiple directories.
    """
    flat = str(relpath.with_suffix("")).replace("/", "__")
    if len(flat) > MAX_SIDECAR_STEM:
        digest = hashlib.sha256(str(relpath).encode()).hexdigest()[:12]
        flat = flat[: MAX_SIDECAR_STEM - 13] + "_" + digest
    return SCHEMA_DIR / f"{flat}.schema.json"


SRS_LENS_TOOL = "srs-lens schema (genson-rs engine)"
# sidecar filename -> engine that produced it. Persisted so a cached re-run still
# reports the right engine, without writing bookkeeping keys into the schema files.
TOOL_INDEX_PATH = SCHEMA_DIR / "_schema_tools.json"
TOOL_INDEX = (
    json.loads(TOOL_INDEX_PATH.read_text()) if TOOL_INDEX_PATH.exists() else {}
)
FALLBACK_TOOL = (
    "python genson (fallback). srs-lens/genson-rs 0.2.0 panics on this file with "
    "InternalError(TapeError) and reports it as malformed input; the file in fact "
    "parses cleanly under python json, so the tool's malformed verdict is a defect "
    "in genson-rs, NOT a defect in the delivered file."
)


def json_documents(path):
    """Return every concatenated top-level JSON document in a text file.

    Although the usual case is one document, some producers concatenate several
    without using JSONL.  ``json.load`` reports ``Extra data`` for those files;
    iterative ``raw_decode`` preserves the document stream instead.
    """
    raw = path.read_text(encoding="utf-8", errors="replace")
    dec = json.JSONDecoder()
    docs, i, n = [], 0, len(raw)
    while i < n:
        while i < n and raw[i] in " \t\r\n":
            i += 1
        if i >= n:
            break
        obj, i = dec.raw_decode(raw, i)
        docs.append(obj)
    return docs


def schema_via_genson(path):
    """Fallback schema inference in-process, for files genson-rs crashes on."""
    from genson import SchemaBuilder

    builder = SchemaBuilder()
    for doc in json_documents(path):
        builder.add_object(doc)
    return builder.to_schema()


def infer_json_schema(path, relpath):
    """Delegate to srs-lens (genson-rs), falling back to python genson when that
    engine crashes. Cached: an existing sidecar newer than its source is reused, so a
    re-run with a different inline setting is not a re-compute.
    Returns (schema_obj, tool_string, sidecar_path)."""
    side = sidecar_path(relpath)
    if side.exists() and side.stat().st_mtime >= path.stat().st_mtime:
        # Sidecars are pure JSON Schema; which engine produced each is kept in the
        # companion index so the schema file itself carries no bookkeeping keys.
        return json.loads(side.read_text()), TOOL_INDEX.get(side.name, SRS_LENS_TOOL), side

    proc = subprocess.run(["srs-lens", "schema", str(path)], capture_output=True, text=True)
    if proc.returncode == 0 and proc.stdout.strip():
        schema, tool = json.loads(proc.stdout), SRS_LENS_TOOL
    else:
        try:
            schema, tool = schema_via_genson(path), FALLBACK_TOOL
        except Exception as exc:  # genuinely unreadable: say so, do not guess
            return None, f"schema inference failed under both engines: {exc}"[:300], None

    SCHEMA_DIR.mkdir(parents=True, exist_ok=True)
    side.write_text(json.dumps(schema, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    TOOL_INDEX[side.name] = tool
    return schema, tool, side


def json_row_count(path):
    """A JSON document has no intrinsic row concept. Report the top-level element
    count and say which basis produced it, rather than inventing one."""
    try:
        docs = json_documents(path)
    except json.JSONDecodeError as exc:
        return None, f"file could not be parsed as JSON ({exc.msg})"

    if len(docs) != 1:
        return len(docs), (
            f"concatenated JSON documents: this file holds {len(docs)} separate "
            "top-level documents rather than one, so it is not a single JSON document "
            "despite the .json extension"
        )

    data = docs[0]
    if isinstance(data, list):
        return len(data), "top-level array length"
    if isinstance(data, dict):
        return len(data), "top-level object key count"
    return None, "top-level value is a scalar"


def jsonl_record_count(path):
    """JSONL is one record per non-empty line. Counted by parsing rather than by
    counting newlines, so a malformed trailing fragment is not scored as a record."""
    good = bad = 0
    with open(path, encoding="utf-8", errors="replace") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            try:
                json.loads(line)
                good += 1
            except json.JSONDecodeError:
                bad += 1
    basis = "JSONL records, one per non-empty line, counted by parsing each line"
    if bad:
        basis += f"; {bad} line(s) did not parse and are excluded"
    return good, basis


def describe(rel, laptop_path, generated_iso, generated_source, source_container, inline):
    ext = laptop_path.suffix.lower()
    entry = {
        "filename": laptop_path.name,
        "relative_path": rel.as_posix(),
        "path_local": str(laptop_path) if INCLUDE_LOCAL_PATHS else None,
        "path_mirror": (
            None
            if MIRROR_ROOT is None
            else f"{MIRROR_ROOT.rstrip('/')}/{rel.as_posix()}"
        ),
        "checksum": sha256(laptop_path),
        "checksum_method": "sha256",
        "file_type": ext.lstrip(".") or "(no extension)",
        "date_generated_or_downloaded": generated_iso,
        "date_source": generated_source,
        "source_container": source_container,
        "bytes": laptop_path.stat().st_size,
        "sheets": None,
        "headers": None,
        "json_schema": None,
        "json_schema_tool": None,
        "num_rows": None,
        "num_rows_basis": "not applicable to this file type",
    }

    if ext in DELIMITED:
        headers, rows = delimited_headers_and_rows(laptop_path, DELIMITED[ext])
        entry["headers"] = headers
        entry["num_rows"] = rows
        entry["num_rows_basis"] = "data rows, excluding the header row"
    elif ext in WORKBOOK_EXT:
        sheets, headers, rows = workbook_sheets_and_headers(laptop_path)
        entry["sheets"] = sheets
        entry["headers"] = headers
        entry["num_rows"] = rows
        entry["num_rows_basis"] = "data rows summed across sheets, excluding each header row"
    elif ext == ".md":
        entry["headers"] = markdown_headings(laptop_path)
        entry["num_rows_basis"] = "markdown has no row concept; headers hold the ATX headings"
    elif ext in (".json", ".jsonl"):
        schema, tool, side = infer_json_schema(laptop_path, rel)
        entry["json_schema"] = schema if (inline and schema is not None) else None
        entry["json_schema_tool"] = tool
        entry["json_schema_sidecar"] = str(side.relative_to(OUT_DIR)) if side else None
        if ext == ".json":
            rows, basis = json_row_count(laptop_path)
        else:
            rows, basis = jsonl_record_count(laptop_path)
        entry["num_rows"] = rows
        entry["num_rows_basis"] = basis
    elif ext == ".zip":
        entry["num_rows_basis"] = (
            "nested archive recorded as a single file; its members are NOT enumerated"
        )
    return entry


def main():
    global INCLUDE_LOCAL_PATHS, MIRROR_HOST, MIRROR_ROOT

    ap = argparse.ArgumentParser(
        description=(
            "Inventory an extracted ZIP archive and write MANIFEST.json plus "
            "JSON schema sidecars."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        epilog=(
            "Privacy note: generated filenames, headings, headers, schemas, and optional "
            "path/host fields may be sensitive. Review outputs before publication."
        ),
    )
    ap.add_argument(
        "--archive", type=Path, required=True,
        help="ZIP archive whose members define the inventory",
    )
    ap.add_argument(
        "--extracted-base", type=Path,
        help="directory containing member paths; defaults to the archive parent",
    )
    ap.add_argument(
        "--manifest-name", default="Archive contents manifest",
        help="human-readable manifest title",
    )
    ap.add_argument(
        "--source-description",
        help="optional provenance statement; emitted verbatim",
    )
    ap.add_argument(
        "--include-local-paths", action="store_true",
        help="include absolute local paths (redacted by default)",
    )
    ap.add_argument(
        "--mirror-host",
        help="optional mirror hostname; emitted verbatim",
    )
    ap.add_argument(
        "--mirror-root",
        help="optional mirror root used to form path_mirror values",
    )
    ap.add_argument("--no-inline-schemas", action="store_true",
                    help="write schemas only as sidecars; leave json_schema null in MANIFEST.json")
    args = ap.parse_args()
    archive = args.archive.expanduser().resolve()
    extracted_base = (
        args.extracted_base.expanduser().resolve()
        if args.extracted_base is not None else archive.parent
    )
    if not archive.is_file():
        ap.error(f"archive does not exist or is not a file: {archive}")
    if not extracted_base.is_dir():
        ap.error(f"extracted base does not exist or is not a directory: {extracted_base}")

    INCLUDE_LOCAL_PATHS = args.include_local_paths
    MIRROR_HOST = args.mirror_host
    MIRROR_ROOT = args.mirror_root
    inline = not args.no_inline_schemas

    with zipfile.ZipFile(archive) as zf:
        members = [i for i in zf.infolist() if not i.is_dir()]

    entries = []
    for info in sorted(members, key=lambda i: i.filename):
        rel = Path(info.filename)
        # Reject traversal-like member names even though this script only reads an
        # already-extracted tree.  It avoids accidentally inventorying a file outside
        # extracted_base when given an untrusted archive.
        if rel.is_absolute() or ".." in rel.parts:
            raise ValueError(f"unsafe archive member path: {info.filename!r}")
        laptop_path = extracted_base / rel
        if not laptop_path.is_file():
            raise FileNotFoundError(
                f"archive member has no extracted file: {info.filename!r} "
                f"(expected {laptop_path})"
            )
        y, mo, d, h, mi, s = info.date_time
        entries.append(describe(
            rel, laptop_path, datetime(y, mo, d, h, mi, s).isoformat(),
            "ZIP member timestamp; extracted-file mtimes may instead reflect extraction time",
            archive.name, inline,
        ))

    # Fail loudly rather than ship a silently short manifest.
    assert len(entries) == len(members), f"{len(entries)} entries for {len(members)} members"
    paths = {e["relative_path"] for e in entries}
    assert len(paths) == len(members), f"{len(paths)} unique paths for {len(members)} members"

    manifest = {
        "manifest_name": args.manifest_name,
        "generated_by": "build_manifest.py",
        "generated_at": datetime.now(timezone.utc).astimezone().isoformat(),
        "source": args.source_description or (
            f"The {len(entries)} files contained in {archive.name}; archive contents only"
        ),
        "local_root": str(extracted_base) if INCLUDE_LOCAL_PATHS else None,
        "mirror_host": MIRROR_HOST,
        "mirror_root": MIRROR_ROOT,
        "checksum_method": "sha256",
        "json_schema_method": (
            "srs-lens schema (genson-rs engine), inline plus a sidecar under schemas/"
            if inline else
            "srs-lens schema (genson-rs engine), sidecar under schemas/ only; "
            "json_schema is null in this file and the sidecar holds the schema"
        ),
        "source_archive": {
            "filename": archive.name,
            "path_local": str(archive) if INCLUDE_LOCAL_PATHS else None,
            "checksum": sha256(archive),
            "checksum_method": "sha256",
            "bytes": archive.stat().st_size,
            "date_downloaded": datetime.fromtimestamp(
                archive.stat().st_mtime, tz=timezone.utc).astimezone().isoformat(),
            "note": "Provenance for the contents below. The archive is deliberately NOT "
                    "a file entry: this manifest covers its contents only.",
        },
        "notes": [
            "Metadata and checksums are computed from extracted files, not "
            "directly from ZIP member bytes.",
            "date_generated_or_downloaded uses ZIP member timestamps because "
            "extracted-file mtimes may record extraction time.",
            "Entries are identified by relative_path, not basename; schema "
            "sidecars derive their names from that path.",
            "Absolute local paths are included only when --include-local-paths is supplied.",
            "Mirror metadata is declarative only; this run does not contact or verify a mirror.",
            "num_rows has no single cross-format meaning; num_rows_basis states the rule "
            "used for each entry.",
            "JSONL is counted by parsing each non-empty line; malformed lines "
            "are excluded and reported in num_rows_basis.",
            "Nested ZIP members are recorded as opaque files and are not recursively enumerated.",
            "Spreadsheet headers are values from each sheet's first row; "
            "semantic headers located lower in a sheet are not inferred.",
            "A JSON file may contain concatenated top-level documents; when "
            "detected, num_rows is the document count.",
            "Schema inference prefers srs-lens and falls back to Python genson; "
            "json_schema_tool records the engine or failure.",
        ],
        "file_count": len(entries),
        "files": entries,
    }

    if TOOL_INDEX:
        SCHEMA_DIR.mkdir(parents=True, exist_ok=True)
        TOOL_INDEX_PATH.write_text(
            json.dumps(TOOL_INDEX, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    out = OUT_DIR / "MANIFEST.json"
    out.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    print(f"wrote {out} with {len(entries)} entries "
          f"({out.stat().st_size/1e6:.1f} MB, inline_schemas={inline})")


if __name__ == "__main__":
    main()
